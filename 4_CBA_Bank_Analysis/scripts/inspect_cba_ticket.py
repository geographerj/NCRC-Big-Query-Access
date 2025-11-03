"""Inspect the CBA ticket file structure"""

import openpyxl
from pathlib import Path

ticket_file = Path(r"C:\Users\edite\OneDrive - Nat'l Community Reinvestment Coaltn\Desktop\DREAM Analysis\PNC+FirstBank Research Ticket_CBA.xlsx")

wb = openpyxl.load_workbook(ticket_file, data_only=True)

print("="*80)
print("INSPECTING CBA TICKET FILE")
print("="*80)

# Check Bank Details sheet
if "Bank Details" in wb.sheetnames:
    print("\n[Bank Details Sheet]")
    ws = wb["Bank Details"]
    print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
    
    print("\nFirst 20 rows:")
    for row_idx in range(1, min(21, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(8, ws.max_column + 1)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                row_data.append(str(cell_value)[:60])
        if row_data:
            print(f"  Row {row_idx}: {row_data}")

# Check LMA Ticket Filters sheet
if "LMA Ticket Filters" in wb.sheetnames:
    print("\n[LMA Ticket Filters Sheet]")
    ws = wb["LMA Ticket Filters"]
    print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
    
    print("\nAll rows:")
    for row_idx in range(1, ws.max_row + 1):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                row_data.append(str(cell_value).strip())
        if row_data:
            print(f"  Row {row_idx}: {row_data}")

