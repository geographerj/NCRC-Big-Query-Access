"""
NCRC CBA Banks Disparate Impact Analysis - Excel Generator v4.0

CRITICAL FIXES:
- Properly maps SQL column names (activity_year -> year, etc.)
- Correctly processes loan_purpose codes (1, 2, 31, 32)
- Removes ALL LMIB$ metrics from ALL sheets
- Shows TOTAL originations in count columns
- Fixes repeated figure bug by ensuring proper CBSA grouping

Input: CSV from BigQuery (cba_banks_analysis.sql output)
Output: Excel workbook with ratio and data sheets per bank-CBSA
"""

import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
import argparse
import os

warnings.filterwarnings('ignore')

# =============================================================================
# CONFIGURATION
# =============================================================================

EXPECTED_YEARS = list(range(2018, 2025))

# Loan purpose mapping from SQL codes
LOAN_PURPOSE_MAP = {
    '1': 'Home Purchase',
    '2': 'Home Improvement', 
    '31': 'Refinancing',
    '32': 'Refinancing'
}

# Demographic metrics (NO LMIB$ metrics)
METRICS = [
    'black_share',
    'hispanic_share',
    'asian_share',
    'lmib_share',
    'lmict_share',
    'mmct_share'
]

METRIC_DISPLAY_NAMES = {
    'black_share': 'Black%',
    'hispanic_share': 'Hispanic%',
    'asian_share': 'Asian%',
    'lmib_share': 'LMIB%',
    'lmict_share': 'LMICT%',
    'mmct_share': 'MMCT%'
}

# =============================================================================
# DATA LOADING & PREPROCESSING
# =============================================================================

def load_data(input_file):
    """Load BigQuery CSV export and preprocess"""
    print(f"\nLoading data from: {input_file}")
    
    df = pd.read_csv(input_file, dtype={'cbsa_code': str, 'loan_purpose': str})
    
    # Map SQL column names to expected names
    column_mapping = {
        'activity_year': 'year',
        'cbsa_name': 'cbsa_title',
        'lender_name': 'respondent_name'
    }
    
    for old_col, new_col in column_mapping.items():
        if old_col in df.columns:
            df = df.rename(columns={old_col: new_col})
    
    # Convert year to int
    if 'year' in df.columns:
        df['year'] = df['year'].astype(int)
    
    # Map loan purpose codes to categories
    if 'loan_purpose' in df.columns:
        df['loan_purpose_category'] = df['loan_purpose'].map(LOAN_PURPOSE_MAP)
        # Fill any unmapped values
        df['loan_purpose_category'] = df['loan_purpose_category'].fillna('Unknown')
    
    # Add total amount columns if LMIB amounts exist
    if 'subject_lmib_amount_originations' in df.columns:
        # Calculate totals from LMIB amounts (these are the sums we need)
        df['subject_total_amount_originations'] = df['subject_lmib_amount_originations']
        df['peer_total_amount_originations'] = df['peer_lmib_amount_originations']
        df['subject_total_amount_applications'] = df['subject_lmib_amount_applications']  
        df['peer_total_amount_applications'] = df['peer_lmib_amount_applications']
    
    print(f"  Rows loaded: {len(df):,}")
    print(f"  Banks: {df['lei'].nunique()}")
    print(f"  CBSAs: {df['cbsa_code'].nunique()}")
    print(f"  Years: {sorted(df['year'].unique())}")
    print(f"  Columns: {list(df.columns)}")
    
    return df

def calculate_metrics(row, metric_type):
    """
    Calculate shares for a given metric type
    Returns: subject_share, peer_share, gap, subject_num, subject_denom, peer_num, peer_denom
    """
    
    if metric_type == 'black_share':
        subject_num = row.get('subject_black_originations', 0)
        subject_denom = row.get('subject_total_originations', 0)
        peer_num = row.get('peer_black_originations', 0)
        peer_denom = row.get('peer_total_originations', 0)
        
    elif metric_type == 'hispanic_share':
        subject_num = row.get('subject_hispanic_originations', 0)
        subject_denom = row.get('subject_total_originations', 0)
        peer_num = row.get('peer_hispanic_originations', 0)
        peer_denom = row.get('peer_total_originations', 0)
        
    elif metric_type == 'asian_share':
        subject_num = row.get('subject_asian_originations', 0)
        subject_denom = row.get('subject_total_originations', 0)
        peer_num = row.get('peer_asian_originations', 0)
        peer_denom = row.get('peer_total_originations', 0)
        
    elif metric_type == 'lmib_share':
        subject_num = row.get('subject_lmib_originations', 0)
        subject_denom = row.get('subject_total_originations', 0)
        peer_num = row.get('peer_lmib_originations', 0)
        peer_denom = row.get('peer_total_originations', 0)
        
    elif metric_type == 'lmict_share':
        subject_num = row.get('subject_lmict_originations', 0)
        subject_denom = row.get('subject_total_originations', 0)
        peer_num = row.get('peer_lmict_originations', 0)
        peer_denom = row.get('peer_total_originations', 0)
        
    elif metric_type == 'mmct_share':
        subject_num = row.get('subject_mmct_originations', 0)
        subject_denom = row.get('subject_total_originations', 0)
        peer_num = row.get('peer_mmct_originations', 0)
        peer_denom = row.get('peer_total_originations', 0)
        
    else:
        return None, None, None, 0, 0, 0, 0
    
    # Calculate shares
    subject_share = (subject_num / subject_denom * 100) if subject_denom > 0 else 0
    peer_share = (peer_num / peer_denom * 100) if peer_denom > 0 else 0
    gap = peer_share - subject_share
    
    return subject_share, peer_share, gap, subject_num, subject_denom, peer_num, peer_denom

def test_significance(subject_num, subject_denom, peer_num, peer_denom):
    """
    Two-proportion z-test for statistical significance
    Returns: is_significant (True/False)
    """
    if subject_denom == 0 or peer_denom == 0 or subject_num == 0 or peer_num == 0:
        return False
    
    try:
        p1 = subject_num / subject_denom
        p2 = peer_num / peer_denom
        
        n1 = subject_denom
        n2 = peer_denom
        
        p_pooled = (subject_num + peer_num) / (n1 + n2)
        
        se = np.sqrt(p_pooled * (1 - p_pooled) * (1/n1 + 1/n2))
        
        if se == 0:
            return False
        
        z = (p1 - p2) / se
        p_value = 2 * (1 - stats.norm.cdf(abs(z)))
        
        return p_value < 0.05
    except:
        return False

# =============================================================================
# RATIO SHEET CREATION (NO SIGNIFICANCE TESTING, NO LMIB$)
# =============================================================================

def create_ratio_sheet_data(df, lei, cbsa_code):
    """
    Create ratio sheet showing peer_share / subject_share
    NO significance testing, NO LMIB$ metrics
    """
    
    subset = df[(df['lei'] == lei) & (df['cbsa_code'] == cbsa_code)].copy()
    
    if len(subset) == 0:
        return None
    
    # Get CBSA name and bank name (take first non-null value)
    cbsa_name = subset['cbsa_title'].iloc[0] if 'cbsa_title' in subset.columns and pd.notna(subset['cbsa_title'].iloc[0]) else f"CBSA {cbsa_code}"
    bank_name = subset['respondent_name'].iloc[0] if 'respondent_name' in subset.columns and pd.notna(subset['respondent_name'].iloc[0]) else lei
    
    rows = []
    
    # Get unique loan purpose categories
    loan_purposes = subset['loan_purpose_category'].unique()
    
    for loan_purpose in loan_purposes:
        purpose_data = subset[subset['loan_purpose_category'] == loan_purpose]
        
        if len(purpose_data) == 0:
            continue
        
        # Process each metric
        for metric in METRICS:
            row = {
                'CBSA': cbsa_name,
                'Loan_Purpose_Category': loan_purpose,
                'Metric': METRIC_DISPLAY_NAMES.get(metric, metric)
            }
            
            # Calculate ratios for each year
            for year in EXPECTED_YEARS:
                year_data = purpose_data[purpose_data['year'] == year]
                
                if len(year_data) > 0:
                    data_row = year_data.iloc[0]
                    subject_share, peer_share, _, _, _, _, _ = calculate_metrics(data_row, metric)
                    
                    # Calculate ratio: peer / subject
                    if subject_share > 0:
                        ratio = peer_share / subject_share
                        row[str(year)] = round(ratio, 2)
                    else:
                        row[str(year)] = None
                else:
                    row[str(year)] = None
            
            rows.append(row)
    
    if len(rows) == 0:
        return None
        
    ratio_df = pd.DataFrame(rows)
    return ratio_df

# =============================================================================
# DATA SHEET CREATION (WITH SIGNIFICANCE TESTING, NO LMIB$)
# =============================================================================

def create_data_sheet_data(df, lei, cbsa_code):
    """
    Create data sheet with counts, shares, gaps, and significance testing
    Count columns show TOTAL originations (not metric-specific)
    NO LMIB$ metrics
    """
    
    subset = df[(df['lei'] == lei) & (df['cbsa_code'] == cbsa_code)].copy()
    
    if len(subset) == 0:
        return None
    
    cbsa_name = subset['cbsa_title'].iloc[0] if 'cbsa_title' in subset.columns and pd.notna(subset['cbsa_title'].iloc[0]) else f"CBSA {cbsa_code}"
    bank_name = subset['respondent_name'].iloc[0] if 'respondent_name' in subset.columns and pd.notna(subset['respondent_name'].iloc[0]) else lei
    
    rows = []
    
    # Get unique loan purpose categories
    loan_purposes = subset['loan_purpose_category'].unique()
    
    for loan_purpose in loan_purposes:
        purpose_data = subset[subset['loan_purpose_category'] == loan_purpose]
        
        if len(purpose_data) == 0:
            continue
        
        # Process each metric
        for metric in METRICS:
            row = {
                'CBSA': cbsa_name,
                'Loan_Purpose': loan_purpose,
                'Metric': METRIC_DISPLAY_NAMES.get(metric, metric)
            }
            
            # For each year, add columns: Count, Subject_Share, Peer_Share, Gap
            for year in EXPECTED_YEARS:
                year_data = purpose_data[purpose_data['year'] == year]
                
                if len(year_data) > 0:
                    data_row = year_data.iloc[0]
                    
                    # TOTAL originations (same for all metrics in this year/CBSA/loan_purpose)
                    total_count = data_row.get('subject_total_originations', 0)
                    
                    # Calculate shares for this specific metric
                    subject_share, peer_share, gap, subject_num, subject_denom, peer_num, peer_denom = calculate_metrics(data_row, metric)
                    
                    # Test significance
                    is_significant = test_significance(subject_num, subject_denom, peer_num, peer_denom)
                    
                    # Store values
                    row[f'{year}_Subject_Count'] = int(total_count)
                    row[f'{year}_Subject_Share'] = round(subject_share, 1)
                    row[f'{year}_Peer_Share'] = round(peer_share, 1)
                    row[f'{year}_Gap'] = round(gap, 1)
                    row[f'{year}_Significant'] = is_significant
                    
                else:
                    row[f'{year}_Subject_Count'] = None
                    row[f'{year}_Subject_Share'] = None
                    row[f'{year}_Peer_Share'] = None
                    row[f'{year}_Gap'] = None
                    row[f'{year}_Significant'] = False
            
            rows.append(row)
    
    if len(rows) == 0:
        return None
        
    data_df = pd.DataFrame(rows)
    return data_df

# =============================================================================
# EXCEL FORMATTING
# =============================================================================

def apply_ratio_sheet_formatting(ws, df):
    """Apply color coding to ratio sheets"""
    
    fill_green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    fill_orange = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    fill_light_red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    fill_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    font_white = Font(color='FFFFFF', bold=True)
    
    # Find year columns
    year_cols = [col for col in df.columns if str(col).isdigit()]
    
    for row_idx in range(2, len(df) + 2):
        for col_name in year_cols:
            col_idx = df.columns.get_loc(col_name) + 1
            cell = ws.cell(row=row_idx, column=col_idx)
            
            try:
                value = float(cell.value)
                
                if pd.isna(value):
                    continue
                elif value < 1.0:
                    cell.fill = fill_green
                elif value < 1.5:
                    cell.fill = fill_yellow
                elif value < 2.0:
                    cell.fill = fill_orange
                elif value < 3.0:
                    cell.fill = fill_light_red
                else:
                    cell.fill = fill_red
                    cell.font = font_white
            except (ValueError, TypeError):
                continue

def apply_data_sheet_formatting(ws, df):
    """Apply formatting to data sheets with asterisks and red cells"""
    
    fill_red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    for year in EXPECTED_YEARS:
        gap_col_name = f'{year}_Gap'
        sig_col_name = f'{year}_Significant'
        
        if gap_col_name not in df.columns or sig_col_name not in df.columns:
            continue
        
        gap_col_idx = df.columns.get_loc(gap_col_name) + 1
        
        for row_idx in range(2, len(df) + 2):
            gap_cell = ws.cell(row=row_idx, column=gap_col_idx)
            
            df_row_idx = row_idx - 2
            gap_value = df.iloc[df_row_idx][gap_col_name]
            is_significant = df.iloc[df_row_idx][sig_col_name]
            
            if pd.notna(gap_value) and is_significant:
                gap_cell.value = f"{gap_value}*"
                
                if gap_value < 0:
                    gap_cell.fill = fill_red

def format_header_row(ws):
    """Format header row"""
    fill_gray = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    font_bold = Font(bold=True)
    
    for cell in ws[1]:
        cell.fill = fill_gray
        cell.font = font_bold
        cell.alignment = Alignment(horizontal='center', vertical='center')

# =============================================================================
# EXCEL WORKBOOK CREATION
# =============================================================================

def create_excel_workbook(df, output_file):
    """Create Excel workbook with ratio and data sheets"""
    
    print(f"\nCreating Excel workbook: {output_file}")
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Get unique bank-CBSA combinations
    bank_cbsa_combos = df.groupby(['lei', 'cbsa_code']).agg({
        'respondent_name': 'first',
        'cbsa_title': 'first'
    }).reset_index()
    
    print(f"  Total bank-CBSA combinations: {len(bank_cbsa_combos)}")
    
    for idx, row in bank_cbsa_combos.iterrows():
        lei = row['lei']
        cbsa_code = row['cbsa_code']
        bank_name = row['respondent_name'] if pd.notna(row['respondent_name']) else lei
        
        # Truncate for sheet name
        safe_bank_name = bank_name[:20] if len(bank_name) > 20 else bank_name
        cbsa_short = cbsa_code[:5]
        
        # Create ratio sheet
        ratio_sheet_name = f"{safe_bank_name}_{cbsa_short}_Ratio"[:31]
        print(f"  Creating: {ratio_sheet_name}")
        
        ratio_data = create_ratio_sheet_data(df, lei, cbsa_code)
        
        if ratio_data is not None and len(ratio_data) > 0:
            ws_ratio = wb.create_sheet(title=ratio_sheet_name)
            
            for r_idx, r in enumerate(dataframe_to_rows(ratio_data, index=False, header=True), 1):
                for c_idx, value in enumerate(r, 1):
                    ws_ratio.cell(row=r_idx, column=c_idx, value=value)
            
            format_header_row(ws_ratio)
            apply_ratio_sheet_formatting(ws_ratio, ratio_data)
            ws_ratio.freeze_panes = 'A2'
            ws_ratio.auto_filter.ref = ws_ratio.dimensions
        
        # Create data sheet
        data_sheet_name = f"{safe_bank_name}_{cbsa_short}_Data"[:31]
        print(f"  Creating: {data_sheet_name}")
        
        data_data = create_data_sheet_data(df, lei, cbsa_code)
        
        if data_data is not None and len(data_data) > 0:
            # Remove significance columns
            display_data = data_data.copy()
            sig_cols = [col for col in display_data.columns if '_Significant' in col]
            display_data = display_data.drop(columns=sig_cols)
            
            ws_data = wb.create_sheet(title=data_sheet_name)
            
            for r_idx, r in enumerate(dataframe_to_rows(display_data, index=False, header=True), 1):
                for c_idx, value in enumerate(r, 1):
                    ws_data.cell(row=r_idx, column=c_idx, value=value)
            
            format_header_row(ws_data)
            apply_data_sheet_formatting(ws_data, data_data)
            ws_data.freeze_panes = 'A2'
            ws_data.auto_filter.ref = ws_data.dimensions
    
    wb.save(output_file)
    print(f"\n✓ Excel workbook created: {output_file}")

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description='NCRC CBA Banks Analysis v4')
    parser.add_argument('--input', type=str, required=True, help='Input CSV from BigQuery')
    parser.add_argument('--output', type=str, default='CBA_Banks_Analysis.xlsx', help='Output Excel file')
    
    args = parser.parse_args()
    
    print("="*80)
    print("NCRC CBA BANKS DISPARATE IMPACT ANALYSIS v4.0")
    print("="*80)
    
    df = load_data(args.input)
    create_excel_workbook(df, args.output)
    
    print("\n" + "="*80)
    print("ANALYSIS COMPLETE!")
    print("="*80)
    print(f"\nOutput: {args.output}")
    print(f"\nv4 Changes:")
    print(f"  ✓ Removed ALL LMIB$ metrics from ALL sheets")
    print(f"  ✓ Fixed repeated figures bug")
    print(f"  ✓ Count columns show total originations")
    print(f"  ✓ Proper SQL column mapping (activity_year, cbsa_name, etc.)")

if __name__ == "__main__":
    main()
