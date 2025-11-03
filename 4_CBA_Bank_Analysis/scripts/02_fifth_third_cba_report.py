"""
Fifth Third Bank CBA Borrower Demographics Report Generator

Creates an Excel report with one sheet per CBSA showing borrower demographic metrics.
Each CBSA sheet shows:
- Loan_Purpose_Category × Metric × Year with Subject Share, Peer Share, and Gap

Metrics: Hispanic%, Black%, Asian%, Native American%, HoPI%, LMIB%, LMICT%

Input: CSV from BigQuery with Fifth Third borrower demographics data
Output: Excel workbook with:
  - One sheet per CBSA (10 CBSAs total)
  - Methods sheet documenting filters and methodology

Author: NCRC Research Department
Date: January 2025
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import warnings
import sys
import os
import glob
from datetime import datetime
import argparse
import re
# scipy not needed - we'll implement z-test with numpy

warnings.filterwarnings('ignore')

# =============================================================================
# CONFIGURATION
# =============================================================================

# Metrics grouped by type for ordered display
METRIC_GROUPS = {
    'borrower_demographics': {
        'asian': 'Asian%',
        'black': 'Black%',
        'hispanic': 'Hispanic%',
        'hopi': 'HoPI%',
        'native_american': 'Native American%',
    },
    'income': {
        'lmib': 'LMIB%',
        'lmict': 'LMICT%',
    },
    'redlining': {
        'mmct_50': 'MMCT 50%',
        'mmct_80': 'MMCT 80%',
        'black_tract_50': 'Black Tract 50%',
        'black_tract_80': 'Black Tract 80%',
        'hispanic_tract_50': 'Hispanic Tract 50%',
        'hispanic_tract_80': 'Hispanic Tract 80%',
        'black_hispanic_tract_50': 'Black+Hispanic Tract 50%',
        'black_hispanic_tract_80': 'Black+Hispanic Tract 80%',
    }
}

# Flattened METRICS dict for backward compatibility
METRICS = {}
for group_dict in METRIC_GROUPS.values():
    METRICS.update(group_dict)

# Data years expected
EXPECTED_YEARS = ['2018', '2019', '2020', '2021', '2022', '2023', '2024']

# Statistical significance threshold
SIGNIFICANCE_LEVEL = 0.05

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def two_proportion_z_test(x1, n1, x2, n2):
    """
    Perform two-proportion z-test
    Returns: (z_statistic, p_value)
    x1, n1: successes and trials for group 1 (subject bank)
    x2, n2: successes and trials for group 2 (peer banks)
    """
    if n1 == 0 or n2 == 0:
        return None, None
    
    p1 = x1 / n1
    p2 = x2 / n2
    
    # Pooled proportion
    p_pool = (x1 + x2) / (n1 + n2)
    
    # Standard error
    se = np.sqrt(p_pool * (1 - p_pool) * (1/n1 + 1/n2))
    
    if se == 0:
        return None, None
    
    # Z-statistic
    z = (p1 - p2) / se
    
    # Two-tailed p-value (approximation using error function)
    # P(|Z| > |z|) = 2 * (1 - Φ(|z|))
    # Using approximation: Φ(x) ≈ 0.5 * (1 + erf(x/√2))
    from math import erf, sqrt
    p_value = 2 * (1 - 0.5 * (1 + erf(abs(z) / sqrt(2))))
    
    return z, p_value

# =============================================================================
# DATA LOADING AND PROCESSING
# =============================================================================

def load_data(input_file, second_file=None):
    """Load CSV data from BigQuery export(s) and merge if needed"""
    print(f"\nLoading data from: {input_file}")
    
    df1 = pd.read_csv(input_file)
    print(f"  Loaded {len(df1):,} rows from first file")
    
    # Convert year to string for consistency
    df1['activity_year'] = df1['activity_year'].astype(str)
    
    # Clean up loan_purpose_category if it's coming as a tuple string
    if 'loan_purpose_category' in df1.columns:
        df1['loan_purpose_category'] = df1['loan_purpose_category'].astype(str)
        # Handle tuple format: ('Home Purchase',) -> Home Purchase
        # Remove leading (' and trailing ',) or ') or )
        df1['loan_purpose_category'] = df1['loan_purpose_category'].str.replace(r"^\('|^\(|',\)$|'\)$|\)$", '', regex=True)
        # Remove remaining quotes
        df1['loan_purpose_category'] = df1['loan_purpose_category'].str.strip().str.strip("'\"")
        # Final cleanup for any edge cases
        df1['loan_purpose_category'] = df1['loan_purpose_category'].apply(lambda x: x.replace("('", "").replace("',)", "").replace("')", "").strip() if isinstance(x, str) else x)
    
    # If second file provided, load and merge
    if second_file:
        print(f"\nLoading data from: {second_file}")
        df2 = pd.read_csv(second_file)
        print(f"  Loaded {len(df2):,} rows from second file")
        
        # Convert year to string for consistency
        df2['activity_year'] = df2['activity_year'].astype(str)
        
        # Clean up loan_purpose_category if it's coming as a tuple string
        if 'loan_purpose_category' in df2.columns:
            df2['loan_purpose_category'] = df2['loan_purpose_category'].astype(str)
            # Handle tuple format: ('Home Purchase',) -> Home Purchase
            # Remove leading (' and trailing ',) or ') or )
            df2['loan_purpose_category'] = df2['loan_purpose_category'].str.replace(r"^\('|^\(|',\)$|'\)$|\)$", '', regex=True)
            # Remove remaining quotes
            df2['loan_purpose_category'] = df2['loan_purpose_category'].str.strip().str.strip("'\"")
            # Final cleanup for any edge cases
            df2['loan_purpose_category'] = df2['loan_purpose_category'].apply(lambda x: x.replace("('", "").replace("',)", "").replace("')", "").strip() if isinstance(x, str) else x)
        
        # Merge on key columns
        merge_cols = ['activity_year', 'cbsa_code', 'loan_purpose_category']
        
        # Check which columns exist in both
        common_cols = [col for col in merge_cols if col in df1.columns and col in df2.columns]
        
        print(f"\nMerging on columns: {common_cols}")
        
        # Identify columns that should not be merged (keep from both)
        df1_only_cols = [col for col in df1.columns if col not in df2.columns]
        df2_only_cols = [col for col in df2.columns if col not in df1.columns]
        shared_cols = [col for col in df1.columns if col in df2.columns and col not in common_cols]
        
        print(f"  Columns only in first file: {len(df1_only_cols)}")
        print(f"  Columns only in second file: {len(df2_only_cols)}")
        print(f"  Shared columns (will be merged): {len(shared_cols)}")
        
        # Merge: use outer join to keep all rows, then combine overlapping columns
        # For overlapping columns like subject_total_originations, take from df1 or combine
        df = pd.merge(df1, df2, on=common_cols, how='outer', suffixes=('', '_y'))
        
        # For duplicate columns (shared_cols), prefer df1 but fill with df2 if null
        # BUT for total_originations, use the maximum to avoid undercounting
        for col in shared_cols:
            if f'{col}_y' in df.columns:
                if 'total_originations' in col:
                    # For totals, use the maximum to ensure we don't undercount
                    df[col] = df[[col, f'{col}_y']].max(axis=1)
                else:
                    # For other columns, prefer df1 but fill with df2 if null
                    df[col] = df[col].fillna(df[f'{col}_y'])
                df = df.drop(columns=[f'{col}_y'])
        
        print(f"  Merged to {len(df):,} rows")
    else:
        df = df1
    
    print(f"\nFinal dataset:")
    print(f"  CBSAs: {sorted(df['cbsa_code'].unique())}")
    print(f"  Years: {sorted(df['activity_year'].unique())}")
    if 'loan_purpose_category' in df.columns:
        print(f"  Loan Purpose Categories: {sorted(df['loan_purpose_category'].unique())}")
    
    return df

def calculate_metrics(row, metric_base):
    """
    Calculate share and gap for a metric
    Returns: (subject_share, peer_share, gap, is_significant)
    """
    subject_total = row.get('subject_total_originations', 0)
    peer_total = row.get('peer_total_originations', 0)
    
    if subject_total == 0 or peer_total == 0:
        return None, None, None, False
    
    subject_count = row.get(f'subject_{metric_base}_originations', 0)
    peer_count = row.get(f'peer_{metric_base}_originations', 0)
    
    # Calculate shares
    subject_share = (subject_count / subject_total * 100) if subject_total > 0 else 0
    peer_share = (peer_count / peer_total * 100) if peer_total > 0 else 0
    
    # Cap percentages at 100% (data quality safeguard)
    # If >100%, it indicates a data issue (double-counting or wrong denominator)
    if subject_share > 100:
        print(f"WARNING: Subject share exceeds 100% for {metric_base}: {subject_share:.1f}% (count: {subject_count}, total: {subject_total})")
        subject_share = 100.0
    if peer_share > 100:
        print(f"WARNING: Peer share exceeds 100% for {metric_base}: {peer_share:.1f}% (count: {peer_count}, total: {peer_total})")
        peer_share = 100.0
    
    gap = subject_share - peer_share
    
    # Statistical significance test
    z_stat, p_value = two_proportion_z_test(
        subject_count, subject_total,
        peer_count, peer_total
    )
    
    is_significant = (p_value is not None and 
                     p_value < SIGNIFICANCE_LEVEL and 
                     gap < 0)  # Only flag negative gaps
    
    return subject_share, peer_share, gap, is_significant

def add_calculated_metrics(df):
    """Add calculated shares and gaps for all metrics"""
    print("\nCalculating shares and gaps...")
    print(f"  Input dataframe shape: {df.shape}")
    print(f"  Sample columns: {list(df.columns[:10])}")
    
    # Check for required columns
    required_cols = ['subject_total_originations', 'peer_total_originations']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        print(f"  WARNING: Missing required columns: {missing_cols}")
    
    for metric_base, metric_name in METRICS.items():
        print(f"  Processing {metric_name} (base: {metric_base})...")
        
        # Check if source columns exist
        subject_col = f'subject_{metric_base}_originations'
        peer_col = f'peer_{metric_base}_originations'
        
        if subject_col not in df.columns or peer_col not in df.columns:
            print(f"    WARNING: Missing columns for {metric_name}: {subject_col}, {peer_col}")
            # Create empty columns
            df[f'{metric_base}_subject_share'] = None
            df[f'{metric_base}_peer_share'] = None
            df[f'{metric_base}_gap'] = None
            df[f'{metric_base}_significant'] = False
            continue
        
        # Calculate for each row
        results = df.apply(lambda row: calculate_metrics(row, metric_base), axis=1)
        
        # Unpack results into separate columns
        df[f'{metric_base}_subject_share'] = results.apply(lambda x: x[0] if x else None)
        df[f'{metric_base}_peer_share'] = results.apply(lambda x: x[1] if x else None)
        df[f'{metric_base}_gap'] = results.apply(lambda x: x[2] if x else None)
        df[f'{metric_base}_significant'] = results.apply(lambda x: x[3] if x else False)
        
        # Count non-null values
        non_null = df[f'{metric_base}_subject_share'].notna().sum()
        print(f"    Created {non_null} non-null values for {metric_name}")
    
    print(f"  Output dataframe shape: {df.shape}")
    return df

# =============================================================================
# CBSA SHEET PREPARATION
# =============================================================================

def prepare_cbsa_sheet_data(df, cbsa_code):
    """
    Prepare data for a CBSA sheet
    Returns three separate DataFrames: demographics, income, redlining
    Format: Loan_Purpose_Category × Metric × Year
    Shows: Subject Share, Peer Share, Gap
    """
    cbsa_data = df[df['cbsa_code'] == cbsa_code].copy()
    
    if len(cbsa_data) == 0:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    demographics_rows = []
    income_rows = []
    redlining_rows = []
    
    # Group by loan purpose category
    print(f"    Grouping by loan_purpose_category...")
    print(f"    Available loan purposes: {cbsa_data['loan_purpose_category'].unique() if 'loan_purpose_category' in cbsa_data.columns else 'COLUMN MISSING'}")
    print(f"    Available years: {sorted(cbsa_data['activity_year'].unique()) if 'activity_year' in cbsa_data.columns else 'COLUMN MISSING'}")
    
    if 'loan_purpose_category' not in cbsa_data.columns:
        print(f"    ERROR: 'loan_purpose_category' column missing!")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    for loan_purpose_cat, group in cbsa_data.groupby(['loan_purpose_category']):
        print(f"    Processing loan purpose: {loan_purpose_cat} ({len(group)} rows)")
        
        # Borrower Demographics Table
        for metric_base, metric_name in METRIC_GROUPS['borrower_demographics'].items():
            row = create_metric_row(loan_purpose_cat, metric_name, metric_base, group)
            demographics_rows.append(row)
        
        # Income Table
        for metric_base, metric_name in METRIC_GROUPS['income'].items():
            row = create_metric_row(loan_purpose_cat, metric_name, metric_base, group)
            income_rows.append(row)
        
        # Redlining Table
        for metric_base, metric_name in METRIC_GROUPS['redlining'].items():
            row = create_metric_row(loan_purpose_cat, metric_name, metric_base, group)
            redlining_rows.append(row)
    
    demographics_df = pd.DataFrame(demographics_rows)
    income_df = pd.DataFrame(income_rows)
    redlining_df = pd.DataFrame(redlining_rows)
    
    # Sort by Loan Purpose Category only (metrics already in correct order)
    if len(demographics_df) > 0:
        demographics_df = demographics_df.sort_values(['Loan_Purpose_Category'])
    if len(income_df) > 0:
        income_df = income_df.sort_values(['Loan_Purpose_Category'])
    if len(redlining_df) > 0:
        redlining_df = redlining_df.sort_values(['Loan_Purpose_Category'])
    
    return demographics_df, income_df, redlining_df

def create_metric_row(loan_purpose_cat, metric_name, metric_base, group):
    """Helper function to create a metric row with all year columns"""
    # Clean up loan_purpose_cat if it's a tuple string
    if isinstance(loan_purpose_cat, str):
        # Remove tuple formatting
        loan_purpose_cat = loan_purpose_cat.replace("('", "").replace("',)", "").replace("')", "").replace("'", "").strip()
    
    row = {
        'Loan_Purpose_Category': loan_purpose_cat,
        'Metric': metric_name
    }
    
    # For each year, add three columns: Subject Share, Peer Share, Gap
    for year in EXPECTED_YEARS:
        # Convert year to string for comparison (activity_year should be string)
        year_str = str(year)
        year_data = group[group['activity_year'].astype(str) == year_str]
        
        if len(year_data) > 0:
            # Check if required columns exist
            subject_col = f'{metric_base}_subject_share'
            peer_col = f'{metric_base}_peer_share'
            gap_col = f'{metric_base}_gap'
            sig_col = f'{metric_base}_significant'
            
            if subject_col in year_data.columns and peer_col in year_data.columns:
                subject_share = year_data[subject_col].iloc[0]
                peer_share = year_data[peer_col].iloc[0]
                gap = year_data[gap_col].iloc[0] if gap_col in year_data.columns else None
                is_significant = year_data[sig_col].iloc[0] if sig_col in year_data.columns else False
                
                # Format values
                row[f'{year}_Subject_Share'] = f"{subject_share:.1f}%" if pd.notna(subject_share) else "N/A"
                row[f'{year}_Peer_Share'] = f"{peer_share:.1f}%" if pd.notna(peer_share) else "N/A"
                
                if pd.notna(gap):
                    gap_str = f"{gap:.1f}pp"
                    if is_significant and gap < 0:
                        gap_str += "*"
                    row[f'{year}_Gap'] = gap_str
                else:
                    row[f'{year}_Gap'] = "N/A"
            else:
                # Columns missing - might be a data issue
                row[f'{year}_Subject_Share'] = "N/A"
                row[f'{year}_Peer_Share'] = "N/A"
                row[f'{year}_Gap'] = "N/A"
        else:
            row[f'{year}_Subject_Share'] = "N/A"
            row[f'{year}_Peer_Share'] = "N/A"
            row[f'{year}_Gap'] = "N/A"
    
    return row

# =============================================================================
# EXCEL WORKBOOK CREATION
# =============================================================================

def clean_excel_value(value):
    """Clean a value to ensure it can be written to Excel"""
    if pd.isna(value) or value is None:
        return None
    
    # Handle actual tuple objects (most important case)
    if isinstance(value, tuple):
        # Extract first element if it's a single-element tuple
        if len(value) == 1:
            value = value[0]
        else:
            # Multiple elements - join them
            value = ", ".join(str(v) for v in value)
        # Recursively clean in case the tuple element is also problematic
        return clean_excel_value(value)
    
    # Handle list objects too
    if isinstance(value, list):
        if len(value) == 1:
            value = value[0]
        else:
            value = ", ".join(str(v) for v in value)
        return clean_excel_value(value)
    
    if isinstance(value, str):
        # Remove tuple formatting if present - check for various formats
        if "('" in value or "',)" in value or value.startswith("('") or value.endswith("',)"):
            # Remove all tuple formatting
            value = value.replace("('", "").replace("',)", "").replace("')", "").replace("('", "")
            # Remove remaining quotes and parentheses
            value = value.strip("'\"()")
            value = value.strip()
        return value
    
    # For other types (int, float, bool), return as-is
    return value

def get_cbsa_short_name(cbsa_full_name):
    """Extract short name from CBSA full name (e.g., 'Chicago-Naperville-Elgin, IL-IN-WI' -> 'Chicago')"""
    if pd.isna(cbsa_full_name):
        return None
    
    # Take the part before the first comma or dash
    short_name = cbsa_full_name.split(',')[0].split('-')[0].strip()
    
    # Remove characters not allowed in Excel sheet names
    # Excel doesn't allow: / \ ? * [ ] 
    short_name = re.sub(r'[\/\\\?\*\[\]]', ' ', short_name)
    short_name = short_name.strip()
    
    return short_name if short_name else None

def create_excel_workbook(df, output_file):
    """Create Excel workbook with one sheet per CBSA"""
    print(f"\nCreating Excel workbook: {output_file}")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # Get all CBSAs with their names
        cbsa_info = df.groupby(['cbsa_code', 'cbsa_name']).first().reset_index()
        print(f"  Creating sheets for {len(cbsa_info)} CBSAs...")
        
        # Create one sheet per CBSA
        for _, row in cbsa_info.iterrows():
            cbsa_code = row['cbsa_code']
            cbsa_full_name = row['cbsa_name']
            cbsa_short_name = get_cbsa_short_name(cbsa_full_name)
            
            # Create sheet name from short name (e.g., "Chicago")
            sheet_name = cbsa_short_name if cbsa_short_name else f"CBSA_{cbsa_code}"
            sheet_name = sheet_name[:31]  # Excel sheet name limit
            
            print(f"  Creating sheet: {sheet_name} ({cbsa_code})")
            
            demographics_df, income_df, redlining_df = prepare_cbsa_sheet_data(df, cbsa_code)
            print(f"    Demographics rows: {len(demographics_df)}, Income rows: {len(income_df)}, Redlining rows: {len(redlining_df)}")
            
            # Create empty sheet first
            ws = writer.book.create_sheet(sheet_name)
            current_row = 1
            
            table_starts = {}  # Track where each table starts
            
            # Check if we have any data at all
            if len(demographics_df) == 0 and len(income_df) == 0 and len(redlining_df) == 0:
                print(f"    WARNING: No data found for CBSA {cbsa_code}")
                ws.cell(row=1, column=1, value="No data available for this CBSA")
                continue
            
            # Write Borrower Demographics table
            if len(demographics_df) > 0:
                table_starts['demographics'] = current_row + 1  # Header row
                ws.cell(row=current_row, column=1, value="BORROWER DEMOGRAPHICS")
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
                current_row += 1
                
                # Clean DataFrame before writing (handle tuple values)
                demographics_df_clean = demographics_df.copy()
                for col in demographics_df_clean.columns:
                    demographics_df_clean[col] = demographics_df_clean[col].apply(clean_excel_value)
                
                # Write header and data
                for r_idx, row_data in enumerate(dataframe_to_rows(demographics_df_clean, index=False, header=True), current_row):
                    for c_idx, value in enumerate(row_data, 1):
                        # Double-check - clean value again for safety
                        clean_value = clean_excel_value(value)
                        ws.cell(row=r_idx, column=c_idx, value=clean_value)
                current_row += len(demographics_df) + 1  # +1 for header
                current_row += 2  # Add spacing between tables
            
            # Write Income table
            if len(income_df) > 0:
                table_starts['income'] = current_row + 1  # Header row
                ws.cell(row=current_row, column=1, value="INCOME METRICS")
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
                current_row += 1
                
                # Clean DataFrame before writing (handle tuple values)
                income_df_clean = income_df.copy()
                for col in income_df_clean.columns:
                    income_df_clean[col] = income_df_clean[col].apply(clean_excel_value)
                
                # Write header and data
                for r_idx, row_data in enumerate(dataframe_to_rows(income_df_clean, index=False, header=True), current_row):
                    for c_idx, value in enumerate(row_data, 1):
                        # Double-check - clean value again for safety
                        clean_value = clean_excel_value(value)
                        ws.cell(row=r_idx, column=c_idx, value=clean_value)
                current_row += len(income_df) + 1  # +1 for header
                current_row += 2  # Add spacing between tables
            
            # Write Redlining table
            if len(redlining_df) > 0:
                table_starts['redlining'] = current_row + 1  # Header row
                ws.cell(row=current_row, column=1, value="REDLINING METRICS")
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
                current_row += 1
                
                # Clean DataFrame before writing (handle tuple values)
                redlining_df_clean = redlining_df.copy()
                for col in redlining_df_clean.columns:
                    redlining_df_clean[col] = redlining_df_clean[col].apply(clean_excel_value)
                
                # Write header and data
                for r_idx, row_data in enumerate(dataframe_to_rows(redlining_df_clean, index=False, header=True), current_row):
                    for c_idx, value in enumerate(row_data, 1):
                        # Double-check - clean value again for safety
                        clean_value = clean_excel_value(value)
                        ws.cell(row=r_idx, column=c_idx, value=clean_value)
                current_row += len(redlining_df) + 1  # +1 for header
            
            # Format all header rows
            format_all_header_rows(ws)
            
            # Apply auto-filter to each table
            apply_auto_filters_simple(ws, demographics_df, income_df, redlining_df, table_starts)
            
            # Freeze first row
            ws.freeze_panes = 'A2'
            
            # Apply coloring to all tables
            apply_cbsa_sheet_coloring_multi(ws, demographics_df, income_df, redlining_df, table_starts)
            
            # Add footnotes with CBSA metadata
            add_cbsa_footnotes(ws, df, cbsa_code, cbsa_full_name, current_row)
        
        # Methodology sheet - create within the writer context
        print("  Creating methodology sheet...")
        ws_methods = writer.book.create_sheet('Methods')
        
        methodology_text = """
FIFTH THIRD BANK COMBINED ANALYSIS - METHODOLOGY

PURPOSE:
This report combines two types of analysis:
1. REDLINING ANALYSIS: Fifth Third Bank's lending patterns in majority-minority and racially concentrated census tracts
   compared to peer lenders. Focuses on identifying potential redlining (disparate geographic impact).
2. BORROWER DEMOGRAPHICS ANALYSIS: Fifth Third Bank's lending patterns to borrowers by race/ethnicity and income
   compared to peer lenders. Focuses on identifying potential disparate impact in lending to minority and low-to-moderate income borrowers.

GEOGRAPHIC SCOPE:
- Top 10 CBSAs (Core Based Statistical Areas) by Fifth Third Bank home purchase loan volume
- Years: 2018-2024
- All census tracts within selected CBSAs

LOAN FILTERS (STANDARD HMDA FILTERS):
- Loan Purpose: Home Purchase only (loan_purpose = '1')
- Owner Occupied: Yes (occupancy_type = '1')
- Reverse Mortgage: Excluded (reverse_mortgage != '1')
- Construction Method: Site-built (construction_method = '1')
- Number of Units: 1-4 units (total_units IN '1','2','3','4')
- Action Taken: Originations only (action_taken = '1')

PEER COMPARISON:
- Peers defined as lenders with 50%-200% of Fifth Third Bank's origination volume in each CBSA-year
- Analysis excludes Fifth Third Bank from peer group
- Each CBSA-year has its own peer group

METRICS ANALYZED:

REDLINING METRICS (Tract-Level Demographics):
1. MMCT 50%: Loans in majority-minority census tracts (>50% non-white population)
2. MMCT 80%: Loans in supermajority-minority census tracts (>80% non-white population)
3. Black Tract 50%: Loans in majority-Black census tracts (>50% Black population)
4. Black Tract 80%: Loans in supermajority-Black census tracts (>80% Black population)
5. Hispanic Tract 50%: Loans in majority-Hispanic census tracts (>50% Hispanic population)
6. Hispanic Tract 80%: Loans in supermajority-Hispanic census tracts (>80% Hispanic population)
7. Black+Hispanic Tract 50%: Loans in majority Black+Hispanic census tracts (>50% combined)
8. Black+Hispanic Tract 80%: Loans in supermajority Black+Hispanic census tracts (>80% combined)

BORROWER DEMOGRAPHIC METRICS:
9. Hispanic%: Share of loans to Hispanic/Latino borrowers (any ethnicity field = 1, 11-14)
10. Black%: Share of loans to Black borrowers (non-Hispanic, race = 3)
11. Asian%: Share of loans to Asian borrowers (non-Hispanic, race = 2 or 21-27)
12. Native American%: Share of loans to Native American borrowers (non-Hispanic, race = 1)
13. HoPI%: Share of loans to Native Hawaiian/Pacific Islander borrowers (non-Hispanic, race = 4 or 41-44)
14. LMIB%: Share of loans to Low-to-Moderate Income Borrowers (borrower income ≤80% of MSA median)
15. LMICT%: Share of loans in Low-to-Moderate Income Census Tracts (tract income ≤80% of MSA median)

RACE/ETHNICITY CLASSIFICATION METHODOLOGY (NCRC):
- Hierarchical classification: Hispanic ethnicity checked FIRST (regardless of race)
- If any of the 5 ethnicity fields (applicant_ethnicity_1 through applicant_ethnicity_5) contains codes 1, 11, 12, 13, or 14, borrower is classified as Hispanic
- If NOT Hispanic, borrower categorized by race using all 5 race fields (applicant_race_1 through applicant_race_5)
- Race categories (non-Hispanic only):
  - Black: Race = 3
  - Asian: Race = 2 or 21-27 (Asian subgroups)
  - Native American: Race = 1
  - HoPI: Race = 4 or 41-44 (Pacific Islander subgroups)
  - White: Race = 5
  - Unknown: No race or ethnicity data provided
- LMIB (Low-to-Moderate Income Borrowers): Borrower income ≤80% of MSA median income
  - Calculated as: (income * 1000) / ffiec_msa_md_median_family_income * 100 ≤ 80%
  - Income field is in thousands, so multiplied by 1000 to get actual income
- LMICT (Low-to-Moderate Income Census Tracts): Tract median income ≤80% of MSA median income

DATA SOURCES:
- HMDA Data: Federal Financial Institutions Examination Council (FFIEC)
  - Years: 2018-2024
  - Table: hdma1-242116.hmda.hmda
  - Borrower demographics from applicant_ethnicity_1-5 and applicant_race_1-5 fields
  - Tract minority percentage from tract_minority_population_percent field
  - LMIB calculated from income field (in thousands) and ffiec_msa_md_median_family_income field
  - LMICT calculated from tract_to_msa_income_percentage field
- Geographic Data: Census tract demographics
  - Table: hdma1-242116.geo.black_hispanic_majority
  - Contains black_pct, hispanic_pct, black_and_hispanic_pct by census tract
- CBSA Crosswalk: County to CBSA mapping
  - Table: hdma1-242116.geo.cbsa_to_county
  - Maps state_code + county_code (GEOID5) to CBSA code and name

CALCULATION METHODS:
- Subject Share: (Fifth Third metric count / Fifth Third total originations) × 100
- Peer Share: (Peer metric count / Peer total originations) × 100
- Gap: Subject Share - Peer Share (percentage points)
- Positive gap means Fifth Third outperforms peers
- Negative gap means Fifth Third underperforms peers

STATISTICAL SIGNIFICANCE:
- Two-proportion z-test performed for each metric
- Asterisk (*) indicates statistically significant negative gap (p < 0.05)
- Only negative gaps are flagged for significance

COLOR CODING:
- Green: Positive gap (bank outperforms peers)
- Yellow: Small negative gap (0 to -2.0pp)
- Orange: Moderate negative gap (-2.0 to -5.0pp)
- Light Red: Large negative gap (-5.0 to -10.0pp)
- Red: Severe negative gap (<-10.0pp)
- Red with white bold text: Severe negative gap + statistically significant

INTERPRETATION:
- Focus on patterns across years, not single-year gaps
- Statistically significant gaps (marked with *) are more reliable
- Large, persistent negative gaps warrant closer review
- REDLINING: Systematic underperformance in minority tracts indicates potential redlining
- DISPARATE IMPACT: Systematic underperformance to minority borrowers indicates potential disparate impact

LIMITATIONS:
- Does not account for differences in business models or market strategies
- Geographic selection biased toward high-volume markets
- Peer matching based solely on application volume
- Does not include all factors affecting lending decisions
- Correlation does not imply causation
- Does not assess individual loan applications for discrimination
- Borrower race/ethnicity classification based on HMDA reporting (self-reported or visual observation)

CONTACT:
NCRC Research Department
National Community Reinvestment Coalition
For questions about methodology or data, contact: research@ncrc.org
"""
        
        # Write methodology text
        for i, line in enumerate(methodology_text.strip().split('\n'), 1):
            ws_methods.cell(row=i, column=1, value=line)
        
        # Format
        for row in ws_methods.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        ws_methods.column_dimensions['A'].width = 120
    
    print(f"\nExcel workbook created: {output_file}")

def format_header_row(worksheet):
    """Apply formatting to header row"""
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def format_all_header_rows(worksheet):
    """Format all header rows in the sheet (each table has its own header)"""
    # Find header rows by looking for "Loan_Purpose_Category" in first column
    for row_idx, row in enumerate(worksheet.iter_rows(), 1):
        if row[0].value == "Loan_Purpose_Category":
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def apply_auto_filters_simple(worksheet, demographics_df, income_df, redlining_df, table_starts):
    """Apply auto-filter to each table separately"""
    # Note: Excel only supports one auto_filter per sheet, so we'll apply it to the largest table
    # Or we can apply it to the first table with data
    
    if len(demographics_df) > 0 and 'demographics' in table_starts:
        header_row = table_starts['demographics']
        data_end_row = header_row + len(demographics_df)
        if demographics_df.shape[1] > 0:
            last_col_letter = get_column_letter(demographics_df.shape[1])
            worksheet.auto_filter.ref = f"A{header_row}:{last_col_letter}{data_end_row}"
    elif len(income_df) > 0 and 'income' in table_starts:
        header_row = table_starts['income']
        data_end_row = header_row + len(income_df)
        if income_df.shape[1] > 0:
            last_col_letter = get_column_letter(income_df.shape[1])
            worksheet.auto_filter.ref = f"A{header_row}:{last_col_letter}{data_end_row}"
    elif len(redlining_df) > 0 and 'redlining' in table_starts:
        header_row = table_starts['redlining']
        data_end_row = header_row + len(redlining_df)
        if redlining_df.shape[1] > 0:
            last_col_letter = get_column_letter(redlining_df.shape[1])
            worksheet.auto_filter.ref = f"A{header_row}:{last_col_letter}{data_end_row}"

def apply_cbsa_sheet_coloring(worksheet, dataframe):
    """Apply color coding to gap cells based on severity (single table version)"""
    # Define colors
    fill_green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    fill_orange = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    fill_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    fill_red_bold = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    
    font_white_bold = Font(color='FFFFFF', bold=True)
    
    # Find gap columns (columns with "_Gap" in header)
    for col_idx, cell in enumerate(worksheet[1], 1):
        if '_Gap' in str(cell.value):
            col_letter = get_column_letter(col_idx)
            
            # Apply color to each gap cell
            for row_idx in range(2, len(dataframe) + 2):
                cell_ref = worksheet[f'{col_letter}{row_idx}']
                gap_value = cell_ref.value
                
                if gap_value and isinstance(gap_value, str) and gap_value != "N/A":
                    # Extract numeric value
                    gap_num = float(gap_value.replace('pp', '').replace('*', ''))
                    
                    # Apply color based on gap severity
                    if gap_num >= 0:
                        cell_ref.fill = fill_green
                    elif gap_num >= -2.0:
                        cell_ref.fill = fill_yellow
                    elif gap_num >= -5.0:
                        cell_ref.fill = fill_orange
                    elif gap_num >= -10.0:
                        cell_ref.fill = fill_red
                    else:
                        cell_ref.fill = fill_red_bold
                        cell_ref.font = font_white_bold

def apply_cbsa_sheet_coloring_multi(worksheet, demographics_df, income_df, redlining_df, table_starts):
    """Apply color coding to gap cells in all three tables"""
    fill_green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    fill_orange = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    fill_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    font_white_bold = Font(color='FFFFFF', bold=True)
    
    # Process each table
    tables = [
        (demographics_df, 'demographics'),
        (income_df, 'income'),
        (redlining_df, 'redlining')
    ]
    
    for df, table_key in tables:
        if len(df) > 0 and table_key in table_starts:
            start_row = table_starts[table_key]
            
            # Find gap columns by examining the header row
            header_row = worksheet[start_row]
            gap_cols = []
            for col_idx, cell in enumerate(header_row, 1):
                if cell.value and '_Gap' in str(cell.value):
                    gap_cols.append((col_idx, get_column_letter(col_idx)))
            
            # Apply coloring to gap cells in this table
            for col_idx, col_letter in gap_cols:
                for row_offset in range(1, len(df) + 1):  # Skip header, start with data rows
                    row_idx = start_row + row_offset
                    cell_ref = worksheet[f'{col_letter}{row_idx}']
                    gap_value = cell_ref.value
                    
                    if gap_value and isinstance(gap_value, str) and gap_value != "N/A":
                        try:
                            gap_num = float(gap_value.replace('pp', '').replace('*', ''))
                            
                            if gap_num >= 0:
                                cell_ref.fill = fill_green
                            elif gap_num >= -2.0:
                                cell_ref.fill = fill_yellow
                            elif gap_num >= -5.0:
                                cell_ref.fill = fill_orange
                            elif gap_num >= -10.0:
                                cell_ref.fill = fill_red
                            else:
                                cell_ref.fill = fill_red
                                cell_ref.font = font_white_bold
                        except (ValueError, AttributeError):
                            pass  # Skip invalid values

def get_counties_for_cbsa(df, cbsa_code):
    """Extract unique counties for a CBSA from the dataframe"""
    cbsa_data = df[df['cbsa_code'] == cbsa_code].copy()
    
    counties = []
    
    # First, check for counties_included column from SQL query (preferred)
    if 'counties_included' in cbsa_data.columns:
        counties_str = cbsa_data['counties_included'].iloc[0] if len(cbsa_data) > 0 else None
        if pd.notna(counties_str) and str(counties_str).strip():
            # Parse comma-separated list
            counties = [c.strip() for c in str(counties_str).split(',') if c.strip()]
            if len(counties) > 0:
                return counties
    
    # Fallback: Try to find county information in the data
    # Check common column names for county info
    county_cols = ['county_name', 'county', 'County Name', 'COUNTY_NAME', 
                   'state_county', 'state_and_county']
    
    # If we have county name directly
    for col in county_cols:
        if col in cbsa_data.columns:
            unique_counties = cbsa_data[col].dropna().unique()
            if len(unique_counties) > 0:
                counties = sorted([str(c) for c in unique_counties])
                break
    
    # If we have state_code and county_code, try to construct names
    if len(counties) == 0 and 'state_code' in cbsa_data.columns and 'county_code' in cbsa_data.columns:
        # Get unique combinations
        if 'county_name' not in cbsa_data.columns:
            # We have codes but not names - we'll show codes
            state_county_pairs = cbsa_data[['state_code', 'county_code']].drop_duplicates()
            counties = [f"{row['state_code']}-{row['county_code']}" 
                       for _, row in state_county_pairs.iterrows()]
            counties = sorted(counties)
        else:
            # If county_name exists, use it
            counties = sorted(cbsa_data['county_name'].dropna().unique().tolist())
    
    return counties

def add_cbsa_footnotes(worksheet, df, cbsa_code, cbsa_full_name, current_row):
    """Add footnotes with CBSA metadata at the bottom of the sheet"""
    cbsa_data = df[df['cbsa_code'] == cbsa_code].copy()
    
    # Use provided current_row or find the last row with data
    if current_row:
        note_start_row = current_row + 3
    else:
        last_data_row = len([row for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row) if any(cell.value for cell in row)])
        note_start_row = last_data_row + 3
    
    # Get loan counts by year
    year_counts = {}
    for year in EXPECTED_YEARS:
        year_data = cbsa_data[cbsa_data['activity_year'] == year]
        if len(year_data) > 0:
            year_counts[year] = year_data['subject_total_originations'].iloc[0]
    
    # Get counties for this CBSA
    counties = get_counties_for_cbsa(df, cbsa_code)
    
    # Write footnotes
    worksheet.cell(row=note_start_row, column=1, value="CBSA Information:")
    worksheet.cell(row=note_start_row, column=1).font = Font(bold=True)
    
    worksheet.cell(row=note_start_row+1, column=1, value=f"Full Name: {cbsa_full_name}")
    worksheet.cell(row=note_start_row+2, column=1, value=f"CBSA Code: {cbsa_code}")
    
    # Add counties right after CBSA name/code
    row_offset = note_start_row + 3
    if len(counties) > 0:
        counties_text = ", ".join(counties)
        # If too long, split across multiple lines
        if len(counties_text) > 100:
            # Split into chunks
            worksheet.cell(row=row_offset, column=1, value="Counties Included:")
            worksheet.cell(row=row_offset, column=1).font = Font(bold=True)
            row_offset += 1
            current_text = []
            current_length = 0
            for county in counties:
                if current_length + len(county) + 2 > 100 and current_text:
                    # Write current line
                    worksheet.cell(row=row_offset, column=1, value="  " + ", ".join(current_text))
                    current_text = [county]
                    current_length = len(county)
                    row_offset += 1
                else:
                    current_text.append(county)
                    current_length += len(county) + 2
            if current_text:
                worksheet.cell(row=row_offset, column=1, value="  " + ", ".join(current_text))
                row_offset += 1
        else:
            worksheet.cell(row=row_offset, column=1, value=f"Counties Included: {counties_text}")
            row_offset += 1
    else:
        worksheet.cell(row=row_offset, column=1, value="Counties Included: (Not available in data)")
        row_offset += 1
    
    note_start_row = row_offset
    
    worksheet.cell(row=note_start_row+1, column=1, value="Fifth Third Bank Loans by Year:")
    worksheet.cell(row=note_start_row+1, column=1).font = Font(bold=True)
    
    year_row = note_start_row + 5
    worksheet.cell(row=year_row, column=1, value="Year")
    worksheet.cell(row=year_row, column=2, value="Home Purchase Loans")
    worksheet.cell(row=year_row, column=1).font = Font(bold=True)
    worksheet.cell(row=year_row, column=2).font = Font(bold=True)
    
    for year in EXPECTED_YEARS:
        year_row += 1
        worksheet.cell(row=year_row, column=1, value=year)
        worksheet.cell(row=year_row, column=2, value=year_counts.get(year, 0))

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Fifth Third Bank CBA Combined Report Generator (Redlining + Demographics)'
    )
    parser.add_argument('--input', type=str, 
                       help='Input CSV file from BigQuery (redlining or demographics)')
    parser.add_argument('--input2', type=str,
                       help='Second CSV file from BigQuery to merge (optional)')
    parser.add_argument('--output', type=str, 
                       default='Fifth_Third_CBA_Report.xlsx',
                       help='Output Excel file name')
    
    args = parser.parse_args()
    
    # Set up folder structure
    reports_dir = os.path.join(os.getcwd(), 'reports')
    data_dir = os.path.join(os.getcwd(), 'data')
    
    # Create folders if they don't exist
    os.makedirs(reports_dir, exist_ok=True)
    os.makedirs(data_dir, exist_ok=True)
    
    print("="*80)
    print("FIFTH THIRD BANK CBA COMBINED REPORT GENERATOR")
    print("Combines Redlining (Tract Demographics) + Borrower Demographics")
    print("="*80)
    print(f"Working Directory: {os.getcwd()}")
    print(f"Reports Folder: {reports_dir}")
    print(f"Data Folder: {data_dir}")
    
    # Find input file(s) - check both current directory and data folder
    if args.input:
        # Check if file exists as-is, or in data folder
        if os.path.exists(args.input):
            input_file = args.input
        elif os.path.exists(os.path.join(data_dir, args.input)):
            input_file = os.path.join(data_dir, args.input)
        elif os.path.exists(os.path.join(os.getcwd(), args.input)):
            input_file = os.path.join(os.getcwd(), args.input)
        else:
            print(f"ERROR: Input file not found: {args.input}")
            print(f"  Checked: {args.input}")
            print(f"  Checked: {os.path.join(data_dir, args.input)}")
            sys.exit(1)
    else:
        # Look for CSV files in current directory and data folder
        csv_files = glob.glob('bquxjob_*.csv') + glob.glob(os.path.join(data_dir, 'bquxjob_*.csv'))
        if csv_files:
            csv_files.sort(key=lambda x: os.path.getmtime(x) if os.path.exists(x) else 0, reverse=True)
            input_file = csv_files[0]
            print(f"Using most recent file: {input_file}")
        else:
            print("ERROR: No input CSV file found!")
            print(f"Expected: bquxjob_*.csv in current directory or {data_dir}")
            sys.exit(1)
    
    # Find second file if not specified
    second_file = None
    if args.input2:
        # Check if file exists as-is, or in data folder
        if os.path.exists(args.input2):
            second_file = args.input2
        elif os.path.exists(os.path.join(data_dir, args.input2)):
            second_file = os.path.join(data_dir, args.input2)
        else:
            print(f"WARNING: Second file not found: {args.input2}")
    else:
        # Try to find a second bquxjob file
        csv_files = glob.glob('bquxjob_*.csv') + glob.glob(os.path.join(data_dir, 'bquxjob_*.csv'))
        if len(csv_files) >= 2:
            csv_files.sort(key=lambda x: os.path.getmtime(x) if os.path.exists(x) else 0, reverse=True)
            # Use second most recent file if it's different from the first
            if csv_files[1] != input_file:
                second_file = csv_files[1]
                print(f"Auto-detected second file: {second_file}")
    
    # Set output file path to reports folder
    if args.output:
        if os.path.isabs(args.output):
            output_file = args.output
        else:
            output_file = os.path.join(reports_dir, args.output)
    else:
        # Default filename with timestamp
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(reports_dir, f'Fifth_Third_CBA_Report_{timestamp}.xlsx')
    
    # Load and process data
    df = load_data(input_file, second_file)
    df = add_calculated_metrics(df)
    
    # Create Excel workbook
    create_excel_workbook(df, output_file)
    
    print("\n" + "="*80)
    print("ANALYSIS COMPLETE!")
    print("="*80)
    print(f"\nOutput file: {output_file}")
    print(f"Saved to: {reports_dir}")
    print("\nThe Excel workbook contains:")
    print("  - One sheet per CBSA showing all metrics (redlining + demographics) by year")
    print("  - Methods sheet documenting filters, data sources, and methodology")
    print("\nFolder Structure:")
    print(f"  - Reports: {reports_dir}")
    print(f"  - Data: {data_dir}")
    print("\nThank you for using the NCRC CBA Banks Analysis tool.")

if __name__ == '__main__':
    main()
