"""
Fifth Third Bank Small Business Lending Report Generator
Matches format of HMDA CBA reports
"""

import sys
import os
import pandas as pd
import numpy as np
from pathlib import Path
import glob
import argparse

# Import the CBA report functions for Excel formatting
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
# Import numbered module using importlib
import importlib.util
_ft_path = os.path.join(os.path.dirname(__file__), "02_fifth_third_cba_report.py")
_spec = importlib.util.spec_from_file_location("ft_cba", _ft_path)
_ft_module = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_ft_module)
# Import the needed functions
from ft_cba import (
    create_excel_workbook,  # Reuse Excel creation with modifications
    two_proportion_z_test,
    SIGNIFICANCE_LEVEL
)

# SB Metrics Configuration
SB_METRICS = {
    'sb_loans': 'SB Loans',
    'lmict': 'LMICT%',
    'avg_sb_lmict_loan_amount': 'Avg SB LMICT Loan Amount',
    'loans_rev_under_1m': 'Loans Rev Under $1m%',
    'avg_loan_amt_rum_sb': 'Avg Loan Amt for RUM SB'
}

def calculate_sb_metrics(row):
    """
    Calculate shares and gaps for SB metrics
    Returns dict with all calculated metrics
    """
    results = {}
    
    subject_total = row.get('subject_total_sb_loans', 0)
    peer_total = row.get('peer_total_sb_loans', 0)
    
    # 1. SB Loans (total count) - compare counts directly
    subject_count = row.get('subject_total_sb_loans', 0)
    peer_count = row.get('peer_total_sb_loans', 0)
    
    # Gap in absolute terms (subject - peer)
    results['sb_loans_subject'] = subject_count
    results['sb_loans_peer'] = peer_count
    results['sb_loans_gap'] = subject_count - peer_count if (subject_total > 0 and peer_total > 0) else None
    
    # 2. LMICT% - percentage share
    subject_lmict = row.get('subject_lmict_loan_count', 0)
    peer_lmict = row.get('peer_lmict_loan_count', 0)
    
    subject_lmict_pct = (subject_lmict / subject_total * 100) if subject_total > 0 else None
    peer_lmict_pct = (peer_lmict / peer_total * 100) if peer_total > 0 else None
    lmict_gap = subject_lmict_pct - peer_lmict_pct if (subject_lmict_pct is not None and peer_lmict_pct is not None) else None
    
    # Statistical significance for LMICT%
    lmict_significant = False
    if subject_total > 0 and peer_total > 0:
        z_stat, p_value = two_proportion_z_test(
            subject_lmict, subject_total,
            peer_lmict, peer_total
        )
        lmict_significant = (p_value is not None and p_value < SIGNIFICANCE_LEVEL and lmict_gap is not None and lmict_gap < 0)
    
    results['lmict_subject'] = subject_lmict_pct
    results['lmict_peer'] = peer_lmict_pct
    results['lmict_gap'] = lmict_gap
    results['lmict_significant'] = lmict_significant
    
    # 3. Avg SB LMICT Loan Amount - dollar amount (not percentage)
    subject_avg_lmict = row.get('subject_avg_sb_lmict_loan_amount', None)
    peer_avg_lmict = row.get('peer_avg_sb_lmict_loan_amount', None)
    
    results['avg_sb_lmict_subject'] = subject_avg_lmict
    results['avg_sb_lmict_peer'] = peer_avg_lmict
    results['avg_sb_lmict_gap'] = (subject_avg_lmict - peer_avg_lmict) if (subject_avg_lmict is not None and peer_avg_lmict is not None) else None
    results['avg_sb_lmict_significant'] = False  # No stat test for dollar amounts
    
    # 4. Loans Rev Under $1m% - percentage share
    subject_rev = row.get('subject_loans_rev_under_1m', 0)
    peer_rev = row.get('peer_loans_rev_under_1m', 0)
    
    subject_rev_pct = (subject_rev / subject_total * 100) if subject_total > 0 else None
    peer_rev_pct = (peer_rev / peer_total * 100) if peer_total > 0 else None
    rev_gap = subject_rev_pct - peer_rev_pct if (subject_rev_pct is not None and peer_rev_pct is not None) else None
    
    # Statistical significance for Loans Rev Under $1m%
    rev_significant = False
    if subject_total > 0 and peer_total > 0:
        z_stat, p_value = two_proportion_z_test(
            subject_rev, subject_total,
            peer_rev, peer_total
        )
        rev_significant = (p_value is not None and p_value < SIGNIFICANCE_LEVEL and rev_gap is not None and rev_gap < 0)
    
    results['loans_rev_under_1m_subject'] = subject_rev_pct
    results['loans_rev_under_1m_peer'] = peer_rev_pct
    results['loans_rev_under_1m_gap'] = rev_gap
    results['loans_rev_under_1m_significant'] = rev_significant
    
    # 5. Avg Loan Amt for RUM SB - dollar amount (not percentage)
    subject_avg_rum = row.get('subject_avg_loan_amt_rum_sb', None)
    peer_avg_rum = row.get('peer_avg_loan_amt_rum_sb', None)
    
    results['avg_loan_amt_rum_sb_subject'] = subject_avg_rum
    results['avg_loan_amt_rum_sb_peer'] = peer_avg_rum
    results['avg_loan_amt_rum_sb_gap'] = (subject_avg_rum - peer_avg_rum) if (subject_avg_rum is not None and peer_avg_rum is not None) else None
    results['avg_loan_amt_rum_sb_significant'] = False  # No stat test for dollar amounts
    
    return results

def add_calculated_sb_metrics(df):
    """Add calculated shares and gaps for all SB metrics"""
    print("\nCalculating SB metrics...")
    print(f"  Input dataframe shape: {df.shape}")
    
    # Calculate metrics for each row
    results_list = df.apply(calculate_sb_metrics, axis=1)
    
    # Add results to dataframe
    for key in results_list.iloc[0].keys():
        df[key] = results_list.apply(lambda x: x.get(key))
    
    print(f"  Output dataframe shape: {df.shape}")
    return df

def create_sb_excel_workbook(df, output_file):
    """
    Create Excel workbook with SB data, matching HMDA report format
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Group by CBSA
    cbsas = sorted(df['cbsa_code'].unique())
    
    print(f"  Creating sheets for {len(cbsas)} CBSAs...")
    
    for cbsa_code in cbsas:
        cbsa_data = df[df['cbsa_code'] == cbsa_code].copy()
        cbsa_name = cbsa_data['cbsa_name'].iloc[0] if len(cbsa_data) > 0 else f"CBSA {cbsa_code}"
        
        # Clean CBSA name for sheet name
        sheet_name = cbsa_name.split(',')[0].strip()[:31]  # Excel sheet name limit
        
        print(f"  Creating sheet: {sheet_name} ({cbsa_code})")
        ws = wb.create_sheet(sheet_name)
        
        # Header
        ws.append([f'Fifth Third Bank - Small Business Lending Analysis'])
        ws.append([f'{cbsa_name} (CBSA {cbsa_code})'])
        ws.append([f'2018-2023'])
        ws.append([])
        
        # Get all years in data
        years = sorted([int(y) for y in cbsa_data['activity_year'].unique() if pd.notna(y)])
        
        # Create column headers: Metric, then for each year: Subject, Peer, Gap
        headers = ['Metric']
        for year in years:
            headers.extend([f'{year} Subject', f'{year} Peer', f'{year} Gap'])
        ws.append(headers)
        
        # Define metrics with their display names and data keys
        metrics = [
            {
                'name': 'SB Loans',
                'subject_key': 'sb_loans_subject',
                'peer_key': 'sb_loans_peer',
                'gap_key': 'sb_loans_gap',
                'format': 'count',
                'significant_key': None
            },
            {
                'name': 'LMICT%',
                'subject_key': 'lmict_subject',
                'peer_key': 'lmict_peer',
                'gap_key': 'lmict_gap',
                'format': 'percent',
                'significant_key': 'lmict_significant'
            },
            {
                'name': 'Avg SB LMICT Loan Amount',
                'subject_key': 'avg_sb_lmict_subject',
                'peer_key': 'avg_sb_lmict_peer',
                'gap_key': 'avg_sb_lmict_gap',
                'format': 'dollar',
                'significant_key': None
            },
            {
                'name': 'Loans Rev Under $1m%',
                'subject_key': 'loans_rev_under_1m_subject',
                'peer_key': 'loans_rev_under_1m_peer',
                'gap_key': 'loans_rev_under_1m_gap',
                'format': 'percent',
                'significant_key': 'loans_rev_under_1m_significant'
            },
            {
                'name': 'Avg Loan Amt RUM SB',
                'subject_key': 'avg_loan_amt_rum_sb_subject',
                'peer_key': 'avg_loan_amt_rum_sb_peer',
                'gap_key': 'avg_loan_amt_rum_sb_gap',
                'format': 'dollar',
                'significant_key': None
            }
        ]
        
        # Create row for each metric
        for metric in metrics:
            row = [metric['name']]
            gap_col_idx = 2  # Start after Metric column
            
            for year in years:
                year_data = cbsa_data[cbsa_data['activity_year'] == year]
                
                if len(year_data) > 0:
                    data = year_data.iloc[0]
                    
                    # Subject value
                    subject_val = data.get(metric['subject_key'])
                    if subject_val is not None:
                        if metric['format'] == 'percent':
                            row.append(f"{subject_val:.2f}%")
                        elif metric['format'] == 'dollar':
                            row.append(f"${subject_val:,.0f}")
                        else:  # count
                            row.append(int(subject_val) if pd.notna(subject_val) else '')
                    else:
                        row.append('')
                    
                    # Peer value
                    peer_val = data.get(metric['peer_key'])
                    if peer_val is not None:
                        if metric['format'] == 'percent':
                            row.append(f"{peer_val:.2f}%")
                        elif metric['format'] == 'dollar':
                            row.append(f"${peer_val:,.0f}")
                        else:  # count
                            row.append(int(peer_val) if pd.notna(peer_val) else '')
                    else:
                        row.append('')
                    
                    # Gap value
                    gap_val = data.get(metric['gap_key'])
                    is_significant = data.get(metric['significant_key'], False) if metric['significant_key'] else False
                    
                    if gap_val is not None:
                        if metric['format'] == 'percent':
                            gap_str = f"{gap_val:.2f}pp"
                            if is_significant and gap_val < 0:
                                gap_str = f"{gap_val:.2f}pp*"
                            row.append(gap_str)
                        elif metric['format'] == 'dollar':
                            row.append(f"${gap_val:,.0f}")
                        else:  # count
                            row.append(int(gap_val) if pd.notna(gap_val) else '')
                    else:
                        row.append('')
                    
                    gap_col_idx += 3  # Move to next year's columns
                else:
                    # No data for this year
                    row.extend(['', '', ''])
                    gap_col_idx += 3
            
            ws.append(row)
            
            # Apply color coding to gap columns for this metric row
            row_idx = ws.max_row
            for year_idx, year in enumerate(years):
                year_data = cbsa_data[cbsa_data['activity_year'] == year]
                
                if len(year_data) > 0:
                    data = year_data.iloc[0]
                    gap_val = data.get(metric['gap_key'])
                    is_significant = data.get(metric['significant_key'], False) if metric['significant_key'] else False
                    
                    # Gap column is 3 columns per year: Metric + (year_idx * 3) + 3 (Subject, Peer, Gap)
                    gap_col = 1 + (year_idx * 3) + 3
                    
                    if gap_val is not None and metric['format'] in ['percent']:
                        if gap_val > 0:
                            ws.cell(row=row_idx, column=gap_col).fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                        elif gap_val >= -2:
                            ws.cell(row=row_idx, column=gap_col).fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
                        elif gap_val >= -5:
                            ws.cell(row=row_idx, column=gap_col).fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
                        else:
                            fill_color = 'FF0000' if is_significant else 'FF6B6B'
                            ws.cell(row=row_idx, column=gap_col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                            if is_significant:
                                ws.cell(row=row_idx, column=gap_col).font = Font(bold=True, color='FFFFFF')
        
        # Summary section: Year-by-year loan counts and amounts
        ws.append([])
        ws.append(['SUMMARY - SUBJECT BANK LOAN COUNTS AND AMOUNTS'])
        ws.append([])
        
        # Create summary table headers
        summary_headers = ['Year', 'Category', 'Loan Count', 'Loan Amount']
        ws.append(summary_headers)
        
        # Format summary header row
        summary_header_row = ws.max_row
        for cell in ws[summary_header_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # For each year, show detailed breakdown
        for year in years:
            year_data = cbsa_data[cbsa_data['activity_year'] == year]
            
            if len(year_data) > 0:
                data = year_data.iloc[0]
                
                # Total SB Loans
                total_sb_loans = data.get('subject_total_sb_loans', 0)
                total_sb_amount = data.get('subject_total_sb_amount', None)
                
                # LMICT loans
                lmict_count = data.get('subject_lmict_loan_count', 0)
                lmict_amount = data.get('subject_lmict_loan_amount', None)
                
                # Revenue Under $1m loans
                rev_count = data.get('subject_loans_rev_under_1m', 0)
                rev_amount = data.get('subject_amount_rev_under_1m', None)
                
                # Write summary rows for this year
                ws.append([
                    year,
                    'Total SB Loans',
                    int(total_sb_loans) if pd.notna(total_sb_loans) and total_sb_loans > 0 else '',
                    f"${total_sb_amount:,.0f}" if total_sb_amount is not None and total_sb_amount > 0 else ''
                ])
                
                ws.append([
                    year,
                    'LMICT Loans',
                    int(lmict_count) if pd.notna(lmict_count) and lmict_count > 0 else '',
                    f"${lmict_amount:,.0f}" if lmict_amount is not None and lmict_amount > 0 else ''
                ])
                
                ws.append([
                    year,
                    'Revenue Under $1m Loans',
                    int(rev_count) if pd.notna(rev_count) and rev_count > 0 else '',
                    f"${rev_amount:,.0f}" if rev_amount is not None and rev_amount > 0 else ''
                ])
                
                ws.append([])  # Blank row between years
        
        # Format header row
        header_row = ws[5]  # Header row (row 5 after title rows)
        for cell in header_row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Create Methods sheet
    print("  Creating Methods sheet...")
    ws_methods = wb.create_sheet('Methods', 0)  # Insert at beginning
    
    methodology_text = """
FIFTH THIRD BANK SMALL BUSINESS LENDING ANALYSIS - METHODOLOGY

PURPOSE:
This report analyzes Fifth Third Bank's small business lending patterns compared to peer lenders
across key markets. Focuses on identifying potential disparities in lending to low-to-moderate 
income communities and small businesses.

GEOGRAPHIC SCOPE:
- Top 10 CBSAs (Core Based Statistical Areas) by Fifth Third Bank small business loan volume
- Years: 2018-2023
- All census tracts within selected CBSAs

DATA SOURCE:
- Community Reinvestment Act (CRA) Small Business Lending Data
- Federal Financial Institutions Examination Council (FFIEC)
- BigQuery Dataset: hdma1-242116.sb.disclosure and hdma1-242116.sb.lenders
- Years: 2018-2023

METRICS ANALYZED:

1. SB LOANS (Total Count):
   - Definition: Total number of small business loans originated
   - Calculation: num_under_100k + num_100k_250k + num_250k_1m
   - Represents loans ≤$1 million in original loan amount
   - Units: Count of loans

2. LMICT% (Low-to-Moderate Income Census Tract):
   - Definition: Percentage of loans in Low-to-Moderate Income Census Tracts
   - Calculation: (Loans in LMICT / Total SB Loans) × 100
   - LMICT Definition: Census tracts with income_group_total = '101', '102', or '1'-'8'
     - '101': Income relative to area median income ≤ 50%
     - '102': Income relative to area median income 50%-80%
     - '1'-'8': Additional LMI tract classifications
   - Units: Percentage

3. AVG SB LMICT LOAN AMOUNT:
   - Definition: Average loan amount for loans in Low-to-Moderate Income Census Tracts
   - Calculation: SUM(Amount of LMICT loans) / COUNT(LMICT loans)
   - Provides insight into loan size in underserved communities
   - Units: Dollars

4. LOANS REV UNDER $1m% (Revenue-Based Small Business Loans):
   - Definition: Percentage of loans to businesses with revenue under $1 million
   - Calculation: (numsbrev_under_1m / Total SB Loans) × 100
   - Identifies focus on very small businesses
   - Units: Percentage

5. AVG LOAN AMT FOR RUM SB (Revenue Under $1M Small Business):
   - Definition: Average loan amount for revenue-based small business loans under $1M
   - Calculation: amtsbrev_under_1m / numsbrev_under_1m
   - Indicates typical loan size for very small businesses
   - Units: Dollars

PEER COMPARISON METHODOLOGY:
- Peer Definition: Lenders with 50%-200% of Fifth Third Bank's small business loan volume 
  in each CBSA-year combination
- Peer Selection: Volume-based matching within each market and year
- Analysis excludes Fifth Third Bank from peer group
- Each CBSA-year has its own independent peer group
- Peer metrics are aggregated across all qualifying peer lenders

CALCULATION METHODS:
- Subject Share: (Fifth Third metric / Fifth Third total SB loans) × 100
- Peer Share: (Peer group metric / Peer group total SB loans) × 100
- Gap: Subject Share - Peer Share (percentage points)
- For dollar amounts (averages): Subject Average - Peer Average
- Positive gap means Fifth Third outperforms peers
- Negative gap means Fifth Third underperforms peers

STATISTICAL SIGNIFICANCE:
- Two-proportion z-test performed for percentage-based metrics (LMICT%, Loans Rev Under $1m%)
- Asterisk (*) indicates statistically significant negative gap (p < 0.05)
- Only negative gaps are flagged for significance
- Dollar amount metrics (averages) do not include statistical testing

COLOR CODING:
- Green: Positive gap (bank outperforms peers)
- Yellow: Small negative gap (0 to -2.0 percentage points)
- Orange: Moderate negative gap (-2.0 to -5.0 percentage points)
- Light Red: Large negative gap (-5.0 to -10.0 percentage points)
- Red: Severe negative gap (<-10.0 percentage points)
- Red with white bold text: Severe negative gap + statistically significant

DATA FILTERS:
- Small Business Loans: Loans ≤$1 million in original amount
- Revenue-Based Loans: Loans to businesses with revenue under $1 million
- Census Tract Income Classification: Based on FFIEC CRA income classifications
- Geographic Matching: CBSA codes matched from msamd field in disclosure table

LIMITATIONS:
- Does not account for differences in business models or market strategies
- Geographic selection biased toward high-volume markets
- Peer matching based solely on loan volume
- Does not include all factors affecting small business lending decisions
- Correlation does not imply causation
- Does not assess individual loan applications for discrimination
- LMICT classifications based on FFIEC CRA definitions, which may not capture all dimensions of need
- Revenue data only available for revenue-based loans, not all small business loans

INTERPRETATION GUIDANCE:
- Focus on patterns across years, not single-year gaps
- Statistically significant gaps (marked with *) are more reliable
- Large, persistent negative gaps warrant closer review
- Lower average loan amounts in LMICT may indicate focus on smaller businesses, which could be positive
- Higher percentage of loans to very small businesses (revenue <$1M) may indicate better outreach to underserved communities

CONTACT:
NCRC Research Department
National Community Reinvestment Coalition
For questions about methodology or data, contact: research@ncrc.org
"""
    
    # Write methodology text
    for i, line in enumerate(methodology_text.strip().split('\n'), 1):
        ws_methods.cell(row=i, column=1, value=line)
    
    # Format Methods sheet
    for row in ws_methods.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    ws_methods.column_dimensions['A'].width = 120
    
    wb.save(output_file)
    print(f"\nExcel workbook created: {output_file}")

def main():
    parser = argparse.ArgumentParser(description='Generate Fifth Third Small Business Lending Report')
    parser.add_argument('--input', help='Input CSV file from BigQuery')
    parser.add_argument('--output', help='Output Excel file path')
    
    args = parser.parse_args()
    
    base_dir = Path(__file__).parent.parent
    reports_dir = base_dir / "reports" / "fifth_third_merger"  # Same folder as other merger reports
    data_dir = base_dir / "data" / "raw"
    
    reports_dir.mkdir(parents=True, exist_ok=True)
    
    # Find input file
    if args.input:
        input_file = args.input if os.path.isabs(args.input) else data_dir / args.input
    else:
        # Auto-detect
        sb_files = list(data_dir.glob("*fifth*third*sb*.csv")) + list(data_dir.glob("*fifth*third*small*business*.csv"))
        if not sb_files:
            print(f"ERROR: No SB data file found in {data_dir}")
            sys.exit(1)
        input_file = sb_files[0]
        print(f"Auto-detected input file: {input_file}")
    
    # Load data
    print(f"\nLoading data from: {input_file}")
    df = pd.read_csv(input_file)
    print(f"  Loaded {len(df)} rows")
    
    # Calculate metrics
    df = add_calculated_sb_metrics(df)
    
    # Create Excel report
    if args.output:
        output_file = args.output
    else:
        output_file = reports_dir / "Fifth_Third_SB_Report.xlsx"
    
    print(f"\nCreating Excel workbook: {output_file}")
    create_sb_excel_workbook(df, str(output_file))
    
    print("\n" + "="*80)
    print("REPORT COMPLETE!")
    print("="*80)

if __name__ == "__main__":
    main()

