"""
Inspect sample goal setting analysis file to understand structure
"""
import pandas as pd
import openpyxl
from pathlib import Path

sample_file = r"C:\Users\edite\OneDrive - Nat'l Community Reinvestment Coaltn\Desktop\DREAM Analysis\scripts\The Huntington National Bank Research Ticket_CBA .xlsx"

print("="*80)
print("INSPECTING SAMPLE GOAL SETTING ANALYSIS FILE")
print("="*80)

# Load workbook
wb = openpyxl.load_workbook(sample_file, data_only=True)

print(f"\nFile: {Path(sample_file).name}")
print(f"\nNumber of sheets: {len(wb.sheetnames)}")
print(f"\nSheet names:")
for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {sheet_name}")

# Inspect each sheet
for sheet_name in wb.sheetnames:
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}")
    
    ws = wb[sheet_name]
    
    # Get dimensions
    print(f"\nDimensions: {ws.max_row} rows x {ws.max_column} columns")
    
    # Read cells directly with openpyxl
    print("\nFirst 20 rows with data:")
    row_count = 0
    for row_idx in range(1, min(21, ws.max_row + 1)):
        row_data = []
        has_data = False
        for col_idx in range(1, min(ws.max_column + 1, 20)):  # Limit to first 20 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            cell_value = cell.value
            if cell_value is not None:
                str_value = str(cell_value).strip()
                if str_value:
                    row_data.append(str_value[:40])  # Limit length
                    has_data = True
                else:
                    row_data.append("")
            else:
                row_data.append("")
        if has_data:
            print(f"Row {row_idx}: {' | '.join([x if x else '[empty]' for x in row_data[:10]])}")
            row_count += 1
            if row_count >= 20:
                break
    
    # Also try reading as DataFrame with openpyxl
    try:
        # Read without pandas Excel, use openpyxl directly
        from openpyxl import load_workbook
        wb2 = load_workbook(sample_file, data_only=True, read_only=True)
        ws2 = wb2[sheet_name]
        
        # Try to find header row and read structure
        print("\n\nTrying to identify data structure...")
        data_rows = []
        for row_idx in range(1, min(100, ws.max_row + 1)):
            row_values = []
            for col_idx in range(1, min(ws.max_column + 1, 30)):
                cell_value = ws2.cell(row=row_idx, column=col_idx).value
                row_values.append(cell_value)
            if any(v is not None for v in row_values):
                data_rows.append(row_values[:15])  # First 15 columns
                if len(data_rows) >= 30:
                    break
        
        if data_rows:
            print("\nSample rows (first 15 columns, showing None as [None]):")
            for idx, row in enumerate(data_rows[:15], 1):
                row_str = ' | '.join([str(v)[:30] if v is not None else '[None]' for v in row])
                print(f"  {idx}: {row_str}")
        
        wb2.close()
        
    except Exception as e:
        print(f"\nCould not read with read_only mode: {e}")
    
    # Read as DataFrame (may fail due to openpyxl version)
    try:
        df = pd.read_excel(sample_file, sheet_name=sheet_name, header=None, nrows=50, engine='openpyxl')
        
        # Show first few rows
        print(f"\n\nDataFrame view (first 15 rows):")
        print(df.head(15).to_string())
        
        # Look for header row (first row with text)
        header_row = None
        for idx, row in df.iterrows():
            if row.notna().sum() > 2:  # Row has more than 2 non-null values
                header_row = idx
                break
        
        if header_row is not None:
            print(f"\nLikely header row: {header_row}")
            # Try reading with header
            df_header = pd.read_excel(sample_file, sheet_name=sheet_name, header=header_row, nrows=30)
            print(f"\nData with header (first 30 rows):")
            print(df_header.head(30).to_string())
            
            print(f"\nColumn names:")
            for col in df_header.columns:
                print(f"  - {col}")
            
            print(f"\nData types:")
            print(df_header.dtypes)
            
            print(f"\nNon-null counts:")
            print(df_header.notna().sum())
        
    except Exception as e:
        print(f"Error reading sheet: {e}")
        # Try reading raw cells
        print("\nFirst 10 rows (raw):")
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(11, ws.max_column + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    row_data.append(str(cell_value)[:30])
                else:
                    row_data.append("")
            if any(row_data):  # Only print if row has data
                print(f"Row {row_idx}: {' | '.join(row_data)}")

print(f"\n{'='*80}")
print("INSPECTION COMPLETE")
print(f"{'='*80}")

