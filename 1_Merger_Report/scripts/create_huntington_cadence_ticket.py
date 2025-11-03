"""
Create Merger Research Ticket for Huntington + Cadence Merger

This script creates an Excel file with the correct structure for the
Goal Setting Analysis workflow.

Usage:
    python scripts/create_huntington_cadence_ticket.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
from pathlib import Path

def create_merger_ticket():
    """Create the merger research ticket Excel file"""
    
    # Bank information
    HUNTINGTON = {
        'name': 'The Huntington National Bank',
        'lei': '2WHM8VNJH63UN14OL754',
        'rssd': '12311',
        'sb_respondent_id': '7745'
    }
    
    CADENCE = {
        'name': 'Cadence Bank',
        'lei': 'Q7C315HKI8VX0SSKBS64',
        'rssd': '606046',
        'sb_respondent_id': '11813'
    }
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Sheet 1: Bank Details
    ws_bank = wb.create_sheet('Bank Details')
    
    # Header row (row 3)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    # Row 3: Headers
    ws_bank.cell(row=3, column=1, value='BANK INFORMATION')
    ws_bank.cell(row=3, column=2, value='ACQUIRER')
    ws_bank.cell(row=3, column=4, value='BANK INFORMATION')
    ws_bank.cell(row=3, column=5, value='TARGET')
    
    for col in [1, 2, 4, 5]:
        cell = ws_bank.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Row 4: BANK
    ws_bank.cell(row=4, column=2, value='BANK')
    ws_bank.cell(row=4, column=3, value=HUNTINGTON['name'])
    ws_bank.cell(row=4, column=5, value='BANK')
    ws_bank.cell(row=4, column=6, value=CADENCE['name'])
    
    # Row 5: LEI CODE
    ws_bank.cell(row=5, column=2, value='LEI CODE')
    ws_bank.cell(row=5, column=3, value=HUNTINGTON['lei'])
    ws_bank.cell(row=5, column=5, value='LEI CODE')
    ws_bank.cell(row=5, column=6, value=CADENCE['lei'])
    
    # Row 6: RSS-ID
    ws_bank.cell(row=6, column=2, value='RSS-ID')
    ws_bank.cell(row=6, column=3, value=HUNTINGTON['rssd'])
    ws_bank.cell(row=6, column=5, value='RSS-ID')
    ws_bank.cell(row=6, column=6, value=CADENCE['rssd'])
    
    # Row 7: SB RESPONDENT ID
    ws_bank.cell(row=7, column=2, value='SB RESPONDENT ID')
    ws_bank.cell(row=7, column=3, value=HUNTINGTON['sb_respondent_id'])
    ws_bank.cell(row=7, column=5, value='SB RESPONDENT ID')
    ws_bank.cell(row=7, column=6, value=CADENCE['sb_respondent_id'])
    
    # Format data cells
    data_font = Font(size=11)
    for row in range(4, 8):
        for col in [2, 3, 5, 6]:
            cell = ws_bank.cell(row=row, column=col)
            cell.font = data_font
    
    # Adjust column widths
    ws_bank.column_dimensions['A'].width = 20
    ws_bank.column_dimensions['B'].width = 18
    ws_bank.column_dimensions['C'].width = 30
    ws_bank.column_dimensions['D'].width = 5
    ws_bank.column_dimensions['E'].width = 18
    ws_bank.column_dimensions['F'].width = 30
    
    # Sheet 2: Assessment Areas (empty - to be filled manually)
    ws_aa = wb.create_sheet('Assessment Areas')
    
    # Headers
    headers = ['Bank Name', 'CBSA Name', 'CBSA Code', 'State Name', 'County Name', 'County Code (GEOID)']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_aa.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add example row
    ws_aa.cell(row=2, column=1, value=HUNTINGTON['name'])
    ws_aa.cell(row=2, column=2, value='[Example CBSA Name]')
    ws_aa.cell(row=2, column=3, value='[Example CBSA Code]')
    ws_aa.cell(row=2, column=4, value='[Example State]')
    ws_aa.cell(row=2, column=5, value='[Example County]')
    ws_aa.cell(row=2, column=6, value='[Example GEOID]')
    
    # Format example row (gray)
    example_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    for col in range(1, 7):
        cell = ws_aa.cell(row=2, column=col)
        cell.fill = example_fill
        cell.font = Font(size=10, italic=True)
    
    # Adjust column widths
    ws_aa.column_dimensions['A'].width = 30
    ws_aa.column_dimensions['B'].width = 30
    ws_aa.column_dimensions['C'].width = 15
    ws_aa.column_dimensions['D'].width = 15
    ws_aa.column_dimensions['E'].width = 25
    ws_aa.column_dimensions['F'].width = 20
    
    # Sheet 3: Notes (optional but helpful)
    ws_notes = wb.create_sheet('Notes')
    ws_notes.cell(row=1, column=1, value='MERGER RESEARCH TICKET - NOTES')
    ws_notes.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_notes.cell(row=3, column=1, value=f'Created: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    ws_notes.cell(row=5, column=1, value='INSTRUCTIONS:')
    ws_notes.cell(row=5, column=1).font = Font(bold=True)
    ws_notes.cell(row=6, column=1, value='1. Fill in Assessment Areas sheet with specific counties for each bank')
    ws_notes.cell(row=7, column=1, value='2. Each row should represent one county in a CBSA assessment area')
    ws_notes.cell(row=8, column=1, value='3. Bank Name should match exactly: "The Huntington National Bank" or "Cadence Bank"')
    ws_notes.cell(row=9, column=1, value='4. County Code (GEOID) will be mapped automatically if county name and state are correct')
    ws_notes.cell(row=11, column=1, value='After filling in assessment areas, run:')
    ws_notes.cell(row=12, column=1, value='python scripts/goal_setting_analysis_main.py "Huntington+Cadence merger research ticket.xlsx"')
    ws_notes.cell(row=12, column=1).font = Font(italic=True)
    
    # Adjust column width
    ws_notes.column_dimensions['A'].width = 100
    
    # Save file
    output_dir = Path.cwd()
    output_file = output_dir / 'Huntington+Cadence merger research ticket.xlsx'
    
    wb.save(str(output_file))
    
    print("="*80)
    print("MERGER RESEARCH TICKET CREATED")
    print("="*80)
    print(f"\nFile saved to: {output_file}")
    print("\nSheets created:")
    print("  1. Bank Details - Contains all bank information (LEI, RSSD, SB Respondent ID)")
    print("  2. Assessment Areas - Template for counties (fill this in manually)")
    print("  3. Notes - Instructions and metadata")
    print("\nNEXT STEPS:")
    print("  1. Fill in the 'Assessment Areas' sheet with the specific counties")
    print("     for each bank's CBSA assessment areas")
    print("  2. Each row = one county in one CBSA for one bank")
    print("  3. After filling in assessment areas, run:")
    print(f'     python scripts/goal_setting_analysis_main.py "{output_file.name}"')
    print("\nBank Information Included:")
    print(f"  Acquirer (Huntington):")
    print(f"    Name: {HUNTINGTON['name']}")
    print(f"    LEI: {HUNTINGTON['lei']}")
    print(f"    RSSD: {HUNTINGTON['rssd']}")
    print(f"    SB Respondent ID: {HUNTINGTON['sb_respondent_id']}")
    print(f"  Target (Cadence):")
    print(f"    Name: {CADENCE['name']}")
    print(f"    LEI: {CADENCE['lei']}")
    print(f"    RSSD: {CADENCE['rssd']}")
    print(f"    SB Respondent ID: {CADENCE['sb_respondent_id']}")
    print("="*80)

if __name__ == "__main__":
    create_merger_ticket()

