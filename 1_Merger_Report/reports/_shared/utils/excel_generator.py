"""
Excel generator for Goal-Setting Analysis.

Creates Excel workbook with all required sheets, preserving formulas
and populating with data from query results.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import pandas as pd
from typing import Dict, List, Optional
from copy import copy
import re


def wrap_formulas_with_iferror(wb: Workbook):
    """
    Process all sheets and wrap formulas containing division operations with IFERROR.
    This prevents #DIV/0! errors throughout the workbook.
    """
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Process each cell that contains a formula
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    formula = str(cell.value)
                    
                    # Check if formula contains division (/) but is not already wrapped in IFERROR
                    if '/' in formula and not formula.upper().startswith('=IFERROR'):
                        # Wrap the entire formula with IFERROR
                        # Handle formulas that already start with =
                        if formula.startswith('='):
                            # Remove the = and wrap with IFERROR
                            inner_formula = formula[1:]
                            # Determine appropriate default value based on formula type
                            # For percentage differences, use 0; for sums/ratios, use 0
                            default_value = '0'
                            
                            # For percentage formulas, the result should be a percentage
                            if '%' in str(cell.number_format).upper() or 'PERCENT' in str(cell.number_format).upper():
                                new_formula = f'=IFERROR({inner_formula},0)'
                            else:
                                new_formula = f'=IFERROR({inner_formula},0)'
                            
                            cell.value = new_formula
                        else:
                            # Formula doesn't start with =, wrap as-is
                            cell.value = f'=IFERROR({formula},0)'


def create_goal_setting_excel(
    output_path: Path,
    bank_a_name: str,
    bank_b_name: str,
    assessment_areas_data: Dict,
    mortgage_goals_data: Optional[Dict] = None,
    bank_a_mortgage_data: Optional[pd.DataFrame] = None,
    bank_b_mortgage_data: Optional[pd.DataFrame] = None,
    bank_a_mortgage_peer_data: Optional[pd.DataFrame] = None,
    bank_b_mortgage_peer_data: Optional[pd.DataFrame] = None,
    bank_a_sb_data: Optional[pd.DataFrame] = None,
    bank_b_sb_data: Optional[pd.DataFrame] = None,
    bank_a_sb_peer_data: Optional[pd.DataFrame] = None,
    bank_b_sb_peer_data: Optional[pd.DataFrame] = None,
    bank_a_branch_data: Optional[pd.DataFrame] = None,
    bank_b_branch_data: Optional[pd.DataFrame] = None,
    years_hmda: List[str] = None,
    years_sb: List[str] = None,
    template_file: Optional[Path] = None
):
    """
    Create complete Goal-Setting Analysis Excel workbook.
    
    IMPORTANT: If template_file is provided, loads the ENTIRE template and populates
    data into it. This preserves all formulas, formatting, and structure exactly.
    
    Args:
        output_path: Path to save Excel file
        bank_a_name: Acquirer bank name (e.g., "PNC Bank", "Bank A")
        bank_b_name: Target bank name (e.g., "FirstBank", "Bank B")
        assessment_areas_data: Dictionary with 'acquirer' and 'target' county lists
        template_file: Path to template Excel file (if provided, loads entire template)
    """
    
    # If template file is provided, load the ENTIRE template
    if template_file and template_file.exists():
        print(f"Loading ENTIRE template from: {template_file}")
        wb = load_workbook(template_file, data_only=False)
        
        # Rename sheets to match bank names (template may have generic "Bank A" and "Bank B")
        # Map template bank names to actual bank names
        template_to_actual = {}
        
        # Determine which template name corresponds to which bank
        # Template may have "Bank A"/"BankB" or specific bank names - map to actual bank names
        # Generic template keys: Bank A, BankB, Bank_A, etc.
        for sheet_name in wb.sheetnames:
            new_name = sheet_name
            renamed = False
            
            # Priority 1: Replace specific legacy names (PNC/FirstBank) first to avoid double replacements
            if 'PNC Bank' in sheet_name or 'PNCBank' in sheet_name:
                # Replace "PNC Bank" or "PNCBank" with bank_a_name
                new_name = sheet_name.replace('PNC Bank', bank_a_name).replace('PNCBank', bank_a_name)
                template_to_actual[sheet_name] = new_name
                renamed = True
            elif 'PNC' in sheet_name or 'Pnc' in sheet_name:
                # Replace standalone "PNC" with bank_a_name
                new_name = sheet_name.replace('PNC', bank_a_name).replace('Pnc', bank_a_name)
                template_to_actual[sheet_name] = new_name
                renamed = True
            elif 'FirstBank' in sheet_name or 'Firstbank' in sheet_name or 'First Bank' in sheet_name:
                # Replace "FirstBank" variations with bank_b_name
                new_name = sheet_name.replace('FirstBank', bank_b_name).replace('Firstbank', bank_b_name).replace('First Bank', bank_b_name)
                template_to_actual[sheet_name] = new_name
                renamed = True
            
            # Priority 2: Check for Bank A/B patterns (only if not already renamed)
            if not renamed:
                if any(pattern in sheet_name for pattern in ['Bank A', 'BankA', 'Bank_A', 'Acquirer']):
                    # Replace Bank A patterns with actual acquirer name
                    for pattern in ['Bank A', 'BankA', 'Bank_A', 'Acquirer']:
                        new_name = new_name.replace(pattern, bank_a_name)
                    template_to_actual[sheet_name] = new_name
                    renamed = True
                elif any(pattern in sheet_name for pattern in ['Bank B', 'BankB', 'Bank_B', 'Target']):
                    # Replace Bank B patterns with actual target name
                    for pattern in ['Bank B', 'BankB', 'Bank_B', 'Target']:
                        new_name = new_name.replace(pattern, bank_b_name)
                    template_to_actual[sheet_name] = new_name
                    renamed = True
            
            # If not renamed, keep original name
            if not renamed and sheet_name not in template_to_actual:
                template_to_actual[sheet_name] = sheet_name
        
        # Rename sheets
        for old_name, new_name in template_to_actual.items():
            if old_name in wb.sheetnames:
                wb[old_name].title = new_name
                print(f"  Renamed sheet: {old_name} -> {new_name}")
        
        print(f"  Template loaded - {len(wb.sheetnames)} sheets, all formulas and formatting preserved")
        
        # Wrap all formulas with IFERROR to prevent division by zero errors
        print("  Wrapping formulas with IFERROR to prevent #DIV/0! errors...")
        wrap_formulas_with_iferror(wb)
        
    else:
        # Create new workbook if no template
        print("WARNING: No template file provided - creating workbook from scratch")
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create all sheets in order (fallback - not ideal)
        print("Creating Excel workbook from scratch...")
        create_notes_sheet(wb, bank_a_name, bank_b_name, years_hmda or [], years_sb or [])
        create_assessment_areas_sheet(wb, bank_a_name, bank_b_name, assessment_areas_data)
        create_mortgage_goals_sheet(wb, bank_a_name, bank_b_name, mortgage_goals_data, template_file)
        create_mortgage_data_sheet(wb, bank_a_name, bank_a_mortgage_data, bank_a_mortgage_peer_data, years_hmda or [])
        create_mortgage_data_sheet(wb, bank_b_name, bank_b_mortgage_data, bank_b_mortgage_peer_data, years_hmda or [])
        create_sb_data_sheet(wb, bank_a_name, bank_a_sb_data, bank_a_sb_peer_data, years_sb or [])
        create_sb_data_sheet(wb, bank_b_name, bank_b_sb_data, bank_b_sb_peer_data, years_sb or [])
        create_branch_data_sheet(wb, bank_a_name, bank_a_branch_data)
        create_branch_data_sheet(wb, bank_b_name, bank_b_branch_data)
    
    # Now populate all sheets with data (this preserves formulas)
    print("\nPopulating sheets with data...")
    
    # 1. Update header rows with actual bank names
    update_header_rows(wb, bank_a_name, bank_b_name)
    
    # 2. Populate Assessment Areas sheet
    populate_assessment_areas_sheet(wb, bank_a_name, bank_b_name, assessment_areas_data)
    
    # 3. Populate Mortgage Goals sheet
    if 'Mortgage Goals' in wb.sheetnames:
        ws = wb['Mortgage Goals']
        if mortgage_goals_data:
            insert_mortgage_goals_data(ws, mortgage_goals_data)
    
    # 4-5. Populate Mortgage Data sheets
    bank_a_mortgage_sheet = f"{bank_a_name}Mortgage Data"
    bank_b_mortgage_sheet = f"{bank_b_name}Mortgage Data"
    
    if bank_a_mortgage_sheet in wb.sheetnames:
        populate_mortgage_data_sheet(
            wb[bank_a_mortgage_sheet], bank_a_name, bank_a_mortgage_data, bank_a_mortgage_peer_data
        )
    
    if bank_b_mortgage_sheet in wb.sheetnames:
        populate_mortgage_data_sheet(
            wb[bank_b_mortgage_sheet], bank_b_name, bank_b_mortgage_data, bank_b_mortgage_peer_data
        )
    
    # 6-7. Populate Small Business sheets
    bank_a_sb_sheet = f"{bank_a_name} SB Lending"
    bank_b_sb_sheet = f"{bank_b_name} SB Lending"
    
    if bank_a_sb_sheet in wb.sheetnames:
        populate_sb_data_sheet(
            wb[bank_a_sb_sheet], bank_a_name, bank_a_sb_data, bank_a_sb_peer_data
        )
    
    if bank_b_sb_sheet in wb.sheetnames:
        populate_sb_data_sheet(
            wb[bank_b_sb_sheet], bank_b_name, bank_b_sb_data, bank_b_sb_peer_data
        )
    
    # 8-9. Populate Branch sheets
    bank_a_branch_sheet = f"{bank_a_name} Branches"
    bank_b_branch_sheet = f"{bank_b_name} Branches"
    
    if bank_a_branch_sheet in wb.sheetnames:
        populate_branch_data_sheet(wb[bank_a_branch_sheet], bank_a_name, bank_a_branch_data)
    
    if bank_b_branch_sheet in wb.sheetnames:
        populate_branch_data_sheet(wb[bank_b_branch_sheet], bank_b_name, bank_b_branch_data)
    
    # Validate workbook has data before saving
    validation_results = validate_excel_data(wb, bank_a_name, bank_b_name, 
                                            mortgage_goals_data, 
                                            bank_a_mortgage_data, bank_b_mortgage_data,
                                            bank_a_sb_data, bank_b_sb_data,
                                            bank_a_branch_data, bank_b_branch_data)
    
    if not validation_results['is_valid']:
        print("\n" + "="*80)
        print("WARNING: EXCEL VALIDATION FAILED")
        print("="*80)
        for warning in validation_results['warnings']:
            print(f"  - {warning}")
        print("\nThe file will still be saved, but please review the warnings above.")
        print("You may need to manually verify the data in the Excel file.")
    
    # Save workbook
    wb.save(output_path)
    print(f"\nExcel workbook saved to: {output_path}")
    
    return wb


def update_header_rows(wb: Workbook, bank_a_name: str, bank_b_name: str):
    """Update header rows (row 1) in all sheets to replace generic bank names with actual bank names."""
    
    # Patterns to replace in headers (longest first for proper replacement)
    replacements = [
        ('@PNC Bank AA', f'@{bank_a_name} AA'),
        ('@PNCBank AA', f'@{bank_a_name} AA'),
        ('@PNC AA', f'@{bank_a_name} AA'),
        ('@Pnc AA', f'@{bank_a_name} AA'),
        ('PNC Bank', bank_a_name),
        ('PNCBank', bank_a_name),
        ('PNC', bank_a_name),
        ('Pnc', bank_a_name),
        ('@FirstBank AA', f'@{bank_b_name} AA'),
        ('@Firstbank AA', f'@{bank_b_name} AA'),
        ('@First Bank AA', f'@{bank_b_name} AA'),
        ('FirstBank', bank_b_name),
        ('Firstbank', bank_b_name),
        ('First Bank', bank_b_name),
        ('Bank A', bank_a_name),
        ('BankA', bank_a_name),
        ('Bank_A', bank_a_name),
        ('Acquirer', bank_a_name),
        ('Bank B', bank_b_name),
        ('BankB', bank_b_name),
        ('Bank_B', bank_b_name),
        ('Target', bank_b_name)
    ]
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Update row 1 (header row)
        for col_idx in range(1, min(ws.max_column + 1, 20)):  # Limit to first 20 columns for performance
            cell = ws.cell(row=1, column=col_idx)
            if cell.value:
                cell_value = str(cell.value)
                updated_value = cell_value
                
                # Replace bank names in header (try longest matches first, only replace once per cell)
                # Track what's been replaced to avoid nested replacements
                replaced = False
                for old_name, new_name in replacements:
                    if not replaced and old_name in updated_value:
                        updated_value = updated_value.replace(old_name, new_name)
                        replaced = True  # Only replace once to avoid nested replacements
                
                if updated_value != cell_value:
                    cell.value = updated_value


def populate_assessment_areas_sheet(wb: Workbook, bank_a_name: str, bank_b_name: str,
                                    assessment_areas_data: Dict):
    """Populate Assessment Areas sheet with county data.
    
    Template structure:
    - Columns A-C: Bank A data (we'll use: Bank Name, CBSA Name, County Name)
    - Columns E-G: Bank B data (we'll use: Bank Name, CBSA Name, County Name)
    - Note: State column removed - only CBSA and county information
    """
    if 'Assessment Areas' not in wb.sheetnames:
        print("  WARNING: Assessment Areas sheet not found in template")
        return
    
    ws = wb['Assessment Areas']
    
    # Clear existing data but keep header row (row 1)
    max_row = ws.max_row
    if max_row > 1:
        ws.delete_rows(2, max_row - 1)
    
    # Update header row to match our structure (Bank Name, CBSA Name, County Name, no state)
    # Column A header: Bank Name (or keep existing if it has bank name)
    # Column B header: CBSA Name (or @BankName AA)
    # Column C header: County Name (not "County State")
    ws.cell(row=1, column=1, value='Bank Name')
    ws.cell(row=1, column=2, value='CBSA Name')
    ws.cell(row=1, column=3, value='CBSA Code')
    ws.cell(row=1, column=4, value='County Name')
    ws.cell(row=1, column=5, value='County Code (GEOID)')
    # Column E-G headers for Bank B (if template has separate columns)
    ws.cell(row=1, column=6, value=None)  # Clear old Bank B column E
    ws.cell(row=1, column=7, value=None)  # Clear old Bank B column F
    ws.cell(row=1, column=8, value=None)  # Clear old Bank B column G
    
    row_num = 2
    
    # Bank A counties (acquirer) - populate columns A-E
    bank_a_key = 'acquirer' if 'acquirer' in assessment_areas_data else 'pnc'
    counties_a = assessment_areas_data.get(bank_a_key, {}).get('counties', assessment_areas_data.get(bank_a_key, []))
    
    for county in counties_a:
        # Ensure all fields have values (no None)
        cbsa_name = county.get('cbsa_name', '') or ''
        cbsa_code = str(county.get('cbsa_code', '')) if county.get('cbsa_code') else ''
        county_name = county.get('county_name', '') or ''
        county_code = str(county.get('county_code', county.get('geoid5', ''))) if county.get('county_code') or county.get('geoid5') else ''
        
        # Column A: Bank Name (always fill)
        ws.cell(row=row_num, column=1, value=bank_a_name)
        # Column B: CBSA Name (always fill)
        ws.cell(row=row_num, column=2, value=cbsa_name)
        # Column C: CBSA Code (always fill)
        ws.cell(row=row_num, column=3, value=cbsa_code)
        # Column D: County Name (always fill)
        ws.cell(row=row_num, column=4, value=county_name)
        # Column E: County Code (always fill)
        ws.cell(row=row_num, column=5, value=county_code)
        row_num += 1
    
    # Bank B counties (target) - populate in same columns, continuing from Bank A
    # All data goes in same columns (A-E), differentiated by Bank Name in column A
    bank_b_key = 'target' if 'target' in assessment_areas_data else 'firstbank'
    counties_b = assessment_areas_data.get(bank_b_key, {}).get('counties', assessment_areas_data.get(bank_b_key, []))
    
    for county in counties_b:
        # Ensure all fields have values (no None)
        cbsa_name = county.get('cbsa_name', '') or ''
        cbsa_code = str(county.get('cbsa_code', '')) if county.get('cbsa_code') else ''
        county_name = county.get('county_name', '') or ''
        county_code = str(county.get('county_code', county.get('geoid5', ''))) if county.get('county_code') or county.get('geoid5') else ''
        
        # Same structure as Bank A - all in columns A-E
        # Column A: Bank Name (always fill)
        ws.cell(row=row_num, column=1, value=bank_b_name)
        # Column B: CBSA Name (always fill)
        ws.cell(row=row_num, column=2, value=cbsa_name)
        # Column C: CBSA Code (always fill)
        ws.cell(row=row_num, column=3, value=cbsa_code)
        # Column D: County Name (always fill)
        ws.cell(row=row_num, column=4, value=county_name)
        # Column E: County Code (always fill)
        ws.cell(row=row_num, column=5, value=county_code)
        row_num += 1
    
    print(f"  Populated: Assessment Areas sheet ({len(counties_a)} + {len(counties_b)} counties, no state columns)")


def populate_mortgage_data_sheet(ws, bank_name: str,
                                subject_data: Optional[pd.DataFrame],
                                peer_data: Optional[pd.DataFrame]):
    """
    Populate Mortgage Data sheet with data (preserves template formulas).
    
    Template structure:
    - Row 2: Grand Total row
    - Row 3+: Data rows with:
      - Column A: State (first row of state only)
      - Column B: CBSA (first row of CBSA only, on "Loans" metric)
      - Column C: Metrics (all rows)
      - Column D: Subject bank data
      - Column E: Peer data
      - Column F: Difference formulas
    """
    if subject_data is None or subject_data.empty:
        print(f"    No subject data for {bank_name}Mortgage Data")
        return
    
    # Find starting row (should be row 3, after Grand Total)
    start_row = 3
    
    # Metrics in template order (template has "Loans" at the END, not the beginning!)
    # Template order: LMICT%, LMIB%, LMIB$, MMCT%, MINB%, Asian%, Black%, Native American%, HoPI%, Hispanic%, Loans
    metrics = [
        'LMICT%',
        'LMIB%',
        'LMIB$',  # Dollar amount on row 5 of each CBSA group
        'MMCT%',
        'MINB%',
        'Asian%',
        'Black%',
        'Native American%',
        'HoPI%',
        'Hispanic%',
        'Loans'  # Loans is LAST in the template, not first
    ]
    
    # Group data by CBSA only (no state grouping)
    grouped = subject_data.groupby('cbsa_code')
    
    current_row = start_row
    
    for group_key, group_data in grouped:
        cbsa_code = group_key
        
        # Aggregate this CBSA's data
        cbsa_subject_agg = group_data.agg({
            'total_loans': 'sum',
            'lmict_loans': 'sum',
            'lmib_loans': 'sum',
            'lmib_amount': 'sum',
            'mmct_loans': 'sum',
            'minb_loans': 'sum',
            'asian_loans': 'sum',
            'black_loans': 'sum',
            'native_american_loans': 'sum',
            'hopi_loans': 'sum',
            'hispanic_loans': 'sum'
        }).to_dict()
        
        total_loans = cbsa_subject_agg.get('total_loans', 0)
        if total_loans > 0:
            cbsa_subject_agg['lmict_percentage'] = (cbsa_subject_agg.get('lmict_loans', 0) / total_loans) * 100
            cbsa_subject_agg['lmib_percentage'] = (cbsa_subject_agg.get('lmib_loans', 0) / total_loans) * 100
            cbsa_subject_agg['mmct_percentage'] = (cbsa_subject_agg.get('mmct_loans', 0) / total_loans) * 100
            cbsa_subject_agg['minb_percentage'] = (cbsa_subject_agg.get('minb_loans', 0) / total_loans) * 100
            cbsa_subject_agg['asian_percentage'] = (cbsa_subject_agg.get('asian_loans', 0) / total_loans) * 100
            cbsa_subject_agg['black_percentage'] = (cbsa_subject_agg.get('black_loans', 0) / total_loans) * 100
            cbsa_subject_agg['native_american_percentage'] = (cbsa_subject_agg.get('native_american_loans', 0) / total_loans) * 100
            cbsa_subject_agg['hopi_percentage'] = (cbsa_subject_agg.get('hopi_loans', 0) / total_loans) * 100
            cbsa_subject_agg['hispanic_percentage'] = (cbsa_subject_agg.get('hispanic_loans', 0) / total_loans) * 100
        
        # Get peer data
        cbsa_peer_agg = None
        if peer_data is not None and not peer_data.empty:
            peer_group = peer_data[peer_data['cbsa_code'] == cbsa_code]
            
            if not peer_group.empty:
                cbsa_peer_agg = peer_group.agg({
                    'total_loans': 'sum',
                    'lmict_loans': 'sum',
                    'lmib_loans': 'sum',
                    'lmib_amount': 'sum',
                    'mmct_loans': 'sum',
                    'minb_loans': 'sum',
                    'asian_loans': 'sum',
                    'black_loans': 'sum',
                    'native_american_loans': 'sum',
                    'hopi_loans': 'sum',
                    'hispanic_loans': 'sum'
                }).to_dict()
                
                peer_total = cbsa_peer_agg.get('total_loans', 0)
                if peer_total > 0:
                    cbsa_peer_agg['lmict_percentage'] = (cbsa_peer_agg.get('lmict_loans', 0) / peer_total) * 100
                    cbsa_peer_agg['lmib_percentage'] = (cbsa_peer_agg.get('lmib_loans', 0) / peer_total) * 100
                    cbsa_peer_agg['mmct_percentage'] = (cbsa_peer_agg.get('mmct_loans', 0) / peer_total) * 100
                    cbsa_peer_agg['minb_percentage'] = (cbsa_peer_agg.get('minb_loans', 0) / peer_total) * 100
                    cbsa_peer_agg['asian_percentage'] = (cbsa_peer_agg.get('asian_loans', 0) / peer_total) * 100
                    cbsa_peer_agg['black_percentage'] = (cbsa_peer_agg.get('black_loans', 0) / peer_total) * 100
                    cbsa_peer_agg['native_american_percentage'] = (cbsa_peer_agg.get('native_american_loans', 0) / peer_total) * 100
                    cbsa_peer_agg['hopi_percentage'] = (cbsa_peer_agg.get('hopi_loans', 0) / peer_total) * 100
                    cbsa_peer_agg['hispanic_percentage'] = (cbsa_peer_agg.get('hispanic_loans', 0) / peer_total) * 100
        
        # Get CBSA name (from subject_data or peer_data if available)
        cbsa_name = cbsa_code
        if 'cbsa_name' in group_data.columns:
            cbsa_name = group_data['cbsa_name'].iloc[0] if not group_data.empty else cbsa_code
        
        # Helper function to safely set cell value (handles merged cells)
        def safe_set_cell(row, col, value):
            """Set cell value, handling merged cells"""
            cell = ws.cell(row=row, column=col)
            # Check if cell is part of a merged range
            if hasattr(cell, 'coordinate'):
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Write to the top-left cell of the merged range
                        top_left = merged_range.min_row, merged_range.min_col
                        ws.cell(row=top_left[0], column=top_left[1]).value = value
                        return
            # Not merged, write directly
            cell.value = value
        
        # Insert data for each metric
        for metric_idx, metric in enumerate(metrics):
            # Column A: CBSA (only on "Loans" row, which is the LAST metric in template)
            # Note: State column removed - only CBSA in column A
            if metric == 'Loans':  # Loans is the last metric in the template - CBSA goes here
                safe_set_cell(current_row, 1, cbsa_name)
            
            # Column C: Metric name (preserve if already exists, otherwise insert)
            existing_metric = ws.cell(row=current_row, column=3).value
            if not existing_metric:
                safe_set_cell(current_row, 3, metric)
            
            # Column D: Subject bank data
            subject_row = pd.Series(cbsa_subject_agg)
            insert_mortgage_metric_data(ws, current_row, metric, subject_row, start_col=4)
            
            # Column E: Peer data
            if cbsa_peer_agg is not None:
                peer_row = pd.Series(cbsa_peer_agg)
                insert_mortgage_metric_data(ws, current_row, metric, peer_row, start_col=5)
            
            # Column F: Difference (preserve formula if exists, otherwise create)
            # Note: LMIB$ should be blank - we don't compare dollar amounts
            existing_formula = ws.cell(row=current_row, column=6)
            if existing_formula.data_type != 'f':
                if metric.endswith('%') and metric != 'LMIB$':  # Only percentages, NOT LMIB$
                    # Wrap with IFERROR to handle any division by zero errors
                    existing_formula.value = f'=IFERROR(D{current_row}-E{current_row},0)'
                elif metric == 'LMIB$':
                    # Leave blank - no difference calculation for LMIB dollar amounts
                    existing_formula.value = None
            
            current_row += 1
    
    print(f"    Populated: {bank_name}Mortgage Data ({current_row - start_row} rows)")


def populate_sb_data_sheet(ws, bank_name: str,
                          subject_data: Optional[pd.DataFrame],
                          peer_data: Optional[pd.DataFrame]):
    """Populate Small Business Data sheet (preserves template formulas)"""
    if subject_data is None or subject_data.empty:
        print(f"    No SB data for {bank_name} SB Lending")
        return
    
    start_row = 3  # After Grand Total row
    
    metrics = [
        'SB Loans',
        '#LMICT',
        'Avg SB LMICT Loan Amount',
        'Loans Rev Under $1m',
        'Avg Loan Amt for RUM SB'
    ]
    
    # Group by CBSA only (no state grouping)
    grouped = subject_data.groupby('cbsa_code')
    
    current_row = start_row
    
    for group_key, group_data in grouped:
        cbsa_code = group_key
        
        # Aggregate SB metrics
        cbsa_subject_agg = {
            'sb_loans_total': group_data['sb_loans_total'].sum(),
            'lmict_count': group_data['lmict_count'].sum(),
            'loans_rev_under_1m_count': group_data['loans_rev_under_1m_count'].sum(),
            'avg_sb_lmict_loan_amount': group_data['avg_sb_lmict_loan_amount'].mean() if group_data['lmict_count'].sum() > 0 else 0,
            'avg_loan_amt_rum_sb': group_data['avg_loan_amt_rum_sb'].mean() if group_data['loans_rev_under_1m_count'].sum() > 0 else 0
        }
        
        # Get peer data
        cbsa_peer_agg = None
        if peer_data is not None and not peer_data.empty:
            peer_group = peer_data[peer_data['cbsa_code'] == cbsa_code]
            
            if not peer_group.empty:
                cbsa_peer_agg = {
                    'sb_loans_total': peer_group['sb_loans_total'].sum(),
                    'lmict_count': peer_group['lmict_count'].sum(),
                    'loans_rev_under_1m_count': peer_group['loans_rev_under_1m_count'].sum(),
                    'avg_sb_lmict_loan_amount': peer_group['avg_sb_lmict_loan_amount'].mean() if peer_group['lmict_count'].sum() > 0 else 0,
                    'avg_loan_amt_rum_sb': peer_group['avg_loan_amt_rum_sb'].mean() if peer_group['loans_rev_under_1m_count'].sum() > 0 else 0
                }
        
        cbsa_name = cbsa_code
        if 'cbsa_name' in group_data.columns:
            cbsa_name = group_data['cbsa_name'].iloc[0] if not group_data.empty else cbsa_code
        
        # Insert data for each metric
        for metric_idx, metric in enumerate(metrics):
            # Helper function to safely set cell value (handles merged cells)
            def safe_set_cell(row, col, value):
                """Set cell value, handling merged cells"""
                cell = ws.cell(row=row, column=col)
                # Check if cell is part of a merged range
                if hasattr(cell, 'coordinate'):
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            # Write to the top-left cell of the merged range
                            top_left = merged_range.min_row, merged_range.min_col
                            ws.cell(row=top_left[0], column=top_left[1]).value = value
                            return
                # Not merged, write directly
                cell.value = value
            
            # Column A: CBSA (only on first metric)
            # Note: State column removed - only CBSA in column A
            if metric_idx == 0:  # SB Loans metric (first metric)
                safe_set_cell(current_row, 1, cbsa_name)
            
            existing_metric = ws.cell(row=current_row, column=3).value
            if not existing_metric:
                safe_set_cell(current_row, 3, metric)
            
            # Column D: Subject data
            subject_row = pd.Series(cbsa_subject_agg)
            insert_sb_metric_data(ws, current_row, metric, subject_row, col=4)
            
            # Column E: Peer data
            if cbsa_peer_agg is not None:
                peer_row = pd.Series(cbsa_peer_agg)
                insert_sb_metric_data(ws, current_row, metric, peer_row, col=5)
            
            # Column F: Difference formula
            # Wrap all formulas with IFERROR to handle division by zero errors
            existing_formula = ws.cell(row=current_row, column=6)
            if existing_formula.data_type != 'f':
                if metric == '#LMICT':
                    # Percentage difference: (LMICT_count/SB_Loans) - (Peer_LMICT_count/Peer_SB_Loans)
                    existing_formula.value = f'=IFERROR((D{current_row}/D{current_row-1})-(E{current_row}/E{current_row-1}),0)'
                elif metric == 'Avg SB LMICT Loan Amount':
                    # Ratio: (Subject_Avg/Peer_Avg) - 1
                    existing_formula.value = f'=IFERROR((D{current_row}/E{current_row})-1,0)'
                elif metric == 'Loans Rev Under $1m':
                    # Percentage difference: (Rev_Under_1m/SB_Loans) - (Peer_Rev_Under_1m/Peer_SB_Loans)
                    existing_formula.value = f'=IFERROR((D{current_row}/D{current_row-2})-(E{current_row}/E{current_row-2}),0)'
                elif metric == 'Avg Loan Amt for RUM SB':
                    # Ratio: (Subject_Avg/Peer_Avg) - 1
                    existing_formula.value = f'=IFERROR((D{current_row}/E{current_row})-1,0)'
            
            current_row += 1
    
    print(f"    Populated: {bank_name} SB Lending ({current_row - start_row} rows)")


def populate_branch_data_sheet(ws, bank_name: str,
                               branch_data: Optional[pd.DataFrame]):
    """Populate Branch Data sheet (preserves template formulas)"""
    if branch_data is None or branch_data.empty:
        print(f"    No branch data for {bank_name} Branches")
        return
    
    start_row = 3  # After Grand Total row
    
    # IMPORTANT: Template order is LMICT, MMCT, Branches (NOT Branches, LMICT, MMCT!)
    metrics = ['LMICT', 'MMCT', 'Branches']
    
    # Group by CBSA only (no state grouping)
    grouped = branch_data.groupby('cbsa_code')
    
    current_row = start_row
    
    for group_key, group_data in grouped:
        cbsa_code = group_key
        
        # Branch metrics (using actual column names from query)
        # For percentages, we need to recalculate from totals, not average percentages
        subject_total = group_data['subject_total_branches'].sum() if 'subject_total_branches' in group_data.columns else 0
        other_total = group_data['market_total_branches'].sum() if 'market_total_branches' in group_data.columns else 0
        
        cbsa_branch_agg = {
            'subject_total_branches': subject_total,
            'other_total_branches': other_total,
            # Calculate percentages from weighted average if needed, or use first value if single row
            'subject_lmi_pct': group_data['subject_pct_lmict'].iloc[0] if len(group_data) > 0 and 'subject_pct_lmict' in group_data.columns else 0,
            'subject_mm_pct': group_data['subject_pct_mmct'].iloc[0] if len(group_data) > 0 and 'subject_pct_mmct' in group_data.columns else 0,
            'other_lmi_pct': group_data['market_pct_lmict'].iloc[0] if len(group_data) > 0 and 'market_pct_lmict' in group_data.columns else 0,
            'other_mm_pct': group_data['market_pct_mmct'].iloc[0] if len(group_data) > 0 and 'market_pct_mmct' in group_data.columns else 0
        }
        
        cbsa_name = cbsa_code
        if 'cbsa_name' in group_data.columns:
            cbsa_name = group_data['cbsa_name'].iloc[0] if not group_data.empty else cbsa_code
        
        for metric_idx, metric in enumerate(metrics):
            # Helper function to safely set cell value (handles merged cells)
            def safe_set_cell(row, col, value):
                """Set cell value, handling merged cells"""
                cell = ws.cell(row=row, column=col)
                # Check if cell is part of a merged range
                if hasattr(cell, 'coordinate'):
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            # Write to the top-left cell of the merged range
                            top_left = merged_range.min_row, merged_range.min_col
                            ws.cell(row=top_left[0], column=top_left[1]).value = value
                            return
                # Not merged, write directly
                cell.value = value
            
            # Column A: CBSA (only on Branches metric)
            # Note: State column removed - only CBSA in column A
            if metric_idx == 2:  # Branches metric (3rd in list: LMICT=0, MMCT=1, Branches=2)
                safe_set_cell(current_row, 1, cbsa_name)
            
            # Find the correct row for this metric in the template
            # Template order: LMICT (idx 0), MMCT (idx 1), Branches (idx 2)
            # So: metric_idx 0=LMICT, 1=MMCT, 2=Branches
            existing_metric = ws.cell(row=current_row, column=3).value
            
            # Only write metric name if cell is empty (preserve template)
            if not existing_metric:
                safe_set_cell(current_row, 3, metric)
            
            # Column D: Subject data - FORCE correct format
            cell_d = ws.cell(row=current_row, column=4)
            # Clear everything first
            cell_d.value = None
            if cell_d.data_type == 'f':
                cell_d.value = None
            
            if metric == 'Branches':
                branch_count = int(cbsa_branch_agg['subject_total_branches'])
                cell_d.data_type = 'n'
                cell_d.value = branch_count
                cell_d.number_format = '0'  # Integer format (no decimals)
            elif metric == 'LMICT':
                # LMICT: Percentage of branches in LMICT
                pct_lmict = float(cbsa_branch_agg['subject_lmi_pct']) if cbsa_branch_agg['subject_lmi_pct'] else 0.0
                pct_value = pct_lmict / 100.0  # Convert 0-100 to 0-1 for Excel
                cell_d.data_type = 'n'
                cell_d.value = pct_value
                cell_d.number_format = '0.00%'  # Percentage format
            elif metric == 'MMCT':
                # MMCT: Percentage of branches in MMCT
                pct_mmct = float(cbsa_branch_agg['subject_mm_pct']) if cbsa_branch_agg['subject_mm_pct'] else 0.0
                pct_value = pct_mmct / 100.0  # Convert 0-100 to 0-1 for Excel
                cell_d.data_type = 'n'
                cell_d.value = pct_value
                cell_d.number_format = '0.00%'  # Percentage format
            
            # Column E: Other/Market data - FORCE correct format
            cell_e = ws.cell(row=current_row, column=5)
            # Clear everything first
            cell_e.value = None
            if cell_e.data_type == 'f':
                cell_e.value = None
            
            if metric == 'Branches':
                branch_count = int(cbsa_branch_agg['other_total_branches'])
                cell_e.data_type = 'n'
                cell_e.value = branch_count
                cell_e.number_format = '0'  # Integer format (no decimals)
            elif metric == 'LMICT':
                # Percentage value
                pct_lmict = float(cbsa_branch_agg['other_lmi_pct']) if cbsa_branch_agg['other_lmi_pct'] else 0.0
                pct_value = pct_lmict / 100.0  # Convert 0-100 to 0-1
                cell_e.data_type = 'n'
                cell_e.value = pct_value
                cell_e.number_format = '0.00%'  # Percentage format
            elif metric == 'MMCT':
                # Percentage value
                pct_mmct = float(cbsa_branch_agg['other_mm_pct']) if cbsa_branch_agg['other_mm_pct'] else 0.0
                pct_value = pct_mmct / 100.0  # Convert 0-100 to 0-1
                cell_e.data_type = 'n'
                cell_e.value = pct_value
                cell_e.number_format = '0.00%'  # Percentage format
            
            # Column F: Difference formula
            # LMICT and MMCT are percentages, so difference = Subject % - Other %
            # Wrap with IFERROR to handle any errors
            cell_f = ws.cell(row=current_row, column=6)
            if cell_f.data_type != 'f':
                if metric == 'Branches':
                    # No difference formula for Branches
                    pass
                elif metric == 'LMICT':
                    # LMICT difference: (Subject LMICT % - Other LMICT %)
                    # Both are already percentages (0-1 range), so simple subtraction
                    cell_f.value = f'=IFERROR(D{current_row}-E{current_row},0)'
                    cell_f.number_format = '0.00%'  # Result is also a percentage difference
                elif metric == 'MMCT':
                    # MMCT difference: (Subject MMCT % - Other MMCT %)
                    cell_f.value = f'=IFERROR(D{current_row}-E{current_row},0)'
                    cell_f.number_format = '0.00%'  # Result is also a percentage difference
            
            current_row += 1
    
    print(f"    Populated: {bank_name} Branches ({current_row - start_row} rows)")


# Keep existing create functions for fallback when no template
def create_notes_sheet(wb: Workbook, bank_a_name: str, bank_b_name: str, 
                      years_hmda: List[str], years_sb: List[str]):
    """Create Notes (Methodology) sheet"""
    # ... (keep existing implementation)
    pass


def create_assessment_areas_sheet(wb: Workbook, bank_a_name: str, bank_b_name: str,
                                  assessment_areas_data: Dict):
    """Create Assessment Areas sheet with CBSA and county listings"""
    # ... (keep existing implementation)
    pass


def create_mortgage_goals_sheet(wb: Workbook, bank_a_name: str, bank_b_name: str,
                              mortgage_goals_data: Optional[Dict],
                              template_file: Optional[Path] = None):
    """Create Mortgage Goals sheet from template, preserving all formulas."""
    # ... (keep existing implementation - this one already loads from template)
    pass


def create_mortgage_data_sheet(wb: Workbook, bank_name: str, 
                                subject_data: Optional[pd.DataFrame],
                                peer_data: Optional[pd.DataFrame],
                                years: List[str]):
    """Create Mortgage Data sheet for one bank."""
    # ... (keep existing implementation)
    pass


def create_sb_data_sheet(wb: Workbook, bank_name: str,
                        subject_data: Optional[pd.DataFrame],
                        peer_data: Optional[pd.DataFrame],
                        years: List[str]):
    """Create Small Business Data sheet for one bank."""
    # ... (keep existing implementation)
    pass


def create_branch_data_sheet(wb: Workbook, bank_name: str,
                             branch_data: Optional[pd.DataFrame]):
    """Create Branch Data sheet for one bank."""
    # ... (keep existing implementation)
    pass


def insert_mortgage_goals_data(ws, mortgage_goals_data: Dict):
    """
    Insert mortgage goals data into columns C, D, E only.
    Preserves all existing formulas in columns F-L and structure in columns A-B.
    
    Template structure:
    - Column A: State names (preserved from template)
    - Column B: Metric names (preserved from template)  
    - Column C: Home Purchase data (INSERT HERE)
    - Column D: Refinance data (INSERT HERE)
    - Column E: Home Equity data (INSERT HERE)
    - Columns F-L: Formulas (PRESERVED - DO NOT MODIFY)
    """
    # Map metric names from template to data fields
    metric_mapping = {
        'Loans': ('total_loans', 'number'),
        '~LMICT': ('lmict_loans', 'number'),
        '~LMIB': ('lmib_loans', 'number'),
        'LMIB$': ('lmib_amount', 'dollar'),
        '~MMCT': ('mmct_loans', 'number'),
        '~MINB': ('minb_loans', 'number'),
        '~Asian': ('asian_loans', 'number'),
        '~Black': ('black_loans', 'number'),
        '~Native American': ('native_american_loans', 'number'),
        '~HoPI': ('hopi_loans', 'number'),
        '~Hispanic': ('hispanic_loans', 'number')
    }
    
    # Loan type column mapping
    loan_type_cols = {
        'home_purchase': 3,  # Column C
        'refinance': 4,       # Column D
        'home_equity': 5      # Column E
    }
    
    # Start from row 3 (row 1 = header, row 2 = Grand Total)
    current_row = 3
    
    # Get all states from the template (column A, starting row 3)
    states_in_template = []
    while current_row <= ws.max_row:
        state_cell = ws.cell(row=current_row, column=1)
        if state_cell.value and state_cell.value != 'Grand Total':
            state = str(state_cell.value).strip()
            if state and state not in states_in_template:
                states_in_template.append(state)
        
        # Check metric column to find next state group
        metric_cell = ws.cell(row=current_row, column=2)
        if metric_cell.value == 'Loans':  # First metric indicates new state group
            current_row += 1
        else:
            current_row += 1
        
        # Safety check - don't loop forever
        if current_row > 1000:
            break
    
    # If we didn't find states in template, try to get them from data
    if not states_in_template:
        for loan_type_data in mortgage_goals_data.values():
            if isinstance(loan_type_data, pd.DataFrame) and 'state' in loan_type_data.columns:
                states_in_template = sorted(loan_type_data['state'].unique().tolist())
                break
    
    # Insert data by state
    current_row = 3
    for state in states_in_template:
        # For each metric in the template order
        for metric_template, (field_name, format_type) in metric_mapping.items():
            # Find the row with this metric for this state
            # (Metric column is B, and we're looking for rows where A=state and B=metric)
            row_found = None
            for r in range(current_row, ws.max_row + 1):
                state_val = ws.cell(row=r, column=1).value
                metric_val = ws.cell(row=r, column=2).value
                if state_val == state and metric_val == metric_template:
                    row_found = r
                    break
            
            if row_found:
                # Insert data for each loan type in columns C, D, E
                for loan_type, target_col in loan_type_cols.items():
                    loan_data = mortgage_goals_data.get(loan_type)
                    if loan_data is not None and isinstance(loan_data, pd.DataFrame):
                        # Get data for this state
                        state_data = loan_data[loan_data['state'] == state]
                        if not state_data.empty and field_name in state_data.columns:
                            # Sum across all CBSAs in this state
                            value = state_data[field_name].sum()
                            
                            if pd.notna(value):
                                cell = ws.cell(row_found, target_col)
                                
                                # Only insert if cell doesn't already have a formula
                                if cell.data_type != 'f':
                                    if format_type == 'number':
                                        cell.value = int(value)
                                    elif format_type == 'dollar':
                                        cell.value = float(value)
                                        cell.number_format = '$#,##0'
        
        # Move to next state (skip 11 metric rows per state)
        current_row += 11
    
    print(f"  Inserted Mortgage Goals data for {len(states_in_template)} states")


def insert_mortgage_metric_data(ws, row_num: int, metric: str, data_row: pd.Series, start_col: int):
    """
    Insert mortgage metric data into a specific row and column.
    
    Args:
        ws: Worksheet
        row_num: Row number to insert into
        metric: Metric name (e.g., 'Loans', 'LMICT% ', 'LMIB%', etc.)
        data_row: Series with aggregated data (total_loans, lmict_loans, percentages, etc.)
        start_col: Starting column (4 = Column D for subject, 5 = Column E for peer)
    """
    cell = ws.cell(row=row_num, column=start_col)
    
    # Only insert if cell doesn't have a formula
    if cell.data_type == 'f':
        return
    
    if metric == 'Loans':
        value = data_row.get('total_loans', 0)
        if pd.notna(value):
            cell.value = int(value)
    
    elif metric == 'LMICT%' or metric == 'LMICT% ':  # Handle with or without trailing space
        value = data_row.get('lmict_percentage', 0)
        if pd.notna(value):
            cell.value = float(value) / 100.0  # Convert percentage to decimal for Excel
            cell.number_format = '0.00%'
    
    elif metric == 'LMIB%':
        value = data_row.get('lmib_percentage', 0)
        if pd.notna(value):
            cell.value = float(value) / 100.0
            cell.number_format = '0.00%'
    
    elif metric == 'LMIB$':
        value = data_row.get('lmib_amount', 0)
        if pd.notna(value):
            cell.value = float(value)
            cell.number_format = '$#,##0'
    
    elif metric == 'MMCT%':
        value = data_row.get('mmct_percentage', 0)
        if pd.notna(value):
            cell.value = float(value) / 100.0
            cell.number_format = '0.00%'
    
    elif metric == 'MINB%':
        value = data_row.get('minb_percentage', 0)
        if pd.notna(value):
            cell.value = float(value) / 100.0
            cell.number_format = '0.00%'
    
    elif metric == 'Asian%':
        value = data_row.get('asian_percentage', 0)
        if pd.notna(value):
            cell.value = float(value) / 100.0
            cell.number_format = '0.00%'
    
    elif metric == 'Black%':
        value = data_row.get('black_percentage', 0)
        if pd.notna(value):
            cell.value = float(value) / 100.0
            cell.number_format = '0.00%'
    
    elif metric == 'Native American%':
        value = data_row.get('native_american_percentage', 0)
        if pd.notna(value):
            cell.value = float(value) / 100.0
            cell.number_format = '0.00%'
    
    elif metric == 'HoPI%':
        value = data_row.get('hopi_percentage', 0)
        if pd.notna(value):
            cell.value = float(value) / 100.0
            cell.number_format = '0.00%'
    
    elif metric == 'Hispanic%':
        value = data_row.get('hispanic_percentage', 0)
        if pd.notna(value):
            cell.value = float(value) / 100.0
            cell.number_format = '0.00%'


def validate_excel_data(wb: Workbook, bank_a_name: str, bank_b_name: str,
                        mortgage_goals_data: Optional[Dict],
                        bank_a_mortgage_data: Optional[pd.DataFrame],
                        bank_b_mortgage_data: Optional[pd.DataFrame],
                        bank_a_sb_data: Optional[pd.DataFrame],
                        bank_b_sb_data: Optional[pd.DataFrame],
                        bank_a_branch_data: Optional[pd.DataFrame],
                        bank_b_branch_data: Optional[pd.DataFrame]) -> Dict:
    """
    Validate that the Excel workbook contains data before saving.
    
    Returns:
        Dictionary with 'is_valid' (bool) and 'warnings' (list of strings)
    """
    warnings = []
    is_valid = True
    
    print("\n" + "="*80)
    print("Validating Excel Data")
    print("="*80)
    
    # Check Mortgage Goals sheet
    if 'Mortgage Goals' in wb.sheetnames:
        ws = wb['Mortgage Goals']
        # Check if data exists in columns C, D, E (starting from row 3)
        data_found = False
        data_rows = 0
        for row in range(3, min(ws.max_row + 1, 500)):  # Check up to row 500
            for col in [3, 4, 5]:  # Columns C, D, E
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and cell.data_type != 'f':
                    # Check if it's a number (data, not formula)
                    try:
                        float(cell.value)
                        data_found = True
                        data_rows += 1
                        break
                    except (ValueError, TypeError):
                        pass
            if data_found:
                break
        
        if not data_found:
            warnings.append(f"Mortgage Goals sheet: No data found in columns C, D, E")
            is_valid = False
        else:
            print(f"  OK: Mortgage Goals sheet has data ({data_rows} rows with data)")
    else:
        warnings.append("Mortgage Goals sheet not found")
        is_valid = False
    
    # Check Assessment Areas sheet
    if 'Assessment Areas' in wb.sheetnames:
        ws = wb['Assessment Areas']
        data_rows = 0
        for row in range(2, min(ws.max_row + 1, 1000)):
            if ws.cell(row=row, column=1).value or ws.cell(row=row, column=2).value:
                data_rows += 1
        
        if data_rows == 0:
            warnings.append(f"Assessment Areas sheet: No data found")
            is_valid = False
        else:
            print(f"  OK: Assessment Areas sheet has data ({data_rows} rows)")
    else:
        warnings.append("Assessment Areas sheet not found")
        is_valid = False
    
    # Check Bank A Mortgage Data sheet
    bank_a_mortgage_sheet = f"{bank_a_name}Mortgage Data"
    if bank_a_mortgage_sheet in wb.sheetnames:
        ws = wb[bank_a_mortgage_sheet]
        data_rows = 0
        for row in range(3, min(ws.max_row + 1, 1000)):
            # Check columns D and E (subject and peer data)
            d_val = ws.cell(row=row, column=4).value
            e_val = ws.cell(row=row, column=5).value
            if (d_val is not None and d_val != '') or (e_val is not None and e_val != ''):
                data_rows += 1
        
        if data_rows == 0:
            warnings.append(f"{bank_a_mortgage_sheet}: No data found in columns D or E")
            is_valid = False
        else:
            print(f"  OK: {bank_a_mortgage_sheet} has data ({data_rows} rows)")
    else:
        warnings.append(f"{bank_a_mortgage_sheet} sheet not found")
        is_valid = False
    
    # Check Bank B Mortgage Data sheet
    bank_b_mortgage_sheet = f"{bank_b_name}Mortgage Data"
    if bank_b_mortgage_sheet in wb.sheetnames:
        ws = wb[bank_b_mortgage_sheet]
        data_rows = 0
        for row in range(3, min(ws.max_row + 1, 1000)):
            d_val = ws.cell(row=row, column=4).value
            e_val = ws.cell(row=row, column=5).value
            if (d_val is not None and d_val != '') or (e_val is not None and e_val != ''):
                data_rows += 1
        
        if data_rows == 0:
            warnings.append(f"{bank_b_mortgage_sheet}: No data found in columns D or E")
            is_valid = False
        else:
            print(f"  OK: {bank_b_mortgage_sheet} has data ({data_rows} rows)")
    else:
        warnings.append(f"{bank_b_mortgage_sheet} sheet not found")
        is_valid = False
    
    # Check Bank A Small Business sheet
    bank_a_sb_sheet = f"{bank_a_name} SB Lending"
    if bank_a_sb_sheet in wb.sheetnames:
        ws = wb[bank_a_sb_sheet]
        data_rows = 0
        for row in range(3, min(ws.max_row + 1, 1000)):
            d_val = ws.cell(row=row, column=4).value
            e_val = ws.cell(row=row, column=5).value
            if (d_val is not None and d_val != '') or (e_val is not None and e_val != ''):
                data_rows += 1
        
        if data_rows == 0:
            warnings.append(f"{bank_a_sb_sheet}: No data found in columns D or E")
            is_valid = False
        else:
            print(f"  OK: {bank_a_sb_sheet} has data ({data_rows} rows)")
    else:
        warnings.append(f"{bank_a_sb_sheet} sheet not found")
    
    # Check Bank B Small Business sheet
    bank_b_sb_sheet = f"{bank_b_name} SB Lending"
    if bank_b_sb_sheet in wb.sheetnames:
        ws = wb[bank_b_sb_sheet]
        data_rows = 0
        for row in range(3, min(ws.max_row + 1, 1000)):
            d_val = ws.cell(row=row, column=4).value
            e_val = ws.cell(row=row, column=5).value
            if (d_val is not None and d_val != '') or (e_val is not None and e_val != ''):
                data_rows += 1
        
        if data_rows == 0:
            warnings.append(f"{bank_b_sb_sheet}: No data found in columns D or E")
            is_valid = False
        else:
            print(f"  OK: {bank_b_sb_sheet} has data ({data_rows} rows)")
    else:
        warnings.append(f"{bank_b_sb_sheet} sheet not found")
    
    # Check Bank A Branch sheet
    bank_a_branch_sheet = f"{bank_a_name} Branches"
    if bank_a_branch_sheet in wb.sheetnames:
        ws = wb[bank_a_branch_sheet]
        data_rows = 0
        for row in range(3, min(ws.max_row + 1, 1000)):
            d_val = ws.cell(row=row, column=4).value
            e_val = ws.cell(row=row, column=5).value
            if (d_val is not None and d_val != '') or (e_val is not None and e_val != ''):
                data_rows += 1
        
        if data_rows == 0:
            warnings.append(f"{bank_a_branch_sheet}: No data found in columns D or E")
            is_valid = False
        else:
            print(f"  OK: {bank_a_branch_sheet} has data ({data_rows} rows)")
    else:
        warnings.append(f"{bank_a_branch_sheet} sheet not found")
    
    # Check Bank B Branch sheet
    bank_b_branch_sheet = f"{bank_b_name} Branches"
    if bank_b_branch_sheet in wb.sheetnames:
        ws = wb[bank_b_branch_sheet]
        data_rows = 0
        for row in range(3, min(ws.max_row + 1, 1000)):
            d_val = ws.cell(row=row, column=4).value
            e_val = ws.cell(row=row, column=5).value
            if (d_val is not None and d_val != '') or (e_val is not None and e_val != ''):
                data_rows += 1
        
        if data_rows == 0:
            warnings.append(f"{bank_b_branch_sheet}: No data found in columns D or E")
            is_valid = False
        else:
            print(f"  OK: {bank_b_branch_sheet} has data ({data_rows} rows)")
    else:
        warnings.append(f"{bank_b_branch_sheet} sheet not found")
    
    if is_valid:
        print("\n  Validation PASSED: All required sheets contain data")
    else:
        print(f"\n  Validation FAILED: {len(warnings)} issue(s) found")
    
    return {
        'is_valid': is_valid,
        'warnings': warnings
    }


def insert_sb_metric_data(ws, row_num: int, metric: str, data_row: pd.Series, col: int):
    """
    Insert Small Business metric data into a specific row and column.
    
    Args:
        ws: Worksheet
        row_num: Row number to insert into
        metric: Metric name (e.g., 'SB Loans', '#LMICT', etc.)
        data_row: Series with aggregated SB data
        col: Column number (4 = Column D for subject, 5 = Column E for peer)
    """
    cell = ws.cell(row=row_num, column=col)
    
    # Only insert if cell doesn't have a formula
    if cell.data_type == 'f':
        return
    
    if metric == 'SB Loans':
        value = data_row.get('sb_loans_total', 0)
        if pd.notna(value):
            cell.value = int(value)
    
    elif metric == '#LMICT':
        value = data_row.get('lmict_count', 0)
        if pd.notna(value):
            cell.value = int(value)
    
    elif metric == 'Avg SB LMICT Loan Amount':
        value = data_row.get('avg_sb_lmict_loan_amount', 0)
        if pd.notna(value) and value > 0:
            cell.value = float(value)
            cell.number_format = '$#,##0'
    
    elif metric == 'Loans Rev Under $1m':
        value = data_row.get('loans_rev_under_1m_count', 0)
        if pd.notna(value):
            cell.value = int(value)
    
    elif metric == 'Avg Loan Amt for RUM SB':
        value = data_row.get('avg_loan_amt_rum_sb', 0)
        if pd.notna(value) and value > 0:
            cell.value = float(value)
            cell.number_format = '$#,##0'
