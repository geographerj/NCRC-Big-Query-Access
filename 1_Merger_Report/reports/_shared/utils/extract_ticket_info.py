"""
Extract detailed information from the merger research ticket Excel file.
Reads bank details, years, and filters.
"""

import openpyxl
from pathlib import Path
import sys
import json

def sanitize_folder_name(name):
    """Sanitize bank name for use in folder name"""
    # Remove special characters, keep only alphanumeric and spaces
    import re
    name = re.sub(r'[^\w\s-]', '', name)
    # Replace spaces with underscores, limit length
    name = name.replace(' ', '_')
    # Remove multiple underscores
    name = re.sub(r'_+', '_', name)
    return name.strip('_')

def extract_ticket_info(excel_file):
    """
    Extract all relevant information from the merger ticket.
    Returns a dictionary with bank info, years, and filters.
    """
    excel_path = Path(excel_file)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    result = {
        'acquirer': {},
        'target': {},
        'years': [],
        'filters': {}
    }
    
    # Parse Notes sheet for bank info and years
    if "Notes" in wb.sheetnames:
        ws = wb["Notes"]
        print("\nParsing 'Notes' sheet...")
        
        # Based on inspection: Bank info is in columns E (name) and F (LEI/RSSD)
        # Row 3: Bank B (Target) name and LEI
        # Row 4: Bank A (Acquirer) name and LEI  
        # Row 13: Bank B (Target) name and RSSD
        # Row 14: Bank A (Acquirer) name and RSSD
        
        for row_idx in range(1, min(20, ws.max_row + 1)):
            # Read columns E and F (5 and 6)
            col_e = ws.cell(row=row_idx, column=5).value
            col_f = ws.cell(row=row_idx, column=6).value
            
            if col_e and col_f:
                name = str(col_e).strip()
                identifier = str(col_f).strip()
                
                # Check for LEI (20 characters, starts with letters/numbers)
                if len(identifier) == 20 and identifier.startswith(('54', 'AD', '57')):
                    # Store the actual bank name from the ticket (generic logic - detects which is acquirer/target)
                    # Typically first bank found is acquirer, second is target, but use Bank Details sheet structure
                    # For template: assume columns are in order: Acquirer then Target
                    if not result['target'].get('lei') and not result['acquirer'].get('lei'):
                        # First LEI found - assume acquirer
                        result['acquirer']['name'] = name
                        result['acquirer']['lei'] = identifier
                        print(f"  Found Acquirer Bank: {name}, LEI: {identifier}")
                    elif not result['target'].get('lei'):
                        # Second LEI found - assume target
                        result['target']['name'] = name
                        result['target']['lei'] = identifier
                        print(f"  Found Target Bank: {name}, LEI: {identifier}")
                
                # Check for RSSD (numeric, typically 6 digits)
                elif identifier.isdigit() and len(identifier) >= 6:
                    # Store the actual bank name from the ticket
                    if not result['target'].get('rssd') and not result['acquirer'].get('rssd'):
                        # First RSSD found - assume acquirer
                        result['acquirer']['rssd'] = identifier
                        print(f"  Found Acquirer Bank RSSD: {identifier}")
                    elif not result['target'].get('rssd'):
                        # Second RSSD found - assume target
                        result['target']['rssd'] = identifier
                        print(f"  Found Target Bank RSSD: {identifier}")
            
            # Check for years in column C (3) - "2020 through 2024"
            col_c = ws.cell(row=row_idx, column=3).value
            if col_c:
                text = str(col_c).lower()
                if 'through' in text and any(char.isdigit() for char in text):
                    import re
                    years_match = re.search(r'(\d{4})\s+through\s+(\d{4})', text)
                    if years_match:
                        start_year = int(years_match.group(1))
                        end_year = int(years_match.group(2))
                        # Use the most recent year range found
                        new_years = list(range(start_year, end_year + 1))
                        if not result.get('years') or end_year > max(result['years']):
                            result['years'] = new_years
                            print(f"  Found years: {result['years']}")
    
    # Parse Assessment Areas sheet to get structure
    if "Assessment Areas" in wb.sheetnames:
        ws = wb["Assessment Areas"]
        print("\nParsing 'Assessment Areas' sheet structure...")
        
        # Check header row
        header_row = []
        for col_idx in range(1, min(9, ws.max_column + 1)):
            cell_value = ws.cell(row=1, column=col_idx).value
            if cell_value:
                header_row.append((col_idx, str(cell_value)))
        
        print(f"  Header columns: {[v for _, v in header_row]}")
        
        # Count rows with data
        data_rows = 0
        for row_idx in range(2, min(100, ws.max_row + 1)):
            if ws.cell(row=row_idx, column=1).value or ws.cell(row=row_idx, column=5).value:
                data_rows += 1
        
        print(f"  Found {data_rows} data rows (sampled first 100)")
    
    # Try to find more bank info from other sheets
    # Sheet names may contain bank names, but we'll extract from Bank Details sheet primarily
    # For template: generic extraction based on Bank Details sheet structure
    
    # Parse "LMA Ticket Filters" sheet for filter settings
    if "LMA Ticket Filters" in wb.sheetnames:
        ws_filters = wb["LMA Ticket Filters"]
        print("\nParsing 'LMA Ticket Filters' sheet...")
        
        # Structure: Row 1 = headers, Row 2+ = data
        # Column A = Filter Type, Column B = Default Setting, Column C = Adjustment
        for row_idx in range(2, min(ws_filters.max_row + 1, 15)):
            filter_type_cell = ws_filters.cell(row=row_idx, column=1).value
            default_setting_cell = ws_filters.cell(row=row_idx, column=2).value
            adjustment_cell = ws_filters.cell(row=row_idx, column=3).value
            
            if not filter_type_cell or not default_setting_cell:
                continue
            
            filter_type = str(filter_type_cell).strip().lower()
            default_setting = str(default_setting_cell).strip()
            adjustment = str(adjustment_cell).strip() if adjustment_cell else ""
            
            # Parse HMDA years - "2020 - 2024" or "5 years of data, use most recent available 5"
            if "hmda" in filter_type and "year" in filter_type:
                import re
                # Try to parse year range from "2020 - 2024" format
                year_range_match = re.search(r'(\d{4})\s*-\s*(\d{4})', default_setting)
                if year_range_match:
                    start_year = int(year_range_match.group(1))
                    end_year = int(year_range_match.group(2))
                    year_range = [str(y) for y in range(start_year, end_year + 1)]  # Convert to strings
                    result['filters']['hmda_years'] = {
                        'default': default_setting,
                        'range': year_range,
                        'adjustment': adjustment
                    }
                    print(f"  HMDA Years: {year_range} (parsed from '{default_setting}')")
                else:
                    # Fallback to Notes sheet years or default
                    result['filters']['hmda_years'] = {
                        'default': default_setting,
                        'adjustment': adjustment,
                        'note': 'Default specifies range, using Notes sheet if available'
                    }
                    if result.get('years'):
                        # Convert years to strings if they're integers
                        years_list = result['years']
                        if isinstance(years_list[0], int) if years_list else False:
                            years_list = [str(y) for y in years_list]
                        result['filters']['hmda_years']['range'] = years_list
                        print(f"  HMDA Years: Using {len(years_list)} years from Notes sheet")
            
            # Parse Small Business years - "2020 - 2023" format
            elif ("small business" in filter_type and "year" in filter_type) or ("sb" in filter_type and "year" in filter_type):
                import re
                # Try to parse year range from "2020 - 2023" format
                year_range_match = re.search(r'(\d{4})\s*-\s*(\d{4})', default_setting)
                if year_range_match:
                    start_year = int(year_range_match.group(1))
                    end_year = int(year_range_match.group(2))
                    year_range = [str(y) for y in range(start_year, end_year + 1)]  # Convert to strings
                    result['filters']['sb_years'] = {
                        'default': default_setting,
                        'range': year_range,
                        'adjustment': adjustment
                    }
                    print(f"  Small Business Years: {year_range} (parsed from '{default_setting}')")
                else:
                    result['filters']['sb_years'] = {
                        'default': default_setting,
                        'adjustment': adjustment,
                        'note': 'Default specifies range, using Notes sheet if available'
                    }
                    if result.get('years'):
                        # Convert years to strings if they're integers
                        years_list = result['years']
                        if isinstance(years_list[0], int) if years_list else False:
                            years_list = [str(y) for y in years_list]
                        result['filters']['sb_years']['range'] = years_list
                        print(f"  Small Business Years: Using {len(years_list)} years from Notes sheet")
            
            # Parse HMDA filters
            elif "hmda" in filter_type:
                if "occupancy" in filter_type:
                    # Extract code if present (e.g., "1 (principal residence only)")
                    import re
                    code_match = re.search(r'(\d+)', default_setting)
                    occupancy_code = code_match.group(1) if code_match else default_setting
                    result['filters']['occupancy'] = {
                        'default': default_setting,
                        'code': occupancy_code,
                        'adjustment': adjustment
                    }
                    print(f"  Occupancy: {default_setting} (code: {occupancy_code})")
                
                elif "action taken" in filter_type:
                    code_match = re.search(r'(\d+)', default_setting)
                    action_code = code_match.group(1) if code_match else default_setting
                    result['filters']['action_taken'] = {
                        'default': default_setting,
                        'code': action_code,
                        'adjustment': adjustment
                    }
                    print(f"  Action Taken: {default_setting} (code: {action_code})")
                
                elif "dwelling" in filter_type or "units" in filter_type:
                    # "Single Family (1-4 Units):Site-Built..."
                    if "1-4" in default_setting.lower() or "site-built" in default_setting.lower():
                        result['filters']['units'] = {
                            'default': default_setting,
                            'adjustment': adjustment
                        }
                        result['filters']['construction_method'] = {
                            'default': 'Site-built (1)',
                            'adjustment': adjustment
                        }
                        print(f"  Dwelling/Units: {default_setting}")
                
                elif "loan purpose" in filter_type:
                    # "home purchase" -> code '1'
                    loan_purpose_map = {
                        'home purchase': '1',
                        'refinance': '31,32',
                        'home improvement': '2',
                        'home equity': '4'
                    }
                    purpose_lower = default_setting.lower()
                    purpose_code = loan_purpose_map.get(purpose_lower, default_setting)
                    result['filters']['loan_purpose'] = {
                        'default': default_setting,
                        'code': purpose_code,
                        'adjustment': adjustment
                    }
                    print(f"  Loan Purpose: {default_setting} (code: {purpose_code})")
    
    # Parse "Bank Details" sheet if it exists
    if "Bank Details" in wb.sheetnames:
        ws_bank = wb["Bank Details"]
        print("\nParsing 'Bank Details' sheet...")
        
        # Structure from inspection:
        # Row 3: Headers ['BANK INFORMATION', 'ACQUIRER', 'BANK INFORMATION', 'TARGET']
        # Row 4+: Data rows where:
        #   Column 1 = Label (Acquirer), Column 2 = Value (Acquirer)
        #   Column 3 = Label (Target), Column 4 = Value (Target)
        
        # Structure: Column 2 = Acquirer Label, Column 3 = Acquirer Value
        #           Column 5 = Target Label, Column 6 = Target Value
        
        # Scan rows 4-25 for bank information
        for row_idx in range(4, min(ws_bank.max_row + 1, 25)):
            # Acquirer info (columns 2 and 3)
            acq_label = ws_bank.cell(row=row_idx, column=2).value
            acq_value = ws_bank.cell(row=row_idx, column=3).value
            
            # Target info (columns 5 and 6)
            tgt_label = ws_bank.cell(row=row_idx, column=5).value
            tgt_value = ws_bank.cell(row=row_idx, column=6).value
            
            # Process Acquirer
            if acq_label and acq_value:
                label = str(acq_label).strip()
                value = str(acq_value).strip()
                label_lower = label.lower()
                
                if label_lower == "bank":
                    if not result['acquirer'].get('name'):
                        result['acquirer']['name'] = value
                        print(f"  Found Acquirer Bank: {value}")
                elif "lei" in label_lower and "code" in label_lower:
                    if len(value) == 20 and not result['acquirer'].get('lei'):
                        result['acquirer']['lei'] = value
                        print(f"  Found Acquirer LEI: {value}")
                elif "rss" in label_lower and "id" in label_lower:
                    # Handle "RSS-ID " with space, value might be int
                    value_str = str(value).strip()
                    if value_str.isdigit() and not result['acquirer'].get('rssd'):
                        result['acquirer']['rssd'] = value_str
                        print(f"  Found Acquirer RSSD: {value_str}")
                elif "respondent id" in label_lower and "sb" in label_lower:
                    if not result['acquirer'].get('sb_respondent_id'):
                        result['acquirer']['sb_respondent_id'] = value
                        print(f"  Found Acquirer SB Respondent ID: {value}")
            
            # Process Target
            if tgt_label and tgt_value:
                label = str(tgt_label).strip()
                value = str(tgt_value).strip()
                label_lower = label.lower()
                
                if label_lower == "bank":
                    if not result['target'].get('name'):
                        result['target']['name'] = value
                        print(f"  Found Target Bank: {value}")
                elif "lei" in label_lower and "code" in label_lower:
                    if len(value) == 20 and not result['target'].get('lei'):
                        result['target']['lei'] = value
                        print(f"  Found Target LEI: {value}")
                elif "rss" in label_lower and "id" in label_lower:
                    value_str = str(value).strip()
                    if value_str.isdigit() and not result['target'].get('rssd'):
                        result['target']['rssd'] = value_str
                        print(f"  Found Target RSSD: {value_str}")
                elif "respondent id" in label_lower and "sb" in label_lower:
                    if not result['target'].get('sb_respondent_id'):
                        result['target']['sb_respondent_id'] = value
                        print(f"  Found Target SB Respondent ID: {value}")
    
    return result

def main():
    """Main function"""
    if len(sys.argv) < 2:
        print("Usage: python extract_ticket_info.py <excel_file>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    
    try:
        info = extract_ticket_info(excel_file)
        
        print("\n" + "="*80)
        print("EXTRACTED TICKET INFORMATION")
        print("="*80)
        
        print(f"\nAcquirer Bank: {info['acquirer']}")
        print(f"Target Bank: {info['target']}")
        print(f"Years: {info['years']}")
        print(f"Filters: {info['filters']}")
        
        # Save to JSON for reference
        output_file = Path(excel_file).parent / "ticket_info_extracted.json"
        with open(output_file, 'w') as f:
            json.dump(info, f, indent=2)
        print(f"\nResults saved to: {output_file}")
        
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()

