"""
Parse assessment areas from the ticket Excel file's "Assessment Areas" sheet.
This may be more reliable than parsing PDFs.
"""

import openpyxl
from pathlib import Path
import sys
import json
import pandas as pd

def parse_assessment_areas_sheet(excel_file, acquirer_name=None, target_name=None):
    """
    Parse the Assessment Areas sheet from the ticket Excel file.
    
    Structure expected:
    Column 1: Bank Name
    Column 2: CBSA Name
    Column 3: CBSA Code
    Column 4: State Name
    Column 5: County Name
    Column 6: County Code (GEOID) - format: ##### (state code ## + county code ###)
    """
    excel_path = Path(excel_file)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    if "Assessment Areas" not in wb.sheetnames:
        raise ValueError("'Assessment Areas' sheet not found in Excel file")
    
    ws = wb["Assessment Areas"]
    
    print(f"\nParsing 'Assessment Areas' sheet...")
    print(f"  Dimensions: {ws.max_row} rows x {ws.max_column} columns")
    
    # Try to get bank names from Bank Details sheet if not provided
    if not acquirer_name or not target_name:
        if "Bank Details" in wb.sheetnames:
            ws_bank = wb["Bank Details"]
            for row_idx in range(4, min(ws_bank.max_row + 1, 25)):
                label = ws_bank.cell(row=row_idx, column=2).value
                value = ws_bank.cell(row=row_idx, column=3).value
                if label and 'name' in str(label).lower() and value:
                    acquirer_name = str(value).strip()
                label = ws_bank.cell(row=row_idx, column=5).value
                value = ws_bank.cell(row=row_idx, column=6).value
                if label and 'name' in str(label).lower() and value:
                    target_name = str(value).strip()
    
    bank_a_areas = []
    bank_b_areas = []
    
    # Read row by row
    for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
        bank_name = ws.cell(row=row_idx, column=1).value  # Column A: Bank Name
        cbsa_name = ws.cell(row=row_idx, column=2).value  # Column B: CBSA Name
        cbsa_code = ws.cell(row=row_idx, column=3).value  # Column C: CBSA Code
        state_name = ws.cell(row=row_idx, column=4).value  # Column D: State Name
        county_name = ws.cell(row=row_idx, column=5).value  # Column E: County Name
        county_code = ws.cell(row=row_idx, column=6).value  # Column F: County Code (GEOID)
        
        # Skip empty rows
        if not bank_name and not county_name:
            continue
        
        # Determine which bank this row belongs to
        bank_name_str = str(bank_name).strip() if bank_name else ""
        
        # Extract state code from GEOID (first 2 digits)
        # GEOID format: ##### (state code ## + county code ###)
        state_code = None
        if county_code:
            county_code_str = str(county_code).zfill(5)  # Ensure 5 digits with leading zeros
            state_code = county_code_str[:2]  # First 2 digits are state code
        
        # If state code not available from GEOID, try to get from state name
        if not state_code and state_name:
            state_code = get_state_code(state_name)
        
        # Build county entry
        county_entry = {
            'state': state_name,
            'state_name': state_name,
            'state_code': state_code,
            'cbsa': cbsa_code,
            'cbsa_name': cbsa_name,
            'county_name': str(county_name).strip() if county_name else None,
            'county_code': str(county_code).zfill(5) if county_code else None,  # Ensure 5 digits
            'geoid5': str(county_code).zfill(5) if county_code else None  # Alias for county_code
        }
        
        # Match bank name to acquirer or target
        # Use fuzzy matching to handle variations
        if acquirer_name and acquirer_name.lower() in bank_name_str.lower():
            bank_a_areas.append(county_entry)
        elif target_name and target_name.lower() in bank_name_str.lower():
            bank_b_areas.append(county_entry)
        else:
            # Fallback: if only one bank has entries, assign accordingly
            # Or use position-based (first bank encountered = acquirer)
            if not acquirer_name and not target_name:
                # Can't determine, skip for now or assign to bank_a
                bank_a_areas.append(county_entry)
            else:
                # Try to match by checking if it's in acquirer first
                if acquirer_name:
                    bank_a_areas.append(county_entry)
                elif target_name:
                    bank_b_areas.append(county_entry)
    
    print(f"\n  Found {len(bank_a_areas)} Bank A (Acquirer) assessment area entries")
    print(f"  Found {len(bank_b_areas)} Bank B (Target) assessment area entries")
    
    # Show samples
    if bank_a_areas:
        print(f"\n  Bank A sample entries (first 5):")
        for entry in bank_a_areas[:5]:
            print(f"    - {entry['county_name']}, {entry.get('state_name', entry.get('state_code', 'N/A'))}, GEOID: {entry.get('county_code', 'N/A')}, CBSA: {entry.get('cbsa', 'N/A')}")
    
    if bank_b_areas:
        print(f"\n  Bank B sample entries (first 5):")
        for entry in bank_b_areas[:5]:
            print(f"    - {entry['county_name']}, {entry.get('state_name', entry.get('state_code', 'N/A'))}, GEOID: {entry.get('county_code', 'N/A')}, CBSA: {entry.get('cbsa', 'N/A')}")
    
    return {
        'bank_a': bank_a_areas,
        'bank_b': bank_b_areas
    }

def parse_county_string(county_string):
    """
    Parse county string like "Dallas County, Alabama" or "Maricopa County, Arizona"
    Returns dict with county_name, state_code, state_name
    """
    if not county_string:
        return None
    
    county_str = str(county_string).strip()
    
    # Pattern: "County Name County, State Name"
    import re
    pattern = r'(.+?)\s+County,\s+(.+)'
    match = re.match(pattern, county_str)
    
    if match:
        county_name = match.group(1).strip()
        state = match.group(2).strip()
        
        # Try to get state code from state name
        state_code = get_state_code(state)
        
        return {
            'county_name': county_name,
            'state_name': state,
            'state_code': state_code
        }
    
    return None

def get_state_code(state_name):
    """Convert state name to 2-letter code"""
    state_map = {
        'Alabama': '01', 'Arizona': '04', 'Arkansas': '05', 'California': '06',
        'Colorado': '08', 'Connecticut': '09', 'Delaware': '10', 'Florida': '12',
        'Georgia': '13', 'Idaho': '16', 'Illinois': '17', 'Indiana': '18',
        'Iowa': '19', 'Kansas': '20', 'Kentucky': '21', 'Louisiana': '22',
        'Maine': '23', 'Maryland': '24', 'Massachusetts': '25', 'Michigan': '26',
        'Minnesota': '27', 'Mississippi': '28', 'Missouri': '29', 'Montana': '30',
        'Nebraska': '31', 'Nevada': '32', 'New Hampshire': '33', 'New Jersey': '34',
        'New Mexico': '35', 'New York': '36', 'North Carolina': '37', 'North Dakota': '38',
        'Ohio': '39', 'Oklahoma': '40', 'Oregon': '41', 'Pennsylvania': '42',
        'Rhode Island': '44', 'South Carolina': '45', 'South Dakota': '46', 'Tennessee': '47',
        'Texas': '48', 'Utah': '49', 'Vermont': '50', 'Virginia': '51',
        'Washington': '53', 'West Virginia': '54', 'Wisconsin': '55', 'Wyoming': '56',
        'District of Columbia': '11'
    }
    
    return state_map.get(state_name, None)

def main():
    """Main function"""
    if len(sys.argv) < 2:
        print("Usage: python parse_assessment_areas_from_ticket.py <ticket_excel_file>")
        print("\nExample:")
        print('  python parse_assessment_areas_from_ticket.py "BankA+BankB merger research ticket.xlsx"')
        sys.exit(1)
    
    excel_file = sys.argv[1]
    
    try:
        results = parse_assessment_areas_sheet(excel_file)
        
        # Save results
        output_file = Path(excel_file).parent / "assessment_areas_from_ticket.json"
        with open(output_file, 'w') as f:
            json.dump(results, f, indent=2)
        
        print(f"\n\nResults saved to: {output_file}")
        print(f"\nNote: County codes (GEOIDs) will need to be looked up using county name + state code.")
        
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()


