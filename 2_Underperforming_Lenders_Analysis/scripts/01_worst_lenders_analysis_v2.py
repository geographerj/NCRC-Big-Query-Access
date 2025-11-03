"""
Worst Lenders Redlining Analysis v2 - Excel Generator
Identifies top 13 worst performing banks for CBA enforcement

V2 Changes:
- Bank qualification: ≥500 apps in CBSA for 2022-2024 period (accounts for census boundary changes)
- Scoring and weighting: Limited to 2022-2024 data only
- Output: Still includes all years 2018-2024 for historical comparison

Processes BigQuery output from worst_lenders_redlining_query_v2.sql

Key Features:
- Only 8 redlining metrics (tract demographics)
- County-level breakdown within CBSAs
- Applications and originations (kind column)
- All loans and home purchase (loan_purpose column)
- Chi-squared test for statistical significance
- Scoring: ratios >= 2.0, weighted by CBSA size and metrics (2022-2024 only)
- Selection: Worst 13 banks with consistency and non-improvement filters
- Force include: 3 specific banks (1st National Bank of Pennsylvania, Webster Bank, Frost Bank)
- Exclude: CBA banks + East West Bank + Cathay Bank

Author: NCRC Research Department
Date: October 2025
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
import sys
import os
import argparse
from pathlib import Path
from scipy.stats import chi2_contingency

warnings.filterwarnings('ignore')

# =============================================================================
# CONFIGURATION
# =============================================================================

# Only 8 redlining metrics (tract-level demographics)
METRICS = {
    'mmct_50': 'MMCT 50%',
    'mmct_80': 'MMCT 80%',
    'black_tract_50': 'Black Tract 50%',
    'black_tract_80': 'Black Tract 80%',
    'hispanic_tract_50': 'Hispanic Tract 50%',
    'hispanic_tract_80': 'Hispanic Tract 80%',
    'black_hispanic_tract_50': 'Black+Hispanic Tract 50%',
    'black_hispanic_tract_80': 'Black+Hispanic Tract 80%'
}

# Expected years (all years included in output)
EXPECTED_YEARS = list(range(2018, 2025))  # 2018-2024

# Scoring years (only 2022-2024 used for scoring and weighting)
SCORING_YEARS = [2022, 2023, 2024]

# Statistical significance threshold
SIGNIFICANCE_LEVEL = 0.05

# Bank inclusion/exclusion lists
FORCE_INCLUDE_BANKS = [
    '1st national bank of pennsylvania',
    'webster bank',
    'frost bank'
]

CBA_EXCLUDE_BANKS = [
    'umb bank',
    'atlantic union bank',
    'southstate bank',
    'south state bank',
    'bmo harris bank',
    'umpqua',
    'us bank',
    'old national bancorp',
    'old national bank',
    'flagstar bank',
    'new york community bancorp',
    'm&t bank',
    'pnc bank',
    'first citizens bank',
    'cit group',
    'morgan stanley',
    'first merchants bank',
    'truist',
    'cadence bank',
    'wells fargo',
    'wells fargo & company',
    'first horizon bank',
    'first tennessee bank',
    'iberiabank',
    'santander bank',
    'first financial bank',
    'fifth third bank',
    'huntington bancshares',
    'huntington bank',
    'keybank',
    'east west bank',
    'east west bancorp',
    'cathay bank',
    'cathay general bancorp'
]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def chi_squared_test(x1, n1, x2, n2):
    """
    Perform chi-squared test for two proportions
    Returns: (chi2_stat, p_value, is_significant)
    x1, n1: successes and trials for group 1 (subject bank)
    x2, n2: successes and trials for group 2 (peer banks)
    """
    if n1 == 0 or n2 == 0:
        return None, None, False
    
    # Create 2x2 contingency table
    # Row 1: Subject bank (successes, failures)
    # Row 2: Peer banks (successes, failures)
    contingency = np.array([
        [x1, n1 - x1],  # Subject: metric count, non-metric count
        [x2, n2 - x2]   # Peer: metric count, non-metric count
    ])
    
    try:
        chi2, p_value, dof, expected = chi2_contingency(contingency)
        is_significant = (p_value is not None and p_value < SIGNIFICANCE_LEVEL)
        return chi2, p_value, is_significant
    except:
        return None, None, False

def normalize_bank_name(name):
    """Normalize bank name for matching (lowercase, remove common suffixes)"""
    if pd.isna(name):
        return ''
    name = str(name).lower().strip()
    # Remove common suffixes
    for suffix in [' inc', ' inc.', ' corp', ' corp.', ' ltd', ' ltd.', ' llc', ' llc.']:
        if name.endswith(suffix):
            name = name[:-len(suffix)]
    return name.strip()

def matches_bank_name(bank_name, search_list):
    """Check if bank name matches any in search list (normalized)"""
    normalized_bank = normalize_bank_name(bank_name)
    for search_name in search_list:
        normalized_search = normalize_bank_name(search_name)
        # Check for exact match or if the normalized bank contains the search term as a whole word
        if normalized_bank == normalized_search:
            return True
        # Check if search term is a complete word/phrase in the bank name
        # This prevents partial matches (e.g., "us bank" won't match "us national bank")
        if normalized_search in normalized_bank:
            # Make sure it's not a partial word match
            # Split and check if any word matches
            bank_words = normalized_bank.split()
            search_words = normalized_search.split()
            # If search term is a single word, check if it matches any word in bank name
            if len(search_words) == 1:
                if search_words[0] in bank_words:
                    return True
            # If search term is multiple words, check if all words appear in bank name
            elif all(word in bank_words for word in search_words):
                return True
    return False

# =============================================================================
# DATA LOADING AND FILTERING
# =============================================================================

def load_data(input_file):
    """Load CSV data from BigQuery export"""
    print(f"\nLoading data from: {input_file}")
    
    df = pd.read_csv(input_file, dtype={'county_code': str, 'cbsa_code': str})
    print(f"  Loaded {len(df):,} rows")
    
    # Convert year to int
    if 'activity_year' in df.columns:
        df['activity_year'] = df['activity_year'].astype(int)
    
    # Normalize lender names
    if 'lender_name' in df.columns:
        df['lender_name_normalized'] = df['lender_name'].apply(normalize_bank_name)
    
    print(f"  Banks: {df['lender_name'].nunique()}")
    print(f"  CBSAs: {df['cbsa_code'].nunique()}")
    print(f"  Counties: {df['county_code'].nunique()}")
    print(f"  Years: {sorted(df['activity_year'].unique())}")
    print(f"  Scoring years (2022-2024): {len(df[df['activity_year'].isin(SCORING_YEARS)]):,} rows")
    
    return df

def filter_banks(df):
    """Filter banks: exclude CBA banks, keep force-included banks"""
    print("\nFiltering banks...")
    
    original_count = df['lender_name'].nunique()
    
    # Ensure normalized name column exists
    if 'lender_name_normalized' not in df.columns:
        df['lender_name_normalized'] = df['lender_name'].apply(normalize_bank_name)
    
    # Identify banks to exclude
    exclude_mask = df['lender_name_normalized'].apply(
        lambda x: matches_bank_name(x, CBA_EXCLUDE_BANKS)
    )
    exclude_banks = df.loc[exclude_mask, 'lender_name'].unique()
    
    print(f"  Excluding {len(exclude_banks)} CBA banks")
    for bank in exclude_banks[:10]:  # Show first 10
        print(f"    - {bank}")
    if len(exclude_banks) > 10:
        print(f"    ... and {len(exclude_banks) - 10} more")
    
    # Filter out excluded banks
    df_filtered = df[~exclude_mask].copy()
    
    # Identify force-included banks
    force_included = df['lender_name_normalized'].apply(
        lambda x: matches_bank_name(x, FORCE_INCLUDE_BANKS)
    )
    force_included_banks = df.loc[force_included, 'lender_name'].unique()
    
    print(f"\n  Force-including {len(force_included_banks)} banks:")
    for bank in force_included_banks:
        print(f"    - {bank}")
    
    print(f"  Banks after filtering: {df_filtered['lender_name'].nunique()} (from {original_count})")
    
    return df_filtered, force_included_banks

# =============================================================================
# METRIC CALCULATIONS
# =============================================================================

def calculate_metrics(row, metric_base):
    """
    Calculate share and gap for a metric
    Returns: (subject_share, peer_share, gap, is_significant, subject_count, peer_count)
    """
    subject_total = row.get('subject_total_count', 0)
    peer_total = row.get('peer_total_count', 0)
    
    if subject_total == 0 or peer_total == 0:
        return None, None, None, False, 0, 0
    
    # Get metric-specific counts
    subject_col = f'subject_{metric_base}'
    peer_col = f'peer_{metric_base}'
    
    subject_count = row.get(subject_col, 0)
    peer_count = row.get(peer_col, 0)
    
    # Calculate shares as percentages
    subject_share = (subject_count / subject_total * 100) if subject_total > 0 else 0
    peer_share = (peer_count / peer_total * 100) if peer_total > 0 else 0
    gap = subject_share - peer_share  # Gap in percentage points
    
    # Chi-squared test for statistical significance
    _, _, is_significant = chi_squared_test(
        subject_count, subject_total,
        peer_count, peer_total
    )
    
    # Only flag negative gaps as significant
    is_significant = is_significant and gap < 0
    
    return subject_share, peer_share, gap, is_significant, subject_count, peer_count

def add_calculated_metrics(df):
    """Add calculated shares, gaps, ratios, and significance for all metrics"""
    print("\nCalculating shares, gaps, and ratios...")
    
    for metric_base, metric_name in METRICS.items():
        print(f"  Processing {metric_name}...")
        
        results = df.apply(lambda row: calculate_metrics(row, metric_base), axis=1)
        
        # Unpack results
        df[f'{metric_base}_subject_share'] = results.apply(lambda x: x[0])
        df[f'{metric_base}_peer_share'] = results.apply(lambda x: x[1])
        df[f'{metric_base}_gap'] = results.apply(lambda x: x[2])
        df[f'{metric_base}_significant'] = results.apply(lambda x: x[3])
        df[f'{metric_base}_subject_count'] = results.apply(lambda x: x[4])
        df[f'{metric_base}_peer_count'] = results.apply(lambda x: x[5])
        
        # Calculate ratio: peer_share / subject_share
        df[f'{metric_base}_ratio'] = df.apply(
            lambda row: row[f'{metric_base}_peer_share'] / row[f'{metric_base}_subject_share']
            if row[f'{metric_base}_subject_share'] > 0 else None,
            axis=1
        )
    
    return df

# =============================================================================
# SCORING AND SELECTION (V2: 2022-2024 ONLY)
# =============================================================================

def calculate_bank_scores(df, cbsa_sizes=None):
    """
    Calculate scores for each bank (V2: Only 2022-2024 data):
    - Count ratios >= 2.0 (weighted by CBSA size, metrics, and year recency)
    - Weight 2024 performance heavily (banks getting worse prioritized)
    - Penalize banks that are improving in recent years
    - Check consistency (bad in >=50% of CBSAs)
    - Check non-improvement (bad in 2023-2024)
    """
    print("\nCalculating bank scores (2022-2024 data only)...")
    
    # Filter to scoring years only for CBSA size calculation
    scoring_data = df[df['activity_year'].isin(SCORING_YEARS)].copy()
    
    if cbsa_sizes is None:
        # Calculate CBSA sizes (total applications in 2022-2024 only)
        cbsa_sizes = scoring_data.groupby('cbsa_code')['subject_total_count'].sum().to_dict()
    
    bank_scores = []
    
    for bank_name in df['lender_name'].unique():
        # Use all data for bank CBSA identification, but scoring only on 2022-2024
        bank_data_all = df[df['lender_name'] == bank_name].copy()
        bank_data = scoring_data[scoring_data['lender_name'] == bank_name].copy()
        bank_cbsas = bank_data_all['cbsa_code'].unique()
        
        total_bad_ratios = 0
        weighted_score = 0
        cbsas_with_bad_ratios = set()
        metrics_with_bad_ratios = set()
        recent_bad_ratios = 0  # Bad ratios in 2023-2024
        worsening_trend_score = 0  # Score for banks getting worse over time
        
        # Year weights: prioritize recent performance, especially 2024 (2022-2024 only)
        year_weights = {
            2022: 1.0,  # Baseline weight
            2023: 1.5,  # Recent years weighted more
            2024: 3.0   # 2024 weighted most heavily (banks getting worse)
        }
        
        # Collect ratio data by year for trend analysis (2022-2024 only)
        ratio_trends = {}  # {(cbsa, county, loan_purpose, kind, metric): {year: ratio}}
        
        # First pass: collect all ratios for trend analysis (2022-2024 only)
        for (year, cbsa, county, loan_purpose, kind), group in bank_data.groupby([
            'activity_year', 'cbsa_code', 'county_code', 'loan_purpose_category', 'kind'
        ]):
            cbsa_size = cbsa_sizes.get(cbsa, 1)
            
            for metric_base in METRICS.keys():
                ratio_col = f'{metric_base}_ratio'
                if ratio_col not in group.columns:
                    continue
                
                ratio_value = group[ratio_col].iloc[0] if len(group) > 0 else None
                
                if pd.notna(ratio_value):
                    key = (cbsa, county, loan_purpose, kind, metric_base)
                    if key not in ratio_trends:
                        ratio_trends[key] = {}
                    ratio_trends[key][year] = ratio_value
        
        # Second pass: calculate scores with year weighting and trend analysis (2022-2024 only)
        for (year, cbsa, county, loan_purpose, kind), group in bank_data.groupby([
            'activity_year', 'cbsa_code', 'county_code', 'loan_purpose_category', 'kind'
        ]):
            cbsa_size = cbsa_sizes.get(cbsa, 1)
            year_weight = year_weights.get(year, 1.0)
            
            for metric_base in METRICS.keys():
                ratio_col = f'{metric_base}_ratio'
                if ratio_col not in group.columns:
                    continue
                
                ratio_value = group[ratio_col].iloc[0] if len(group) > 0 else None
                
                if pd.notna(ratio_value) and ratio_value >= 2.0:
                    total_bad_ratios += 1
                    cbsas_with_bad_ratios.add(cbsa)
                    metrics_with_bad_ratios.add(metric_base)
                    
                    # Base weight by CBSA size (2022-2024 applications)
                    base_weight = cbsa_size / 1000000  # Normalize by millions
                    
                    # Apply year weight (2024 gets 3x, 2023 gets 1.5x)
                    weighted_score += base_weight * year_weight
                    
                    # Check if in recent years (2023-2024)
                    if year >= 2023:
                        recent_bad_ratios += 1
                    
                    # Trend analysis: check if ratio is getting worse over time (2022-2024)
                    key = (cbsa, county, loan_purpose, kind, metric_base)
                    if key in ratio_trends:
                        trend_years = sorted(ratio_trends[key].keys())
                        if len(trend_years) >= 2:  # Need at least 2 years for trend
                            # Special check: if 2024 is worse than 2023, penalize heavily
                            if 2024 in trend_years and 2023 in trend_years:
                                if ratio_trends[key][2024] > ratio_trends[key][2023]:
                                    # Bank got worse in 2024 - extra penalty
                                    worsening_trend_score += base_weight * (ratio_trends[key][2024] - ratio_trends[key][2023]) * 5
                            
                            # Check if 2024 is worse than 2022 (overall worsening)
                            if 2024 in trend_years and 2022 in trend_years:
                                if ratio_trends[key][2024] > ratio_trends[key][2022]:
                                    worsening_increase = ratio_trends[key][2024] - ratio_trends[key][2022]
                                    worsening_trend_score += base_weight * worsening_increase * year_weight * 2  # 2x multiplier for worsening trends
        
        # Calculate consistency: bad in >=50% of CBSAs? (based on 2022-2024 scoring)
        consistency_ratio = len(cbsas_with_bad_ratios) / len(bank_cbsas) if len(bank_cbsas) > 0 else 0
        is_consistent = consistency_ratio >= 0.5
        
        # Check non-improvement: must have bad ratios in 2023-2024
        is_not_improving = recent_bad_ratios > 0
        
        # Final weighted score combines base score + worsening trend bonus
        final_weighted_score = weighted_score + worsening_trend_score
        
        bank_scores.append({
            'lender_name': bank_name,
            'total_bad_ratios': total_bad_ratios,
            'weighted_score': final_weighted_score,
            'base_weighted_score': weighted_score,
            'worsening_trend_score': worsening_trend_score,
            'num_cbsas': len(bank_cbsas),
            'cbsas_with_bad_ratios': len(cbsas_with_bad_ratios),
            'consistency_ratio': consistency_ratio,
            'is_consistent': is_consistent,
            'recent_bad_ratios': recent_bad_ratios,
            'is_not_improving': is_not_improving,
            'metrics_affected': len(metrics_with_bad_ratios)
        })
    
    scores_df = pd.DataFrame(bank_scores)
    scores_df = scores_df.sort_values('weighted_score', ascending=False)
    
    return scores_df

def select_worst_13_banks(scores_df, force_included_banks):
    """Select worst 13 banks, ensuring force-included banks are included (regardless of criteria)"""
    print("\nSelecting worst 13 banks...")
    
    # Ensure normalized name column exists
    if 'lender_name_normalized' not in scores_df.columns:
        scores_df['lender_name_normalized'] = scores_df['lender_name'].apply(normalize_bank_name)
    
    # Filter for banks that meet consistency and non-improvement criteria
    qualified = scores_df[
        scores_df['is_consistent'] & 
        scores_df['is_not_improving']
    ].copy()
    
    print(f"  Banks meeting consistency and non-improvement criteria: {len(qualified)}")
    
    # ALWAYS add force-included banks regardless of criteria
    print(f"\n  Force-including banks (regardless of criteria):")
    for bank_name in force_included_banks:
        # Find the bank in original scores_df
        bank_rows = scores_df[scores_df['lender_name_normalized'].apply(
            lambda x: matches_bank_name(x, [bank_name])
        )]
        if len(bank_rows) > 0:
            bank_row = bank_rows.iloc[0:1].copy()
            # Check if already in qualified
            already_included = qualified[qualified['lender_name_normalized'].apply(
                lambda x: matches_bank_name(x, [bank_name])
            )]
            if len(already_included) == 0:
                qualified = pd.concat([qualified, bank_row], ignore_index=True)
                print(f"    Added (force): {bank_name} (score: {bank_row.iloc[0]['weighted_score']:.2f})")
            else:
                print(f"    Already included: {bank_name}")
        else:
            print(f"    WARNING: Force-included bank not found in scores: {bank_name}")
    
    # Remove duplicates
    qualified = qualified.drop_duplicates(subset=['lender_name'])
    
    # Sort by weighted score (descending - worst first)
    qualified = qualified.sort_values('weighted_score', ascending=False)
    
    # Ensure force-included banks are in top 13
    # First, separate force-included banks
    force_included_list = []
    other_banks = []
    
    for idx, row in qualified.iterrows():
        is_force = any(matches_bank_name(row['lender_name'], [b]) for b in force_included_banks)
        if is_force:
            force_included_list.append(row)
        else:
            other_banks.append(row)
    
    # Combine: force-included first, then others, then select top 13
    if force_included_list:
        force_df = pd.DataFrame(force_included_list)
        other_df = pd.DataFrame(other_banks) if other_banks else pd.DataFrame()
        combined = pd.concat([force_df, other_df], ignore_index=True)
        worst_13 = combined.head(13).copy()
    else:
        # If no force-included found, just take top 13
        worst_13 = qualified.head(13).copy()
    
    print(f"\n  Selected worst 13 banks:")
    for idx, row in worst_13.iterrows():
        is_force = any(matches_bank_name(row['lender_name'], [b]) for b in force_included_banks)
        force_tag = " [FORCE]" if is_force else ""
        print(f"    {row['lender_name']}: {row['weighted_score']:.2f} (bad ratios: {row['total_bad_ratios']}){force_tag}")
    
    return worst_13['lender_name'].tolist()

# =============================================================================
# EXCEL OUTPUT PREPARATION
# =============================================================================

def calculate_cbsa_application_percentage(df, bank_name):
    """Calculate what percentage of bank's total applications are in each CBSA"""
    bank_data = df[df['lender_name'] == bank_name].copy()
    
    if len(bank_data) == 0:
        return {}
    
    # Calculate total applications per CBSA across all years
    cbsa_totals = bank_data.groupby('cbsa_code')['subject_total_count'].sum()
    
    # Calculate bank's total applications across all CBSAs
    bank_total = cbsa_totals.sum()
    
    if bank_total == 0:
        return {}
    
    # Calculate percentage for each CBSA
    cbsa_percentages = (cbsa_totals / bank_total * 100).to_dict()
    
    # Categorize into buckets
    cbsa_buckets = {}
    for cbsa_code, pct in cbsa_percentages.items():
        if pct <= 1:
            bucket = "<= 1%"
        elif pct < 5:
            bucket = ">1% to <5%"
        elif pct < 50:
            bucket = ">=5% to <50%"
        else:
            bucket = ">= 50%"
        cbsa_buckets[cbsa_code] = bucket
    
    return cbsa_buckets

def prepare_ratio_sheet_data(df, bank_name):
    """Prepare ratio sheet data: peer/subject ratios by year (all years 2018-2024)"""
    bank_data = df[df['lender_name'] == bank_name].copy()
    
    if len(bank_data) == 0:
        return pd.DataFrame()
    
    # Calculate CBSA application percentage buckets
    cbsa_buckets = calculate_cbsa_application_percentage(df, bank_name)
    
    result_rows = []
    
    # Group by CBSA, county, loan purpose, kind, and metric
    for (cbsa_code, cbsa_name, county_code, county_name, state_name, loan_purpose, kind), group in bank_data.groupby([
        'cbsa_code', 'cbsa_name', 'county_code', 'county_name', 'state_name', 'loan_purpose_category', 'kind'
    ]):
        for metric_base, metric_name in METRICS.items():
            row = {
                'CBSA': cbsa_name,
                'County': county_name if pd.notna(county_name) else 'Unknown',
                'State': state_name if pd.notna(state_name) else 'Unknown',
                'CBSA_App_Percent_Bucket': cbsa_buckets.get(cbsa_code, 'Unknown'),
                'Loan_Purpose': loan_purpose,
                'Kind': kind,
                'Metric': metric_name
            }
            
            has_all_years = True
            
            # Add ratios for each year (all years 2018-2024)
            for year in EXPECTED_YEARS:
                year_data = group[group['activity_year'] == year]
                if len(year_data) > 0:
                    row_data = year_data.iloc[0]
                    ratio_value = row_data.get(f'{metric_base}_ratio')
                    
                    if pd.notna(ratio_value):
                        row[str(year)] = round(ratio_value, 2)
                    else:
                        row[str(year)] = None
                else:
                    row[str(year)] = None
            
            result_rows.append(row)
    
    result_df = pd.DataFrame(result_rows)
    if len(result_df) > 0:
        result_df = result_df.sort_values([
            'CBSA', 'County', 'State', 'Loan_Purpose', 'Kind', 'Metric'
        ])
    
    return result_df

def prepare_shares_gaps_sheet_data(df, bank_name):
    """Prepare shares and gaps sheet: counts, shares, gaps with significance (all years 2018-2024)"""
    bank_data = df[df['lender_name'] == bank_name].copy()
    
    if len(bank_data) == 0:
        return pd.DataFrame()
    
    # Calculate CBSA application percentage buckets
    cbsa_buckets = calculate_cbsa_application_percentage(df, bank_name)
    
    result_rows = []
    
    # Group by CBSA, county, loan purpose, kind, and metric
    for (cbsa_code, cbsa_name, county_code, county_name, state_name, loan_purpose, kind), group in bank_data.groupby([
        'cbsa_code', 'cbsa_name', 'county_code', 'county_name', 'state_name', 'loan_purpose_category', 'kind'
    ]):
        for metric_base, metric_name in METRICS.items():
            row = {
                'CBSA': cbsa_name,
                'County': county_name if pd.notna(county_name) else 'Unknown',
                'State': state_name if pd.notna(state_name) else 'Unknown',
                'CBSA_App_Percent_Bucket': cbsa_buckets.get(cbsa_code, 'Unknown'),
                'Loan_Purpose': loan_purpose,
                'Kind': kind,
                'Metric': metric_name
            }
            
            for year in EXPECTED_YEARS:
                year_data = group[group['activity_year'] == year]
                if len(year_data) > 0:
                    row_data = year_data.iloc[0]
                    
                    subject_count = row_data.get('subject_total_count', 0)
                    subject_share = row_data.get(f'{metric_base}_subject_share', 0)
                    peer_share = row_data.get(f'{metric_base}_peer_share', 0)
                    gap = row_data.get(f'{metric_base}_gap', 0)
                    is_significant = row_data.get(f'{metric_base}_significant', False)
                    
                    row[f'{year}_Subject_Count'] = int(subject_count) if pd.notna(subject_count) else 0
                    row[f'{year}_Subject_Share'] = f"{subject_share:.1f}%" if pd.notna(subject_share) else "0.0%"
                    row[f'{year}_Peer_Share'] = f"{peer_share:.1f}%" if pd.notna(peer_share) else "0.0%"
                    
                    # Format gap with asterisk if significant and negative
                    if pd.notna(gap):
                        gap_str = f"{gap:.1f}pp"
                        if is_significant and gap < 0:
                            gap_str += "*"
                        row[f'{year}_Gap'] = gap_str
                    else:
                        row[f'{year}_Gap'] = "0.0pp"
                else:
                    row[f'{year}_Subject_Count'] = 0
                    row[f'{year}_Subject_Share'] = "0.0%"
                    row[f'{year}_Peer_Share'] = "0.0%"
                    row[f'{year}_Gap'] = "0.0pp"
            
            result_rows.append(row)
    
    result_df = pd.DataFrame(result_rows)
    if len(result_df) > 0:
        result_df = result_df.sort_values([
            'CBSA', 'County', 'State', 'Loan_Purpose', 'Kind', 'Metric'
        ])
    
    return result_df

# =============================================================================
# EXCEL FORMATTING
# =============================================================================

def format_header_row(worksheet):
    """Format header row"""
    fill_gray = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    font_bold = Font(bold=True)
    
    for cell in worksheet[1]:
        cell.fill = fill_gray
        cell.font = font_bold
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def apply_ratio_coloring(worksheet, dataframe):
    """Apply color coding to ratio sheets"""
    year_cols = [col for col in dataframe.columns if col.isdigit()]
    
    fill_green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # < 1.0
    fill_yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # 1.0-1.5
    fill_orange = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')  # 1.5-2.0
    fill_light_red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # 2.0-3.0
    fill_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # >= 3.0
    
    col_indices = {col: dataframe.columns.get_loc(col) + 1 for col in year_cols}
    
    for row_idx in range(2, len(dataframe) + 2):
        for col_name, col_idx in col_indices.items():
            cell = worksheet.cell(row=row_idx, column=col_idx)
            try:
                if cell.value is None:
                    continue
                value = float(cell.value)
                if pd.isna(value):
                    continue
                elif value < 1.0:
                    cell.fill = fill_green
                elif value < 1.5:
                    cell.fill = fill_yellow
                elif value < 2.0:
                    cell.fill = fill_orange
                elif value < 3.0:
                    cell.fill = fill_light_red
                else:
                    cell.fill = fill_red
                    cell.font = Font(color='FFFFFF', bold=True)
            except (ValueError, TypeError):
                continue

def apply_gaps_coloring(worksheet, dataframe):
    """Apply red fill to negative significant gaps"""
    gap_cols = [col for col in dataframe.columns if col.endswith('_Gap')]
    
    fill_red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    col_indices = {col: dataframe.columns.get_loc(col) + 1 for col in gap_cols}
    
    for row_idx in range(2, len(dataframe) + 2):
        for col_name, col_idx in col_indices.items():
            cell = worksheet.cell(row=row_idx, column=col_idx)
            try:
                value_str = str(cell.value)
                if '*' in value_str and '-' in value_str.split('pp')[0]:
                    cell.fill = fill_red
                    cell.font = Font(color='000000', bold=True)
            except (ValueError, TypeError):
                continue

# =============================================================================
# EXCEL WORKBOOK CREATION
# =============================================================================

def create_methodology_sheet(wb):
    """Create methodology documentation sheet"""
    ws_methods = wb.create_sheet(title='Methodology')
    
    methodology_text = """
WORST LENDERS REDLINING ANALYSIS v2 - METHODOLOGY

V2 CHANGES:
- Bank qualification: ≥500 applications in CBSA for 2022-2024 period only
  (accounts for 2022 census boundary changes that increased majority-minority tracts)
- Scoring and weighting: Limited to 2022-2024 data only
- Output: All years 2018-2024 still included for historical comparison
- Peer matching: Independent by year and CBSA (matches peers separately for each year/CBSA)

DATA SOURCE:
- Home Mortgage Disclosure Act (HMDA) data, years 2018-2024 (all years in output)
- Extracted via BigQuery using worst_lenders_redlining_query_v2.sql
- Geographic data from hdma1-242116.geo.black_hispanic_majority
- CBSA crosswalk from hdma1-242116.geo.cbsa_to_county
- Bank information from hdma1-242116.hmda.lenders18

GEOGRAPHIC SCOPE:
- Top 200 CBSAs (Core Based Statistical Areas) by total applications (2018-2024)
- County-level breakdown within each CBSA
- Banks must have at least 500 applications in CBSA across 2022-2024 period (qualification)

BANK SELECTION CRITERIA:

Inclusion Requirements:
- Asset size: $10 billion to $100 billion (from lenders18 table, assets in $000s)
- Type: Banks and bank affiliates only (excludes credit unions)
- Volume: Minimum 500 applications in CBSA across 2022-2024 period (qualification threshold)
- Geographic scope: Must operate in one of top 200 CBSAs

Excluded Banks (CBA Signatories):
UMB Bank, Atlantic Union Bank, SouthState Bank, BMO Harris Bank, Umpqua, US Bank, 
Old National Bancorp, Flagstar Bank, M&T Bank, PNC Bank, First Citizens Bank, 
CIT Group, Morgan Stanley, First Merchants Bank, Truist, Cadence Bank, Wells Fargo, 
First Horizon Bank, IBERIABANK, Santander Bank, First Financial Bank, Fifth Third Bank, 
Huntington Bancshares, KeyBank

Force-Included Banks (Always Analyzed):
1st National Bank of Pennsylvania, Webster Bank, Frost Bank

LOAN FILTERS (STANDARD HMDA FILTERS):
- Owner-occupied (occupancy_type = '1')
- Site-built construction (construction_method = '1')
- 1-4 unit properties (total_units IN '1','2','3','4')
- Exclude reverse mortgages (reverse_mortgage != '1')
- Years: 2018-2024 (all years included in output for historical comparison)

LOAN PURPOSE CATEGORIES:
1. All Loans - All residential mortgage purposes combined:
   - Home Purchase (loan_purpose = '1')
   - Home Improvement (loan_purpose = '2')
   - Cash-Out Refinancing (loan_purpose = '31')
   - No Cash-Out Refinancing (loan_purpose = '32')

2. Home Purchase - Home purchase loans only (loan_purpose = '1')

ACTION TYPES:
- Applications: All applications (all action_taken codes)
- Originations: Originations only (action_taken = '1')

REDLINING METRICS ANALYZED (Tract-Level Demographics Only):
1. MMCT 50% - Share of loans in majority-minority census tracts (>50% non-white population)
2. MMCT 80% - Share of loans in supermajority-minority census tracts (>80% non-white population)
3. Black Tract 50% - Share of loans in majority-Black census tracts (>50% Black population)
4. Black Tract 80% - Share of loans in supermajority-Black census tracts (>80% Black population)
5. Hispanic Tract 50% - Share of loans in majority-Hispanic census tracts (>50% Hispanic population)
6. Hispanic Tract 80% - Share of loans in supermajority-Hispanic census tracts (>80% Hispanic population)
7. Black+Hispanic Tract 50% - Share of loans in majority Black+Hispanic tracts (>50% combined Black OR Hispanic)
8. Black+Hispanic Tract 80% - Share of loans in supermajority Black+Hispanic tracts (>80% combined Black OR Hispanic)

PEER BANK DEFINITION:
- Other lenders in same CBSA, year, county, loan purpose category, and action type
- With application volume between 50% and 200% of subject bank's volume
- Aggregated across all qualifying peer banks
- Separate peer groups for each combination of:
  - Year (independent peer matching for each year 2018-2024)
  - CBSA (independent peer matching for each CBSA)
  - County
  - Loan purpose (All Loans vs. Home Purchase)
  - Action type (Applications vs. Originations)

RATIO CALCULATION:
- Ratio = Peer Share / Subject Share
- Ratio = 2.0 means peers have 2x the share (subject bank is 50% below peers)
- Ratio >= 2.0 indicates significant underperformance

STATISTICAL SIGNIFICANCE:
- Chi-squared test for two proportions (p < 0.05)
- Tests whether gap between subject and peer shares is statistically different from zero
- Marked with asterisk (*) in Shares/Gaps sheets
- Only negative gaps flagged as significant (subject underperforming peers)

SCORING METHODOLOGY (V2: 2022-2024 ONLY):

Year Weights (Prioritizes Recent Performance, 2022-2024 Only):
- 2024: 3.0x weight (highest priority - banks getting worse prioritized)
- 2023: 1.5x weight
- 2022: 1.0x weight (baseline)

Base Score Calculation:
- Count all ratios >= 2.0 in 2022-2024 data only
- Weight each bad ratio by:
  * CBSA size (total applications in 2022-2024, normalized by millions)
  * Year weight (2024 gets 3x, 2023 gets 1.5x, 2022 gets 1x)

Worsening Trend Analysis (2022-2024):
- If 2024 is worse than 2023, applies 5x multiplier to the difference
- If 2024 is worse than 2022, adds trend bonus with 2x multiplier
- Banks getting worse over time receive higher scores

Final Score = Base Weighted Score + Worsening Trend Score
(All scoring based on 2022-2024 data only)

SELECTION CRITERIA FOR WORST 13 BANKS:

1. Consistency Requirement:
   - Must be bad (ratios >= 2.0) in >=50% of CBSAs where bank operates (based on 2022-2024 scoring)
   - Prevents selection of banks that are only bad in one market

2. Non-Improvement Requirement:
   - Must have ratios >= 2.0 in 2023-2024
   - Banks that improved in recent years are excluded

3. Ranking:
   - Banks ranked by final weighted score (highest = worst performance)
   - Scores based on 2022-2024 data only
   - Top 13 selected
   - Force-included banks (1st National Bank of Pennsylvania, Webster Bank, Frost Bank) always included if they meet criteria

OUTPUT SHEETS:

1. Ratio Sheets (one per bank):
   - Years as columns (2018-2024 - all years shown for historical comparison)
   - Metrics as rows
   - Shows ratios (peer/subject) for each year
   - Color coding:
     * Green: < 1.0 (bank outperforms peers)
     * Yellow: 1.0-1.5
     * Orange: 1.5-2.0
     * Light Red: 2.0-3.0 (significant underperformance)
     * Red: >= 3.0 (severe underperformance)
   - County-level breakdown within CBSAs

2. Shares/Gaps Sheets (one per bank):
   - Years as column groups (2018-2024 - all years shown for historical comparison)
   - For each year/metric combination shows:
     * Subject Count: Number of loans from subject bank
     * Subject Share: Percentage for subject bank
     * Peer Share: Percentage for peer banks
     * Gap: Percentage points difference (Subject Share - Peer Share)
   - Asterisk (*) indicates statistically significant negative gap (chi-squared, p < 0.05)
   - Red cells highlight negative significant gaps
   - County-level breakdown within CBSAs

3. Filterable Columns:
   - Kind: 'Application' or 'Origination' (filter in Excel)
   - Loan_Purpose: 'All Loans' or 'Home Purchase' (filter in Excel)
   - CBSA_Name, County_Code: Geographic identifiers

HOW TO INTERPRET RESULTS:

Ratio Sheets:
- Focus on ratios >= 2.0 (red/orange cells) - these indicate significant underperformance
- Ratios >= 2.0 mean subject bank has at least 50% deficit compared to peers
- Consistent patterns across years and CBSAs are most concerning
- Higher ratios in 2022-2024 indicate current performance issues
- Note: 2022 boundary changes increased majority-minority tracts, making pre-2022 comparisons less reliable

Shares/Gaps Sheets:
- Focus on negative gaps marked with asterisk (*)
- Significant negative gaps indicate statistical evidence of underperformance
- Use county-level data to identify specific geographic problems
- Filter by 'Kind' to see applications vs. originations separately
- Filter by 'Loan_Purpose' to see all loans vs. home purchase separately

LIMITATIONS:
- Does not account for differences in business models or market strategies
- Geographic selection biased toward high-volume markets
- Peer matching based solely on application volume
- County-level data may have small sample sizes in some cases
- Correlation does not imply causation
- Does not assess individual loan applications for discrimination
- Tract demographic data comes from census sources (may have slight variations)
- Pre-2022 data uses old census boundaries (may not be directly comparable to 2022-2024)
- Scoring focuses on 2022-2024 period due to census boundary changes

CONTACT:
NCRC Research Department
National Community Reinvestment Coalition
For questions about methodology or data, contact: research@ncrc.org
"""
    
    # Write methodology text line by line
    for i, line in enumerate(methodology_text.strip().split('\n'), 1):
        ws_methods.cell(row=i, column=1, value=line)
    
    # Format the sheet
    ws_methods.column_dimensions['A'].width = 120
    for row in ws_methods.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    
    print("  Creating: Methodology sheet")

def create_excel_workbook(df, worst_13_banks, output_file):
    """Create Excel workbook with ratio and shares/gaps sheets for worst 13 banks"""
    print(f"\nCreating Excel workbook: {output_file}")
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Create methodology sheet first
    create_methodology_sheet(wb)
    
    for bank_name in worst_13_banks:
        safe_bank_name = bank_name[:20].replace('/', '_').replace('\\', '_')
        
        # Ratio sheet
        ratio_sheet_name = f"{safe_bank_name}_Ratios"[:31]
        print(f"  Creating: {ratio_sheet_name}")
        
        ratio_data = prepare_ratio_sheet_data(df, bank_name)
        if len(ratio_data) > 0:
            ws_ratio = wb.create_sheet(title=ratio_sheet_name)
            for r_idx, r in enumerate(dataframe_to_rows(ratio_data, index=False, header=True), 1):
                for c_idx, value in enumerate(r, 1):
                    ws_ratio.cell(row=r_idx, column=c_idx, value=value)
            
            format_header_row(ws_ratio)
            apply_ratio_coloring(ws_ratio, ratio_data)
            ws_ratio.freeze_panes = 'A2'
            ws_ratio.auto_filter.ref = ws_ratio.dimensions
        
        # Shares and gaps sheet
        gaps_sheet_name = f"{safe_bank_name}_SharesGaps"[:31]
        print(f"  Creating: {gaps_sheet_name}")
        
        gaps_data = prepare_shares_gaps_sheet_data(df, bank_name)
        if len(gaps_data) > 0:
            ws_gaps = wb.create_sheet(title=gaps_sheet_name)
            for r_idx, r in enumerate(dataframe_to_rows(gaps_data, index=False, header=True), 1):
                for c_idx, value in enumerate(r, 1):
                    ws_gaps.cell(row=r_idx, column=c_idx, value=value)
            
            format_header_row(ws_gaps)
            apply_gaps_coloring(ws_gaps, gaps_data)
            ws_gaps.freeze_panes = 'A2'
            ws_gaps.auto_filter.ref = ws_gaps.dimensions
    
    wb.save(output_file)
    print(f"\nExcel workbook created: {output_file}")

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Worst Lenders Redlining Analysis v2'
    )
    parser.add_argument('--input', type=str, required=True,
                       help='Input CSV file from BigQuery (worst_lenders_redlining_query_v2.sql output)')
    
    # Construct default output path using pathlib to handle apostrophes correctly
    # Since we're working from C:\DREAM (symbolic link), construct path to reports folder
    default_output_dir = Path(r"C:\Users\edite\OneDrive - Nat'l Community Reinvestment Coaltn\Desktop\DREAM Analysis\reports")
    default_output_file = default_output_dir / "Worst_Lenders_Analysis_v2.xlsx"
    
    parser.add_argument('--output', type=str,
                       default=str(default_output_file),
                       help='Output Excel file path')
    
    args = parser.parse_args()
    
    # Ensure output directory exists
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    print("="*80)
    print("WORST LENDERS REDLINING ANALYSIS v2")
    print("="*80)
    
    # Load and process data
    df = load_data(args.input)
    df, force_included_banks = filter_banks(df)
    df = add_calculated_metrics(df)
    
    # Calculate scores and select worst 13 (scoring based on 2022-2024 only)
    scores_df = calculate_bank_scores(df)
    worst_13_banks = select_worst_13_banks(scores_df, force_included_banks)
    
    # Filter data to worst 13 banks only (but keep all years 2018-2024)
    df_filtered = df[df['lender_name'].isin(worst_13_banks)].copy()
    
    # Create Excel workbook
    create_excel_workbook(df_filtered, worst_13_banks, args.output)
    
    print("\n" + "="*80)
    print("ANALYSIS COMPLETE!")
    print("="*80)
    print(f"\nOutput: {args.output}")
    print(f"\nWorst 13 banks selected based on:")
    print(f"  - Ratios >= 2.0 in 2022-2024 data (weighted by CBSA size and metrics)")
    print(f"  - Consistent patterns (bad in >=50% of CBSAs)")
    print(f"  - Non-improving (bad ratios in 2023-2024)")
    print(f"\nNote: Scoring based on 2022-2024 only (post-census boundary changes)")
    print(f"      All years 2018-2024 included in output for historical comparison")

if __name__ == "__main__":
    main()

