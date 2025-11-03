"""Debug Bank Details sheet parsing"""

import openpyxl

ticket_file = r"C:\Users\edite\OneDrive - Nat'l Community Reinvestment Coaltn\Desktop\DREAM Analysis\PNC+FirstBank Research Ticket_CBA.xlsx"

wb = openpyxl.load_workbook(ticket_file, data_only=True)
ws = wb["Bank Details"]

print("="*80)
print("DEBUGGING BANK DETAILS SHEET")
print("="*80)

print("\nRows 3-10, Columns 1-4:")
for row_idx in range(3, 11):
    row_data = []
    for col_idx in range(1, 5):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell_value = cell.value
        cell_type = type(cell_value).__name__
        row_data.append(f"{cell_value} ({cell_type})" if cell_value else "None")
    print(f"Row {row_idx}: {row_data}")

print("\n" + "="*80)
print("Testing specific cells:")
print("="*80)
print(f"Row 4, Col 1: {repr(ws.cell(4, 1).value)}")
print(f"Row 4, Col 2: {repr(ws.cell(4, 2).value)}")
print(f"Row 4, Col 3: {repr(ws.cell(4, 3).value)}")
print(f"Row 4, Col 4: {repr(ws.cell(4, 4).value)}")

print(f"\nRow 6, Col 1: {repr(ws.cell(6, 1).value)}")
print(f"Row 6, Col 2: {repr(ws.cell(6, 2).value)}")
print(f"Row 6, Col 3: {repr(ws.cell(6, 3).value)}")
print(f"Row 6, Col 4: {repr(ws.cell(6, 4).value)}")

print(f"\nRow 7, Col 1: {repr(ws.cell(7, 1).value)}")
print(f"Row 7, Col 2: {repr(ws.cell(7, 2).value)}")
print(f"Row 7, Col 3: {repr(ws.cell(7, 3).value)}")
print(f"Row 7, Col 4: {repr(ws.cell(7, 4).value)}")

print(f"\nRow 8, Col 1: {repr(ws.cell(8, 1).value)}")
print(f"Row 8, Col 2: {repr(ws.cell(8, 2).value)}")
print(f"Row 8, Col 3: {repr(ws.cell(8, 3).value)}")
print(f"Row 8, Col 4: {repr(ws.cell(8, 4).value)}")

