"""
Parse assessment areas from the ticket Excel file's "Assessment Areas" sheet.
This may be more reliable than parsing PDFs.
"""

import openpyxl
from pathlib import Path
import sys
import json
import pandas as pd

def parse_assessment_areas_sheet(excel_file):
    """
    Parse the Assessment Areas sheet from the ticket Excel file.
    """
    excel_path = Path(excel_file)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    if "Assessment Areas" not in wb.sheetnames:
        raise ValueError("'Assessment Areas' sheet not found in Excel file")
    
    ws = wb["Assessment Areas"]
    
    print(f"\nParsing 'Assessment Areas' sheet...")
    print(f"  Dimensions: {ws.max_row} rows x {ws.max_column} columns")
    
    # Based on inspection, structure is:
    # Columns A-C: PNC data (State, @PNC AA, County State)
    # Columns E-G: FirstBank data (State, @FirstBank AA, County State)
    
    pnc_areas = []
    firstbank_areas = []
    
    # Read row by row
    for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
        # PNC columns (A-C)
        state_pnc = ws.cell(row=row_idx, column=1).value
        cbsa_pnc = ws.cell(row=row_idx, column=2).value  # @PNC AA column
        county_pnc = ws.cell(row=row_idx, column=3).value  # County State column
        
        # FirstBank columns (E-G)
        state_fb = ws.cell(row=row_idx, column=5).value
        cbsa_fb = ws.cell(row=row_idx, column=6).value  # @FirstBank AA column
        county_fb = ws.cell(row=row_idx, column=7).value  # County State column
        
        # Parse PNC data
        if state_pnc or cbsa_pnc or county_pnc:
            if county_pnc:
                # Parse county name and state from "County Name, State" format
                county_info = parse_county_string(county_pnc)
                if county_info:
                    pnc_areas.append({
                        'state': state_pnc,
                        'cbsa': cbsa_pnc,
                        'county_name': county_info.get('county_name'),
                        'state_code': county_info.get('state_code'),
                        'state_name': county_info.get('state_name'),
                        'full_county_string': county_pnc
                    })
        
        # Parse FirstBank data
        if state_fb or cbsa_fb or county_fb:
            if county_fb:
                county_info = parse_county_string(county_fb)
                if county_info:
                    firstbank_areas.append({
                        'state': state_fb,
                        'cbsa': cbsa_fb,
                        'county_name': county_info.get('county_name'),
                        'state_code': county_info.get('state_code'),
                        'state_name': county_info.get('state_name'),
                        'full_county_string': county_fb
                    })
    
    print(f"\n  Found {len(pnc_areas)} PNC assessment area entries")
    print(f"  Found {len(firstbank_areas)} FirstBank assessment area entries")
    
    # Show samples
    if pnc_areas:
        print(f"\n  PNC sample entries (first 5):")
        for entry in pnc_areas[:5]:
            print(f"    - {entry['county_name']}, {entry.get('state_name', entry.get('state_code', 'N/A'))}, CBSA: {entry.get('cbsa', 'N/A')}")
    
    if firstbank_areas:
        print(f"\n  FirstBank sample entries (first 5):")
        for entry in firstbank_areas[:5]:
            print(f"    - {entry['county_name']}, {entry.get('state_name', entry.get('state_code', 'N/A'))}, CBSA: {entry.get('cbsa', 'N/A')}")
    
    return {
        'pnc': pnc_areas,
        'firstbank': firstbank_areas
    }

def parse_county_string(county_string):
    """
    Parse county string like "Dallas County, Alabama" or "Maricopa County, Arizona"
    Returns dict with county_name, state_code, state_name
    """
    if not county_string:
        return None
    
    county_str = str(county_string).strip()
    
    # Pattern: "County Name County, State Name"
    import re
    pattern = r'(.+?)\s+County,\s+(.+)'
    match = re.match(pattern, county_str)
    
    if match:
        county_name = match.group(1).strip()
        state = match.group(2).strip()
        
        # Try to get state code from state name
        state_code = get_state_code(state)
        
        return {
            'county_name': county_name,
            'state_name': state,
            'state_code': state_code
        }
    
    return None

def get_state_code(state_name):
    """Convert state name to 2-letter code"""
    state_map = {
        'Alabama': '01', 'Arizona': '04', 'Arkansas': '05', 'California': '06',
        'Colorado': '08', 'Connecticut': '09', 'Delaware': '10', 'Florida': '12',
        'Georgia': '13', 'Idaho': '16', 'Illinois': '17', 'Indiana': '18',
        'Iowa': '19', 'Kansas': '20', 'Kentucky': '21', 'Louisiana': '22',
        'Maine': '23', 'Maryland': '24', 'Massachusetts': '25', 'Michigan': '26',
        'Minnesota': '27', 'Mississippi': '28', 'Missouri': '29', 'Montana': '30',
        'Nebraska': '31', 'Nevada': '32', 'New Hampshire': '33', 'New Jersey': '34',
        'New Mexico': '35', 'New York': '36', 'North Carolina': '37', 'North Dakota': '38',
        'Ohio': '39', 'Oklahoma': '40', 'Oregon': '41', 'Pennsylvania': '42',
        'Rhode Island': '44', 'South Carolina': '45', 'South Dakota': '46', 'Tennessee': '47',
        'Texas': '48', 'Utah': '49', 'Vermont': '50', 'Virginia': '51',
        'Washington': '53', 'West Virginia': '54', 'Wisconsin': '55', 'Wyoming': '56',
        'District of Columbia': '11'
    }
    
    return state_map.get(state_name, None)

def main():
    """Main function"""
    if len(sys.argv) < 2:
        print("Usage: python parse_assessment_areas_from_ticket.py <ticket_excel_file>")
        print("\nExample:")
        print('  python parse_assessment_areas_from_ticket.py "PNC+FirstBank merger research ticket.xlsx"')
        sys.exit(1)
    
    excel_file = sys.argv[1]
    
    try:
        results = parse_assessment_areas_sheet(excel_file)
        
        # Save results
        output_file = Path(excel_file).parent / "assessment_areas_from_ticket.json"
        with open(output_file, 'w') as f:
            json.dump(results, f, indent=2)
        
        print(f"\n\nResults saved to: {output_file}")
        print(f"\nNote: County codes (GEOIDs) will need to be looked up using county name + state code.")
        
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()


