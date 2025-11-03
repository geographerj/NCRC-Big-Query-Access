"""
Parse Cadence Bank Assessment Areas from PDF
File: "412025 List of Census Tracts - Public File.pdf"

This script extracts census tract information and maps to counties and CBSAs.
"""

import pdfplumber
from pathlib import Path
import sys
import json
import re
import pandas as pd

def parse_cadence_assessment_areas_pdf(pdf_path):
    """
    Parse Cadence Bank assessment areas PDF.
    This appears to be a list of census tracts.
    We need to extract tract information and map to counties/CBSAs.
    """
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF file not found: {pdf_path}")
    
    print(f"\nParsing Cadence Assessment Areas PDF: {pdf_path}")
    
    try:
        import pdfplumber
    except ImportError:
        raise ImportError("Please install pdfplumber: pip install pdfplumber")
    
    assessment_areas = []
    tracts_data = []
    
    with pdfplumber.open(pdf_path) as pdf:
        print(f"  Total pages: {len(pdf.pages)}")
        
        # Extract text and tables from all pages
        full_text = ""
        all_tables = []
        
        for page_num, page in enumerate(pdf.pages, 1):
            text = page.extract_text()
            if text:
                full_text += "\n" + text
            
            # Extract tables (census tract data often in tables)
            tables = page.extract_tables()
            if tables:
                for table_idx, table in enumerate(tables):
                    if table and len(table) > 0:
                        all_tables.append({
                            'page': page_num,
                            'table': table_idx,
                            'data': table
                        })
        
        print(f"  Extracted {len(pdf.pages)} pages of text")
        print(f"  Found {len(all_tables)} tables")
        
        # Try to extract census tract information
        # Census tract GEOID format: SSCCCTTTTTT (11 digits)
        # SS = State code (2 digits)
        # CCC = County code (3 digits)
        # TTTTTT = Tract code (6 digits)
        
        tract_pattern = r'\b(\d{11})\b'  # 11-digit GEOID
        tract_matches = re.finditer(tract_pattern, full_text)
        
        tracts_found = []
        for match in tract_matches:
            geoid = match.group(1)
            # Extract state and county from GEOID
            state_code = geoid[:2]
            county_code = geoid[2:5]
            tract_code = geoid[5:11]
            
            tracts_found.append({
                'full_geoid': geoid,
                'state_code': state_code,
                'county_code': county_code,
                'tract_code': tract_code,
                'geoid5': state_code + county_code  # 5-digit county GEOID
            })
        
        print(f"  Found {len(tracts_found)} census tract GEOIDs")
        
        # Also try to extract from tables
        for table_info in all_tables:
            table = table_info['data']
            if len(table) > 1:  # Has header + data
                # Look for GEOID column or tract information
                header_row = table[0] if table else []
                
                # Find columns that might contain GEOIDs or tract info
                geoid_col_idx = None
                county_col_idx = None
                state_col_idx = None
                
                for col_idx, header in enumerate(header_row):
                    if header:
                        header_lower = str(header).lower()
                        if 'geoid' in header_lower or 'tract' in header_lower:
                            geoid_col_idx = col_idx
                        if 'county' in header_lower:
                            county_col_idx = col_idx
                        if 'state' in header_lower:
                            state_col_idx = col_idx
                
                # Extract data rows - try all columns in case we didn't find a header
                for row in table[1:]:
                    # Check all cells in the row for 11-digit GEOIDs
                    for cell_idx, cell in enumerate(row):
                        if cell:
                            cell_str = str(cell).strip()
                            # Check if it's an 11-digit GEOID (with or without spaces/dashes)
                            cell_clean = re.sub(r'[\s\-]', '', cell_str)
                            if re.match(r'^\d{11}$', cell_clean):
                                state_code = cell_clean[:2]
                                county_code = cell_clean[2:5]
                                tract_code = cell_clean[5:11]
                                
                                tracts_found.append({
                                    'full_geoid': cell_clean,
                                    'state_code': state_code,
                                    'county_code': county_code,
                                    'tract_code': tract_code,
                                    'geoid5': state_code + county_code
                                })
                            
                            # Also try extracting from longer strings that might contain GEOID
                            geoid_match = re.search(r'(\d{11})', cell_clean)
                            if geoid_match:
                                geoid_val = geoid_match.group(1)
                                state_code = geoid_val[:2]
                                county_code = geoid_val[2:5]
                                tract_code = geoid_val[5:11]
                                
                                tracts_found.append({
                                    'full_geoid': geoid_val,
                                    'state_code': state_code,
                                    'county_code': county_code,
                                    'tract_code': tract_code,
                                    'geoid5': state_code + county_code
                                })
        
        # Load county crosswalk to map GEOID5 to county names and CBSAs
        crosswalk_path = Path(__file__).parent.parent / 'data' / 'reference' / 'CBSA_to_County_Mapping.csv'
        
        if crosswalk_path.exists():
            crosswalk_df = pd.read_csv(crosswalk_path)
            crosswalk_df['Geoid5'] = crosswalk_df['Geoid5'].astype(str).str.zfill(5)
            
            # Create mapping from GEOID5 to county info
            geoid5_to_county = {}
            for _, row in crosswalk_df.iterrows():
                geoid5 = str(row['Geoid5']).zfill(5)
                if geoid5 not in geoid5_to_county:
                    geoid5_to_county[geoid5] = {
                        'county_name': row.get('County Name', ''),
                        'state_code': geoid5[:2],
                        'state_name': row.get('State', ''),
                        'cbsa_code': str(row.get('Cbsa Code', '')),
                        'cbsa_name': row.get('Cbsa', '')
                    }
            
            # Group tracts by county (GEOID5)
            counties_with_tracts = {}
            for tract in tracts_found:
                geoid5 = tract['geoid5']
                if geoid5 in geoid5_to_county:
                    if geoid5 not in counties_with_tracts:
                        counties_with_tracts[geoid5] = {
                            'county_info': geoid5_to_county[geoid5],
                            'tract_count': 0,
                            'tracts': []
                        }
                    counties_with_tracts[geoid5]['tract_count'] += 1
                    counties_with_tracts[geoid5]['tracts'].append(tract)
                else:
                    # Still add if we can't map it
                    if geoid5 not in counties_with_tracts:
                        counties_with_tracts[geoid5] = {
                            'county_info': {
                                'county_name': f'Unknown County (GEOID: {geoid5})',
                                'state_code': geoid5[:2],
                                'state_name': '',
                                'cbsa_code': '',
                                'cbsa_name': ''
                            },
                            'tract_count': 0,
                            'tracts': []
                        }
                    counties_with_tracts[geoid5]['tract_count'] += 1
            
            # Convert to assessment areas format
            for geoid5, data in counties_with_tracts.items():
                county_info = data['county_info']
                
                assessment_areas.append({
                    'bank_name': 'Cadence Bank',
                    'cbsa_name': county_info.get('cbsa_name', ''),
                    'cbsa_code': county_info.get('cbsa_code', ''),
                    'state_code': county_info.get('state_code', ''),
                    'state_name': county_info.get('state_name', ''),
                    'county_name': county_info.get('county_name', ''),
                    'county_code': geoid5,
                    'tract_count': data['tract_count']
                })
            
            print(f"\n  Mapped to {len(assessment_areas)} unique counties")
            print(f"  Total census tracts: {len(tracts_found)}")
            
            # Show sample counties
            if assessment_areas:
                print(f"\n  Sample counties (first 10):")
                for aa in assessment_areas[:10]:
                    print(f"    - {aa['county_name']}, {aa['state_name']}, CBSA: {aa['cbsa_code']} ({aa['cbsa_name']}), Tracts: {aa['tract_count']}")
        else:
            print(f"\n  WARNING: County crosswalk not found at {crosswalk_path}")
            print(f"  Cannot map GEOID5 to county names and CBSAs")
            print(f"  Found {len(set(t['geoid5'] for t in tracts_found))} unique counties by GEOID5")
    
    return assessment_areas

def add_to_excel_ticket(ticket_file, assessment_areas):
    """Add assessment areas to the Excel ticket file"""
    from openpyxl import load_workbook
    
    ticket_path = Path(ticket_file)
    if not ticket_path.exists():
        print(f"\nERROR: Ticket file not found: {ticket_path}")
        return False
    
    print(f"\nAdding {len(assessment_areas)} assessment area entries to ticket...")
    
    wb = load_workbook(ticket_path)
    
    if "Assessment Areas" not in wb.sheetnames:
        ws = wb.create_sheet("Assessment Areas")
        # Add headers
        headers = ['Bank Name', 'CBSA Name', 'CBSA Code', 'State Name', 'County Name', 'County Code (GEOID)']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
    else:
        ws = wb["Assessment Areas"]
    
    # Find next empty row
    next_row = ws.max_row + 1
    if ws.max_row == 1:
        next_row = 2  # Start after header
    
    # Add assessment areas
    for aa in assessment_areas:
        ws.cell(row=next_row, column=1, value=aa['bank_name'])
        ws.cell(row=next_row, column=2, value=aa['cbsa_name'])
        ws.cell(row=next_row, column=3, value=aa['cbsa_code'])
        ws.cell(row=next_row, column=4, value=aa['state_name'])
        ws.cell(row=next_row, column=5, value=aa['county_name'])
        ws.cell(row=next_row, column=6, value=aa['county_code'])
        next_row += 1
    
    wb.save(ticket_path)
    print(f"  Saved {len(assessment_areas)} entries to 'Assessment Areas' sheet")
    
    return True

def main():
    if len(sys.argv) < 2:
        pdf_path = Path(__file__).parent.parent / "412025 List of Census Tracts - Public File.pdf"
        if not pdf_path.exists():
            print("Usage: python parse_cadence_assessment_areas.py <pdf_file>")
            print(f"\nOr place PDF at: {pdf_path}")
            sys.exit(1)
    else:
        pdf_path = Path(sys.argv[1])
    
    # Parse PDF
    assessment_areas = parse_cadence_assessment_areas_pdf(pdf_path)
    
    # Save to JSON for inspection
    output_json = pdf_path.parent / "cadence_assessment_areas_parsed.json"
    with open(output_json, 'w') as f:
        json.dump(assessment_areas, f, indent=2)
    print(f"\n  Saved parsed data to: {output_json}")
    
    # Add to Excel ticket if it exists
    ticket_file = Path(__file__).parent.parent / "Huntington+Cadence merger research ticket.xlsx"
    if ticket_file.exists():
        add_to_excel_ticket(ticket_file, assessment_areas)
        print(f"\n[SUCCESS] Added assessment areas to ticket: {ticket_file}")
    else:
        print(f"\n⚠️  Ticket file not found: {ticket_file}")
        print(f"   Assessment areas saved to JSON: {output_json}")
        print(f"   You can manually add them to the ticket later")
    
        # If no GEOIDs found in text, try extracting directly from tables more thoroughly
        if len(tracts_found) == 0 and all_tables:
            print("\n  No GEOIDs found in text search. Trying direct table extraction...")
            
            for table_info in all_tables[:10]:  # Check first 10 tables as sample
                table = table_info['data']
                if not table or len(table) < 2:
                    continue
                
                # Try to extract from first few rows
                for row_idx, row in enumerate(table[:20]):  # Check first 20 rows
                    for cell in row:
                        if cell:
                            cell_str = str(cell).strip()
                            # Look for patterns like "48-201-001234" or "48201001234" or "48201 001234"
                            patterns = [
                                r'(\d{2})[-\s]?(\d{3})[-\s]?(\d{6})',  # State-County-Tract with separators
                                r'(\d{11})',  # Direct 11-digit
                                r'(\d{5})\s+(\d{6})',  # GEOID5 + Tract separately
                            ]
                            
                            for pattern in patterns:
                                matches = re.finditer(pattern, cell_str)
                                for match in matches:
                                    if len(match.groups()) == 3:
                                        # Format: SS-CCC-TTTTTT
                                        state_code = match.group(1)
                                        county_code = match.group(2)
                                        tract_code = match.group(3)
                                        geoid = state_code + county_code + tract_code
                                        
                                        if len(geoid) == 11:
                                            tracts_found.append({
                                                'full_geoid': geoid,
                                                'state_code': state_code,
                                                'county_code': county_code,
                                                'tract_code': tract_code,
                                                'geoid5': state_code + county_code
                                            })
                                    elif len(match.groups()) == 1:
                                        # Direct 11-digit or 5+6 format
                                        if len(match.group(1)) == 11:
                                            geoid = match.group(1)
                                            tracts_found.append({
                                                'full_geoid': geoid,
                                                'state_code': geoid[:2],
                                                'county_code': geoid[2:5],
                                                'tract_code': geoid[5:11],
                                                'geoid5': geoid[:5]
                                            })
            
            print(f"  Found {len(tracts_found)} tracts after table extraction")
        
        print("\n" + "="*80)
        print("PARSING COMPLETE")
        print("="*80)
    print(f"\nTotal assessment area entries: {len(assessment_areas)}")
    print(f"Unique counties: {len(set(aa['county_code'] for aa in assessment_areas))}")
    print(f"Unique CBSAs: {len(set(aa['cbsa_code'] for aa in assessment_areas if aa['cbsa_code']))}")

if __name__ == "__main__":
    main()

