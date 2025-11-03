"""
Extract Huntington assessment areas from CRA PE document
This is a large document (352 pages) so we'll search for assessment area tables
"""

import pdfplumber
from pathlib import Path
import pandas as pd
import json
import re
from collections import defaultdict

def extract_huntington_assessment_areas(pdf_path):
    """Extract assessment areas from Huntington CRA PE"""
    pdf_path = Path(pdf_path)
    
    print(f"\nExtracting Huntington assessment areas from: {pdf_path}")
    
    assessment_areas = []
    seen_counties = set()
    counties_found = []
    
    with pdfplumber.open(pdf_path) as pdf:
        print(f"Total pages: {len(pdf.pages)}")
        
        # Strategy: Look for pages with assessment area information
        # CRA PE documents often have assessment area summaries in early pages
        # and then detailed tables throughout
        
        # Check first 50 pages more thoroughly, then sample rest
        pages_to_check_thoroughly = list(range(1, min(51, len(pdf.pages) + 1)))
        pages_to_sample = list(range(51, len(pdf.pages) + 1, 10))  # Every 10th page
        
        all_pages_to_check = pages_to_check_thoroughly + pages_to_sample
        
        for page_num in all_pages_to_check:
            page = pdf.pages[page_num - 1]  # 0-indexed
            text = page.extract_text()
            tables = page.extract_tables()
            
            # Extract from text - look for county patterns
            if text:
                # Pattern 1: "County Name, State" or "County Name County, State"
                # Be more specific - only match if it looks like a county (has "County" or is a known county name)
                county_patterns = [
                    r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+County,\s+([A-Z]{2})',  # Must have "County"
                ]
                
                # Also try: County Name, State Code (numeric) - these are more reliable
                county_with_state_code = r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+County,\s+([A-Z]{2})|([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+County\s+\((\d{2})\)'
                
                for pattern in county_patterns:
                    matches = re.finditer(pattern, text)
                    for match in matches:
                        county_name = match.group(1).strip()
                        state_code = match.group(2).strip()
                        
                        # Try to find CBSA code nearby
                        # Look for 5-digit number within 200 characters
                        match_end = match.end()
                        nearby_text = text[max(0, match.start()-100):match_end+100]
                        cbsa_match = re.search(r'\b(\d{5})\b', nearby_text)
                        cbsa_code = cbsa_match.group(1) if cbsa_match else ""
                        
                        if cbsa_code != '99999':
                            counties_found.append({
                                'county_name': county_name,
                                'state_code': state_code,
                                'cbsa_code': cbsa_code,
                                'page': page_num,
                                'source': 'text'
                            })
            
            # Extract from tables
            for table in tables:
                if not table or len(table) < 2:
                    continue
                
                # Look for tables with county/CBSA information
                header_row = table[0] if table else []
                header_text = ' '.join([str(cell).lower() if cell else '' for cell in header_row])
                
                # Check if this looks like assessment area table
                has_county = 'county' in header_text
                has_cbsa = 'cbsa' in header_text or 'msa' in header_text
                has_state = 'state' in header_text
                
                if has_county and (has_cbsa or has_state):
                    # Try to identify column indices
                    county_col = None
                    state_col = None
                    cbsa_col = None
                    geoid_col = None
                    
                    for col_idx, header in enumerate(header_row):
                        if header:
                            header_lower = str(header).lower()
                            if 'county' in header_lower and 'name' in header_lower:
                                county_col = col_idx
                            elif 'state' in header_lower and ('code' in header_lower or 'number' in header_lower):
                                state_col = col_idx
                            elif 'cbsa' in header_lower or 'msa' in header_lower:
                                cbsa_col = col_idx
                            elif 'geoid' in header_lower or 'fips' in header_lower:
                                geoid_col = col_idx
                    
                    # Extract data rows
                    for row in table[1:]:
                        if len(row) < max([c for c in [county_col, state_col, cbsa_col, geoid_col] if c is not None] or [0]):
                            continue
                        
                        county_name = None
                        state_code = None
                        cbsa_code = None
                        geoid5 = None
                        
                        if county_col is not None and county_col < len(row):
                            county_name = str(row[county_col]).strip().replace(' County', '') if row[county_col] else None
                        
                        if state_col is not None and state_col < len(row):
                            state_cell = str(row[state_col]).strip() if row[state_col] else ""
                            # Extract 2-digit state code
                            state_match = re.search(r'(\d{2})', state_cell)
                            if state_match:
                                state_code = state_match.group(1).zfill(2)
                        
                        if cbsa_col is not None and cbsa_col < len(row):
                            cbsa_cell = str(row[cbsa_col]).strip() if row[cbsa_col] else ""
                            cbsa_match = re.search(r'(\d{5})', cbsa_cell)
                            if cbsa_match:
                                cbsa_code = cbsa_match.group(1)
                        
                        if geoid_col is not None and geoid_col < len(row):
                            geoid_cell = str(row[geoid_col]).strip() if row[geoid_col] else ""
                            # Could be 5-digit or 11-digit
                            geoid_clean = re.sub(r'[\s\-]', '', geoid_cell)
                            if len(geoid_clean) >= 5:
                                geoid5 = geoid_clean[:5]
                                if not state_code:
                                    state_code = geoid5[:2]
                        
                        # If we have county name and state, try to match
                        if county_name and (state_code or geoid5):
                            counties_found.append({
                                'county_name': county_name,
                                'state_code': state_code if state_code else (geoid5[:2] if geoid5 else None),
                                'cbsa_code': cbsa_code if cbsa_code else '',
                                'geoid5': geoid5,
                                'page': page_num,
                                'source': 'table'
                            })
            
            if page_num % 50 == 0:
                print(f"  Processed {page_num} pages, found {len(counties_found)} county mentions...")
    
    print(f"\nFound {len(counties_found)} county mentions in document")
    
    # Load crosswalk to map counties to GEOIDs and CBSAs
    crosswalk_path = Path(__file__).parent.parent / 'data' / 'reference' / 'CBSA_to_County_Mapping.csv'
    
    if crosswalk_path.exists():
        crosswalk_df = pd.read_csv(crosswalk_path)
        crosswalk_df['Geoid5'] = crosswalk_df['Geoid5'].astype(str).str.zfill(5)
        
        # Get actual column names from crosswalk
        geoid_col = 'Geoid5'
        state_code_col = 'State Code'
        county_state_col = 'County State'  # Format: "County Name State"
        cbsa_code_col = 'Cbsa Code'
        cbsa_name_col = 'Cbsa'
        
        # Create mappings
        geoid5_to_info = {}
        county_state_to_info = {}
        
        for _, row in crosswalk_df.iterrows():
            geoid5 = str(row[geoid_col]).zfill(5)
            county_state = str(row.get(county_state_col, '')).strip()  # "County Name State"
            state_code = str(row.get(state_code_col, '')).strip()
            cbsa_code = str(row.get(cbsa_code_col, '')).strip()
            cbsa_name = str(row.get(cbsa_name_col, '')).strip()
            
            # Extract county name and state name from "County State" column
            # Format: "Autauga County Alabama" -> county="Autauga", state="Alabama"
            county_name = None
            state_name = None
            if county_state:
                parts = county_state.rsplit(' ', 1)  # Split on last space
                if len(parts) == 2:
                    county_name = parts[0].replace(' County', '').strip()
                    state_name = parts[1].strip()
            
            if geoid5 not in geoid5_to_info:
                geoid5_to_info[geoid5] = {
                    'county_name': county_name or '',
                    'state_code': state_code,
                    'state_name': state_name or '',
                    'cbsa_code': cbsa_code,
                    'cbsa_name': cbsa_name
                }
            
            # Also create name+state lookup
            if county_name and state_code:
                key = f"{county_name}|{state_code}"
                if key not in county_state_to_info:
                    county_state_to_info[key] = {
                        'county_code': geoid5,
                        'cbsa_code': cbsa_code,
                        'cbsa_name': cbsa_name,
                        'state_name': state_name or ''
                    }
        
        # Show sample of what was found before mapping
        print(f"\nSample counties found (before mapping):")
        for county in counties_found[:10]:
            print(f"  - {county.get('county_name', 'N/A')}, State: {county.get('state_code', 'N/A')}, CBSA: {county.get('cbsa_code', 'N/A')}, GEOID5: {county.get('geoid5', 'N/A')}")
        
        # Map counties found
        for county_info in counties_found:
            mapped_info = None
            
            # Try GEOID5 first
            if county_info.get('geoid5'):
                geoid5 = str(county_info['geoid5']).zfill(5)
                if geoid5 in geoid5_to_info:
                    mapped_info = geoid5_to_info[geoid5].copy()
                    mapped_info['county_code'] = geoid5
            
            # Try county name + state code (with variations)
            if not mapped_info and county_info.get('county_name') and county_info.get('state_code'):
                county_name = str(county_info['county_name']).replace(' County', '').strip()
                state_code = str(county_info['state_code']).strip()
                
                # If state_code is numeric (like "17" for Illinois), use as-is
                # If it's text (like "IL"), need to convert
                state_code_numeric = None
                if state_code.isdigit():
                    state_code_numeric = state_code.zfill(2)
                else:
                    # Try to find state code from crosswalk
                    # Check actual column names in crosswalk
                    state_col = None
                    state_code_col = None
                    for col in crosswalk_df.columns:
                        if 'state' in col.lower() and 'code' not in col.lower():
                            state_col = col
                        elif 'state' in col.lower() and 'code' in col.lower():
                            state_code_col = col
                    
                    # Use state_code_col from crosswalk
                    state_code_col = 'State Code'
                    county_state_col = 'County State'
                    
                    # Search for state in County State column (e.g., "County Alabama")
                    state_rows = crosswalk_df[crosswalk_df[county_state_col].str.contains(f' {state_code.upper()}$', na=False, regex=True)]
                    if len(state_rows) > 0:
                        state_code_numeric = str(state_rows.iloc[0].get(state_code_col, '')).strip().zfill(2)
                
                if state_code_numeric:
                    # Try exact match
                    key = f"{county_name}|{state_code_numeric}"
                    if key in county_state_to_info:
                        mapped_info = county_state_to_info[key].copy()
                        mapped_info['county_name'] = county_name
                        mapped_info['state_code'] = state_code_numeric
                    else:
                        # Try with "County" suffix
                        key_with_county = f"{county_name} County|{state_code_numeric}"
                        if key_with_county in county_state_to_info:
                            mapped_info = county_state_to_info[key_with_county].copy()
                            mapped_info['county_name'] = county_name
                            mapped_info['state_code'] = state_code_numeric
            
            # Use CBSA from county_info if available
            if mapped_info and county_info.get('cbsa_code'):
                mapped_info['cbsa_code'] = county_info['cbsa_code']
                # Try to get CBSA name from crosswalk
                cbsa_code = county_info['cbsa_code']
                cbsa_rows = crosswalk_df[crosswalk_df['Cbsa Code'].astype(str) == str(cbsa_code)]
                if len(cbsa_rows) > 0:
                    mapped_info['cbsa_name'] = str(cbsa_rows.iloc[0].get('Cbsa', mapped_info.get('cbsa_name', '')))
            
            # If still no match but we have GEOID5 components, try to construct it
            if not mapped_info and county_info.get('state_code') and county_info.get('county_name'):
                # Try to find county in crosswalk by state and name
                state_code_numeric = str(county_info['state_code']).strip()
                if state_code_numeric.isdigit():
                    state_code_numeric = state_code_numeric.zfill(2)
                    county_name = str(county_info['county_name']).replace(' County', '').strip()
                    
                    # Search in crosswalk using actual column names
                    state_code_col = 'State Code'
                    county_state_col = 'County State'
                    
                    # Filter by state code
                    state_rows = crosswalk_df[crosswalk_df[state_code_col].astype(str) == state_code_numeric]
                    
                    # Search for county name in County State column
                    # Format: "County Name State" - look for county name at start
                    county_rows = state_rows[state_rows[county_state_col].str.replace(' County', '').str.split().str[0].str.lower() == county_name.lower()]
                    
                    if len(county_rows) > 0:
                        row = county_rows.iloc[0]
                        # Extract state name from County State column
                        county_state = str(row.get('County State', ''))
                        state_name = county_state.rsplit(' ', 1)[1] if ' ' in county_state else ''
                        
                        mapped_info = {
                            'county_name': county_name,
                            'state_code': state_code_numeric,
                            'state_name': state_name,
                            'cbsa_code': str(row.get('Cbsa Code', county_info.get('cbsa_code', ''))),
                            'cbsa_name': str(row.get('Cbsa', '')),
                            'county_code': str(row['Geoid5']).zfill(5)
                        }
            
            if mapped_info and mapped_info.get('county_code'):
                key = (mapped_info.get('county_code'), mapped_info.get('cbsa_code'))
                if key not in seen_counties:
                    seen_counties.add(key)
                    assessment_areas.append({
                        'bank_name': 'The Huntington National Bank',
                        'cbsa_name': mapped_info.get('cbsa_name', ''),
                        'cbsa_code': mapped_info.get('cbsa_code', ''),
                        'state_code': mapped_info.get('state_code', ''),
                        'state_name': mapped_info.get('state_name', ''),
                        'county_name': mapped_info.get('county_name', county_info.get('county_name', '')),
                        'county_code': mapped_info.get('county_code', '')
                    })
        
        print(f"\nMapped to {len(assessment_areas)} unique county assessment areas for Huntington")
        
        # Show sample
        if assessment_areas:
            print("\nSample counties (first 10):")
            for aa in assessment_areas[:10]:
                print(f"  - {aa['county_name']}, {aa['state_name']}, CBSA: {aa['cbsa_code']} ({aa['cbsa_name']})")
    else:
        print("WARNING: Crosswalk not found, cannot map counties")
    
    return assessment_areas

def update_excel_ticket(assessment_areas):
    """Add Huntington assessment areas to Excel ticket"""
    from openpyxl import load_workbook
    
    ticket_file = Path(__file__).parent.parent / "Huntington+Cadence merger research ticket.xlsx"
    
    if not ticket_file.exists():
        print(f"ERROR: Ticket file not found: {ticket_file}")
        return False
    
    wb = load_workbook(ticket_file)
    
    if "Assessment Areas" not in wb.sheetnames:
        print("ERROR: Assessment Areas sheet not found")
        return False
    
    ws = wb["Assessment Areas"]
    
    # Find next empty row
    next_row = ws.max_row + 1
    
    # Add Huntington assessment areas
    for aa in assessment_areas:
        ws.cell(row=next_row, column=1, value=aa['bank_name'])
        ws.cell(row=next_row, column=2, value=aa['cbsa_name'])
        ws.cell(row=next_row, column=3, value=aa['cbsa_code'])
        ws.cell(row=next_row, column=4, value=aa['state_name'])
        ws.cell(row=next_row, column=5, value=aa['county_name'])
        ws.cell(row=next_row, column=6, value=aa['county_code'])
        next_row += 1
    
    wb.save(ticket_file)
    print(f"\n[SUCCESS] Added {len(assessment_areas)} Huntington assessment area entries to Excel ticket")
    return True

def main():
    pdf_path = Path(__file__).parent.parent / "7745.pdf"
    
    if not pdf_path.exists():
        print(f"ERROR: PDF not found: {pdf_path}")
        return
    
    # Extract assessment areas
    assessment_areas = extract_huntington_assessment_areas(pdf_path)
    
    # Save to JSON
    output_json = pdf_path.parent / "huntington_assessment_areas_final.json"
    with open(output_json, 'w') as f:
        json.dump(assessment_areas, f, indent=2)
    print(f"\nSaved to JSON: {output_json}")
    
    # Update Excel ticket
    update_excel_ticket(assessment_areas)
    
    print(f"\n{'='*80}")
    print("HUNTINGTON EXTRACTION COMPLETE")
    print(f"{'='*80}")
    print(f"Total counties: {len(assessment_areas)}")
    print(f"Unique CBSAs: {len(set(aa['cbsa_code'] for aa in assessment_areas if aa['cbsa_code']))}")
    print(f"States: {sorted(set(aa['state_name'] for aa in assessment_areas if aa['state_name']))}")

if __name__ == "__main__":
    main()

