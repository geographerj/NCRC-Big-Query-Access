"""
Parse Huntington National Bank CRA Performance Evaluation (PE) PDF
File: "7745.pdf"

CRA PE documents typically contain assessment areas information.
This script extracts CBSA/county information from the PE document.
"""

import pdfplumber
from pathlib import Path
import sys
import json
import re
import pandas as pd

def parse_huntington_cra_pe(pdf_path):
    """
    Parse Huntington CRA Performance Evaluation PDF.
    CRA PE documents often contain assessment area information in tables.
    """
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF file not found: {pdf_path}")
    
    print(f"\nParsing Huntington CRA PE PDF: {pdf_path}")
    
    try:
        import pdfplumber
    except ImportError:
        raise ImportError("Please install pdfplumber: pip install pdfplumber")
    
    assessment_areas = []
    counties_found = []
    
    with pdfplumber.open(pdf_path) as pdf:
        print(f"  Total pages: {len(pdf.pages)}")
        
        # Extract text from all pages
        full_text = ""
        all_tables = []
        
        for page_num, page in enumerate(pdf.pages, 1):
            text = page.extract_text()
            if text:
                full_text += "\n" + text
            
            # Extract tables (assessment areas often in tables)
            tables = page.extract_tables()
            if tables:
                for table_idx, table in enumerate(tables):
                    if table and len(table) > 0:
                        all_tables.append({
                            'page': page_num,
                            'table': table_idx,
                            'data': table
                        })
        
        print(f"  Extracted {len(pdf.pages)} pages of text")
        print(f"  Found {len(all_tables)} tables")
        
        # Look for assessment area patterns in text
        # CRA PE documents often mention "Assessment Area" or list CBSAs/MSAs
        
        # Pattern 1: CBSA codes (5 digits)
        cbsa_pattern = r'\b(CBSA|MSA|Metropolitan Statistical Area)[\s:]*(\d{5})\b'
        cbsa_matches = re.finditer(cbsa_pattern, full_text, re.IGNORECASE)
        
        # Pattern 2: County names with states
        county_pattern = r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+County,\s+([A-Z]{2})'
        county_matches = re.finditer(county_pattern, full_text)
        
        # Pattern 3: County names in "County, State" format
        county_pattern2 = r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*),\s+([A-Z]{2}|[A-Z][a-z]+)'
        
        # Extract CBSA codes
        cbsa_codes_found = []
        for match in cbsa_matches:
            cbsa_code = match.group(2)
            if cbsa_code != '99999':  # Skip placeholder
                cbsa_codes_found.append(cbsa_code)
        
        # Extract counties
        counties_from_text = []
        for match in county_matches:
            county_name = match.group(1).strip()
            state_code = match.group(2).strip()
            counties_from_text.append({
                'county_name': county_name,
                'state_code': state_code
            })
        
        print(f"  Found {len(cbsa_codes_found)} CBSA codes in text")
        print(f"  Found {len(counties_from_text)} counties in text")
        
        # Try extracting from tables
        for table_info in all_tables:
            table = table_info['data']
            if not table or len(table) < 2:
                continue
            
            # Look for header row that might indicate assessment area data
            header_row = table[0] if table else []
            header_text = ' '.join([str(cell) if cell else '' for cell in header_row]).lower()
            
            # Check if this table contains assessment area information
            if any(keyword in header_text for keyword in ['assessment', 'area', 'cbsa', 'msa', 'county', 'state', 'geoid']):
                print(f"  Found potential assessment area table on page {table_info['page']}")
                
                # Try to identify column indices
                geoid_col = None
                county_col = None
                state_col = None
                cbsa_col = None
                
                for col_idx, header in enumerate(header_row):
                    if header:
                        header_lower = str(header).lower()
                        if 'geoid' in header_lower or 'geoid10' in header_lower or 'tract' in header_lower:
                            geoid_col = col_idx
                        if 'county' in header_lower:
                            county_col = col_idx
                        if 'state' in header_lower and 'code' in header_lower:
                            state_col = col_idx
                        if 'cbsa' in header_lower or 'msa' in header_lower:
                            cbsa_col = col_idx
                
                # Extract data rows
                for row in table[1:]:
                    row_data = {}
                    
                    # Extract GEOID if available
                    if geoid_col is not None and geoid_col < len(row):
                        geoid_cell = str(row[geoid_col]) if row[geoid_col] else ""
                        # Could be 11-digit tract GEOID or 5-digit county GEOID
                        geoid_clean = re.sub(r'[\s\-]', '', geoid_cell)
                        if re.match(r'^\d{5}$', geoid_clean):
                            row_data['county_code'] = geoid_clean
                        elif re.match(r'^\d{11}$', geoid_clean):
                            row_data['county_code'] = geoid_clean[:5]  # First 5 digits
                    
                    # Extract county name
                    if county_col is not None and county_col < len(row):
                        county_cell = str(row[county_col]) if row[county_col] else ""
                        # Clean up county name (remove "County" suffix if present)
                        county_name = county_cell.replace(' County', '').strip()
                        if county_name:
                            row_data['county_name'] = county_name
                    
                    # Extract state
                    if state_col is not None and state_col < len(row):
                        state_cell = str(row[state_col]) if row[state_col] else ""
                        if state_cell:
                            row_data['state_code'] = str(state_cell).strip()[:2]
                    
                    # Extract CBSA code
                    if cbsa_col is not None and cbsa_col < len(row):
                        cbsa_cell = str(row[cbsa_col]) if row[cbsa_col] else ""
                        cbsa_clean = re.sub(r'[\s\-]', '', str(cbsa_cell))
                        if re.match(r'^\d{5}$', cbsa_clean):
                            row_data['cbsa_code'] = cbsa_clean
                    
                    # If we got county information, add it
                    if row_data.get('county_name') or row_data.get('county_code'):
                        counties_found.append(row_data)
        
        # Load county crosswalk to map to CBSAs if we have county names/codes
        crosswalk_path = Path(__file__).parent.parent / 'data' / 'reference' / 'CBSA_to_County_Mapping.csv'
        
        if crosswalk_path.exists() and (counties_found or counties_from_text):
            crosswalk_df = pd.read_csv(crosswalk_path)
            crosswalk_df['Geoid5'] = crosswalk_df['Geoid5'].astype(str).str.zfill(5)
            
            # Create mapping
            geoid5_to_info = {}
            county_name_state_to_info = {}
            
            for _, row in crosswalk_df.iterrows():
                geoid5 = str(row['Geoid5']).zfill(5)
                county_name = str(row.get('County Name', '')).strip()
                state_code = str(row.get('State Code', '')).strip()
                state_name = str(row.get('State', '')).strip()
                cbsa_code = str(row.get('Cbsa Code', '')).strip()
                cbsa_name = str(row.get('Cbsa', '')).strip()
                
                if geoid5 not in geoid5_to_info:
                    geoid5_to_info[geoid5] = {
                        'county_name': county_name,
                        'state_code': state_code,
                        'state_name': state_name,
                        'cbsa_code': cbsa_code,
                        'cbsa_name': cbsa_name
                    }
                
                # Also create name+state lookup
                key = f"{county_name}|{state_code}"
                if key not in county_name_state_to_info:
                    county_name_state_to_info[key] = {
                        'county_code': geoid5,
                        'cbsa_code': cbsa_code,
                        'cbsa_name': cbsa_name,
                        'state_name': state_name
                    }
            
            # Map counties found
            for county_info in counties_found + counties_from_text:
                mapped_info = None
                
                # Try to match by GEOID5 first
                if county_info.get('county_code'):
                    geoid5 = str(county_info['county_code']).zfill(5)
                    if geoid5 in geoid5_to_info:
                        mapped_info = geoid5_to_info[geoid5].copy()
                        mapped_info['county_code'] = geoid5
                
                # Try to match by county name + state
                elif county_info.get('county_name') and county_info.get('state_code'):
                    county_name = str(county_info['county_name']).replace(' County', '').strip()
                    state_code = str(county_info['state_code']).strip()[:2]
                    key = f"{county_name}|{state_code}"
                    
                    if key in county_name_state_to_info:
                        mapped_info = county_name_state_to_info[key].copy()
                        mapped_info['county_name'] = county_name
                        mapped_info['state_code'] = state_code
                
                if mapped_info:
                    assessment_areas.append({
                        'bank_name': 'The Huntington National Bank',
                        'cbsa_name': mapped_info.get('cbsa_name', ''),
                        'cbsa_code': mapped_info.get('cbsa_code', ''),
                        'state_code': mapped_info.get('state_code', ''),
                        'state_name': mapped_info.get('state_name', ''),
                        'county_name': mapped_info.get('county_name', county_info.get('county_name', '')),
                        'county_code': mapped_info.get('county_code', county_info.get('county_code', ''))
                    })
            
            print(f"\n  Mapped to {len(assessment_areas)} assessment area entries")
            
            # Show sample
            if assessment_areas:
                print(f"\n  Sample entries (first 10):")
                for aa in assessment_areas[:10]:
                    print(f"    - {aa['county_name']}, {aa['state_name']}, CBSA: {aa['cbsa_code']} ({aa['cbsa_name']})")
        else:
            print(f"\n  WARNING: County crosswalk not found or no counties extracted")
            print(f"  Found {len(counties_found)} counties from tables")
            print(f"  Found {len(counties_from_text)} counties from text")
    
    # Remove duplicates
    seen = set()
    unique_areas = []
    for aa in assessment_areas:
        key = (aa['county_code'], aa['bank_name'])
        if key not in seen:
            seen.add(key)
            unique_areas.append(aa)
    
    return unique_areas

def add_to_excel_ticket(ticket_file, assessment_areas):
    """Add assessment areas to the Excel ticket file"""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    
    ticket_path = Path(ticket_file)
    if not ticket_path.exists():
        print(f"\nERROR: Ticket file not found: {ticket_path}")
        return False
    
    print(f"\nAdding {len(assessment_areas)} assessment area entries to ticket...")
    
    wb = load_workbook(ticket_path)
    
    if "Assessment Areas" not in wb.sheetnames:
        ws = wb.create_sheet("Assessment Areas")
        # Add headers
        headers = ['Bank Name', 'CBSA Name', 'CBSA Code', 'State Name', 'County Name', 'County Code (GEOID)']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
    else:
        ws = wb["Assessment Areas"]
    
    # Find next empty row
    next_row = ws.max_row + 1
    if ws.max_row == 1:
        next_row = 2  # Start after header
    
    # Add assessment areas
    for aa in assessment_areas:
        ws.cell(row=next_row, column=1, value=aa['bank_name'])
        ws.cell(row=next_row, column=2, value=aa['cbsa_name'])
        ws.cell(row=next_row, column=3, value=aa['cbsa_code'])
        ws.cell(row=next_row, column=4, value=aa['state_name'])
        ws.cell(row=next_row, column=5, value=aa['county_name'])
        ws.cell(row=next_row, column=6, value=aa['county_code'])
        next_row += 1
    
    wb.save(ticket_path)
    print(f"  Saved {len(assessment_areas)} entries to 'Assessment Areas' sheet")
    
    return True

def main():
    if len(sys.argv) < 2:
        pdf_path = Path(__file__).parent.parent / "7745.pdf"
        if not pdf_path.exists():
            print("Usage: python parse_huntington_cra_pe.py <pdf_file>")
            print(f"\nOr place PDF at: {pdf_path}")
            sys.exit(1)
    else:
        pdf_path = Path(sys.argv[1])
    
    # Parse PDF
    assessment_areas = parse_huntington_cra_pe(pdf_path)
    
    # Save to JSON for inspection
    output_json = pdf_path.parent / "huntington_assessment_areas_parsed.json"
    with open(output_json, 'w') as f:
        json.dump(assessment_areas, f, indent=2)
    print(f"\n  Saved parsed data to: {output_json}")
    
    # Add to Excel ticket if it exists
    ticket_file = Path(__file__).parent.parent / "Huntington+Cadence merger research ticket.xlsx"
    if ticket_file.exists():
        add_to_excel_ticket(ticket_file, assessment_areas)
        print(f"\n[SUCCESS] Added assessment areas to ticket: {ticket_file}")
    else:
        print(f"\n[WARNING] Ticket file not found: {ticket_file}")
        print(f"   Assessment areas saved to JSON: {output_json}")
        print(f"   You can manually add them to the ticket later")
    
    print("\n" + "="*80)
    print("PARSING COMPLETE")
    print("="*80)
    print(f"\nTotal assessment area entries: {len(assessment_areas)}")
    print(f"Unique counties: {len(set(aa['county_code'] for aa in assessment_areas if aa['county_code']))}")
    print(f"Unique CBSAs: {len(set(aa['cbsa_code'] for aa in assessment_areas if aa['cbsa_code']))}")

if __name__ == "__main__":
    main()

