"""
Thorough extraction of assessment areas from PDFs
Uses multiple strategies to extract data from both PDFs
"""

import pdfplumber
from pathlib import Path
import sys
import json
import re
import pandas as pd
from collections import defaultdict

def extract_from_cadence_pdf(pdf_path):
    """Extract census tract GEOIDs from Cadence PDF"""
    print(f"\n{'='*80}")
    print("EXTRACTING FROM CADENCE PDF (Census Tracts)")
    print(f"{'='*80}")
    
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        print(f"ERROR: PDF not found: {pdf_path}")
        return []
    
    tracts_found = []
    all_table_data = []
    
    with pdfplumber.open(pdf_path) as pdf:
        print(f"Total pages: {len(pdf.pages)}")
        
        # Strategy 1: Extract from tables
        for page_num, page in enumerate(pdf.pages[:20], 1):  # Check first 20 pages
            tables = page.extract_tables()
            
            for table_idx, table in enumerate(tables):
                if not table or len(table) < 2:
                    continue
                
                # Save raw table data for inspection
                all_table_data.append({
                    'page': page_num,
                    'table': table_idx,
                    'data': table
                })
                
                # Look through all cells for GEOIDs
                for row_idx, row in enumerate(table):
                    for col_idx, cell in enumerate(row):
                        if cell:
                            cell_str = str(cell).strip()
                            
                            # Try various GEOID patterns
                            patterns = [
                                r'(\d{2})-(\d{3})-(\d{6})',  # SS-CCC-TTTTTT
                                r'(\d{11})',  # Direct 11-digit
                                r'(\d{5})\s+(\d{6})',  # GEOID5 + Tract
                                r'(\d{2})\s+(\d{3})\s+(\d{6})',  # SS CCC TTTTTT
                            ]
                            
                            for pattern in patterns:
                                matches = re.finditer(pattern, cell_str)
                                for match in matches:
                                    if len(match.groups()) == 3:
                                        state_code = match.group(1)
                                        county_code = match.group(2)
                                        tract_code = match.group(3)
                                        geoid = state_code + county_code + tract_code
                                        
                                        if len(geoid) == 11:
                                            tracts_found.append({
                                                'full_geoid': geoid,
                                                'state_code': state_code,
                                                'county_code': county_code,
                                                'tract_code': tract_code,
                                                'geoid5': state_code + county_code,
                                                'source': f'Page {page_num}, Table {table_idx}, Row {row_idx}, Col {col_idx}'
                                            })
                                    elif len(match.groups()) == 1:
                                        if len(match.group(1)) == 11:
                                            geoid = match.group(1)
                                            tracts_found.append({
                                                'full_geoid': geoid,
                                                'state_code': geoid[:2],
                                                'county_code': geoid[2:5],
                                                'tract_code': geoid[5:11],
                                                'geoid5': geoid[:5],
                                                'source': f'Page {page_num}, Table {table_idx}, Row {row_idx}, Col {col_idx}'
                                            })
            
            # Strategy 2: Extract from text (for pages without good tables)
            text = page.extract_text()
            if text:
                # Look for GEOID patterns in text
                geoid_patterns = [
                    r'\b(\d{2})-(\d{3})-(\d{6})\b',  # SS-CCC-TTTTTT
                    r'\b(\d{11})\b',  # Direct 11-digit
                ]
                
                for pattern in geoid_patterns:
                    matches = re.finditer(pattern, text)
                    for match in matches:
                        if len(match.groups()) == 3:
                            geoid = ''.join(match.groups())
                        else:
                            geoid = match.group(1)
                        
                        if len(geoid) == 11:
                            tracts_found.append({
                                'full_geoid': geoid,
                                'state_code': geoid[:2],
                                'county_code': geoid[2:5],
                                'tract_code': geoid[5:11],
                                'geoid5': geoid[:5],
                                'source': f'Page {page_num}, Text'
                            })
        
        print(f"\nExtracted {len(tracts_found)} census tract GEOIDs")
        
        # Save sample table data for inspection
        if all_table_data:
            sample_file = pdf_path.parent / "cadence_tables_sample.json"
            with open(sample_file, 'w') as f:
                json.dump(all_table_data[:5], f, indent=2)
            print(f"Saved sample table data to: {sample_file}")
    
    return tracts_found

def extract_from_huntington_pdf(pdf_path):
    """Extract assessment area info from Huntington CRA PE"""
    print(f"\n{'='*80}")
    print("EXTRACTING FROM HUNTINGTON CRA PE PDF")
    print(f"{'='*80}")
    
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        print(f"ERROR: PDF not found: {pdf_path}")
        return []
    
    assessment_areas = []
    counties_found = []
    cbsa_codes_found = []
    
    with pdfplumber.open(pdf_path) as pdf:
        print(f"Total pages: {len(pdf.pages)}")
        
        # Focus on early pages which often contain assessment area summaries
        pages_to_check = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 15, 16, 17, 20, 25, 30, 35, 40]
        
        for page_num in pages_to_check:
            if page_num > len(pdf.pages):
                continue
                
            page = pdf.pages[page_num - 1]  # 0-indexed
            text = page.extract_text()
            tables = page.extract_tables()
            
            if text:
                # Look for CBSA codes
                cbsa_patterns = [
                    r'\bCBSA[:\s]+(\d{5})\b',
                    r'\bMSA[:\s]+(\d{5})\b',
                    r'\bMetropolitan[:\s]+(\d{5})\b',
                    r'\((\d{5})\)',  # (16980)
                ]
                
                for pattern in cbsa_patterns:
                    matches = re.finditer(pattern, text, re.IGNORECASE)
                    for match in matches:
                        cbsa_code = match.group(1)
                        if cbsa_code != '99999':
                            cbsa_codes_found.append({
                                'cbsa_code': cbsa_code,
                                'page': page_num,
                                'context': text[max(0, match.start()-50):match.end()+50]
                            })
                
                # Look for county patterns
                county_patterns = [
                    r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+County,\s+([A-Z]{2})',
                    r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*),\s+([A-Z]{2})\s+County',
                ]
                
                for pattern in county_patterns:
                    matches = re.finditer(pattern, text)
                    for match in matches:
                        county_name = match.group(1).strip()
                        state_code = match.group(2).strip()
                        counties_found.append({
                            'county_name': county_name,
                            'state_code': state_code,
                            'page': page_num
                        })
            
            # Extract from tables
            if tables:
                for table_idx, table in enumerate(tables):
                    if not table or len(table) < 2:
                        continue
                    
                    # Check header row for keywords
                    header_row = table[0] if table else []
                    header_text = ' '.join([str(cell).lower() if cell else '' for cell in header_row])
                    
                    if any(keyword in header_text for keyword in ['county', 'cbsa', 'msa', 'state', 'assessment', 'area']):
                        print(f"  Found potential assessment area table on page {page_num}, table {table_idx}")
                        
                        # Try to extract data
                        for row in table[1:]:  # Skip header
                            row_text = ' '.join([str(cell) if cell else '' for cell in row])
                            
                            # Look for county names
                            county_match = re.search(r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+County', row_text)
                            state_match = re.search(r'([A-Z]{2})', row_text)
                            cbsa_match = re.search(r'(\d{5})', row_text)
                            
                            if county_match:
                                county_name = county_match.group(1)
                                state_code = state_match.group(1) if state_match else ''
                                cbsa_code = cbsa_match.group(1) if cbsa_match else ''
                                
                                counties_found.append({
                                    'county_name': county_name,
                                    'state_code': state_code,
                                    'cbsa_code': cbsa_code,
                                    'page': page_num,
                                    'table': table_idx,
                                    'row_text': row_text[:100]
                                })
        
        print(f"\nFound {len(cbsa_codes_found)} CBSA codes")
        print(f"Found {len(counties_found)} county mentions")
        
        # Show samples
        if cbsa_codes_found:
            print(f"\nSample CBSA codes:")
            for item in cbsa_codes_found[:5]:
                print(f"  - {item['cbsa_code']} (Page {item['page']})")
        
        if counties_found:
            print(f"\nSample counties:")
            for item in counties_found[:10]:
                print(f"  - {item['county_name']}, {item.get('state_code', 'N/A')} (Page {item['page']})")
    
    return counties_found, cbsa_codes_found

def map_to_assessment_areas(cadence_tracts, huntington_counties):
    """Map extracted data to assessment area format"""
    print(f"\n{'='*80}")
    print("MAPPING TO ASSESSMENT AREAS")
    print(f"{'='*80}")
    
    assessment_areas = []
    
    # Load crosswalk
    crosswalk_path = Path(__file__).parent.parent / 'data' / 'reference' / 'CBSA_to_County_Mapping.csv'
    
    if not crosswalk_path.exists():
        print(f"WARNING: Crosswalk not found: {crosswalk_path}")
        return []
    
    crosswalk_df = pd.read_csv(crosswalk_path)
    crosswalk_df['Geoid5'] = crosswalk_df['Geoid5'].astype(str).str.zfill(5)
    
    # Create mappings
    geoid5_to_info = {}
    for _, row in crosswalk_df.iterrows():
        geoid5 = str(row['Geoid5']).zfill(5)
        if geoid5 not in geoid5_to_info:
            geoid5_to_info[geoid5] = {
                'county_name': str(row.get('County Name', '')).strip(),
                'state_code': str(row.get('State Code', '')).strip(),
                'state_name': str(row.get('State', '')).strip(),
                'cbsa_code': str(row.get('Cbsa Code', '')).strip(),
                'cbsa_name': str(row.get('Cbsa', '')).strip()
            }
    
    # Process Cadence tracts
    if cadence_tracts:
        print(f"\nProcessing {len(cadence_tracts)} Cadence census tracts...")
        
        # Group by GEOID5 (county)
        counties_with_tracts = defaultdict(int)
        for tract in cadence_tracts:
            geoid5 = tract['geoid5']
            counties_with_tracts[geoid5] += 1
        
        # Map to assessment areas
        for geoid5, tract_count in counties_with_tracts.items():
            if geoid5 in geoid5_to_info:
                info = geoid5_to_info[geoid5]
                assessment_areas.append({
                    'bank_name': 'Cadence Bank',
                    'cbsa_name': info['cbsa_name'],
                    'cbsa_code': info['cbsa_code'],
                    'state_code': info['state_code'],
                    'state_name': info['state_name'],
                    'county_name': info['county_name'],
                    'county_code': geoid5,
                    'tract_count': tract_count
                })
        
        print(f"  Mapped to {len(counties_with_tracts)} unique counties for Cadence")
    
    # Process Huntington counties
    if huntington_counties:
        print(f"\nProcessing {len(huntington_counties)} Huntington county mentions...")
        
        # Try to match by county name + state
        county_name_state_to_info = {}
        for _, row in crosswalk_df.iterrows():
            county_name = str(row.get('County Name', '')).strip().replace(' County', '')
            state_code = str(row.get('State Code', '')).strip()
            geoid5 = str(row['Geoid5']).zfill(5)
            key = f"{county_name}|{state_code}"
            
            if key not in county_name_state_to_info:
                county_name_state_to_info[key] = {
                    'county_code': geoid5,
                    'cbsa_code': str(row.get('Cbsa Code', '')).strip(),
                    'cbsa_name': str(row.get('Cbsa', '')).strip(),
                    'state_name': str(row.get('State', '')).strip()
                }
        
        # Match Huntington counties
        matched_counties = set()
        for county_info in huntington_counties:
            county_name = str(county_info['county_name']).replace(' County', '').strip()
            state_code = str(county_info.get('state_code', '')).strip()
            
            if county_name and state_code:
                key = f"{county_name}|{state_code}"
                if key in county_name_state_to_info:
                    info = county_name_state_to_info[key]
                    county_key = (info['county_code'], 'The Huntington National Bank')
                    if county_key not in matched_counties:
                        matched_counties.add(county_key)
                        assessment_areas.append({
                            'bank_name': 'The Huntington National Bank',
                            'cbsa_name': info['cbsa_name'],
                            'cbsa_code': info['cbsa_code'],
                            'state_code': state_code,
                            'state_name': info['state_name'],
                            'county_name': county_name,
                            'county_code': info['county_code']
                        })
        
        print(f"  Matched {len(matched_counties)} unique counties for Huntington")
    
    return assessment_areas

def main():
    cadence_pdf = Path(__file__).parent.parent / "412025 List of Census Tracts - Public File.pdf"
    huntington_pdf = Path(__file__).parent.parent / "7745.pdf"
    
    print("="*80)
    print("THOROUGH ASSESSMENT AREA EXTRACTION")
    print("="*80)
    
    # Extract from both PDFs
    cadence_tracts = extract_from_cadence_pdf(cadence_pdf) if cadence_pdf.exists() else []
    huntington_counties, huntington_cbsas = extract_from_huntington_pdf(huntington_pdf) if huntington_pdf.exists() else ([], [])
    
    # Map to assessment areas
    assessment_areas = map_to_assessment_areas(cadence_tracts, huntington_counties)
    
    # Remove duplicates
    seen = set()
    unique_areas = []
    for aa in assessment_areas:
        key = (aa['bank_name'], aa['county_code'])
        if key not in seen:
            seen.add(key)
            unique_areas.append(aa)
    
    # Save results
    output_json = Path(__file__).parent.parent / "assessment_areas_extracted_final.json"
    with open(output_json, 'w') as f:
        json.dump({
            'cadence_tracts': cadence_tracts[:10],  # Sample
            'huntington_counties': huntington_counties[:20],  # Sample
            'assessment_areas': unique_areas
        }, f, indent=2)
    
    print(f"\n{'='*80}")
    print("EXTRACTION COMPLETE")
    print(f"{'='*80}")
    print(f"\nTotal assessment area entries: {len(unique_areas)}")
    print(f"  - Cadence Bank: {len([a for a in unique_areas if a['bank_name'] == 'Cadence Bank'])} counties")
    print(f"  - Huntington: {len([a for a in unique_areas if a['bank_name'] == 'The Huntington National Bank'])} counties")
    print(f"\nResults saved to: {output_json}")
    
    # Update Excel ticket
    if unique_areas:
        ticket_file = Path(__file__).parent.parent / "Huntington+Cadence merger research ticket.xlsx"
        if ticket_file.exists():
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = load_workbook(ticket_file)
            if "Assessment Areas" in wb.sheetnames:
                ws = wb["Assessment Areas"]
                # Clear existing data (keep headers)
                if ws.max_row > 1:
                    ws.delete_rows(2, ws.max_row)
            else:
                ws = wb.create_sheet("Assessment Areas")
                headers = ['Bank Name', 'CBSA Name', 'CBSA Code', 'State Name', 'County Name', 'County Code (GEOID)']
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
            
            # Add assessment areas
            for row_idx, aa in enumerate(unique_areas, start=2):
                ws.cell(row=row_idx, column=1, value=aa['bank_name'])
                ws.cell(row=row_idx, column=2, value=aa['cbsa_name'])
                ws.cell(row=row_idx, column=3, value=aa['cbsa_code'])
                ws.cell(row=row_idx, column=4, value=aa['state_name'])
                ws.cell(row=row_idx, column=5, value=aa['county_name'])
                ws.cell(row=row_idx, column=6, value=aa['county_code'])
            
            wb.save(ticket_file)
            print(f"\n[SUCCESS] Updated Excel ticket with {len(unique_areas)} assessment area entries")
        else:
            print(f"\n[WARNING] Ticket file not found: {ticket_file}")
    else:
        print("\n[WARNING] No assessment areas extracted. May need manual entry or improved extraction.")

if __name__ == "__main__":
    main()

