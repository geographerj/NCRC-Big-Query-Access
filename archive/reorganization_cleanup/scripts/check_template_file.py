"""Check what's in the provided Excel file"""

import openpyxl
from pathlib import Path

template_file = r"C:\Users\edite\OneDrive - Nat'l Community Reinvestment Coaltn\Desktop\DREAM Analysis\data\raw\PNC+FirstBank merger research ticket.xlsx"

try:
    wb = openpyxl.load_workbook(template_file, data_only=False)
    
    print("="*80)
    print("FILE INSPECTION")
    print("="*80)
    print(f"\nFile: {Path(template_file).name}")
    print(f"Total sheets: {len(wb.sheetnames)}")
    print("\nSheet names:")
    for i, name in enumerate(wb.sheetnames, 1):
        ws = wb[name]
        print(f"  {i}. {name} ({ws.max_row} rows x {ws.max_column} cols)")
    
    # Check for template-like sheets (Mortgage Goals, Assessment Areas, etc.)
    template_sheets = ['Mortgage Goals', 'Assessment Areas', 'Notes', 'Methodology']
    found_template_sheets = [s for s in wb.sheetnames if any(t in s for t in template_sheets)]
    
    if found_template_sheets:
        print("\n" + "="*80)
        print("TEMPLATE SHEETS FOUND:")
        print("="*80)
        for sheet_name in found_template_sheets:
            print(f"\n  {sheet_name}")
            ws = wb[sheet_name]
            print(f"    Rows: {ws.max_row}, Columns: {ws.max_column}")
            
            # Check for formulas
            formula_count = 0
            for row in ws.iter_rows(max_row=min(100, ws.max_row), max_col=min(50, ws.max_column)):
                for cell in row:
                    if cell.data_type == 'f':
                        formula_count += 1
            
            print(f"    Formulas (first 100 rows): {formula_count}")
            
            # Show first few headers
            print("    Headers (row 1):")
            for col in range(1, min(11, ws.max_column + 1)):
                cell = ws.cell(1, col)
                if cell.value:
                    print(f"      Col {col}: {cell.value}")
    else:
        print("\nThis appears to be the MERGER TICKET, not the TEMPLATE.")
        print("Template should have sheets like: 'Mortgage Goals', 'Assessment Areas', 'Notes', etc.")
        print("\nDo you have the actual TEMPLATE file (the final report with formulas)?")
    
except PermissionError:
    print("\nERROR: File is open in Excel. Please close it and try again.")
except Exception as e:
    print(f"\nERROR: {e}")
    import traceback
    traceback.print_exc()

