"""Check for LMA Ticket Filters sheet in the Excel file"""

import openpyxl
from pathlib import Path

ticket_file = Path(r"C:\Users\edite\OneDrive - Nat'l Community Reinvestment Coaltn\Desktop\DREAM Analysis\PNC+FirstBank Research Ticket_CBA.xlsx")

if not ticket_file.exists():
    print(f"File not found: {ticket_file}")
    exit(1)

wb = openpyxl.load_workbook(ticket_file, data_only=True)

print("="*80)
print("CHECKING FOR LMA TICKET FILTERS SHEET")
print("="*80)

print(f"\nAll sheet names ({len(wb.sheetnames)} total):")
for i, name in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {name}")

# Check for variations
filter_sheets = [s for s in wb.sheetnames if 'filter' in s.lower() or 'lma' in s.lower() or 'ticket' in s.lower()]
print(f"\nSheets with 'filter', 'lma', or 'ticket' in name:")
if filter_sheets:
    for sheet in filter_sheets:
        print(f"  - {sheet}")
        ws = wb[sheet]
        print(f"    Dimensions: {ws.max_row} rows x {ws.max_column} columns")
        print(f"    First 5 rows:")
        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(6, ws.max_column + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    row_data.append(str(cell_value)[:40])
            if row_data:
                print(f"      Row {row_idx}: {row_data}")
else:
    print("  (none found)")

# Check if "LMA Ticket Filters" exists exactly
if "LMA Ticket Filters" in wb.sheetnames:
    print("\n[FOUND] 'LMA Ticket Filters' sheet exists!")
    ws = wb["LMA Ticket Filters"]
    print(f"  Dimensions: {ws.max_row} rows x {ws.max_column} columns")
    
    # Show structure
    print("\n  Sheet structure (first 15 rows):")
    for row_idx in range(1, min(16, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(6, ws.max_column + 1)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            row_data.append(str(cell_value)[:50] if cell_value else "")
        if any(row_data):
            print(f"    Row {row_idx}: {row_data}")
else:
    print("\n[NOT FOUND] 'LMA Ticket Filters' sheet does not exist in local file")
    print("  Note: You provided a SharePoint link - the file may need to be downloaded/synced")

