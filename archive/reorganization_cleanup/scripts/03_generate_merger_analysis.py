"""
Complete Merger Analysis: Fifth Third + Comerica

Generates:
1. Comerica CBA Report (same format as Fifth Third)
2. Branch footprint analysis (where each bank operates)
3. Overlap detection (markets where both operate)
4. HHI analysis pre/post merger in overlapping markets

Usage:
    python scripts/generate_merger_analysis.py
"""

import sys
import os
import pandas as pd
from pathlib import Path
from datetime import datetime

# Handle module imports
import importlib.util

def load_module(module_name, file_path):
    spec = importlib.util.spec_from_file_location(module_name, file_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module

base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

bigquery_client = load_module(
    "bigquery_client",
    os.path.join(base_dir, "Lending and Branch Analysis", "utils", "bigquery_client.py")
)
branch_queries = load_module(
    "branch_queries",
    os.path.join(base_dir, "Lending and Branch Analysis", "queries", "branch_queries.py")
)
hhi_analysis = load_module(
    "hhi_analysis",
    os.path.join(base_dir, "Lending and Branch Analysis", "utils", "hhi_analysis.py")
)
branch_matching = load_module(
    "branch_matching",
    os.path.join(base_dir, "Lending and Branch Analysis", "utils", "branch_matching.py")
)

create_client = bigquery_client.create_client

# Bank information with RSSD
FIFTH_THIRD = {
    'name': 'Fifth Third Bank',
    'lei': 'QFROUN1UWUYU0DVIWD51',
    'rssd': '723112'
}

COMERICA = {
    'name': 'Comerica Bank',
    'lei': '70WY0ID1N53Q4254VH70',
    'rssd': '60143'
}

def get_branches_by_rssd(rssd: str, year: int = 2025):
    """Get branch data using RSSD"""
    creds_path = os.path.join(base_dir, "config", "credentials", "hdma1-242116-74024e2eb88f.json")
    
    if os.path.exists(creds_path):
        client = create_client(key_path=creds_path)
    else:
        client = create_client()
    
    query = branch_queries.get_branches_by_rssd(
        rssd=rssd,
        year=year,
        sod_table='hdma1-242116.branches.sod25'
    )
    
    df = client.execute_query(query)
    
    # Deduplicate
    if 'uninumbr' in df.columns and 'year' in df.columns:
        df = branch_matching.deduplicate_branches(df, year_col='year', unique_col='uninumbr')
    
    return df

def analyze_branch_footprints_and_overlap():
    """Analyze branch footprints and find overlap"""
    print(f"\n{'='*80}")
    print("BRANCH FOOTPRINT ANALYSIS")
    print(f"{'='*80}\n")
    
    # Get branches for both banks (2025 data from sod25)
    print(f"Getting Fifth Third branches (RSSD: {FIFTH_THIRD['rssd']})...")
    ft_branches = get_branches_by_rssd(FIFTH_THIRD['rssd'], year=2025)
    print(f"  Found {len(ft_branches):,} branches in 2025")
    
    print(f"\nGetting Comerica branches (RSSD: {COMERICA['rssd']})...")
    com_branches = get_branches_by_rssd(COMERICA['rssd'], year=2025)
    print(f"  Found {len(com_branches):,} branches in 2025")
    
    # Summary by CBSA
    # Note: sod25 doesn't have cbsa_code directly - we'll need to map from geoid5 or add CBSA mapping
    # For now, check if we have geoid5 which can be mapped to CBSA
    footprint_data = {}
    
    # Check if we have CBSA codes (might not be in sod25)
    has_cbsa = 'cbsa_code' in ft_branches.columns and 'cbsa_code' in com_branches.columns
    
    if not has_cbsa and 'geoid5' in ft_branches.columns:
        # Map geoid5 (state+county) to CBSA using the crosswalk
        print("  Mapping geoid5 to CBSA codes...")
        crosswalk_path = os.path.join(base_dir, "data", "reference", "CBSA_to_County_Mapping.csv")
        if os.path.exists(crosswalk_path):
            crosswalk = pd.read_csv(crosswalk_path)
            # Clean geoid5 - ensure it's string and pad if needed
            crosswalk['Geoid5'] = crosswalk['Geoid5'].astype(str).str.zfill(5)
            ft_branches['geoid5'] = ft_branches['geoid5'].astype(str).str.zfill(5)
            com_branches['geoid5'] = com_branches['geoid5'].astype(str).str.zfill(5)
            
            # Merge CBSA codes
            ft_branches = ft_branches.merge(
                crosswalk[['Geoid5', 'Cbsa Code']].rename(columns={'Geoid5': 'geoid5', 'Cbsa Code': 'cbsa_code'}),
                on='geoid5',
                how='left'
            )
            com_branches = com_branches.merge(
                crosswalk[['Geoid5', 'Cbsa Code']].rename(columns={'Geoid5': 'geoid5', 'Cbsa Code': 'cbsa_code'}),
                on='geoid5',
                how='left'
            )
            
            # Filter out rows without CBSA (rural areas with code 99999)
            ft_branches = ft_branches[(ft_branches['cbsa_code'].notna()) & (ft_branches['cbsa_code'] != '99999')]
            com_branches = com_branches[(com_branches['cbsa_code'].notna()) & (com_branches['cbsa_code'] != '99999')]
            
            print(f"  Mapped {len(ft_branches)} Fifth Third branches and {len(com_branches)} Comerica branches to CBSAs")
            has_cbsa = True
    
    if has_cbsa:
        # Fifth Third by CBSA
        ft_by_cbsa = ft_branches.groupby('cbsa_code').agg({
            'uninumbr': 'count',
            'deposits': 'sum',
            'br_lmi': lambda x: (x == 1).sum() if 'br_lmi' in ft_branches.columns else 0,
            'cr_minority': lambda x: (x == 1).sum() if 'cr_minority' in ft_branches.columns else 0
        }).reset_index()
        ft_by_cbsa.columns = ['cbsa_code', 'ft_branch_count', 'ft_total_deposits', 'ft_branches_lmi', 'ft_branches_minority']
        
        # Comerica by CBSA
        com_by_cbsa = com_branches.groupby('cbsa_code').agg({
            'uninumbr': 'count',
            'deposits': 'sum',
            'br_lmi': lambda x: (x == 1).sum() if 'br_lmi' in com_branches.columns else 0,
            'cr_minority': lambda x: (x == 1).sum() if 'cr_minority' in com_branches.columns else 0
        }).reset_index()
        com_by_cbsa.columns = ['cbsa_code', 'com_branch_count', 'com_total_deposits', 'com_branches_lmi', 'com_branches_minority']
        
        # Find overlap (markets where both have branches)
        overlap = ft_by_cbsa.merge(com_by_cbsa, on='cbsa_code', how='inner')
        overlap = overlap.sort_values('ft_branch_count', ascending=False)
        
        footprint_data = {
            'fifth_third_branches': ft_branches,
            'comerica_branches': com_branches,
            'fifth_third_by_cbsa': ft_by_cbsa,
            'comerica_by_cbsa': com_by_cbsa,
            'overlap': overlap
        }
        
        print(f"\n{'='*80}")
        print("OVERLAP SUMMARY")
        print(f"{'='*80}")
        print(f"Markets where both banks have branches: {len(overlap)}")
        print(f"Fifth Third only markets: {len(ft_by_cbsa) - len(overlap)}")
        print(f"Comerica only markets: {len(com_by_cbsa) - len(overlap)}")
        
        if len(overlap) > 0:
            print(f"\nTop 10 overlapping markets:")
            for idx, row in overlap.head(10).iterrows():
                print(f"  CBSA {row['cbsa_code']}: FT={row['ft_branch_count']} branches, COM={row['com_branch_count']} branches")
    else:
        footprint_data = {
            'fifth_third_branches': ft_branches,
            'comerica_branches': com_branches
        }
        print("\nWARNING: No CBSA codes in branch data - cannot analyze overlap")
    
    return footprint_data

def calculate_hhi_for_overlapping_markets(overlapping_cbsas: list, year: int = 2025):
    """Calculate HHI pre and post merger for overlapping markets using branch deposits"""
    print(f"\n{'='*80}")
    print("HHI ANALYSIS FOR OVERLAPPING MARKETS")
    print("Using branch deposits (not lending data)")
    print(f"{'='*80}\n")
    
    creds_path = os.path.join(base_dir, "config", "credentials", "hdma1-242116-74024e2eb88f.json")
    
    if os.path.exists(creds_path):
        client = create_client(key_path=creds_path)
    else:
        client = create_client()
    
    results = []
    
    # Get CBSA names from crosswalk
    crosswalk_path = os.path.join(base_dir, "data", "reference", "CBSA_to_County_Mapping.csv")
    cbsa_names = {}
    if os.path.exists(crosswalk_path):
        crosswalk_df = pd.read_csv(crosswalk_path)
        # Get unique CBSA codes and names
        cbsa_info = crosswalk_df[['Cbsa Code', 'Cbsa']].drop_duplicates()
        cbsa_info['Cbsa Code'] = cbsa_info['Cbsa Code'].astype(str)
        cbsa_names = dict(zip(cbsa_info['Cbsa Code'], cbsa_info['Cbsa']))
    
    for cbsa in overlapping_cbsas:
        # Ensure CBSA code is string (not float)
        cbsa_str = str(int(float(cbsa))) if isinstance(cbsa, (int, float)) else str(cbsa)
        cbsa_name = cbsa_names.get(cbsa_str, f"CBSA {cbsa_str}")
        
        print(f"  Analyzing CBSA {cbsa_str} ({cbsa_name})...")
        
        # Get all branch deposits in this CBSA for the year (using RSSD to identify banks)
        # Need to join with CBSA crosswalk via geoid5 to filter by CBSA
        # Note: deposits_000s is in thousands - multiply by 1000
        query = f"""
        SELECT 
            s.rssd,
            SUM(CAST(s.deposits_000s AS FLOAT64) * 1000) as total_deposits,
            COUNT(DISTINCT s.uninumbr) as branch_count
        FROM `hdma1-242116.branches.sod25` s
        INNER JOIN `hdma1-242116.geo.cbsa_to_county` c
            ON CAST(s.geoid5 AS STRING) = CAST(c.geoid5 AS STRING)
        WHERE CAST(s.year AS STRING) = CAST({year} AS STRING)
          AND CAST(c.cbsa_code AS STRING) = '{cbsa_str}'
          AND c.cbsa_code != '99999'
          AND s.deposits_000s IS NOT NULL
          AND s.deposits_000s > 0
          AND s.rssd IS NOT NULL
          AND s.rssd != ''
        GROUP BY s.rssd
        ORDER BY total_deposits DESC
        """
        
        df = client.execute_query(query)
        
        if len(df) == 0:
            print(f"    No branch deposit data found for {cbsa_str}")
            continue
        
        # Calculate market shares pre-merger (based on deposits)
        total_deposits = df['total_deposits'].sum()
        df['market_share_pct'] = (df['total_deposits'] / total_deposits) * 100
        
        # Pre-merger HHI
        hhi_pre = hhi_analysis.calculate_hhi(df['market_share_pct'], as_percentages=True)
        
        # Get shares for merging banks (using RSSD)
        # Convert rssd to string for comparison
        df['rssd'] = df['rssd'].astype(str)
        ft_row = df[df['rssd'] == FIFTH_THIRD['rssd']]
        com_row = df[df['rssd'] == COMERICA['rssd']]
        
        ft_share = ft_row['market_share_pct'].values[0] if len(ft_row) > 0 else 0
        com_share = com_row['market_share_pct'].values[0] if len(com_row) > 0 else 0
        ft_deposits = ft_row['total_deposits'].values[0] if len(ft_row) > 0 else 0
        com_deposits = com_row['total_deposits'].values[0] if len(com_row) > 0 else 0
        
        # Post-merger: combine the two banks
        df_post = df[~df['rssd'].isin([FIFTH_THIRD['rssd'], COMERICA['rssd']])].copy()
        combined_deposits = ft_deposits + com_deposits
        
        combined_row = pd.DataFrame([{
            'rssd': f"{FIFTH_THIRD['rssd']}+{COMERICA['rssd']}",
            'total_deposits': combined_deposits,
            'branch_count': (ft_row['branch_count'].values[0] if len(ft_row) > 0 else 0) + 
                           (com_row['branch_count'].values[0] if len(com_row) > 0 else 0)
        }])
        df_post = pd.concat([df_post, combined_row], ignore_index=True)
        
        # Recalculate shares post-merger
        total_post = df_post['total_deposits'].sum()
        df_post['market_share_pct'] = (df_post['total_deposits'] / total_post) * 100
        
        # Post-merger HHI
        hhi_post = hhi_analysis.calculate_hhi(df_post['market_share_pct'], as_percentages=True)
        
        hhi_change = hhi_post - hhi_pre
        antitrust_concern = (hhi_pre > 1800) and (hhi_change > 100)
        
        results.append({
            'cbsa_code': cbsa_str,
            'cbsa_name': cbsa_name,
            'year': year,
            'hhi_pre_merger': hhi_pre,
            'hhi_post_merger': hhi_post,
            'hhi_change': hhi_change,
            'concentration_category_pre': hhi_analysis.categorize_market_concentration(hhi_pre),
            'concentration_category_post': hhi_analysis.categorize_market_concentration(hhi_post),
            'fifth_third_deposit_share_pct': ft_share,
            'comerica_deposit_share_pct': com_share,
            'combined_deposit_share_pct': ft_share + com_share,
            'fifth_third_deposits': ft_deposits,
            'comerica_deposits': com_deposits,
            'combined_deposits': combined_deposits,
            'total_deposits_market': total_deposits,
            'total_banks': len(df),
            'antitrust_concern': antitrust_concern
        })
        
        status = "*** ANTITRUST CONCERN ***" if antitrust_concern else "OK"
        print(f"    HHI: {hhi_pre:.0f} -> {hhi_post:.0f} (+{hhi_change:.0f}) {status}")
    
    return pd.DataFrame(results)

def create_comprehensive_report(footprint_data: dict, hhi_data: pd.DataFrame, output_file: str):
    """Create comprehensive Excel report"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    
    print(f"\n{'='*80}")
    print("CREATING COMPREHENSIVE EXCEL REPORT")
    print(f"{'='*80}\n")
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Sheet 1: Executive Summary
    ws_summary = wb.create_sheet('Executive Summary')
    ws_summary.append(['Fifth Third + Comerica Merger Analysis'])
    ws_summary.append(['Report Generated:', datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append([])
    ws_summary.append(['Fifth Third Bank', '', 'Comerica Bank'])
    ws_summary.append(['Name', FIFTH_THIRD['name'], COMERICA['name']])
    ws_summary.append(['LEI', FIFTH_THIRD['lei'], COMERICA['lei']])
    ws_summary.append(['RSSD', FIFTH_THIRD['rssd'], COMERICA['rssd']])
    ws_summary.append(['Total Branches (2025)', 
                       len(footprint_data['fifth_third_branches']), 
                       len(footprint_data['comerica_branches'])])
    
    if 'overlap' in footprint_data:
        ws_summary.append(['Markets with Overlap', len(footprint_data['overlap']), ''])
        ws_summary.append(['Fifth Third Only Markets', len(footprint_data['fifth_third_by_cbsa']) - len(footprint_data['overlap']), ''])
        ws_summary.append(['Comerica Only Markets', '', len(footprint_data['comerica_by_cbsa']) - len(footprint_data['overlap'])])
    
    # Sheet 2: Branch Overlap by CBSA
    if 'overlap' in footprint_data:
        ws_overlap = wb.create_sheet('Branch Overlap by CBSA')
        ws_overlap.append(['Markets Where Both Banks Have Branches'])
        ws_overlap.append([])
        
        for r in dataframe_to_rows(footprint_data['overlap'], index=False, header=True):
            ws_overlap.append(r)
        
        # Format header
        for cell in ws_overlap[3]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Sheet 3: HHI Analysis
    if len(hhi_data) > 0:
        ws_hhi = wb.create_sheet('HHI Analysis')
        ws_hhi.append(['HERFINDAHL-HIRSCHMAN INDEX (HHI) ANALYSIS'])
        ws_hhi.append(['Pre- and Post-Merger Market Concentration'])
        ws_hhi.append(['Markets with Antitrust Concerns Highlighted in RED'])
        ws_hhi.append([])
        
        for r in dataframe_to_rows(hhi_data.sort_values('hhi_change', ascending=False), index=False, header=True):
            ws_hhi.append(r)
        
        # Format and highlight antitrust concerns
        header_row = 5
        for row_idx in range(header_row + 1, len(hhi_data) + header_row + 1):
            concern_idx = list(hhi_data.columns).index('antitrust_concern') + 1
            if ws_hhi.cell(row=row_idx, column=concern_idx).value == True:
                # Highlight entire row in red
                for col_idx in range(1, len(list(hhi_data.columns)) + 1):
                    cell = ws_hhi.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')
        
        # Format header
        for cell in ws_hhi[header_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Sheet 4: Fifth Third Branch Summary by CBSA
    if 'fifth_third_by_cbsa' in footprint_data:
        ws_ft = wb.create_sheet('Fifth Third Branches by CBSA')
        ws_ft.append(['Fifth Third Bank Branch Summary by CBSA (2025)'])
        ws_ft.append([])
        
        for r in dataframe_to_rows(footprint_data['fifth_third_by_cbsa'].sort_values('ft_branch_count', ascending=False), 
                                   index=False, header=True):
            ws_ft.append(r)
        
        # Format header
        for cell in ws_ft[3]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Sheet 5: Comerica Branch Summary by CBSA
    if 'comerica_by_cbsa' in footprint_data:
        ws_com = wb.create_sheet('Comerica Branches by CBSA')
        ws_com.append(['Comerica Bank Branch Summary by CBSA (2025)'])
        ws_com.append([])
        
        for r in dataframe_to_rows(footprint_data['comerica_by_cbsa'].sort_values('com_branch_count', ascending=False),
                                   index=False, header=True):
            ws_com.append(r)
        
        # Format header
        for cell in ws_com[3]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Sheet 6: All Fifth Third Branches (detailed)
    ws_ft_detail = wb.create_sheet('Fifth Third All Branches')
    ws_ft_detail.append(['Fifth Third Bank - All Branch Details (2025)'])
    ws_ft_detail.append([])
    
    # Select relevant columns for detailed view
    detail_cols = ['institution_name', 'branch_name', 'address', 'city', 'state', 'zip_code', 
                   'cbsa_code', 'county_code', 'deposits', 'br_lmi', 'cr_minority']
    ft_detail = footprint_data['fifth_third_branches'][[col for col in detail_cols if col in footprint_data['fifth_third_branches'].columns]]
    
    for r in dataframe_to_rows(ft_detail, index=False, header=True):
        ws_ft_detail.append(r)
    
    # Format header
    for cell in ws_ft_detail[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Sheet 7: All Comerica Branches (detailed)
    ws_com_detail = wb.create_sheet('Comerica All Branches')
    ws_com_detail.append(['Comerica Bank - All Branch Details (2025)'])
    ws_com_detail.append([])
    
    com_detail = footprint_data['comerica_branches'][[col for col in detail_cols if col in footprint_data['comerica_branches'].columns]]
    
    for r in dataframe_to_rows(com_detail, index=False, header=True):
        ws_com_detail.append(r)
    
    # Format header
    for cell in ws_com_detail[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Save
    wb.save(output_file)
    print(f"Report saved to: {output_file}")

def main():
    print("\n" + "="*80)
    print("FIFTH THIRD + COMERICA MERGER ANALYSIS")
    print("Branch Footprints + Overlap + HHI Analysis")
    print("="*80)
    
    # Step 1: Analyze branch footprints
    footprint_data = analyze_branch_footprints_and_overlap()
    
    # Step 2: Get overlapping markets (ensure they're strings, not floats)
    overlapping_cbsas = []
    if 'overlap' in footprint_data:
        overlapping_cbsas = [str(int(float(c))) if isinstance(c, (int, float)) else str(c) 
                            for c in footprint_data['overlap']['cbsa_code'].tolist()]
        
        print(f"\n{'='*80}")
        print(f"Found {len(overlapping_cbsas)} overlapping markets for HHI analysis")
        print(f"{'='*80}")
        
        # Step 3: Calculate HHI for overlapping markets (using 2025 branch deposits)
        hhi_results = calculate_hhi_for_overlapping_markets(overlapping_cbsas, year=2025)
        
        # Step 4: Create comprehensive report
        output_dir = Path("reports/merger_analysis")
        output_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = output_dir / f"FifthThird_Comerica_Merger_Analysis_{timestamp}.xlsx"
        
        create_comprehensive_report(footprint_data, hhi_results, str(output_file))
        
        # Summary
        print(f"\n{'='*80}")
        print("ANALYSIS COMPLETE!")
        print(f"{'='*80}")
        print(f"\nReport saved to: {output_file}")
        
        if len(hhi_results) > 0:
            concerns = hhi_results[hhi_results['antitrust_concern'] == True]
            print(f"\nKEY FINDINGS:")
            print(f"  Total overlapping markets analyzed: {len(hhi_results)}")
            print(f"  Markets with antitrust concerns: {len(concerns)}")
            
            if len(concerns) > 0:
                print(f"\n  MARKETS REQUIRING ATTENTION:")
                for _, row in concerns.iterrows():
                    print(f"    CBSA {row['cbsa_code']}:")
                    print(f"      Pre-merger HHI: {row['hhi_pre_merger']:.0f} ({row['concentration_category_pre']})")
                    print(f"      Post-merger HHI: {row['hhi_post_merger']:.0f} ({row['concentration_category_post']})")
                    print(f"      HHI Increase: +{row['hhi_change']:.0f}")
                    print(f"      Combined Market Share: {row['combined_share_pct']:.1f}%")
                    print()
    else:
        print("\nWARNING: Could not identify overlapping markets - check branch data")
    
    print("\n" + "="*80)
    print("NEXT STEPS:")
    print("="*80)
    print("1. Review the HHI Analysis sheet for antitrust concerns")
    print("2. Review Branch Overlap sheet to see where both banks operate")
    print("3. Generate Comerica CBA report (same format as Fifth Third)")
    print("4. Use findings for regulatory filings and advocacy")
    print("="*80)

if __name__ == "__main__":
    main()

