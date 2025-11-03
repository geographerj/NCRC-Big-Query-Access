"""Document the complete structure of the Local Market Analysis Excel template"""

import openpyxl
from pathlib import Path
import json

template_file = Path(r"C:\Users\edite\OneDrive - Nat'l Community Reinvestment Coaltn\Desktop\DREAM Analysis\data\raw\PNC+FirstBank merger research ticket.xlsx")

print("="*80)
print("LOCAL MARKET ANALYSIS TEMPLATE STRUCTURE DOCUMENTATION")
print("="*80)

if not template_file.exists():
    print(f"\nERROR: Template file not found: {template_file}")
    print("Please ensure the file exists and provide the correct path.")
    exit(1)

try:
    wb = openpyxl.load_workbook(template_file, data_only=False)
except PermissionError:
    print(f"\nERROR: Cannot open file - it is likely open in Excel.")
    print(f"Please close the Excel file and run this script again.")
    print(f"\nFile path: {template_file}")
    print(f"\nAlternatively, you can save a copy of the template and update the path in this script.")
    exit(1)
except Exception as e:
    print(f"\nERROR: Could not open file: {e}")
    print(f"\nFile path: {template_file}")
    exit(1)

print(f"\nFile: {template_file.name}")
print(f"Total sheets: {len(wb.sheetnames)}")

# Document structure for each sheet
template_structure = {
    'file_path': str(template_file),
    'sheets': {}
}

print("\n" + "="*80)
print("SHEET-BY-SHEET STRUCTURE:")
print("="*80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}")
    print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
    
    sheet_info = {
        'name': sheet_name,
        'rows': ws.max_row,
        'columns': ws.max_column,
        'headers': {},
        'column_structure': {},
        'sample_data': []
    }
    
    # Analyze columns A through M (first 13 columns)
    print(f"\nColumn Analysis (Columns A-M):")
    print("-" * 80)
    
    for col_idx in range(1, min(14, ws.max_column + 1)):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        
        # Get header (row 1)
        header_cell = ws.cell(row=1, column=col_idx)
        header_value = header_cell.value if header_cell.value else ""
        
        # Get sample data (rows 2-15)
        sample_values = []
        for row_idx in range(2, min(16, ws.max_row + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                if cell.data_type == 'f':
                    sample_values.append(f"FORMULA: {str(cell.value)[:50]}")
                else:
                    sample_values.append(str(cell.value)[:50])
        
        # Determine column type
        formula_count = sum(1 for row_idx in range(2, min(ws.max_row + 1, 100)) 
                           if ws.cell(row=row_idx, column=col_idx).data_type == 'f')
        data_count = sum(1 for row_idx in range(2, min(ws.max_row + 1, 100)) 
                        if ws.cell(row=row_idx, column=col_idx).value is not None 
                        and ws.cell(row=row_idx, column=col_idx).data_type != 'f')
        
        col_type = "FORMULAS" if formula_count > data_count else "DATA"
        
        print(f"\n  Column {col_letter} ({col_idx}):")
        print(f"    Header: {header_value}")
        print(f"    Type: {col_type} ({formula_count} formulas, {data_count} data cells in first 100 rows)")
        
        if sample_values:
            print(f"    Sample values (rows 2-15):")
            for val in sample_values[:5]:
                print(f"      - {val}")
            if len(sample_values) > 5:
                print(f"      ... ({len(sample_values) - 5} more)")
        
        sheet_info['headers'][col_letter] = str(header_value) if header_value else ""
        sheet_info['column_structure'][col_letter] = {
            'type': col_type,
            'formula_count': formula_count,
            'data_count': data_count,
            'sample_values': sample_values[:10]
        }
    
    # Show first 10 rows for key sheets
    if sheet_name in ['Mortgage Goals', 'PNCMortgage Data', 'FirstBankMortgage Data', 
                      'PNCSmall Business Data', 'FirstBankSmall Business Data',
                      'PNCBranch Data', 'FirstBankBranch Data', 'Assessment Areas', 'Notes']:
        print(f"\nFirst 10 rows (Columns A-F):")
        print("-" * 80)
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(7, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                if cell.value is not None:
                    if cell.data_type == 'f':
                        val = f"={str(cell.value)[:30]}"
                    else:
                        val = str(cell.value)[:30]
                    row_data.append(f"{col_letter}:{val}")
                else:
                    row_data.append(f"{col_letter}:")
            
            if any(ws.cell(row=row_idx, column=col).value for col in range(1, 7)):
                print(f"Row {row_idx}: {' | '.join(row_data)}")
        
        sheet_info['sample_data'] = []
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_dict = {}
            for col_idx in range(1, min(14, ws.max_column + 1)):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    if cell.data_type == 'f':
                        row_dict[col_letter] = f"={str(cell.value)}"
                    else:
                        row_dict[col_letter] = str(cell.value)
            if row_dict:
                sheet_info['sample_data'].append(row_dict)
    
    template_structure['sheets'][sheet_name] = sheet_info

# Save structure to JSON file
output_file = Path("reports/Local Markets Analyses/_shared/goal_setting_template_structure.json")
output_file.parent.mkdir(parents=True, exist_ok=True)

with open(output_file, 'w') as f:
    json.dump(template_structure, f, indent=2, default=str)

print(f"\n{'='*80}")
print("STRUCTURE DOCUMENTATION SAVED")
print(f"{'='*80}")
print(f"\nSaved to: {output_file}")
print(f"\nThis file can be used as a reference for:")
print("  - Understanding column structures for each sheet")
print("  - Identifying which columns contain formulas vs data")
print("  - Determining where to insert data in each sheet")

