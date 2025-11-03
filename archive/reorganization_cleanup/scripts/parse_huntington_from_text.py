"""
Parse Huntington assessment areas from the provided text
This extracts counties from the structured assessment area listing
"""

import re
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def parse_huntington_aa_text(text):
    """Parse assessment areas from the provided text"""
    
    assessment_areas = []
    current_cbsa_name = None
    current_cbsa_code = None
    current_md = None
    current_state_abbr = None  # Track state for counties without prefix
    
    lines = text.split('\n')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Check for MMSA/MSA line - contains CBSA name and possibly code
        # Format: "Chicago-Naperville-Elgin, IL-IN-WI (Chicago MMSA) *"
        mmsa_match = re.search(r'^([A-Za-z][^\(]+?)\s*\(([^\)]+)\)', line)
        if mmsa_match:
            cbsa_full_name = mmsa_match.group(1).strip()
            # Try to extract CBSA code if present
            cbsa_code_match = re.search(r'\(([^\)]+)\)', line)
            # For now, we'll need to map CBSA name to code using crosswalk
            
            # Clean up CBSA name
            current_cbsa_name = cbsa_full_name.split(',')[0].strip()
            current_cbsa_code = None  # Will map later
            current_md = None
            continue
        
        # Check for Metropolitan Division (MD)
        # Format: "Chicago-Naperville-Evanston, IL MD"
        md_match = re.search(r'^([A-Za-z][^,]+),\s+([A-Z]{2})\s+MD', line)
        if md_match:
            current_md = md_match.group(1).strip()
            continue
        
        # Check for state header (e.g., "Colorado", "Florida", "Indiana", "South Dakota", "West Virginia")
        # This helps set context for counties that follow
        # Match single word states or multi-word states
        state_names = [
            'Colorado', 'Florida', 'Indiana', 'Michigan', 'Minnesota', 'Ohio',
            'Pennsylvania', 'South Dakota', 'West Virginia', 'Wisconsin'
        ]
        if line in state_names:
            # Map state name to abbreviation
            state_to_abbr = {
                'Colorado': 'CO', 'Florida': 'FL', 'Indiana': 'IN', 'Michigan': 'MI',
                'Minnesota': 'MN', 'Ohio': 'OH', 'Pennsylvania': 'PA', 
                'South Dakota': 'SD', 'West Virginia': 'WV', 'Wisconsin': 'WI'
            }
            current_state_abbr = state_to_abbr.get(line, '')
            # Don't reset current_cbsa_name here - it might still be valid
            continue
        
        # Check for MSA line (may have counties on same line or next line)
        # Format: "Boulder, CO MSA (Boulder MSA)" or "Denver-Aurora-Lakewood, CO MSA (Denver MSA)"
        msa_match = re.search(r'^([A-Za-z][^,]+),\s+([A-Z]{2})\s+MSA', line)
        if msa_match:
            msa_name = msa_match.group(1).strip()
            state_abbr = msa_match.group(2)
            current_cbsa_name = msa_name
            current_state_abbr = state_abbr  # Track state for counties without prefix
            current_md = None
            continue
        
        # Check for non-metro areas
        # Format: "Michigan non-metro** (MI Non-MSA)" or "Ohio non-metro (OH Non-MSA)"
        nonmetro_match = re.search(r'^([A-Za-z]+)\s+non-metro', line, re.IGNORECASE)
        if nonmetro_match:
            state_name = nonmetro_match.group(1)
            # Map state name to abbreviation
            state_to_abbr = {
                'Michigan': 'MI', 'Minnesota': 'MN', 'Ohio': 'OH', 
                'Pennsylvania': 'PA', 'West Virginia': 'WV'
            }
            current_state_abbr = state_to_abbr.get(state_name, '')
            current_cbsa_name = f"{state_name} Non-MSA"
            current_md = None
            continue
        
        # Check for county list by state
        # Format: "IL: Cook, DuPage, McHenry, Will"
        state_counties_match = re.search(r'^([A-Z]{2}):\s+(.+)', line)
        if state_counties_match:
            state_code_abbr = state_counties_match.group(1)  # IL, OH, etc.
            counties_str = state_counties_match.group(2).strip()
            
            # Split counties (comma-separated)
            counties = [c.strip() for c in counties_str.split(',')]
            
            for county_name in counties:
                if county_name and len(county_name) > 1:
                    assessment_areas.append({
                        'state_abbr': state_code_abbr,
                        'county_name': county_name,
                        'cbsa_name': current_cbsa_name,
                        'md': current_md
                    })
        
        # Check for single county on a line (for Limited reviews) or counties without state prefix
        # Format: "Boulder" or "El Paso" or "Collier" or "Wayne" or "Boone, Hamilton, Hendricks"
        # Only if it looks like county names (capital letter, not a header word)
        if not state_counties_match and not mmsa_match and not md_match and current_cbsa_name and current_state_abbr:
            # Check if line is county name(s) - capitalized word(s), not a header word
            header_words = {'Full', 'Limited', 'MMSAs', 'States', 'Ratings', 'Type', 'Review', 'Other', 
                          'Information', 'Metropolitan', 'Divisions', 'Counties', 'Appendix', 'Charter', 'Number'}
            state_names_lower = [s.lower() for s in ['Colorado', 'Florida', 'Indiana', 'Michigan', 'Minnesota', 
                                                      'Ohio', 'Pennsylvania', 'South Dakota', 'West Virginia', 'Wisconsin']]
            
            if (line not in header_words and line and line[0].isupper() and 
                not line.startswith('Appendix') and line.lower() not in state_names_lower):
                # Check if line contains multiple counties (comma-separated)
                if ',' in line:
                    counties = [c.strip() for c in line.split(',')]
                else:
                    counties = [line.strip()]
                
                # Only add if we have state context
                for county_name in counties:
                    # Remove "County" suffix if present (e.g., "St. Clair County")
                    county_name = county_name.replace(' County', '').strip()
                    # Skip if it's clearly not a county (like "Detroit-Dearborn-Livonia, MI MD")
                    # Skip if it's a state name
                    if (county_name and len(county_name) > 1 and ' MD' not in county_name and 
                        ' MSA' not in county_name and county_name.lower() not in state_names_lower):
                        assessment_areas.append({
                            'state_abbr': current_state_abbr,
                            'county_name': county_name,
                            'cbsa_name': current_cbsa_name,
                            'md': current_md
                        })
    
    return assessment_areas

def map_to_crosswalk(assessment_areas_raw):
    """Map counties to GEOIDs and CBSA codes using crosswalk"""
    
    crosswalk_path = Path(__file__).parent.parent / 'data' / 'reference' / 'CBSA_to_County_Mapping.csv'
    
    if not crosswalk_path.exists():
        print("ERROR: Crosswalk not found")
        return []
    
    crosswalk_df = pd.read_csv(crosswalk_path)
    
    # Create state abbreviation to state code mapping
    state_abbr_to_code = {}
    state_name_to_code = {}
    for _, row in crosswalk_df.iterrows():
        county_state = str(row.get('County State', '')).strip()
        state_code = str(row.get('State Code', '')).strip()
        
        # Extract state name from "County State" (last word)
        if county_state:
            state_name = county_state.rsplit(' ', 1)[-1] if ' ' in county_state else ''
            # Create mapping (state name -> state code)
            if state_name and state_name not in state_name_to_code:
                state_name_to_code[state_name] = state_code
    
    # Also need state abbreviation mapping - common mappings
    state_abbr_map = {
        'IL': '17', 'OH': '39', 'WI': '55', 'KY': '21', 'IN': '18',
        'WV': '54', 'PA': '42', 'CO': '08', 'FL': '12', 'MI': '26',
        'MN': '27', 'SD': '46'
    }
    
    mapped_areas = []
    
    for aa in assessment_areas_raw:
        state_abbr = aa.get('state_abbr', '')
        county_name = aa.get('county_name', '').strip()
        cbsa_name = aa.get('cbsa_name', '')
        
        # Get state code
        state_code = state_abbr_map.get(state_abbr, '')
        if not state_code:
            # Try to find from crosswalk
            for name, code in state_name_to_code.items():
                if state_abbr.lower() in name.lower() or name.lower().startswith(state_abbr.lower()):
                    state_code = code
                    break
        
        if not state_code:
            print(f"  WARNING: Could not find state code for {state_abbr}")
            continue
        
        # Search for county in crosswalk
        # State Code can be int or string, convert to int for comparison
        try:
            state_code_int = int(state_code)
        except (ValueError, TypeError):
            state_code_int = None
        
        if state_code_int is None:
            print(f"  WARNING: Invalid state code for {state_abbr}: {state_code}")
            continue
        
        state_rows = crosswalk_df[crosswalk_df['State Code'].astype(int) == state_code_int]
        
        # Try to match county name - crosswalk format is "County Name State"
        county_name_upper = county_name.upper()
        
        # Try exact match: "County Name County State" or "County Name State"
        county_rows = state_rows[
            state_rows['County State'].str.upper().str.startswith(county_name_upper + ' COUNTY ', na=False) |
            state_rows['County State'].str.upper().str.startswith(county_name_upper + ' ', na=False)
        ]
        
        if len(county_rows) == 0:
            # Try contains match
            county_rows = state_rows[
                state_rows['County State'].str.upper().str.contains(' ' + county_name_upper + ' COUNTY ', na=False) |
                state_rows['County State'].str.upper().str.contains(' ' + county_name_upper + ' ', na=False)
            ]
        
        if len(county_rows) == 0:
            # Try matching first word(s) - for cases like "El Paso" matching "El Paso County"
            county_words = county_name_upper.split()
            if len(county_words) > 0:
                first_word = county_words[0]
                county_rows = state_rows[
                    state_rows['County State'].str.upper().str.startswith(first_word + ' ', na=False)
                ]
        
        if len(county_rows) > 0:
            row = county_rows.iloc[0]
            geoid5 = str(row['Geoid5']).zfill(5)
            cbsa_code = str(row.get('Cbsa Code', '')).strip()
            cbsa_name_from_crosswalk = str(row.get('Cbsa', '')).strip()
            
            # Use CBSA name from input if available, otherwise from crosswalk
            final_cbsa_name = cbsa_name if cbsa_name else cbsa_name_from_crosswalk
            
            # Extract state name from County State
            county_state = str(row.get('County State', ''))
            state_name = county_state.rsplit(' ', 1)[1] if ' ' in county_state else ''
            
            mapped_areas.append({
                'bank_name': 'The Huntington National Bank',
                'cbsa_name': final_cbsa_name,
                'cbsa_code': cbsa_code,
                'state_code': str(state_code_int).zfill(2),
                'state_name': state_name,
                'county_name': county_name.replace(' County', '').strip(),
                'county_code': geoid5
            })
        else:
            print(f"  WARNING: Could not map {county_name}, {state_abbr}")
    
    return mapped_areas

def main():
    text = """List of AAs and Type of Examination

Ratings and AAs

Type of Review

Other Information (Metropolitan Divisions and/or Counties)

MMSAs

Chicago-Naperville-Elgin, IL-IN-WI (Chicago MMSA) *

Full

Chicago-Naperville-Evanston, IL MD

IL: Cook, DuPage, McHenry, Will

Elgin, IL MD

IL: DeKalb, Kane, Kendall

Lake County-Kenosha County, IL-WI MD

IL: Lake

WI: Kenosha

Cincinnati, OH-KY-IN (Cincinnati MSA)

Full

KY: Boone, Kenton

OH: Butler, Clermont, Hamilton, Warren

Weirton-Steubenville, WV-OH (Weirton MMSA)

Full

OH: Jefferson

WV: Hancock

Youngstown-Warren-Boardman, OH-PA (Youngstown MMSA)

Full

OH: Mahoning, Trumbull

PA: Mercer

States

Colorado

Boulder, CO MSA (Boulder MSA)

Limited

Boulder

Colorado Springs, CO MSA (Colorado Springs MSA)

Limited

El Paso

Denver-Aurora-Lakewood, CO MSA (Denver MSA)

Full

Adams, Arapahoe, Broomfield, Denver, Douglas, Jefferson

Florida

Naples-Marco Island, FL MSA (Naples MSA)

Full

Collier

Indiana

Indianapolis-Carmel-Anderson, IN MSA (Indianapolis MSA)

Full

Boone, Hamilton, Hendricks, Johnson, Madison, Marion

Lafayette-West Lafayette, IN MSA (Lafayette MSA)

Limited

Tippecanoe

Michigan

Ann Arbor, MI MSA (Ann Arbor MSA)

Limited

Washtenaw

Battle Creek, MI MSA (Battle Creek MSA)

Limited

Calhoun

Bay City, MI MSA (Bay City MSA)

Limited

Bay

Detroit-Warren-Dearborn, MI MSA (Detroit MSA)

Full

Detroit-Dearborn-Livonia, MI MD

Wayne

Warren-Troy-Farmington Hills, MI MD

Lapeer, Livingston, Macomb, Oakland, St. Clair County

Flint, MI MSA (Flint MSA)

Full

Genesee

Grand Rapids-Kentwood, MI MSA (Grand Rapids MSA)

Limited

Ionia, Kent, Montcalm, Ottawa

Jackson, MI MSA (Jackson MSA)

Limited

Jackson

Kalamazoo-Portage, MI MSA (Kalamazoo MSA)

Limited

Kalamazoo

Lansing-East Lansing, MI MSA (Lansing MSA)

Limited

Clinton, Eaton, Ingham, Shiawassee

Midland, MI MSA (Midland MSA)

Limited

Midland

Monroe, MI MSA (Monroe MSA)

Limited

Monroe

Muskegon, MI MSA (Muskegon MSA)

Limited

Muskegon

Niles, MI MSA (Niles MSA)

Limited

Berrien

Saginaw, MI MSA (Saginaw MSA)

Full

Saginaw

South Bend-Mishawaka, IN-MI MSA (South Bend MSA)

Limited

Cass

Michigan non-metro** (MI Non-MSA)

Limited

Allegan, Alpena, Antrim, Arenac, Barry, Branch, Charlevoix, Chippewa, Clare, Crawford, Delta, Dickinson, Emmet, Gladwin, Grand Traverse, Gratiot, Hillsdale, Houghton, Huron, Iosco, Isabella, Kalkaska, Leelanau, Lenawee, Manistee, Marquette, Mason, Mecosta, Missaukee, Montmorency, Newaygo, Oceana, Ogemaw, Osceola, Oscoda, Otsego, Presque Isle, Roscommon, Sanilac, St. Joseph, Tuscola, Van Buren, Wexford

Minnesota

Duluth, MN-WI MSA (Duluth MSA)

Limited

St. Louis

Mankato, MN MSA (Mankato MSA)

Limited

Blue Earth

Minneapolis-St. Paul-Bloomington, MN-WI MSA (Minneapolis MSA)

Full

Anoka, Carver, Dakota, Hennepin, Ramsey, Scott, Sherburne, Washington, Wright

St. Cloud, MN MSA (St. Cloud MSA)

Limited

Benton, Stearns

Minnesota non-metro (MN Non-MSA)

Limited

Rice

Ohio

Akron, OH MSA (Akron MSA)

Limited

Portage, Summit

Canton, OH MSA (Canton MSA)

Limited

Carroll, Stark

Cleveland-Elyria, OH MSA (Cleveland MSA)

Full

Cuyahoga, Geauga, Lake, Lorain, Medina

Columbus, OH MSA (Columbus MSA)

Full

Delaware, Fairfield, Franklin, Licking, Madison, Pickaway, Union

Dayton-Kettering, OH, MSA (Dayton MSA)

Limited

Greene, Miami, Montgomery

Lima, OH MSA (Lima MSA)

Limited

Allen

Mansfield, OH MSA (Mansfield MSA)

Limited

Richland

Springfield, OH MSA (Springfield MSA)

Limited

Clark

Toledo, OH MSA (Toledo MSA)

Full

Fulton, Lucas, Ottawa, Wood

Wheeling, WV-OH MSA (Wheeling MSA)

Limited

Belmont

Ohio non-metro (OH Non-MSA)

Limited

Ashland, Ashtabula, Columbiana, Crawford, Defiance, Erie, Fayette, Guernsey, Hancock, Hardin, Harrison, Henry, Huron, Knox, Logan, Marion, Muskingum, Putnam, Ross, Sandusky, Seneca, Tuscarawas, Washington, Wayne, Williams, Wyandot

Pennsylvania

Erie, PA MSA (Erie MSA)

Limited

Erie

Pittsburgh, PA MSA (Pittsburgh MSA)

Full

Allegheny, Beaver, Butler, Washington, Westmoreland

Pennsylvania non-metro (PA Non-MSA)

Limited

Lawrence

South Dakota

Sioux Falls, SD MSA (Sioux Falls MSA)

Full

Minnehaha

West Virginia

Charleston, WV MSA (Charleston MSA)

Limited

Kanawha

Huntington-Ashland, WV-KY-OH MSA (Huntington MSA)

Limited

Cabell, Putnam, Wayne

Morgantown, WV MSA (Morgantown MSA)

Limited

Monongalia

Parkersburg-Vienna, WV MSA (Parkersburg MSA)

Limited

Wood

West Virginia non-metro (WV Non-MSA)

Full

Harrison, Lewis, Marion, Randolph, Ritchie

Wisconsin

Milwaukee-Waukesha, WI MSA (Milwaukee MSA)

Full

Milwaukee, Waukesha

Racine, WI MSA (Racine MSA)

Limited

Racine"""
    
    print("="*80)
    print("PARSING HUNTINGTON ASSESSMENT AREAS FROM TEXT")
    print("="*80)
    
    # Parse text
    assessment_areas_raw = parse_huntington_aa_text(text)
    
    print(f"\nFound {len(assessment_areas_raw)} county mentions in text")
    
    # Show sample
    if assessment_areas_raw:
        print("\nSample counties found:")
        for aa in assessment_areas_raw[:10]:
            print(f"  - {aa['county_name']}, {aa['state_abbr']}, CBSA: {aa['cbsa_name']}")
    
    # Map to crosswalk
    assessment_areas = map_to_crosswalk(assessment_areas_raw)
    
    # Remove duplicates
    seen = set()
    unique_areas = []
    for aa in assessment_areas:
        key = (aa['county_code'], aa['cbsa_code'])
        if key not in seen:
            seen.add(key)
            unique_areas.append(aa)
    
    print(f"\nMapped to {len(unique_areas)} unique county assessment areas")
    
    if unique_areas:
        print("\nSample mapped counties (first 10):")
        for aa in unique_areas[:10]:
            print(f"  - {aa['county_name']}, {aa['state_name']}, CBSA: {aa['cbsa_code']} ({aa['cbsa_name']})")
    
    # Update Excel ticket
    ticket_file = Path(__file__).parent.parent / "Huntington+Cadence merger research ticket.xlsx"
    
    if ticket_file.exists():
        wb = load_workbook(ticket_file)
        
        if "Assessment Areas" in wb.sheetnames:
            ws = wb["Assessment Areas"]
            
            # Find next empty row
            next_row = ws.max_row + 1
            
            # Add Huntington assessment areas
            for aa in unique_areas:
                ws.cell(row=next_row, column=1, value=aa['bank_name'])
                ws.cell(row=next_row, column=2, value=aa['cbsa_name'])
                ws.cell(row=next_row, column=3, value=aa['cbsa_code'])
                ws.cell(row=next_row, column=4, value=aa['state_name'])
                ws.cell(row=next_row, column=5, value=aa['county_name'])
                ws.cell(row=next_row, column=6, value=aa['county_code'])
                next_row += 1
            
            wb.save(ticket_file)
            print(f"\n[SUCCESS] Added {len(unique_areas)} Huntington assessment area entries to Excel ticket")
        else:
            print("ERROR: Assessment Areas sheet not found")
    else:
        print(f"ERROR: Ticket file not found: {ticket_file}")
    
    print(f"\n{'='*80}")
    print("PARSING COMPLETE")
    print(f"{'='*80}")
    print(f"\nTotal Huntington counties: {len(unique_areas)}")
    print(f"Unique CBSAs: {len(set(aa['cbsa_code'] for aa in unique_areas if aa['cbsa_code']))}")
    print(f"States: {sorted(set(aa['state_name'] for aa in unique_areas if aa['state_name']))}")

if __name__ == "__main__":
    main()

