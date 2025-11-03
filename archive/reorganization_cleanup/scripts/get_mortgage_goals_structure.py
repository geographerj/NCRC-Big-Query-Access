"""Get complete Mortgage Goals sheet structure"""

import openpyxl
from pathlib import Path

template_file = r"C:\Users\edite\OneDrive - Nat'l Community Reinvestment Coaltn\Desktop\DREAM Analysis\data\raw\PNC+FirstBank merger research ticket.xlsx"

wb = openpyxl.load_workbook(template_file, data_only=False)
ws = wb['Mortgage Goals']

print("="*80)
print("MORTGAGE GOALS - COMPLETE STRUCTURE")
print("="*80)

# Show first 15 rows to understand structure
print("\nFirst 15 rows:")
print("-"*80)
for row_idx in range(1, min(16, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(13, ws.max_column + 1)):
        cell = ws.cell(row_idx, col_idx)
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        
        if cell.data_type == 'f':
            row_data.append(f"{col_letter}{row_idx}:{cell.value[:40]}")
        elif cell.value is not None:
            val_str = str(cell.value)[:30]
            row_data.append(f"{col_letter}{row_idx}:{val_str}")
    
    if row_data:
        print(f"Row {row_idx}: {' | '.join(row_data[:8])}")

# Show formula patterns for key columns
print("\n" + "-"*80)
print("FORMULA PATTERNS BY COLUMN:")
print("-"*80)

formula_examples = {}
for col_idx in range(6, 13):  # Columns F-L (formula columns)
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    examples = []
    for row_idx in range(2, min(20, ws.max_row + 1)):
        cell = ws.cell(row_idx, col_idx)
        if cell.data_type == 'f' and len(examples) < 3:
            examples.append(cell.value)
    
    if examples:
        formula_examples[col_letter] = examples[0]

for col, formula in formula_examples.items():
    print(f"  Column {col}: {formula}")

# Identify data insertion columns
print("\n" + "-"*80)
print("DATA INSERTION POINTS:")
print("-"*80)
print("  Column C (3): Home Purchase - INSERT DATA HERE")
print("  Column D (4): Refinance - INSERT DATA HERE")
print("  Column E (5): Home Equity - INSERT DATA HERE")
print("\n  All other columns (F-L) are FORMULAS - DO NOT MODIFY")

