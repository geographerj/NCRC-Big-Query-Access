"""
Detailed inspection of Excel file structure to understand exact layout,
formula locations, and data insertion points for goal-setting analysis.
"""

import openpyxl
from pathlib import Path
import sys
import json

def inspect_sheet_structure(ws, sheet_name, max_rows=100):
    """
    Detailed inspection of a worksheet to identify:
    - Data insertion points (columns D and E for bank data)
    - Formula locations
    - Header structure
    - Row/column patterns
    """
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}")
    
    print(f"\nDimensions: {ws.max_row} rows x {ws.max_column} columns")
    
    structure = {
        'sheet_name': sheet_name,
        'dimensions': (ws.max_row, ws.max_column),
        'data_columns': {},  # Which columns have data vs formulas
        'formula_cells': [],
        'header_rows': [],
        'data_start_row': None,
        'structure_pattern': []
    }
    
    # Scan first 50 rows to understand structure
    for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
        row_info = {
            'row': row_idx,
            'cells': [],
            'has_formula': False,
            'has_data': False,
            'row_type': None  # 'header', 'total', 'data', 'empty'
        }
        
        for col_idx in range(1, min(11, ws.max_column + 1)):  # Check first 10 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            
            cell_info = {
                'col': col_idx,
                'col_letter': col_letter,
                'value': str(cell.value)[:50] if cell.value is not None else None,
                'data_type': cell.data_type,
                'has_formula': cell.data_type == 'f'
            }
            
            if cell.data_type == 'f':
                cell_info['formula'] = cell.value
                row_info['has_formula'] = True
                structure['formula_cells'].append({
                    'cell': cell.coordinate,
                    'row': row_idx,
                    'col': col_idx,
                    'formula': cell.value
                })
            
            if cell.value is not None and cell.data_type != 'f':
                row_info['has_data'] = True
            
            row_info['cells'].append(cell_info)
        
        # Determine row type
        if row_idx <= 3:
            if row_info['has_data']:
                row_info['row_type'] = 'header'
                structure['header_rows'].append(row_idx)
        elif 'Total' in str([c['value'] for c in row_info['cells'] if c['value']]):
            row_info['row_type'] = 'total'
        elif row_info['has_data'] or row_info['has_formula']:
            if structure['data_start_row'] is None:
                structure['data_start_row'] = row_idx
            row_info['row_type'] = 'data'
        else:
            row_info['row_type'] = 'empty'
        
        structure['structure_pattern'].append(row_info)
    
    # Identify data columns (D and E typically)
    print(f"\nColumn Analysis:")
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        numeric_count = 0
        formula_count = 0
        
        for row_idx in range(1, min(101, ws.max_row + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                numeric_count += 1
            if cell.data_type == 'f':
                formula_count += 1
        
        if numeric_count > 5 or formula_count > 5:
            structure['data_columns'][col_letter] = {
                'numeric_cells': numeric_count,
                'formula_cells': formula_count
            }
            print(f"  Column {col_letter}: {numeric_count} numeric, {formula_count} formulas")
    
    # Show sample structure
    print(f"\nSample Structure (first 15 rows):")
    for row_info in structure['structure_pattern'][:15]:
        if row_info['row_type'] != 'empty':
            row_cells = [f"{c['col_letter']}:{c['value'][:20] if c['value'] else ''}" 
                        for c in row_info['cells'][:6] if c['value']]
            print(f"  Row {row_info['row']} ({row_info['row_type']}): {', '.join(row_cells)}")
    
    return structure

def inspect_workbook(excel_file):
    """Inspect entire workbook structure"""
    excel_path = Path(excel_file)
    
    print("="*80)
    print("DETAILED EXCEL STRUCTURE INSPECTION")
    print("="*80)
    
    wb = openpyxl.load_workbook(excel_file, data_only=False)  # Keep formulas
    
    print(f"\nFile: {excel_path.name}")
    print(f"Total sheets: {len(wb.sheetnames)}")
    
    all_structures = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        structure = inspect_sheet_structure(ws, sheet_name)
        all_structures[sheet_name] = structure
    
    # Save structure to JSON
    # Convert to JSON-serializable format
    json_structure = {}
    for sheet_name, struct in all_structures.items():
        json_structure[sheet_name] = {
            'dimensions': struct['dimensions'],
            'data_start_row': struct['data_start_row'],
            'data_columns': struct['data_columns'],
            'formula_count': len(struct['formula_cells']),
            'header_rows': struct['header_rows'],
            'sample_formulas': [{'cell': f['cell'], 'formula': f['formula']} 
                               for f in struct['formula_cells'][:20]]
        }
    
    output_file = excel_path.parent / f"{excel_path.stem}_structure.json"
    with open(output_file, 'w') as f:
        json.dump(json_structure, f, indent=2)
    
    print(f"\n\nStructure saved to: {output_file}")
    
    return all_structures

def main():
    if len(sys.argv) < 2:
        print("Usage: python inspect_excel_structure_detailed.py <excel_file>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    
    try:
        inspect_workbook(excel_file)
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()


