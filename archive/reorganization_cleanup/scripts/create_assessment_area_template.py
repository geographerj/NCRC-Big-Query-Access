"""
Create a template Excel file for manual entry of assessment areas.

Since PDF parsing can be unreliable, this creates a template where users
can manually enter the counties included in each assessment area.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
import sys

def create_assessment_area_template(output_file="assessment_area_template.xlsx"):
    """Create Excel template for entering assessment area counties"""
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Sheet 1: Instructions
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        "ASSESSMENT AREA ENTRY TEMPLATE",
        "",
        "Purpose:",
        "This template is for entering which counties are included in each assessment area.",
        "Assessment areas are typically CBSAs/MSAs, but only specific counties within them are included.",
        "",
        "How to Use:",
        "1. For each assessment area (CBSA/MSA), list the counties that are included",
        "2. Fill in the Assessment Areas sheet with:",
        "   - CBSA Code (5-digit code)",
        "   - CBSA Name",
        "   - County GEOID (5-digit FIPS code)",
        "   - County Name",
        "   - State Code (2-letter)",
        "   - State Name",
        "",
        "Data Sources:",
        "- Assessment area PDFs from the bank",
        "- Bank's CRA public file",
        "- Bank's regulatory filings",
        "",
        "Example:",
        "If Chicago CBSA (16980) includes Cook, DuPage, and Lake counties:",
        "CBSA Code: 16980",
        "CBSA Name: Chicago-Naperville-Elgin, IL-IN-WI",
        "County GEOID: 17031, County Name: Cook County, State: IL",
        "County GEOID: 17043, County Name: DuPage County, State: IL",
        "County GEOID: 17097, County Name: Lake County, State: IL",
    ]
    
    for idx, line in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=idx, column=1, value=line)
        if idx == 1:
            cell.font = Font(bold=True, size=14)
        elif line.endswith(":"):
            cell.font = Font(bold=True)
    
    ws_instructions.column_dimensions['A'].width = 80
    
    # Sheet 2: Assessment Areas Data Entry
    ws_data = wb.create_sheet("Assessment Areas")
    
    headers = [
        "CBSA_Code",
        "CBSA_Name", 
        "County_GEOID",
        "County_Name",
        "State_Code",
        "State_Name"
    ]
    
    # Write headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Set column widths
    ws_data.column_dimensions['A'].width = 12  # CBSA Code
    ws_data.column_dimensions['B'].width = 40  # CBSA Name
    ws_data.column_dimensions['C'].width = 12  # County GEOID
    ws_data.column_dimensions['D'].width = 30  # County Name
    ws_data.column_dimensions['E'].width = 12  # State Code
    ws_data.column_dimensions['F'].width = 20  # State Name
    
    # Add example rows
    examples = [
        ["16980", "Chicago-Naperville-Elgin, IL-IN-WI", "17031", "Cook County", "IL", "Illinois"],
        ["16980", "Chicago-Naperville-Elgin, IL-IN-WI", "17043", "DuPage County", "IL", "Illinois"],
        ["16980", "Chicago-Naperville-Elgin, IL-IN-WI", "17097", "Lake County", "IL", "Illinois"],
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')  # Gray for examples
    
    ws_data.freeze_panes = 'A2'
    ws_data.auto_filter.ref = 'A1:F1000'
    
    # Save
    wb.save(output_file)
    print(f"\nTemplate created: {output_file}")
    print("\nInstructions:")
    print("1. Fill in the 'Assessment Areas' sheet with county data")
    print("2. Delete the example rows (gray rows)")
    print("3. Save the file")
    print("4. Use this file as input for the goal-setting analysis")
    
    return output_file

def load_assessment_areas_from_excel(excel_file):
    """Load assessment areas from Excel template"""
    print(f"\nLoading assessment areas from: {excel_file}")
    
    df = pd.read_excel(excel_file, sheet_name="Assessment Areas")
    
    # Remove example rows (those with gray fill or specific patterns)
    # User should delete these, but we can filter if needed
    df = df.dropna(subset=['CBSA_Code', 'County_GEOID'])  # Remove empty rows
    
    print(f"  Loaded {len(df)} county-assessment area combinations")
    print(f"  Unique CBSAs: {df['CBSA_Code'].nunique()}")
    print(f"  Unique counties: {df['County_GEOID'].nunique()}")
    
    # Convert GEOID to string and pad if needed (should be 5 digits)
    df['County_GEOID'] = df['County_GEOID'].astype(str).str.zfill(5)
    
    # Group by CBSA
    counties_by_cbsa = {}
    for cbsa_code in df['CBSA_Code'].unique():
        cbsa_data = df[df['CBSA_Code'] == cbsa_code]
        counties_by_cbsa[str(cbsa_code)] = {
            'cbsa_name': cbsa_data['CBSA_Name'].iloc[0] if 'CBSA_Name' in cbsa_data.columns else None,
            'counties': cbsa_data[['County_GEOID', 'County_Name', 'State_Code', 'State_Name']].to_dict('records')
        }
    
    return {
        'counties_by_cbsa': counties_by_cbsa,
        'all_counties': df[['County_GEOID', 'County_Name', 'CBSA_Code', 'CBSA_Name', 'State_Code', 'State_Name']].to_dict('records'),
        'dataframe': df
    }

def main():
    """Main function"""
    if len(sys.argv) < 2:
        print("Usage:")
        print("  Create template:  python create_assessment_area_template.py --create [output_file.xlsx]")
        print("  Load from Excel:  python create_assessment_area_template.py <excel_file.xlsx>")
        print("\nExample:")
        print("  python create_assessment_area_template.py --create assessment_area_template.xlsx")
        print("  python create_assessment_area_template.py my_assessment_areas.xlsx")
        sys.exit(1)
    
    if sys.argv[1] == '--create':
        output_file = sys.argv[2] if len(sys.argv) > 2 else "assessment_area_template.xlsx"
        create_assessment_area_template(output_file)
    else:
        excel_file = sys.argv[1]
        results = load_assessment_areas_from_excel(excel_file)
        
        print("\n" + "="*80)
        print("ASSESSMENT AREAS SUMMARY")
        print("="*80)
        print(f"\nCBSAs: {len(results['counties_by_cbsa'])}")
        for cbsa, data in results['counties_by_cbsa'].items():
            print(f"  CBSA {cbsa} ({data['cbsa_name']}): {len(data['counties'])} counties")
        
        print(f"\nTotal counties: {len(results['all_counties'])}")
        print("\nFirst 10 counties:")
        for county in results['all_counties'][:10]:
            print(f"  GEOID {county['County_GEOID']}: {county['County_Name']}, {county['State_Code']} (CBSA {county['CBSA_Code']})")

if __name__ == "__main__":
    main()


