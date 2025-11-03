from openpyxl import load_workbook

wb = load_workbook('Fifth_Third_CBA_Report.xlsx')
print("Sheets in workbook:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")


