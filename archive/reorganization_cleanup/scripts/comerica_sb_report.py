"""
Comerica Bank Small Business Lending Report Generator
Matches format of HMDA CBA reports
"""

import sys
import os
import pandas as pd
import numpy as np
from pathlib import Path
import glob
import argparse

# Import SB metrics calculation and Excel workbook creation
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from fifth_third_sb_report import (
    calculate_sb_metrics,
    add_calculated_sb_metrics,
    create_sb_excel_workbook
)

def create_comerica_excel_workbook(df, output_file):
    """Create Comerica Excel workbook, reusing the SB workbook function but updating lender name"""
    # Temporarily replace lender_name in dataframe for header
    df_copy = df.copy()
    if 'lender_name' in df_copy.columns:
        df_copy['lender_name'] = 'Comerica Bank'
    
    # Create workbook using the shared function
    create_sb_excel_workbook(df_copy, output_file)
    
    # Now update the headers and Methods sheet in the workbook
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    
    wb = load_workbook(output_file)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Update first row (title)
        if ws.cell(1, 1).value and 'Fifth Third' in str(ws.cell(1, 1).value):
            ws.cell(1, 1).value = 'Comerica Bank - Small Business Lending Analysis'
        
        # Update Methods sheet with Comerica-specific text
        if sheet_name == 'Methods':
            # Update methodology text for Comerica
            methodology_text = """
COMERICA BANK SMALL BUSINESS LENDING ANALYSIS - METHODOLOGY

PURPOSE:
This report analyzes Comerica Bank's small business lending patterns compared to peer lenders
across key markets. Focuses on identifying potential disparities in lending to low-to-moderate 
income communities and small businesses.

GEOGRAPHIC SCOPE:
- Top 10 CBSAs (Core Based Statistical Areas) by Comerica Bank small business loan volume
- Years: 2018-2023
- All census tracts within selected CBSAs

DATA SOURCE:
- Community Reinvestment Act (CRA) Small Business Lending Data
- Federal Financial Institutions Examination Council (FFIEC)
- BigQuery Dataset: hdma1-242116.sb.disclosure and hdma1-242116.sb.lenders
- Years: 2018-2023

METRICS ANALYZED:

1. SB LOANS (Total Count):
   - Definition: Total number of small business loans originated
   - Calculation: num_under_100k + num_100k_250k + num_250k_1m
   - Represents loans ≤$1 million in original loan amount
   - Units: Count of loans

2. LMICT% (Low-to-Moderate Income Census Tract):
   - Definition: Percentage of loans in Low-to-Moderate Income Census Tracts
   - Calculation: (Loans in LMICT / Total SB Loans) × 100
   - LMICT Definition: Census tracts with income_group_total = '101', '102', or '1'-'8'
     - '101': Income relative to area median income ≤ 50%
     - '102': Income relative to area median income 50%-80%
     - '1'-'8': Additional LMI tract classifications
   - Units: Percentage

3. AVG SB LMICT LOAN AMOUNT:
   - Definition: Average loan amount for loans in Low-to-Moderate Income Census Tracts
   - Calculation: SUM(Amount of LMICT loans) / COUNT(LMICT loans)
   - Provides insight into loan size in underserved communities
   - Units: Dollars

4. LOANS REV UNDER $1m% (Revenue-Based Small Business Loans):
   - Definition: Percentage of loans to businesses with revenue under $1 million
   - Calculation: (numsbrev_under_1m / Total SB Loans) × 100
   - Identifies focus on very small businesses
   - Units: Percentage

5. AVG LOAN AMT FOR RUM SB (Revenue Under $1M Small Business):
   - Definition: Average loan amount for revenue-based small business loans under $1M
   - Calculation: amtsbrev_under_1m / numsbrev_under_1m
   - Indicates typical loan size for very small businesses
   - Units: Dollars

PEER COMPARISON METHODOLOGY:
- Peer Definition: Lenders with 50%-200% of Comerica Bank's small business loan volume 
  in each CBSA-year combination
- Peer Selection: Volume-based matching within each market and year
- Analysis excludes Comerica Bank from peer group
- Each CBSA-year has its own independent peer group
- Peer metrics are aggregated across all qualifying peer lenders

CALCULATION METHODS:
- Subject Share: (Comerica metric / Comerica total SB loans) × 100
- Peer Share: (Peer group metric / Peer group total SB loans) × 100
- Gap: Subject Share - Peer Share (percentage points)
- For dollar amounts (averages): Subject Average - Peer Average
- Positive gap means Comerica outperforms peers
- Negative gap means Comerica underperforms peers

STATISTICAL SIGNIFICANCE:
- Two-proportion z-test performed for percentage-based metrics (LMICT%, Loans Rev Under $1m%)
- Asterisk (*) indicates statistically significant negative gap (p < 0.05)
- Only negative gaps are flagged for significance
- Dollar amount metrics (averages) do not include statistical testing

COLOR CODING:
- Green: Positive gap (bank outperforms peers)
- Yellow: Small negative gap (0 to -2.0 percentage points)
- Orange: Moderate negative gap (-2.0 to -5.0 percentage points)
- Light Red: Large negative gap (-5.0 to -10.0 percentage points)
- Red: Severe negative gap (<-10.0 percentage points)
- Red with white bold text: Severe negative gap + statistically significant

DATA FILTERS:
- Small Business Loans: Loans ≤$1 million in original amount
- Revenue-Based Loans: Loans to businesses with revenue under $1 million
- Census Tract Income Classification: Based on FFIEC CRA income classifications
- Geographic Matching: CBSA codes matched from msamd field in disclosure table

LIMITATIONS:
- Does not account for differences in business models or market strategies
- Geographic selection biased toward high-volume markets
- Peer matching based solely on loan volume
- Does not include all factors affecting small business lending decisions
- Correlation does not imply causation
- Does not assess individual loan applications for discrimination
- LMICT classifications based on FFIEC CRA definitions, which may not capture all dimensions of need
- Revenue data only available for revenue-based loans, not all small business loans

INTERPRETATION GUIDANCE:
- Focus on patterns across years, not single-year gaps
- Statistically significant gaps (marked with *) are more reliable
- Large, persistent negative gaps warrant closer review
- Lower average loan amounts in LMICT may indicate focus on smaller businesses, which could be positive
- Higher percentage of loans to very small businesses (revenue <$1M) may indicate better outreach to underserved communities

CONTACT:
NCRC Research Department
National Community Reinvestment Coalition
For questions about methodology or data, contact: research@ncrc.org
"""
            # Clear existing content and write new
            ws.delete_rows(1, ws.max_row)
            for i, line in enumerate(methodology_text.strip().split('\n'), 1):
                ws.cell(row=i, column=1, value=line)
                ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    
    wb.save(output_file)

def main():
    parser = argparse.ArgumentParser(description='Generate Comerica Small Business Lending Report')
    parser.add_argument('--input', help='Input CSV file from BigQuery')
    parser.add_argument('--output', help='Output Excel file path')
    
    args = parser.parse_args()
    
    base_dir = Path(__file__).parent.parent
    reports_dir = base_dir / "reports" / "fifth_third_merger"  # Same folder as other merger reports
    data_dir = base_dir / "data" / "raw"
    
    reports_dir.mkdir(parents=True, exist_ok=True)
    
    # Find input file
    if args.input:
        input_file = args.input if os.path.isabs(args.input) else data_dir / args.input
    else:
        # Auto-detect Comerica SB file
        sb_files = list(data_dir.glob("*comerica*sb*.csv")) + list(data_dir.glob("*comerica*small*business*.csv"))
        if not sb_files:
            print(f"ERROR: No Comerica SB data file found in {data_dir}")
            sys.exit(1)
        input_file = sb_files[0]
        print(f"Auto-detected input file: {input_file}")
    
    # Load data
    print(f"\nLoading data from: {input_file}")
    df = pd.read_csv(input_file)
    print(f"  Loaded {len(df)} rows")
    
    # Calculate metrics
    df = add_calculated_sb_metrics(df)
    
    # Create Excel report
    if args.output:
        output_file = args.output
    else:
        output_file = reports_dir / "Comerica_SB_Report.xlsx"
    
    print(f"\nCreating Excel workbook: {output_file}")
    create_comerica_excel_workbook(df, str(output_file))
    
    print("\n" + "="*80)
    print("REPORT COMPLETE!")
    print("="*80)

if __name__ == "__main__":
    main()

