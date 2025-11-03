"""
Parse the LMA Ticket Excel file to extract bank information and filter settings.

Reads the "Bank Details" and "LMA Ticket Filters" sheets to extract:
- Acquirer bank info (name, LEI, RSSD, etc.)
- Target bank info (name, LEI, RSSD, etc.)
- Filter settings (years, HMDA filters, etc.)
"""

import pandas as pd
import openpyxl
from pathlib import Path
import sys
import json

def parse_ticket_excel(excel_file):
    """
    Parse the LMA ticket Excel file.
    
    Returns a dictionary with:
    - acquirer: Bank information for acquirer
    - target: Bank information for target
    - filters: Filter settings
    - assessment_areas_note: Any notes about assessment areas
    """
    print(f"\nParsing ticket Excel: {excel_file}")
    
    excel_path = Path(excel_file)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    print(f"  Sheets: {wb.sheetnames}")
    
    result = {
        'acquirer': {},
        'target': {},
        'filters': {},
        'assessment_areas_note': None
    }
    
    # Parse "Bank Details" sheet
    if "Bank Details" in wb.sheetnames:
        ws = wb["Bank Details"]
        print("\n  Parsing 'Bank Details' sheet...")
        
        # Read all cells to find bank information
        # Based on inspection, structure is:
        # Row 3: Headers for ACQUIRER and TARGET
        # Row 4+: Bank info
        
        for row_idx in range(1, ws.max_row + 1):
            row_values = []
            for col_idx in range(1, min(ws.max_column + 1, 20)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_values.append(str(cell_value) if cell_value is not None else "")
            
            # Look for specific fields
            if any("bank" in str(v).lower() for v in row_values):
                # Check if this is acquirer or target column
                is_acquirer = False
                is_target = False
                
                # Check header row (row 3)
                if row_idx == 3:
                    if "ACQUIRER" in ' '.join(row_values):
                        is_acquirer = True
                    if "TARGET" in ' '.join(row_values):
                        is_target = True
                
                # If we found headers, read data rows
                if row_idx >= 4:
                    # Determine which column set (acquirer is columns 1-3, target is columns 4-6 approximately)
                    for col_idx in range(1, min(len(row_values), 20)):
                        label = str(row_values[col_idx]).strip()
                        value = str(row_values[col_idx + 1]).strip() if col_idx + 1 < len(row_values) else ""
                        
                        # Determine if acquirer or target based on position
                        # Acquirer info typically in columns B-C, Target in columns E-F
                        is_acquirer_col = col_idx <= 4
                        is_target_col = col_idx >= 4
                        
                        if label and value and value not in ["", "None", "nan"]:
                            label_lower = label.lower()
                            
                            if "bank" in label_lower and "information" not in label_lower:
                                if is_acquirer_col and not result['acquirer'].get('name'):
                                    result['acquirer']['name'] = value
                                elif is_target_col and not result['target'].get('name'):
                                    result['target']['name'] = value
                            
                            elif "lei" in label_lower or "lei code" in label_lower:
                                if is_acquirer_col:
                                    result['acquirer']['lei'] = value
                                elif is_target_col:
                                    result['target']['lei'] = value
                            
                            elif "rssd" in label_lower or "rss-id" in label_lower:
                                if is_acquirer_col:
                                    result['acquirer']['rssd'] = value
                                elif is_target_col:
                                    result['target']['rssd'] = value
                            
                            elif "respondent id" in label_lower and "sb" in label_lower:
                                if is_acquirer_col:
                                    result['acquirer']['sb_respondent_id'] = value
                                elif is_target_col:
                                    result['target']['sb_respondent_id'] = value
                            
                            elif "location" in label_lower:
                                if is_acquirer_col:
                                    result['acquirer']['location'] = value
                                elif is_target_col:
                                    result['target']['location'] = value
                            
                            elif "assessment area" in label_lower and "***" in value:
                                result['assessment_areas_note'] = value
    
    # Parse "LMA Ticket Filters" sheet
    if "LMA Ticket Filters" in wb.sheetnames:
        ws = wb["LMA Ticket Filters"]
        print("\n  Parsing 'LMA Ticket Filters' sheet...")
        
        # Read as table (headers in row 1)
        df_filters = pd.read_excel(excel_file, sheet_name="LMA Ticket Filters", header=0)
        
        for idx, row in df_filters.iterrows():
            filter_type = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
            default_setting = str(row.iloc[1]) if len(row) > 1 and pd.notna(row.iloc[1]) else ""
            adjustment = str(row.iloc[2]) if len(row) > 2 and pd.notna(row.iloc[2]) else ""
            
            if filter_type and "years" in filter_type.lower():
                if "hmda" in filter_type.lower():
                    result['filters']['hmda_years'] = {
                        'default': default_setting,
                        'adjustment': adjustment
                    }
                elif "small business" in filter_type.lower() or "sb" in filter_type.lower():
                    result['filters']['sb_years'] = {
                        'default': default_setting,
                        'adjustment': adjustment
                    }
            
            elif "occupancy" in filter_type.lower():
                result['filters']['occupancy'] = {
                    'default': default_setting,
                    'adjustment': adjustment
                }
            
            elif "action taken" in filter_type.lower():
                result['filters']['action_taken'] = {
                    'default': default_setting,
                    'adjustment': adjustment
                }
            
            elif "units" in filter_type.lower() and "filter" in filter_type.lower():
                result['filters']['units'] = {
                    'default': default_setting,
                    'adjustment': adjustment
                }
            
            elif "construction" in filter_type.lower():
                result['filters']['construction_method'] = {
                    'default': default_setting,
                    'adjustment': adjustment
                }
            
            elif "loan purpose" in filter_type.lower():
                result['filters']['loan_purpose'] = {
                    'default': default_setting,
                    'adjustment': adjustment
                }
    
    # Clean up empty values
    for bank_type in ['acquirer', 'target']:
        result[bank_type] = {k: v for k, v in result[bank_type].items() if v and v != ""}
    
    return result

def print_ticket_results(results):
    """Print parsed ticket results"""
    print("\n" + "="*80)
    print("TICKET PARSING RESULTS")
    print("="*80)
    
    print("\nACQUIRER BANK:")
    for key, value in results['acquirer'].items():
        print(f"  {key}: {value}")
    
    print("\nTARGET BANK:")
    for key, value in results['target'].items():
        print(f"  {key}: {value}")
    
    print("\nFILTERS:")
    for key, value in results['filters'].items():
        print(f"  {key}:")
        print(f"    Default: {value.get('default', 'N/A')}")
        print(f"    Adjustment: {value.get('adjustment', 'N/A')}")
    
    if results['assessment_areas_note']:
        print(f"\nAssessment Areas Note: {results['assessment_areas_note']}")

def main():
    """Main function for command-line usage"""
    if len(sys.argv) < 2:
        print("Usage: python parse_ticket_excel.py <ticket_excel_file>")
        print("\nExample:")
        print('  python parse_ticket_excel.py "The Huntington National Bank Research Ticket_CBA .xlsx"')
        sys.exit(1)
    
    excel_file = sys.argv[1]
    
    try:
        results = parse_ticket_excel(excel_file)
        print_ticket_results(results)
        
        # Optionally save to JSON
        if len(sys.argv) > 2 and sys.argv[2] == '--save':
            output_file = Path(excel_file).stem + "_parsed.json"
            with open(output_file, 'w') as f:
                json.dump(results, f, indent=2)
            print(f"\nResults saved to: {output_file}")
        
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()


