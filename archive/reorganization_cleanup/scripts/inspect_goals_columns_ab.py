"""Inspect columns A and B structure in Mortgage Goals sheet"""

import openpyxl
from pathlib import Path

template_file = r"C:\Users\edite\OneDrive - Nat'l Community Reinvestment Coaltn\Desktop\DREAM Analysis\data\raw\PNC+FirstBank merger research ticket.xlsx"

wb = openpyxl.load_workbook(template_file, data_only=False)
ws = wb['Mortgage Goals']

print("="*80)
print("MORTGAGE GOALS - COLUMNS A & B STRUCTURE")
print("="*80)

# Show first 50 rows to understand the pattern
print("\nRows 1-50 (Columns A & B only):")
print("-"*80)
print(f"{'Row':<6} {'Column A':<30} {'Column B':<30}")
print("-"*80)

for row_idx in range(1, min(51, ws.max_row + 1)):
    col_a = ws.cell(row_idx, 1).value
    col_b = ws.cell(row_idx, 2).value
    
    col_a_str = str(col_a)[:28] if col_a is not None else ''
    col_b_str = str(col_b)[:28] if col_b is not None else ''
    
    print(f"{row_idx:<6} {col_a_str:<30} {col_b_str:<30}")

# Check if there's a pattern with CBSA/county in column A or B
print("\n" + "-"*80)
print("Checking for CBSA/County patterns:")
print("-"*80)

cbsa_patterns = []
county_patterns = []
for row_idx in range(2, min(100, ws.max_row + 1)):
    col_a = ws.cell(row_idx, 1).value
    col_b = ws.cell(row_idx, 2).value
    
    if col_a and isinstance(col_a, str):
        if any(x in col_a.upper() for x in ['CBSA', 'MSA', 'METRO', 'AREA']):
            cbsa_patterns.append((row_idx, col_a))
        if 'COUNTY' in col_a.upper():
            county_patterns.append((row_idx, col_a))
    
    if col_b and isinstance(col_b, str):
        if any(x in col_b.upper() for x in ['CBSA', 'MSA', 'METRO', 'AREA']):
            cbsa_patterns.append((row_idx, col_b))
        if 'COUNTY' in col_b.upper():
            county_patterns.append((row_idx, col_b))

if cbsa_patterns:
    print("\nCBSA patterns found:")
    for row, val in cbsa_patterns[:10]:
        print(f"  Row {row}: {val}")

if county_patterns:
    print("\nCounty patterns found:")
    for row, val in county_patterns[:10]:
        print(f"  Row {row}: {val}")

# Analyze structure
print("\n" + "-"*80)
print("STRUCTURE ANALYSIS:")
print("-"*80)

# Check if column A has hierarchy (State, then CBSA, then County)
state_rows = []
for row_idx in range(2, min(100, ws.max_row + 1)):
    col_a = ws.cell(row_idx, 1).value
    if col_a and isinstance(col_a, str) and col_a not in ['Grand Total'] and not col_a.startswith('~'):
        state_rows.append((row_idx, col_a))

print(f"\nColumn A entries (non-metric, non-Grand Total) in first 100 rows:")
for row_idx, val in state_rows[:20]:
    print(f"  Row {row_idx}: {val}")

