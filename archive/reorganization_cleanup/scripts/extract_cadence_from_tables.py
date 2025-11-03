"""
Extract Cadence assessment areas from the table structure
The PDF has columns: State, State Number, Assessment Area, MSA Number, County Name, County Number, Census Tract
"""

import pdfplumber
from pathlib import Path
import pandas as pd
import json
from collections import defaultdict

def extract_cadence_assessment_areas(pdf_path):
    """Extract assessment areas from Cadence PDF using table structure"""
    pdf_path = Path(pdf_path)
    
    print(f"\nExtracting Cadence assessment areas from: {pdf_path}")
    
    assessment_areas = []
    seen_counties = set()
    
    with pdfplumber.open(pdf_path) as pdf:
        print(f"Total pages: {len(pdf.pages)}")
        
        for page_num, page in enumerate(pdf.pages, 1):
            tables = page.extract_tables()
            
            for table in tables:
                if not table or len(table) < 2:
                    continue
                
                # Check if header row matches expected structure
                header = table[0] if table else []
                header_text = ' '.join([str(cell).lower() if cell else '' for cell in header])
                
                # Look for assessment area table
                if 'assessment area' in header_text and 'msa number' in header_text:
                    # Find column indices
                    state_col = None
                    state_num_col = None
                    aa_col = None
                    msa_col = None
                    county_col = None
                    county_num_col = None
                    
                    for idx, cell in enumerate(header):
                        if cell:
                            cell_lower = str(cell).lower()
                            if 'state' in cell_lower and 'number' not in cell_lower:
                                state_col = idx
                            elif 'state number' in cell_lower or ('state' in cell_lower and 'number' in cell_lower):
                                state_num_col = idx
                            elif 'assessment area' in cell_lower:
                                aa_col = idx
                            elif 'msa number' in cell_lower:
                                msa_col = idx
                            elif 'county name' in cell_lower:
                                county_col = idx
                            elif 'county number' in cell_lower:
                                county_num_col = idx
                    
                    # Extract data rows
                    for row in table[1:]:
                        if len(row) < max([c for c in [state_col, state_num_col, aa_col, msa_col, county_col, county_num_col] if c is not None] or [0]):
                            continue
                        
                        state = str(row[state_col]).strip() if state_col is not None and state_col < len(row) and row[state_col] else ""
                        state_num = str(row[state_num_col]).strip() if state_num_col is not None and state_num_col < len(row) and row[state_num_col] else ""
                        aa_name = str(row[aa_col]).strip() if aa_col is not None and aa_col < len(row) and row[aa_col] else ""
                        msa_code = str(row[msa_col]).strip() if msa_col is not None and msa_col < len(row) and row[msa_col] else ""
                        county_name = str(row[county_col]).strip() if county_col is not None and county_col < len(row) and row[county_col] else ""
                        county_num = str(row[county_num_col]).strip() if county_num_col is not None and county_num_col < len(row) and row[county_num_col] else ""
                        
                        # Create GEOID5: State Number (2 digits) + County Number (3 digits)
                        if state_num and county_num:
                            # Pad state number to 2 digits, county number to 3 digits
                            state_code_padded = str(state_num).zfill(2)
                            county_code_padded = str(county_num).zfill(3)
                            geoid5 = state_code_padded + county_code_padded
                            
                            # Create unique key
                            key = (geoid5, aa_name, msa_code)
                            
                            if key not in seen_counties and geoid5 and msa_code:
                                seen_counties.add(key)
                                
                                assessment_areas.append({
                                    'bank_name': 'Cadence Bank',
                                    'state_name': state,
                                    'state_code': state_code_padded,
                                    'cbsa_name': aa_name,
                                    'cbsa_code': msa_code,
                                    'county_name': county_name,
                                    'county_code': geoid5
                                })
                
                # Alternative: If columns don't match exactly, try to infer from structure
                # The table might have different header format
                elif len(table) > 5 and len(table[0]) >= 7:
                    # Try to extract assuming standard order
                    for row in table[1:]:
                        if len(row) >= 7:
                            try:
                                state = str(row[0]).strip() if row[0] else ""
                                state_num = str(row[1]).strip() if row[1] else ""
                                aa_name = str(row[2]).strip() if row[2] else ""
                                msa_code = str(row[3]).strip() if row[3] else ""
                                county_name = str(row[4]).strip() if row[4] else ""
                                county_num = str(row[5]).strip() if row[5] else ""
                                
                                # Validate: state_num should be 2-digit, msa_code should be 5-digit
                                if state_num and len(state_num) <= 2 and county_num and len(county_num) <= 3 and msa_code and len(msa_code) == 5:
                                    state_code_padded = str(state_num).zfill(2)
                                    county_code_padded = str(county_num).zfill(3)
                                    geoid5 = state_code_padded + county_code_padded
                                    
                                    key = (geoid5, aa_name, msa_code)
                                    
                                    if key not in seen_counties:
                                        seen_counties.add(key)
                                        assessment_areas.append({
                                            'bank_name': 'Cadence Bank',
                                            'state_name': state,
                                            'state_code': state_code_padded,
                                            'cbsa_name': aa_name,
                                            'cbsa_code': msa_code,
                                            'county_name': county_name,
                                            'county_code': geoid5
                                        })
                            except (IndexError, ValueError, TypeError):
                                continue
            
            if page_num % 20 == 0:
                print(f"  Processed {page_num} pages, found {len(assessment_areas)} unique counties so far...")
    
    print(f"\nExtracted {len(assessment_areas)} unique county assessment areas for Cadence")
    
    # Show sample
    if assessment_areas:
        print("\nSample counties (first 10):")
        for aa in assessment_areas[:10]:
            print(f"  - {aa['county_name']}, {aa['state_name']}, CBSA: {aa['cbsa_code']} ({aa['cbsa_name']})")
    
    return assessment_areas

def map_cbsa_names(assessment_areas):
    """Enrich with CBSA names from crosswalk if not already present"""
    crosswalk_path = Path(__file__).parent.parent / 'data' / 'reference' / 'CBSA_to_County_Mapping.csv'
    
    if not crosswalk_path.exists():
        return assessment_areas
    
    crosswalk_df = pd.read_csv(crosswalk_path)
    
    # Create mapping from CBSA code to name
    cbsa_to_name = {}
    for _, row in crosswalk_df.iterrows():
        cbsa_code = str(row.get('Cbsa Code', '')).strip()
        cbsa_name = str(row.get('Cbsa', '')).strip()
        if cbsa_code and cbsa_name:
            cbsa_to_name[cbsa_code] = cbsa_name
    
    # Update CBSA names if missing or improve them
    for aa in assessment_areas:
        cbsa_code = aa.get('cbsa_code', '')
        if cbsa_code in cbsa_to_name:
            # Keep existing name if it looks good, otherwise use crosswalk name
            existing_name = aa.get('cbsa_name', '')
            crosswalk_name = cbsa_to_name[cbsa_code]
            
            # Use crosswalk name if existing is empty or very short
            if not existing_name or len(existing_name) < 3:
                aa['cbsa_name'] = crosswalk_name
    
    return assessment_areas

def update_excel_ticket(assessment_areas):
    """Update Excel ticket with assessment areas"""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    ticket_file = Path(__file__).parent.parent / "Huntington+Cadence merger research ticket.xlsx"
    
    if not ticket_file.exists():
        print(f"ERROR: Ticket file not found: {ticket_file}")
        return False
    
    wb = load_workbook(ticket_file)
    
    if "Assessment Areas" in wb.sheetnames:
        ws = wb["Assessment Areas"]
        # Clear existing data (keep headers)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
    else:
        ws = wb.create_sheet("Assessment Areas")
        headers = ['Bank Name', 'CBSA Name', 'CBSA Code', 'State Name', 'County Name', 'County Code (GEOID)']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border_thin
    
    # Add assessment areas
    for row_idx, aa in enumerate(assessment_areas, start=2):
        ws.cell(row=row_idx, column=1, value=aa['bank_name'])
        ws.cell(row=row_idx, column=2, value=aa['cbsa_name'])
        ws.cell(row=row_idx, column=3, value=aa['cbsa_code'])
        ws.cell(row=row_idx, column=4, value=aa['state_name'])
        ws.cell(row=row_idx, column=5, value=aa['county_name'])
        ws.cell(row=row_idx, column=6, value=aa['county_code'])
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 25
    
    wb.save(ticket_file)
    print(f"\n[SUCCESS] Updated Excel ticket with {len(assessment_areas)} Cadence assessment area entries")
    return True

def main():
    pdf_path = Path(__file__).parent.parent / "412025 List of Census Tracts - Public File.pdf"
    
    if not pdf_path.exists():
        print(f"ERROR: PDF not found: {pdf_path}")
        return
    
    # Extract assessment areas
    assessment_areas = extract_cadence_assessment_areas(pdf_path)
    
    # Map CBSA names
    assessment_areas = map_cbsa_names(assessment_areas)
    
    # Save to JSON
    output_json = pdf_path.parent / "cadence_assessment_areas_final.json"
    with open(output_json, 'w') as f:
        json.dump(assessment_areas, f, indent=2)
    print(f"\nSaved to JSON: {output_json}")
    
    # Update Excel ticket
    update_excel_ticket(assessment_areas)
    
    print(f"\n{'='*80}")
    print("CADENCE EXTRACTION COMPLETE")
    print(f"{'='*80}")
    print(f"Total counties: {len(assessment_areas)}")
    print(f"Unique CBSAs: {len(set(aa['cbsa_code'] for aa in assessment_areas if aa['cbsa_code']))}")
    print(f"States: {sorted(set(aa['state_name'] for aa in assessment_areas if aa['state_name']))}")

if __name__ == "__main__":
    main()

