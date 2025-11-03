"""
Fifth Third + Comerica Merger Analysis
- Branch footprint analysis (where each bank has branches)
- Overlap detection (markets where both operate)
- HHI analysis pre- and post-merger

Focus: 2025 branch data (sod25 table)
"""

import sys
import os
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime

# Handle module imports with spaces
import importlib.util

def load_module(module_name, file_path):
    spec = importlib.util.spec_from_file_location(module_name, file_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module

base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

bigquery_client = load_module(
    "bigquery_client",
    os.path.join(base_dir, "Lending and Branch Analysis", "utils", "bigquery_client.py")
)
branch_queries = load_module(
    "branch_queries",
    os.path.join(base_dir, "Lending and Branch Analysis", "queries", "branch_queries.py")
)
hhi_analysis = load_module(
    "hhi_analysis",
    os.path.join(base_dir, "Lending and Branch Analysis", "utils", "hhi_analysis.py")
)
branch_matching = load_module(
    "branch_matching",
    os.path.join(base_dir, "Lending and Branch Analysis", "utils", "branch_matching.py")
)

create_client = bigquery_client.create_client

# Bank information
FIFTH_THIRD = {
    'name': 'Fifth Third Bank',
    'lei': 'QFROUN1UWUYU0DVIWD51',
    'rssd': '723112'
}

COMERICA = {
    'name': 'Comerica Bank',
    'lei': '70WY0ID1N53Q4254VH70',
    'rssd': '60143'
}

def get_branches_by_rssd(rssd: str, year: int = 2025):
    """Get branch data using RSSD"""
    creds_path = os.path.join(base_dir, "config", "credentials", "hdma1-242116-74024e2eb88f.json")
    
    if os.path.exists(creds_path):
        client = create_client(key_path=creds_path)
    else:
        client = create_client()
    
    query = branch_queries.get_branches_by_rssd(
        rssd=rssd,
        year=year,
        sod_table='sod25'
    )
    
    df = client.execute_query(query)
    
    # Deduplicate
    if 'uninumbr' in df.columns and 'year' in df.columns:
        df = branch_matching.deduplicate_branches(df, year_col='year', unique_col='uninumbr')
    
    return df

def get_top_cbsas(lei: str, year: int = 2024, limit: int = 20):
    """Get top CBSAs by lending volume"""
    creds_path = os.path.join(base_dir, "config", "credentials", "hdma1-242116-74024e2eb88f.json")
    
    if os.path.exists(creds_path):
        client = create_client(key_path=creds_path)
    else:
        client = create_client()
    
    query = f"""
    SELECT 
        cbsa_code,
        COUNTIF(action_taken = '1') as originations,
        SUM(CASE WHEN action_taken = '1' THEN loan_amount END) as loan_volume
    FROM `hdma1-242116.hmda.hmda`
    WHERE activity_year = {year}
      AND lei = '{lei}'
      AND action_taken = '1'
      AND cbsa_code IS NOT NULL
      AND cbsa_code != ''
    GROUP BY cbsa_code
    ORDER BY originations DESC
    LIMIT {limit}
    """
    
    return client.execute_query(query)

def analyze_branch_footprints(year: int = 2025):
    """Analyze branch footprints for both banks"""
    print(f"\n{'='*80}")
    print("ANALYZING BRANCH FOOTPRINTS")
    print(f"{'='*80}\n")
    
    # Get branches for both banks
    print(f"Getting Fifth Third branches (RSSD: {FIFTH_THIRD['rssd']})...")
    ft_branches = get_branches_by_rssd(FIFTH_THIRD['rssd'], year=year)
    print(f"  Found {len(ft_branches):,} branches")
    
    print(f"\nGetting Comerica branches (RSSD: {COMERICA['rssd']})...")
    com_branches = get_branches_by_rssd(COMERICA['rssd'], year=year)
    print(f"  Found {len(com_branches):,} branches")
    
    # Summary by CBSA
    if 'cbsa_code' in ft_branches.columns and 'cbsa_code' in com_branches.columns:
        ft_by_cbsa = ft_branches.groupby('cbsa_code').agg({
            'uninumbr': 'count',
            'deposits': 'sum'
        }).reset_index()
        ft_by_cbsa.columns = ['cbsa_code', 'ft_branch_count', 'ft_total_deposits']
        
        com_by_cbsa = com_branches.groupby('cbsa_code').agg({
            'uninumbr': 'count',
            'deposits': 'sum'
        }).reset_index()
        com_by_cbsa.columns = ['cbsa_code', 'com_branch_count', 'com_total_deposits']
        
        # Find overlap
        overlap = ft_by_cbsa.merge(com_by_cbsa, on='cbsa_code', how='inner')
        
        print(f"\n{'='*80}")
        print("OVERLAP ANALYSIS")
        print(f"{'='*80}\n")
        print(f"Markets where both banks have branches: {len(overlap)}")
        print(f"Fifth Third only: {len(ft_by_cbsa) - len(overlap)}")
        print(f"Comerica only: {len(com_by_cbsa) - len(overlap)}")
        
        return {
            'fifth_third_branches': ft_branches,
            'comerica_branches': com_branches,
            'fifth_third_by_cbsa': ft_by_cbsa,
            'comerica_by_cbsa': com_by_cbsa,
            'overlap': overlap
        }
    
    return {
        'fifth_third_branches': ft_branches,
        'comerica_branches': com_branches
    }

def calculate_hhi_for_markets(cbsa_codes: list, year: int = 2024):
    """Calculate HHI pre- and post-merger for specified markets"""
    print(f"\n{'='*80}")
    print("CALCULATING HHI FOR OVERLAPPING MARKETS")
    print(f"{'='*80}\n")
    
    creds_path = os.path.join(base_dir, "config", "credentials", "hdma1-242116-74024e2eb88f.json")
    
    if os.path.exists(creds_path):
        client = create_client(key_path=creds_path)
    else:
        client = create_client()
    
    cbsa_str = "', '".join(cbsa_codes)
    
    results = []
    
    for cbsa in cbsa_codes:
        print(f"  Analyzing CBSA {cbsa}...")
        
        # Get all lending in this CBSA
        query = f"""
        SELECT 
            lei,
            COUNTIF(action_taken = '1') as originations,
            SUM(CASE WHEN action_taken = '1' THEN loan_amount END) as loan_amount
        FROM `hdma1-242116.hmda.hmda`
        WHERE activity_year = {year}
          AND cbsa_code = '{cbsa}'
          AND action_taken = '1'
        GROUP BY lei
        """
        
        df = client.execute_query(query)
        
        if len(df) == 0:
            continue
        
        # Calculate market shares pre-merger
        total_originations = df['originations'].sum()
        df['market_share_pct'] = (df['originations'] / total_originations) * 100
        
        # Pre-merger HHI
        hhi_pre = hhi_analysis.calculate_hhi(df['market_share_pct'], as_percentages=True)
        
        # Get shares for merging banks
        ft_share = df[df['lei'] == FIFTH_THIRD['lei']]['market_share_pct'].values[0] if len(df[df['lei'] == FIFTH_THIRD['lei']]) > 0 else 0
        com_share = df[df['lei'] == COMERICA['lei']]['market_share_pct'].values[0] if len(df[df['lei'] == COMERICA['lei']]) > 0 else 0
        
        # Post-merger: combine the two banks
        df_post = df[~df['lei'].isin([FIFTH_THIRD['lei'], COMERICA['lei']])].copy()
        combined_originations = df[df['lei'].isin([FIFTH_THIRD['lei'], COMERICA['lei']])]['originations'].sum()
        combined_row = pd.DataFrame([{
            'lei': f"{FIFTH_THIRD['lei']}+{COMERICA['lei']}",
            'originations': combined_originations,
            'loan_amount': df[df['lei'].isin([FIFTH_THIRD['lei'], COMERICA['lei']])]['loan_amount'].sum()
        }])
        df_post = pd.concat([df_post, combined_row], ignore_index=True)
        
        # Recalculate shares post-merger
        total_post = df_post['originations'].sum()
        df_post['market_share_pct'] = (df_post['originations'] / total_post) * 100
        
        # Post-merger HHI
        hhi_post = hhi_analysis.calculate_hhi(df_post['market_share_pct'], as_percentages=True)
        
        results.append({
            'cbsa_code': cbsa,
            'year': year,
            'hhi_pre_merger': hhi_pre,
            'hhi_post_merger': hhi_post,
            'hhi_change': hhi_post - hhi_pre,
            'concentration_pre': hhi_analysis.categorize_market_concentration(hhi_pre),
            'concentration_post': hhi_analysis.categorize_market_concentration(hhi_post),
            'fifth_third_share': ft_share,
            'comerica_share': com_share,
            'combined_share': ft_share + com_share,
            'total_lenders': len(df),
            'total_originations': total_originations,
            'antitrust_concern': (hhi_pre > 1800) and ((hhi_post - hhi_pre) > 100)
        })
    
    return pd.DataFrame(results)

def create_combined_report(footprint_data: dict, hhi_data: pd.DataFrame, output_file: str):
    """Create Excel report with branch analysis and HHI"""
    print(f"\n{'='*80}")
    print("CREATING COMBINED REPORT")
    print(f"{'='*80}\n")
    
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Sheet 1: Executive Summary
    ws_summary = wb.create_sheet('Executive Summary')
    ws_summary.append(['Fifth Third + Comerica Merger Analysis'])
    ws_summary.append([])
    ws_summary.append(['Fifth Third Bank', '', 'Comerica Bank'])
    ws_summary.append(['LEI', FIFTH_THIRD['lei'], COMERICA['lei']])
    ws_summary.append(['RSSD', FIFTH_THIRD['rssd'], COMERICA['rssd']])
    ws_summary.append(['Total Branches (2025)', 
                       len(footprint_data['fifth_third_branches']), 
                       len(footprint_data['comerica_branches'])])
    
    # Sheet 2: Branch Overlap
    ws_overlap = wb.create_sheet('Branch Overlap by CBSA')
    if 'overlap' in footprint_data:
        for r in dataframe_to_rows(footprint_data['overlap'], index=False, header=True):
            ws_overlap.append(r)
    
    # Sheet 3: Fifth Third Branches
    ws_ft = wb.create_sheet('Fifth Third Branches')
    if 'cbsa_code' in footprint_data['fifth_third_branches'].columns:
        ft_summary = footprint_data['fifth_third_branches'].groupby('cbsa_code').agg({
            'uninumbr': 'count',
            'deposits': 'sum',
            'br_lmi': lambda x: (x == 1).sum(),
            'cr_minority': lambda x: (x == 1).sum()
        }).reset_index()
        ft_summary.columns = ['cbsa_code', 'branch_count', 'total_deposits', 'branches_in_lmi_tracts', 'branches_in_minority_tracts']
        for r in dataframe_to_rows(ft_summary, index=False, header=True):
            ws_ft.append(r)
    
    # Sheet 4: Comerica Branches
    ws_com = wb.create_sheet('Comerica Branches')
    if 'cbsa_code' in footprint_data['comerica_branches'].columns:
        com_summary = footprint_data['comerica_branches'].groupby('cbsa_code').agg({
            'uninumbr': 'count',
            'deposits': 'sum',
            'br_lmi': lambda x: (x == 1).sum(),
            'cr_minority': lambda x: (x == 1).sum()
        }).reset_index()
        com_summary.columns = ['cbsa_code', 'branch_count', 'total_deposits', 'branches_in_lmi_tracts', 'branches_in_minority_tracts']
        for r in dataframe_to_rows(com_summary, index=False, header=True):
            ws_com.append(r)
    
    # Sheet 5: HHI Analysis
    ws_hhi = wb.create_sheet('HHI Analysis')
    ws_hhi.append(['MARKET CONCENTRATION ANALYSIS'])
    ws_hhi.append(['Pre- and Post-Merger HHI for Overlapping Markets'])
    ws_hhi.append([])
    for r in dataframe_to_rows(hhi_data.sort_values('hhi_change', ascending=False), index=False, header=True):
        ws_hhi.append(r)
    
    # Format HHI sheet
    if len(hhi_data) > 0:
        # Color code antitrust concerns
        for row_idx, row in enumerate(dataframe_to_rows(hhi_data, index=False, header=True), start=1):
            if row_idx == 1:
                # Header row
                for cell in ws_hhi[row_idx]:
                    cell.font = Font(bold=True)
            else:
                # Check if antitrust concern
                if row_idx <= len(hhi_data) + 1:  # Account for header
                    concern_col = list(hhi_data.columns).index('antitrust_concern') + 1
                    if ws_hhi.cell(row=row_idx, column=concern_col).value == True:
                        # Highlight row in red
                        for col in range(1, len(row) + 1):
                            cell = ws_hhi.cell(row=row_idx, column=col)
                            cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                            cell.font = Font(bold=True, color='FFFFFF')
    
    # Sheet 6: All Fifth Third Branches (detailed)
    ws_ft_detail = wb.create_sheet('Fifth Third All Branches')
    for r in dataframe_to_rows(footprint_data['fifth_third_branches'], index=False, header=True):
        ws_ft_detail.append(r)
    
    # Sheet 7: All Comerica Branches (detailed)
    ws_com_detail = wb.create_sheet('Comerica All Branches')
    for r in dataframe_to_rows(footprint_data['comerica_branches'], index=False, header=True):
        ws_com_detail.append(r)
    
    # Save
    wb.save(output_file)
    print(f"Report saved to: {output_file}")

def main():
    print("\n" + "="*80)
    print("FIFTH THIRD + COMERICA MERGER ANALYSIS")
    print("Branch Footprints + HHI Analysis")
    print("="*80)
    
    # Analyze branch footprints (2025 data)
    footprint_data = analyze_branch_footprints(year=2025)
    
    # Get overlapping CBSAs
    if 'overlap' in footprint_data:
        overlapping_cbsas = footprint_data['overlap']['cbsa_code'].tolist()
        print(f"\n{'='*80}")
        print(f"Found {len(overlapping_cbsas)} overlapping markets")
        print(f"{'='*80}")
        
        # Calculate HHI for overlapping markets
        hhi_results = calculate_hhi_for_markets(overlapping_cbsas, year=2024)
        
        # Create combined report
        output_dir = Path("reports/merger_analysis")
        output_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = output_dir / f"FifthThird_Comerica_Merger_Analysis_{timestamp}.xlsx"
        
        create_combined_report(footprint_data, hhi_results, str(output_file))
        
        print(f"\n{'='*80}")
        print("ANALYSIS COMPLETE!")
        print(f"{'='*80}")
        print(f"\nReport saved to: {output_file}")
        print(f"\nKey Findings:")
        
        if len(hhi_results) > 0:
            concerns = hhi_results[hhi_results['antitrust_concern'] == True]
            print(f"  Markets with antitrust concerns: {len(concerns)}")
            if len(concerns) > 0:
                print("\n  Markets requiring attention:")
                for _, row in concerns.iterrows():
                    print(f"    CBSA {row['cbsa_code']}: HHI {row['hhi_pre_merger']:.0f} -> {row['hhi_post_merger']:.0f} (+{row['hhi_change']:.0f})")
    else:
        print("\nWARNING: Could not analyze overlap (missing CBSA codes in branch data)")

if __name__ == "__main__":
    main()

