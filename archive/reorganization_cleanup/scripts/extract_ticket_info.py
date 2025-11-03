"""
Extract detailed information from the merger research ticket Excel file.
Reads bank details, years, and filters.
"""

import openpyxl
from pathlib import Path
import sys
import json

def sanitize_folder_name(name):
    """Sanitize bank name for use in folder name"""
    # Remove special characters, keep only alphanumeric and spaces
    import re
    name = re.sub(r'[^\w\s-]', '', name)
    # Replace spaces with underscores, limit length
    name = name.replace(' ', '_')
    # Remove multiple underscores
    name = re.sub(r'_+', '_', name)
    return name.strip('_')

def extract_ticket_info(excel_file):
    """
    Extract all relevant information from the merger ticket.
    Returns a dictionary with bank info, years, and filters.
    """
    excel_path = Path(excel_file)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    result = {
        'acquirer': {},
        'target': {},
        'years': [],
        'filters': {}
    }
    
    # Parse Notes sheet for bank info and years
    if "Notes" in wb.sheetnames:
        ws = wb["Notes"]
        print("\nParsing 'Notes' sheet...")
        
        # Read cell by cell to find bank information
        for row_idx in range(1, min(20, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(11, ws.max_column + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    row_data.append((col_idx, str(cell_value)))
            
            # Look for bank information
            row_text = ' '.join([v for _, v in row_data]).lower()
            
            # Check for years (e.g., "2020 through 2024")
            if 'through' in row_text and any(char.isdigit() for char in row_text):
                # Extract year range
                import re
                years_match = re.search(r'(\d{4})\s+through\s+(\d{4})', row_text)
                if years_match:
                    start_year = int(years_match.group(1))
                    end_year = int(years_match.group(2))
                    result['years'] = list(range(start_year, end_year + 1))
                    print(f"  Found years: {result['years']}")
            
            # Check for bank name and LEI
            if 'firstbank' in row_text or 'first bank' in row_text:
                # Look for LEI in nearby cells
                for col_idx, value in row_data:
                    if 'lei' in str(value).lower() or value.startswith('549300'):
                        # LEI might be in next column or same row
                        pass
                
                # Check if this is acquirer or target - look at context
                # In Notes sheet, both banks might be listed
                if not result['target'].get('name'):
                    result['target']['name'] = 'FirstBank'  # Common name
                    # Look for LEI in row
                    for col_idx, value in row_data:
                        if len(value) == 20 and value.startswith('549300'):
                            result['target']['lei'] = value
                            print(f"  Found Target Bank: FirstBank, LEI: {value}")
            
            # Look for PNC
            if 'pnc' in row_text:
                if not result['acquirer'].get('name'):
                    result['acquirer']['name'] = 'PNC Bank'
                    print(f"  Found Acquirer Bank: PNC Bank")
    
    # Parse Assessment Areas sheet to get structure
    if "Assessment Areas" in wb.sheetnames:
        ws = wb["Assessment Areas"]
        print("\nParsing 'Assessment Areas' sheet structure...")
        
        # Check header row
        header_row = []
        for col_idx in range(1, min(9, ws.max_column + 1)):
            cell_value = ws.cell(row=1, column=col_idx).value
            if cell_value:
                header_row.append((col_idx, str(cell_value)))
        
        print(f"  Header columns: {[v for _, v in header_row]}")
        
        # Count rows with data
        data_rows = 0
        for row_idx in range(2, min(100, ws.max_row + 1)):
            if ws.cell(row=row_idx, column=1).value or ws.cell(row=row_idx, column=5).value:
                data_rows += 1
        
        print(f"  Found {data_rows} data rows (sampled first 100)")
    
    # Try to find more bank info from other sheets
    # Check sheet names for bank names
    for sheet_name in wb.sheetnames:
        if 'PNC' in sheet_name and not result['acquirer'].get('name'):
            result['acquirer']['name'] = 'PNC'
        if 'FirstBank' in sheet_name and not result['target'].get('name'):
            result['target']['name'] = 'FirstBank'
    
    return result

def main():
    """Main function"""
    if len(sys.argv) < 2:
        print("Usage: python extract_ticket_info.py <excel_file>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    
    try:
        info = extract_ticket_info(excel_file)
        
        print("\n" + "="*80)
        print("EXTRACTED TICKET INFORMATION")
        print("="*80)
        
        print(f"\nAcquirer Bank: {info['acquirer']}")
        print(f"Target Bank: {info['target']}")
        print(f"Years: {info['years']}")
        print(f"Filters: {info['filters']}")
        
        # Save to JSON for reference
        output_file = Path(excel_file).parent / "ticket_info_extracted.json"
        with open(output_file, 'w') as f:
            json.dump(info, f, indent=2)
        print(f"\nResults saved to: {output_file}")
        
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()

