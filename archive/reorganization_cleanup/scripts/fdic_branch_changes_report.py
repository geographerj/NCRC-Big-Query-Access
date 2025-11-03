"""
FDIC Branch Openings and Closings Weekly Report Generator

Fetches branch structure changes (openings and closings) from FDIC OSCR
and creates a formatted Excel spreadsheet with all branch details.

Event Codes:
- 711: Branch Closings
- 721: Branch Openings
"""

import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import argparse
import sys
import os
import json

# Try importing web scraping libraries (optional dependencies)
try:
    import requests
    from bs4 import BeautifulSoup
    HAS_SCRAPING = True
except ImportError:
    HAS_SCRAPING = False
    print("Note: requests and beautifulsoup4 not installed. Install with: pip install requests beautifulsoup4")

def get_branch_changes_web(start_date, end_date, event_codes=['711', '721'], use_effective_date=False):
    """
    Fetch branch changes by parsing FDIC web interface
    Uses the search form structure
    
    Parameters:
    - use_effective_date: If True, uses EFFDATE (Effective Date), otherwise PROCDATE (Process Date)
    """
    if not HAS_SCRAPING:
        print("ERROR: Web scraping libraries not available")
        print("Install with: pip install requests beautifulsoup4")
        return pd.DataFrame()
    
    # Convert dates to FDIC format (MM/DD/YYYY)
    if isinstance(start_date, datetime):
        start_str = start_date.strftime('%m/%d/%Y')
    else:
        start_str = datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d/%Y')
    
    if isinstance(end_date, datetime):
        end_str = end_date.strftime('%m/%d/%Y')
    else:
        end_str = datetime.strptime(end_date, '%Y-%m-%d').strftime('%m/%d/%Y')
    
    # Build URL (based on the structure from the provided URL)
    base_url = "https://banks.data.fdic.gov/bankfind-suite/oscr"
    
    # Event code filter
    event_filter = ' OR '.join(event_codes)
    
    # Use process date (PROCDATE) - default
    date_type = 'EFFDATE' if use_effective_date else 'PROCDATE'
    
    params = {
        'startDate': start_str,
        'endDate': end_str,
        'eventCode': event_filter,
        'searchDateRadio': date_type,
        'resultLimit': 10000,
        'pageNumber': 1,
        'locationsExpand': 'false'
    }
    
    print(f"\nFetching branch changes from FDIC website...")
    print(f"  Date range: {start_str} to {end_str}")
    print(f"  Date type: {date_type} ({'Effective Date' if use_effective_date else 'Process Date'})")
    print(f"  Event codes: {', '.join(event_codes)}")
    print(f"  URL: {base_url}")
    
    try:
        response = requests.get(base_url, params=params, timeout=30)
        response.raise_for_status()
        
        print(f"  Response status: {response.status_code}")
        print(f"  Response length: {len(response.text)} characters")
        
        # Check if we got redirected or got an error page
        if 'error' in response.text.lower() or 'no results' in response.text.lower():
            print("  Website returned no results or error message")
            # Continue to try parsing anyway
        
        # Parse HTML response
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Look for data tables
        tables = soup.find_all('table')
        if tables:
            print(f"  Found {len(tables)} table(s) in response")
            try:
                # Try parsing HTML tables
                dfs = pd.read_html(response.text)
                if dfs:
                    print(f"  Successfully parsed {len(dfs)} table(s)")
                    df = dfs[0]  # Use first table
                    print(f"  Table shape: {df.shape}")
                    print(f"  Columns: {list(df.columns)}")
                    return df
            except Exception as e:
                print(f"  Error parsing HTML table: {e}")
        
        # Try to find JSON data embedded in script tags
        scripts = soup.find_all('script')
        print(f"  Found {len(scripts)} script tag(s)")
        for script in scripts:
            if script.string and ('data' in script.string.lower() or 'results' in script.string.lower()):
                try:
                    # Extract JSON from script
                    json_str = script.string
                    # Find JSON object
                    start = json_str.find('{')
                    end = json_str.rfind('}') + 1
                    if start >= 0 and end > start:
                        data = json.loads(json_str[start:end])
                        if 'data' in data or 'results' in data:
                            records = data.get('data') or data.get('results', [])
                            if records:
                                print(f"  Found {len(records)} records in JSON")
                                return pd.DataFrame(records)
                except Exception as e:
                    print(f"  Error parsing JSON from script: {e}")
        
        # Save response HTML for debugging
        debug_file = Path("fdic_response_debug.html")
        with open(debug_file, 'w', encoding='utf-8') as f:
            f.write(response.text)
        print(f"  Saved response HTML to {debug_file} for debugging")
        
        print("  Could not find data in expected format")
        return pd.DataFrame()
        
    except Exception as e:
        print(f"ERROR: Failed to fetch data: {e}")
        return pd.DataFrame()

def load_from_csv(csv_file):
    """Load branch changes data from CSV file (manual export)"""
    print(f"\nLoading data from CSV: {csv_file}")
    try:
        df = pd.read_csv(csv_file)
        print(f"  Loaded {len(df)} records")
        print(f"  Columns: {list(df.columns)}")
        return df
    except Exception as e:
        print(f"ERROR: Failed to load CSV: {e}")
        return pd.DataFrame()

def write_notes_sheet(ws, start_date, end_date):
    """Write Notes/Info sheet explaining the report"""
    from openpyxl.styles import Font
    
    notes_content = [
        ["FDIC BANK BRANCH CHANGES REPORT"],
        [],
        ["REPORT INFORMATION"],
        [f"Report Period: {start_date} to {end_date}"],
        [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"],
        [],
        ["WHAT THIS REPORT COVERS"],
        ["This report tracks bank branch openings and closings as reported to the FDIC."],
        ["Event Codes:"],
        ["  • 711: Branch Closings"],
        ["  • 721: Branch Openings"],
        [],
        ["HOW IT WORKS"],
        ["1. Data Source: Federal Deposit Insurance Corporation (FDIC) Office Structure Changes Report (OSCR)"],
        ["2. Search Method: Process Date (when the change was processed by FDIC)"],
        ["3. Date Range: Sunday to Saturday (weekly report period)"],
        ["4. Update Schedule: FDIC files update on Thursdays"],
        ["5. Report Schedule: Generated weekly on Fridays for the previous week"],
        [],
        ["WHEN IT RUNS"],
        ["• Automatic: Every Friday at 9:00 AM UTC via GitHub Actions"],
        ["• Manual: Can be triggered manually from GitHub Actions or run locally"],
        ["• Command: python scripts/fdic_branch_changes_report.py --start-date YYYY-MM-DD --end-date YYYY-MM-DD"],
        [],
        ["SHEET DESCRIPTIONS"],
        [],
        ["1. All Changes"],
        ["   Shows all branch openings and closings for the report period with:"],
        ["   • Bank Name"],
        ["   • Branch Address"],
        ["   • Service Type"],
        ["   • County and State"],
        ["   • Effective Date"],
        [],
        ["2. Trend Analysis"],
        ["   Analyzes branch closure trends over the past year"],
        ["   Compares current week's closures to historical baseline"],
        ["   Identifies trends, patterns, and significant changes"],
        ["   Requires baseline report for full analysis"],
        [],
        ["DATA FIELDS"],
        [],
        ["Key fields included in this report:"],
        ["• Institution Name / Bank Name"],
        ["• Branch Address (street, city, zip)"],
        ["• County Name and State"],
        ["• Service Type (e.g., Full Service, Limited Service)"],
        ["• Effective Date (when the change takes effect)"],
        ["• Process Date (when FDIC processed the change)"],
        ["• Event Code (711 = Closing, 721 = Opening)"],
        [],
        ["NOTES"],
        ["• Data is sourced from FDIC regulatory filings"],
        ["• Process dates may differ from effective dates"],
        ["• Some branches may have multiple events recorded"],
        ["• Trend analysis compares to baseline report covering past year"],
        [],
        ["CONTACT"],
        ["For questions about this report or methodology, contact:"],
        ["NCRC Research Department"],
        ["National Community Reinvestment Coalition"]
    ]
    
    for i, row in enumerate(notes_content, 1):
        ws.append(row)
        # Format title row
        if i == 1:
            ws.cell(i, 1).font = Font(bold=True, size=14)
        # Format section headers
        elif len(row) > 0 and row[0] and isinstance(row[0], str) and row[0].isupper() and not row[0].startswith('•'):
            ws.cell(i, 1).font = Font(bold=True, size=11)
    
    # Set column width
    ws.column_dimensions['A'].width = 80

def write_detailed_changes_sheet(ws, df, start_date, end_date):
    """Write detailed changes sheet with key fields"""
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Header
    ws.append(["FDIC Branch Changes - All Openings and Closings"])
    ws.append([f"Report Period: {start_date} to {end_date}"])
    ws.append([])
    
    # Identify key columns - check for FDIC-specific column names first, then generic
    col_mapping = {}
    for col in df.columns:
        col_lower = col.lower().strip()
        col_upper = col.upper().strip()
        
        # Bank/Institution name - FDIC uses INSTNAME
        if col_upper == 'INSTNAME':
            col_mapping['bank_name'] = col
        elif ('bank' in col_lower or 'institution' in col_lower) and 'name' in col_lower:
            if 'bank_name' not in col_mapping:
                col_mapping['bank_name'] = col
        
        # Address components - FDIC uses OFF_PADDR, OFF_PADDR2, OFF_PCITY, OFF_PSTATE, OFF_PZIP5
        if col_upper == 'OFF_PADDR':
            col_mapping['off_paddr'] = col
        elif col_upper == 'OFF_PADDR2':
            col_mapping['off_paddr2'] = col
        elif col_upper == 'OFF_PCITY':
            col_mapping['off_pcity'] = col
        elif col_upper == 'OFF_PSTATE':
            col_mapping['off_pstate'] = col
        elif col_upper == 'OFF_PZIP5':
            col_mapping['off_pzip5'] = col
        
        # Service type - FDIC uses OFF_SERVTYPE_DESC
        if col_upper == 'OFF_SERVTYPE_DESC':
            col_mapping['service_type'] = col
        elif 'service' in col_lower and 'type' in col_lower and 'desc' in col_lower:
            if 'service_type' not in col_mapping:
                col_mapping['service_type'] = col
        
        # County - FDIC uses OFF_CNTYNAME
        if col_upper == 'OFF_CNTYNAME':
            col_mapping['county'] = col
        elif 'county' in col_lower and 'name' in col_lower:
            if 'county' not in col_mapping:
                col_mapping['county'] = col
        
        # State - FDIC uses OFF_PSTATE or STATE
        if col_upper == 'OFF_PSTATE':
            col_mapping['state'] = col
        elif col_upper == 'STATE':
            if 'state' not in col_mapping:
                col_mapping['state'] = col
        elif 'state' in col_lower and 'code' not in col_lower and 'postal' not in col_lower:
            if 'state' not in col_mapping:
                col_mapping['state'] = col
        
        # Effective date - FDIC uses EFFDATE
        if col_upper == 'EFFDATE':
            col_mapping['effective_date'] = col
        elif 'effective' in col_lower and 'date' in col_lower:
            if 'effective_date' not in col_mapping:
                col_mapping['effective_date'] = col
        
        # Event code - check for Event_Code (from our processing) or CHANGECODE
        if col_upper == 'EVENT_CODE' or col == 'Event_Code':
            col_mapping['event_code'] = col
        elif col_upper == 'CHANGECODE':
            if 'event_code' not in col_mapping:
                col_mapping['event_code'] = col
        elif ('event' in col_lower and 'code' in col_lower) or col_lower == 'eventcd':
            if 'event_code' not in col_mapping:
                col_mapping['event_code'] = col
    
    # Create header row with preferred columns
    # Check if we have process_date mapped
    has_process_date = 'process_date' in col_mapping and col_mapping['process_date']
    date_header = 'Date (Process Date)' if has_process_date else 'Effective Date'
    headers = ['Event Type', 'Bank Name', 'Branch Address', 'Service Type', 'County', 'State', date_header]
    ws.append(headers)
    header_row = ws.max_row
    
    # Write data
    for _, row in df.iterrows():
        # Determine event type
        event_code_col = col_mapping.get('event_code', '')
        event_val = str(row.get(event_code_col, '')) if event_code_col else ''
        
        if '721' in event_val:
            event_type = 'Opening'
        elif '711' in event_val:
            event_type = 'Closing'
        else:
            event_type = 'Unknown'
        
        # Extract values with safe fallbacks
        def safe_get(key, default=''):
            col = col_mapping.get(key, '')
            if not col:
                return default
            val = row.get(col, default)
            if pd.isna(val):
                return default
            val_str = str(val)
            return val_str if val_str and val_str != 'nan' and val_str != 'None' else default
        
        bank_name = safe_get('bank_name')
        
        # Build full address from components
        addr_parts = []
        addr1 = safe_get('off_paddr')
        addr2 = safe_get('off_paddr2')
        city = safe_get('off_pcity')
        state = safe_get('state') or safe_get('off_pstate')
        zip5 = safe_get('off_pzip5')
        
        if addr1:
            addr_parts.append(addr1)
        if addr2:
            addr_parts.append(addr2)
        if city:
            addr_parts.append(city)
        if state:
            addr_parts.append(state)
        if zip5:
            addr_parts.append(zip5)
        
        address = ', '.join(addr_parts) if addr_parts else ''
        service_type = safe_get('service_type')
        county = safe_get('county')
        effective_date = safe_get('effective_date')
        
        # Format effective date if it's a timestamp
        if effective_date and 'T' in str(effective_date):
            try:
                from datetime import datetime
                dt = datetime.fromisoformat(str(effective_date).replace('Z', '+00:00'))
                effective_date = dt.strftime('%Y-%m-%d')
            except:
                pass
        
        # Also get PROCDATE for display if available (more relevant for weekly reports)
        process_date = safe_get('process_date')
        if process_date and 'T' in str(process_date):
            try:
                from datetime import datetime
                dt = datetime.fromisoformat(str(process_date).replace('Z', '+00:00'))
                process_date = dt.strftime('%Y-%m-%d')
            except:
                pass
        
        # Use PROCDATE if available, otherwise EFFDATE
        display_date = process_date if process_date else effective_date
        
        ws.append([
            event_type,
            bank_name,
            address,
            service_type,
            county,
            state,
            display_date
        ])
    
    # Format header
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header
    ws.freeze_panes = f'A{header_row + 1}'

def write_trend_analysis_sheet(ws, current_data, start_date, end_date, baseline_data=None):
    """Write trend analysis comparing current week to past year baseline"""
    from openpyxl.styles import Font, PatternFill, Alignment
    
    ws.append(["FDIC Branch Closure Trend Analysis"])
    ws.append([f"Current Report Period: {start_date} to {end_date}"])
    ws.append([])
    
    # Count current week closures
    event_col = None
    for col in current_data.columns:
        col_lower = col.lower()
        if ('event' in col_lower and 'code' in col_lower) or col_lower == 'eventcd':
            event_col = col
            break
    
    current_closings = 0
    current_openings = 0
    if event_col:
        current_closings = len(current_data[current_data[event_col].astype(str).str.contains('711', na=False, case=False)])
        current_openings = len(current_data[current_data[event_col].astype(str).str.contains('721', na=False, case=False)])
    
    # Weekly summary
    ws.append(["CURRENT WEEK SUMMARY"])
    ws.append(["Metric", "Count"])
    ws.append(["Branch Closings (Current Week)", current_closings])
    ws.append(["Branch Openings (Current Week)", current_openings])
    ws.append(["Net Change (Openings - Closings)", current_openings - current_closings])
    ws.append([])
    
    # Load baseline data if available
    baseline_file = None
    if baseline_data is None:
        # Try to find baseline file
        reports_dir = Path(r"C:\Users\edite\OneDrive - Nat'l Community Reinvestment Coaltn\Desktop\DREAM Analysis\Weekly BankFind Branch Changes Reports")
        baseline_files = list(reports_dir.glob("BASELINE_*.xlsx"))
        if baseline_files:
            baseline_file = baseline_files[0]
            print(f"  Found baseline file: {baseline_file.name}")
            # Try to load baseline data (would need to extract from Excel)
    
    if baseline_data is not None and len(baseline_data) > 0:
        # Analyze baseline data
        baseline_event_col = None
        for col in baseline_data.columns:
            col_lower = col.lower()
            if ('event' in col_lower and 'code' in col_lower) or col_lower == 'eventcd':
                baseline_event_col = col
                break
        
        if baseline_event_col:
            baseline_closings = len(baseline_data[baseline_data[baseline_event_col].astype(str).str.contains('711', na=False, case=False)])
            baseline_openings = len(baseline_data[baseline_data[baseline_event_col].astype(str).str.contains('721', na=False, case=False)])
            
            # Calculate actual number of weeks in baseline data
            baseline_weeks = 52  # Default assumption
            if 'PROCDATE' in baseline_data.columns:
                try:
                    baseline_data_copy = baseline_data.copy()
                    baseline_data_copy['PROCDATE'] = pd.to_datetime(baseline_data_copy['PROCDATE'], errors='coerce')
                    valid_dates = baseline_data_copy['PROCDATE'].dropna()
                    if len(valid_dates) > 0:
                        min_date = valid_dates.min()
                        max_date = valid_dates.max()
                        # Calculate weeks: (max_date - min_date).days / 7
                        days_span = (max_date - min_date).days
                        baseline_weeks = max(1, days_span / 7.0)  # At least 1 week
                        ws.append([f"Baseline Date Range: {min_date.date()} to {max_date.date()}"])
                        ws.append([f"Baseline Period: {baseline_weeks:.1f} weeks"])
                        ws.append([])
                except Exception as e:
                    print(f"  Warning: Could not calculate baseline weeks: {e}")
            
            ws.append(["COMPARISON TO BASELINE AVERAGES"])
            ws.append(["Metric", "Current Week", "Baseline Avg/Week", "Change", "% Change"])
            
            # Calculate averages based on actual weeks covered
            avg_weekly_closings = baseline_closings / baseline_weeks
            avg_weekly_openings = baseline_openings / baseline_weeks
            
            closing_change = current_closings - avg_weekly_closings
            closing_pct = (closing_change / avg_weekly_closings * 100) if avg_weekly_closings > 0 else 0
            
            opening_change = current_openings - avg_weekly_openings
            opening_pct = (opening_change / avg_weekly_openings * 100) if avg_weekly_openings > 0 else 0
            
            ws.append([
                "Average Weekly Closings",
                current_closings,
                f"{avg_weekly_closings:.2f}",
                f"{closing_change:+.2f}",
                f"{closing_pct:+.1f}%"
            ])
            
            ws.append([
                "Average Weekly Openings",
                current_openings,
                f"{avg_weekly_openings:.2f}",
                f"{opening_change:+.2f}",
                f"{opening_pct:+.1f}%"
            ])
            
            # Net change comparison
            avg_net_change = avg_weekly_openings - avg_weekly_closings
            current_net_change = current_openings - current_closings
            net_change_diff = current_net_change - avg_net_change
            ws.append([
                "Net Change (Openings - Closings)",
                current_net_change,
                f"{avg_net_change:.2f}",
                f"{net_change_diff:+.2f}",
                ""
            ])
            ws.append([])
            
            # Total baseline stats
            ws.append(["BASELINE PERIOD TOTALS"])
            ws.append(["Total Closings", baseline_closings])
            ws.append(["Total Openings", baseline_openings])
            ws.append(["Net Change (Total)", baseline_openings - baseline_closings])
            ws.append([])
            
            # Trend indicators
            ws.append(["TREND INDICATORS"])
            if current_closings > avg_weekly_closings * 1.2:
                trend_closing = "↑ Above Average (20%+ higher)"
            elif current_closings > avg_weekly_closings:
                trend_closing = "↑ Slightly Above Average"
            elif current_closings < avg_weekly_closings * 0.8:
                trend_closing = "↓ Below Average (20%+ lower)"
            else:
                trend_closing = "→ Near Average"
            
            if current_openings > avg_weekly_openings * 1.2:
                trend_opening = "↑ Above Average (20%+ higher)"
            elif current_openings > avg_weekly_openings:
                trend_opening = "↑ Slightly Above Average"
            elif current_openings < avg_weekly_openings * 0.8:
                trend_opening = "↓ Below Average (20%+ lower)"
            else:
                trend_opening = "→ Near Average"
            
            ws.append(["Closings Trend", trend_closing])
            ws.append(["Openings Trend", trend_opening])
            ws.append([])
    
    # Trend indicators
    ws.append(["TREND INTERPRETATION"])
    
    # Calculate year-ago date for reference
    try:
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
        year_ago_start = (end_date_obj - timedelta(days=365)).strftime('%Y-%m-%d')
        year_ago_end = (end_date_obj - timedelta(days=358)).strftime('%Y-%m-%d')
    except:
        year_ago_start = "One Year Ago"
        year_ago_end = ""
    
    ws.append([f"Comparison Period: Same week one year ago ({year_ago_start} to {year_ago_end})"])
    ws.append([])
    
    # Interpretation - compare to baseline average if available
    if baseline_data is not None and len(baseline_data) > 0 and baseline_event_col:
        # Calculate baseline average
        baseline_closings_total = len(baseline_data[baseline_data[baseline_event_col].astype(str).str.contains('711', na=False, case=False)])
        
        baseline_weeks = 52  # Default
        if 'PROCDATE' in baseline_data.columns:
            try:
                baseline_data_copy = baseline_data.copy()
                baseline_data_copy['PROCDATE'] = pd.to_datetime(baseline_data_copy['PROCDATE'], errors='coerce')
                valid_dates = baseline_data_copy['PROCDATE'].dropna()
                if len(valid_dates) > 0:
                    min_date = valid_dates.min()
                    max_date = valid_dates.max()
                    days_span = (max_date - min_date).days
                    baseline_weeks = max(1, days_span / 7.0)
            except:
                pass
        
        avg_weekly_closings = baseline_closings_total / baseline_weeks if baseline_weeks > 0 else 0
        
        # Compare to baseline average
        if avg_weekly_closings > 0:
            if current_closings > avg_weekly_closings * 1.5:
                trend = "VERY HIGH - Significantly above average closure activity"
            elif current_closings > avg_weekly_closings * 1.2:
                trend = "HIGH - Above average closure activity"
            elif current_closings > avg_weekly_closings * 0.8:
                trend = "MODERATE - Near average closure activity"
            elif current_closings > avg_weekly_closings * 0.5:
                trend = "LOW - Below average closure activity"
            else:
                trend = "VERY LOW - Significantly below average closure activity"
            
            ws.append([f"Current Week Closure Trend: {trend}"])
            ws.append([f"(Baseline Average: {avg_weekly_closings:.1f} closures/week)"])
        else:
            ws.append([f"Current Week Closure Trend: {current_closings} closures (no baseline for comparison)"])
    else:
        # Fallback to simple thresholds if no baseline
        if current_closings > 15:
            trend = "HIGH - Multiple closures this week"
        elif current_closings > 10:
            trend = "MODERATE - Some closure activity"
        elif current_closings > 5:
            trend = "LOW - Minimal closure activity"
        else:
            trend = "VERY LOW - Few or no closures"
        ws.append([f"Current Week Closure Trend: {trend}"])
    ws.append([])
    
    if baseline_data is None or len(baseline_data) == 0:
        ws.append(["NOTE: Baseline data not found."])
        ws.append(["To enable trend analysis, create a baseline report:"])
        ws.append(["  python scripts/create_baseline_report.py --csv-file <past_year_export>.csv"])
        ws.append([])
        ws.append(["This will establish a historical baseline for comparison."])
    else:
        ws.append(["This analysis compares current week to past year baseline patterns."])
        ws.append(["Trends help identify periods of increased or decreased branch activity."])
    
    # Format section headers
    header_rows = [1, 5, 12, 20]
    for row_num in header_rows:
        if row_num <= ws.max_row:
            try:
                ws.cell(row_num, 1).font = Font(bold=True, size=12)
            except:
                pass
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

def create_excel_report(df, output_file, start_date, end_date, baseline_data=None):
    """
    Create formatted Excel report from branch changes data
    
    Parameters:
    - df: Current week's data
    - baseline_data: Optional DataFrame with historical data for trend analysis
    """
    if len(df) == 0:
        print("\nNo data to create report")
        return
    
    print(f"\nCreating Excel report: {output_file}")
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create Notes/Info sheet first
    ws_notes = wb.create_sheet("Notes", 0)
    write_notes_sheet(ws_notes, start_date, end_date)
    print("  Created Notes sheet")
    
    # Create All Changes sheet (openings and closings combined with key fields)
    ws_all = wb.create_sheet("All Changes")
    write_detailed_changes_sheet(ws_all, df, start_date, end_date)
    print(f"  Created All Changes sheet: {len(df)} records")
    
    # Create Trend Analysis sheet
    ws_trend = wb.create_sheet("Trend Analysis")
    write_trend_analysis_sheet(ws_trend, df, start_date, end_date, baseline_data)
    print("  Created Trend Analysis sheet")
    
    wb.save(output_file)
    print(f"\nExcel report saved: {output_file}")

def load_baseline_data():
    """Load baseline data from baseline CSV file if available"""
    # Check if running in GitHub Actions
    if os.getenv('GITHUB_ACTIONS'):
        base_dir = Path(__file__).parent.parent
        reports_dir = base_dir / "reports" / "branch_changes"
    else:
        reports_dir = Path(r"C:\Users\edite\OneDrive - Nat'l Community Reinvestment Coaltn\Desktop\DREAM Analysis\Weekly BankFind Branch Changes Reports")
    
    # Look for baseline CSV file (if baseline was exported as CSV)
    baseline_csv = reports_dir / "BASELINE_branch_changes.csv"
    if baseline_csv.exists():
        print(f"\nLoading baseline data from CSV: {baseline_csv}")
        try:
            df = load_from_csv(str(baseline_csv))
            print(f"  Loaded {len(df)} baseline records for trend analysis")
            return df
        except Exception as e:
            print(f"  Error loading baseline: {e}")
    
    return None

def main():
    parser = argparse.ArgumentParser(
        description='Generate FDIC Branch Changes Weekly Report',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Weekly report (last 7 days) - runs on Friday for previous week
  python run_weekly_branch_report.py
  
  # Test with specific week (Oct 19-25, 2025)
  python run_weekly_branch_report.py --test
  
  # Custom date range
  python fdic_branch_changes_report.py --start-date 2024-10-01 --end-date 2024-10-31
  
  # From CSV export (manual download from FDIC website)
  python fdic_branch_changes_report.py --csv-file branch_changes.csv
  
IMPORTANT: The FDIC website uses JavaScript to load data. For automated reports:
1. Use Selenium (requires ChromeDriver) - more complex but automated
2. Use manual CSV export (most reliable):
   - Visit: https://banks.data.fdic.gov/bankfind-suite/oscr
   - Set: Start Date, End Date, Event Code (711 OR 721), Search Date = Process Date
   - Export results as CSV
   - Run: python fdic_branch_changes_report.py --csv-file <exported_file>.csv

Reports are saved to:
C:\\Users\\edite\\OneDrive - Nat'l Community Reinvestment Coaltn\\Desktop\\DREAM Analysis\\Weekly BankFind Branch Changes Reports
        """
    )
    
    parser.add_argument('--start-date', help='Start date (YYYY-MM-DD). Defaults to 7 days ago')
    parser.add_argument('--end-date', help='End date (YYYY-MM-DD). Defaults to today')
    parser.add_argument('--output', help='Output Excel file path')
    parser.add_argument('--event-codes', nargs='+', default=['711', '721'],
                       help='Event codes (711=closing, 721=opening). Default: 711 721')
    parser.add_argument('--csv-file', help='Load data from CSV file instead of web scraping')
    parser.add_argument('--baseline-csv', help='CSV file with baseline data for trend analysis')
    
    args = parser.parse_args()
    
    # Set output file
    # Check if running in GitHub Actions or use local folder
    if os.getenv('GITHUB_ACTIONS'):
        # In GitHub Actions, use relative path
        base_dir = Path(__file__).parent.parent
        reports_dir = base_dir / "reports" / "branch_changes"
    else:
        # Local: Use the specific folder in DREAM Analysis as requested
        reports_dir = Path(r"C:\Users\edite\OneDrive - Nat'l Community Reinvestment Coaltn\Desktop\DREAM Analysis\Weekly BankFind Branch Changes Reports")
    
    reports_dir.mkdir(parents=True, exist_ok=True)
    
    print("="*80)
    print("FDIC BRANCH CHANGES WEEKLY REPORT")
    print("="*80)
    
    # Load data
    if args.csv_file:
        # Load from CSV
        df = load_from_csv(args.csv_file)
        
        # Calculate date range from data if not specified
        if args.start_date:
            start_date = datetime.strptime(args.start_date, '%Y-%m-%d')
        else:
            start_date = datetime.now() - timedelta(days=7)
        
        if args.end_date:
            end_date = datetime.strptime(args.end_date, '%Y-%m-%d')
        else:
            end_date = datetime.now()
    else:
        # Fetch from web
        if args.end_date:
            end_date = datetime.strptime(args.end_date, '%Y-%m-%d')
        else:
            end_date = datetime.now()
        
        if args.start_date:
            start_date = datetime.strptime(args.start_date, '%Y-%m-%d')
        else:
            start_date = end_date - timedelta(days=7)  # Last week
        
        df = get_branch_changes_web(start_date, end_date, args.event_codes, use_effective_date=False)
    
    # Load baseline data for trend analysis
    baseline_data = None
    if args.baseline_csv:
        baseline_data = load_from_csv(args.baseline_csv)
    else:
        baseline_data = load_baseline_data()
    
    # Set output filename
    if args.output:
        output_file = Path(args.output)
    else:
        start_str = start_date.strftime('%Y%m%d')
        end_str = end_date.strftime('%Y%m%d')
        filename = f"Branch_Changes_{start_str}_to_{end_str}.xlsx"
        output_file = reports_dir / filename
    
    if len(df) > 0:
        # Create Excel report
        create_excel_report(df, str(output_file), 
                          start_date.strftime('%Y-%m-%d'),
                          end_date.strftime('%Y-%m-%d'),
                          baseline_data)
        
        print("\n" + "="*80)
        print("REPORT COMPLETE!")
        print("="*80)
        print(f"\nReport saved to: {output_file}")
    else:
        print("\n" + "="*80)
        print("NO DATA RETRIEVED")
        print("="*80)
        print("\nOptions:")
        print("1. Try manual export from FDIC website:")
        print("   - Visit: https://banks.data.fdic.gov/bankfind-suite/oscr")
        print("   - Set date range and event codes (711 OR 721)")
        print("   - Search Date: Process Date")
        print("   - Export results as CSV")
        print("   - Run: python fdic_branch_changes_report.py --csv-file <exported_file>.csv")
        print("\n2. Check if FDIC API structure has changed")
        print("3. Verify date range contains branch changes")

if __name__ == "__main__":
    main()
