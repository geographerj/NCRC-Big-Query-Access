"""Inspect Mortgage Goals sheet formulas in detail"""

import openpyxl
from pathlib import Path

template_file = r"C:\Users\edite\OneDrive - Nat'l Community Reinvestment Coaltn\Desktop\DREAM Analysis\data\raw\PNC+FirstBank merger research ticket.xlsx"

wb = openpyxl.load_workbook(template_file, data_only=False)

if 'Mortgage Goals' in wb.sheetnames:
    ws = wb['Mortgage Goals']
    
    print("="*80)
    print("MORTGAGE GOALS SHEET - DETAILED INSPECTION")
    print("="*80)
    
    print(f"\nDimensions: {ws.max_row} rows x {ws.max_column} columns")
    
    # Show headers
    print("\n" + "-"*80)
    print("HEADERS (Row 1):")
    print("-"*80)
    headers = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        if cell.value:
            headers.append((col, cell.value))
            print(f"  Col {col} ({openpyxl.utils.get_column_letter(col)}): {cell.value}")
    
    # Show row 2 structure (sample data row)
    print("\n" + "-"*80)
    print("ROW 2 (Sample Data Row):")
    print("-"*80)
    for col in range(1, min(13, ws.max_column + 1)):
        cell = ws.cell(2, col)
        col_letter = openpyxl.utils.get_column_letter(col)
        if cell.data_type == 'f':
            print(f"  {col_letter}2: {cell.value} (FORMULA)")
        elif cell.value is not None:
            print(f"  {col_letter}2: {cell.value}")
    
    # Find formula patterns
    print("\n" + "-"*80)
    print("FORMULA PATTERNS (First 50 formulas):")
    print("-"*80)
    
    formula_count = 0
    for row_idx in range(2, min(100, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.data_type == 'f':
                formula_count += 1
                if formula_count <= 50:
                    print(f"\n  {cell.coordinate}: {cell.value}")
                elif formula_count == 51:
                    print(f"\n  ... (showing first 50 of {formula_count} formulas)")
    
    print(f"\nTotal formulas found: {formula_count}")
    
    # Identify which columns are formulas vs data
    print("\n" + "-"*80)
    print("COLUMN TYPES (Columns 1-12):")
    print("-"*80)
    for col_idx in range(1, min(13, ws.max_column + 1)):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        formula_count_col = 0
        data_count_col = 0
        
        for row_idx in range(2, min(100, ws.max_row + 1)):
            cell = ws.cell(row_idx, col_idx)
            if cell.data_type == 'f':
                formula_count_col += 1
            elif cell.value is not None:
                data_count_col += 1
        
        col_type = "FORMULAS" if formula_count_col > data_count_col else "DATA"
        print(f"  {col_letter} ({col_idx}): {formula_count_col} formulas, {data_count_col} data cells -> {col_type}")

else:
    print("ERROR: Mortgage Goals sheet not found")

