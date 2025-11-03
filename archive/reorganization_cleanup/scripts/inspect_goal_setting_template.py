"""
Inspect the goal-setting analysis template to understand structure, formulas, and data requirements.
This template must be preserved exactly with formulas intact, and populated with SQL data.
"""

import openpyxl
from openpyxl.formula.translate import Translator
import pandas as pd
from pathlib import Path
import sys

def inspect_template_excel(excel_file):
    """
    Inspect the template Excel file to understand:
    - Sheet structure
    - Formulas and their dependencies
    - Data locations
    - Expected data format
    """
    print("="*80)
    print("INSPECTING GOAL-SETTING ANALYSIS TEMPLATE")
    print("="*80)
    
    excel_path = Path(excel_file)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    # Load workbook with formulas preserved
    # NOTE: File must be closed in Excel before we can read it
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=False)  # False = keep formulas
    except PermissionError as e:
        print(f"\nERROR: Cannot open file - it is likely open in Excel or another program.")
        print(f"       Please close the Excel file and try again.")
        print(f"\n       File path: {excel_path}")
        raise
    
    print(f"\nFile: {excel_path.name}")
    print(f"Total sheets: {len(wb.sheetnames)}")
    print(f"\nSheet names:")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {name}")
    
    # Inspect each sheet
    sheet_details = {}
    
    for sheet_name in wb.sheetnames:
        print(f"\n{'='*80}")
        print(f"SHEET: {sheet_name}")
        print(f"{'='*80}")
        
        ws = wb[sheet_name]
        print(f"\nDimensions: {ws.max_row} rows x {ws.max_column} columns")
        
        # Find formulas
        formulas = []
        data_cells = []
        headers = []
        
        # Scan first 100 rows and 50 columns for structure
        for row_idx in range(1, min(101, ws.max_row + 1)):
            row_formulas = []
            row_data = []
            row_has_formula = False
            
            for col_idx in range(1, min(51, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Check for formulas
                if cell.data_type == 'f':  # Formula
                    row_formulas.append({
                        'cell': cell.coordinate,
                        'formula': cell.value,
                        'row': row_idx,
                        'col': col_idx
                    })
                    row_has_formula = True
                
                # Check for data (non-empty, non-formula)
                elif cell.value is not None:
                    if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.strip()):
                        row_data.append({
                            'cell': cell.coordinate,
                            'value': str(cell.value)[:50],
                            'row': row_idx,
                            'col': col_idx
                        })
                
                # Check if this looks like a header row
                if row_idx <= 5:
                    if cell.value and isinstance(cell.value, str) and len(cell.value) > 0:
                        headers.append({
                            'cell': cell.coordinate,
                            'value': str(cell.value),
                            'row': row_idx,
                            'col': col_idx
                        })
            
            if row_formulas:
                formulas.extend(row_formulas)
            
            if row_data and row_idx <= 20:  # First 20 rows of data
                data_cells.extend(row_data[:10])  # First 10 cells per row
        
        # Print findings
        if headers:
            print(f"\nHeader cells found (first 15):")
            for h in headers[:15]:
                print(f"  {h['cell']} (R{h['row']}C{h['col']}): {h['value']}")
        
        if formulas:
            print(f"\nFormulas found: {len(formulas)} (showing first 20)")
            for f in formulas[:20]:
                print(f"  {f['cell']}: {f['formula']}")
        
        if data_cells:
            print(f"\nSample data cells (first 15):")
            for d in data_cells[:15]:
                print(f"  {d['cell']}: {d['value']}")
        
        # Try to identify data structure
        print(f"\nAnalyzing data structure...")
        
        # Look for patterns that indicate data insertion points
        # Common patterns: blank cells in structured areas, placeholder text, etc.
        potential_data_areas = []
        
        # Check for numeric columns (likely data columns)
        numeric_cols = {}
        for col_idx in range(1, min(21, ws.max_column + 1)):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            numeric_count = 0
            for row_idx in range(1, min(101, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    numeric_count += 1
            if numeric_count > 5:
                numeric_cols[col_letter] = numeric_count
        
        if numeric_cols:
            print(f"  Columns with numeric data: {list(numeric_cols.keys())[:10]}")
        
        sheet_details[sheet_name] = {
            'formulas': formulas,
            'headers': headers,
            'dimensions': (ws.max_row, ws.max_column),
            'numeric_columns': numeric_cols
        }
        
        # Try reading as DataFrame to see structure (first 30 rows)
        try:
            # Read first portion to understand structure
            df_sample = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, nrows=30, engine='openpyxl')
            if not df_sample.empty:
                print(f"\n  DataFrame sample (first 10 rows, first 10 cols):")
                print(df_sample.iloc[:10, :10].to_string())
        except Exception as e:
            print(f"  Could not read as DataFrame: {e}")
    
    return sheet_details

def identify_formula_dependencies(formulas, sheet_name):
    """Identify what cells/ranges formulas reference"""
    dependencies = []
    
    for formula_info in formulas:
        formula = formula_info['formula']
        # Look for cell references (A1, B2, etc.) or ranges (A1:B10, Sheet1!A1, etc.)
        import re
        # Pattern for cell references
        refs = re.findall(r'([A-Z]+[0-9]+(?::[A-Z]+[0-9]+)?)', formula)
        # Pattern for sheet references
        sheet_refs = re.findall(r"'?([^'!]+)'?!([A-Z]+[0-9]+)", formula)
        
        if refs or sheet_refs:
            dependencies.append({
                'cell': formula_info['cell'],
                'formula': formula,
                'references': refs,
                'sheet_references': sheet_refs
            })
    
    return dependencies

def main():
    """Main function"""
    if len(sys.argv) < 2:
        print("Usage: python inspect_goal_setting_template.py <excel_file>")
        print("\nExample:")
        print('  python inspect_goal_setting_template.py "PNC+FirstBank merger research ticket.xlsx"')
        sys.exit(1)
    
    excel_file = sys.argv[1]
    
    try:
        sheet_details = inspect_template_excel(excel_file)
        
        # Summary
        print(f"\n{'='*80}")
        print("SUMMARY")
        print("="*80)
        
        total_formulas = sum(len(details['formulas']) for details in sheet_details.values())
        print(f"\nTotal formulas across all sheets: {total_formulas}")
        
        for sheet_name, details in sheet_details.items():
            print(f"\n{sheet_name}:")
            print(f"  Formulas: {len(details['formulas'])}")
            print(f"  Headers: {len(details['headers'])}")
            print(f"  Dimensions: {details['dimensions'][0]} rows x {details['dimensions'][1]} cols")
        
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()

