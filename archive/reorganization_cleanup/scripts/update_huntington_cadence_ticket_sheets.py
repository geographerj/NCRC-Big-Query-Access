"""
Update Huntington + Cadence Merger Ticket with Methods and Assessment Areas Sheets

This script ensures the ticket has:
1. Methods/Notes sheet with complete methodology documentation
2. Assessment Areas sheet properly formatted (ready for manual entry)
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime

def create_methodology_text():
    """Create comprehensive methodology documentation"""
    return """
HUNTINGTON NATIONAL BANK + CADENCE BANK MERGER ANALYSIS - METHODOLOGY

PURPOSE:
This Local Market Analysis compares two banks planning to merge (The Huntington National Bank and Cadence Bank) 
and sets statewide performance goals for their post-merger performance. This analysis evaluates each bank's lending 
patterns relative to their peers and establishes benchmarks for improvement.

BANKS ANALYZED:
- Acquirer: The Huntington National Bank (LEI: 2WHM8VNJH63UN14OL754, RSSD: 12311, SB Respondent ID: 7745)
- Target: Cadence Bank (LEI: Q7C315HKI8VX0SSKBS64, RSSD: 606046, SB Respondent ID: 11813)

DATA SOURCES:
1. HMDA (Home Mortgage Disclosure Act) Data:
   - Table: hdma1-242116.hmda.hmda
   - Years: 2020-2024 (or as specified in merger ticket filters)
   - Includes: Mortgage loan applications and originations

2. Small Business (SB) Data:
   - Table: hdma1-242116.sb.sb
   - Years: 2019-2023 (or as specified in merger ticket filters)
   - Includes: Small business loans and lending activity

3. Branch Data (Summary of Deposits):
   - Table: hdma1-242116.branches.sod25
   - Year: 2025 only
   - Includes: Branch locations, deposits, and census tract characteristics

GEOGRAPHIC SCOPE:
- Assessment Areas: Specific counties within CBSAs/MSAs as defined by each bank
- Key Point: Assessment areas include ONLY specific counties listed (not all counties in a CBSA)
- All data filtered to assessment area counties using GEOID5 codes

LOAN FILTERS (HMDA):
- Action Taken: Originations only (action_taken = '1')
- Occupancy Type: Owner-occupied (occupancy_type = '1')
- Reverse Mortgage: Excluded (reverse_mortgage != '1')
- Construction Method: Site-built (construction_method = '1')
- Number of Units: 1-4 units (total_units IN '1','2','3','4')
- Loan Purpose: Varies by analysis:
  * Home Purchase: loan_purpose = '1'
  * Refinance: loan_purpose IN ('31', '32')
  * Home Equity: loan_purpose IN ('2', '4')

METRICS ANALYZED:

HMDA MORTGAGE METRICS (per State + CBSA):
1. Loans - Total number of originations (count)
2. LMICT% - Low/Moderate Income Census Tract % (tract income ≤80% of MSA median)
3. LMIB% - Low/Moderate Income Borrower % (borrower income ≤80% of MSA median)
4. LMIB$ - LMIB lending in Dollars ($#,###)
5. MMCT% - Majority-Minority Census Tract % (>50% minority population)
6. MINB% - Combined Minority Borrower % (Hispanic OR any non-White race)
7. Asian% - Asian borrower % (non-Hispanic)
8. Black% - Black borrower % (non-Hispanic)
9. Native American% - Native American borrower % (non-Hispanic)
10. HoPI% - Hawaiian or Pacific Islander borrower % (non-Hispanic)
11. Hispanic% - Hispanic borrower % (any ethnicity)

SMALL BUSINESS METRICS (per State + CBSA):
1. SB Loans - Total number of small business loans (count)
2. #LMICT - Count of loans in Low/Moderate Income Census Tracts
3. Avg SB LMICT Loan Amount - Average loan amount for LMICT loans ($#,###)
4. Loans Rev Under $1m - Loans to businesses with revenue under $1 million (count)
5. Avg Loan Amt for RUM SB - Average loan amount for revenue under $1M loans ($#,###)

BRANCH METRICS (per State + CBSA, 2025 only):
1. Total Branches - Count of branches in assessment area counties
2. % in LMICT - Percentage of branches in Low/Moderate Income Census Tracts
3. % in MMCT - Percentage of branches in Majority-Minority Census Tracts

RACE/ETHNICITY CLASSIFICATION METHODOLOGY (NCRC):
- Hierarchical classification: Hispanic ethnicity checked FIRST (regardless of race)
- If any of the 5 ethnicity fields (applicant_ethnicity_1 through applicant_ethnicity_5) contains 
  codes 1, 11, 12, 13, or 14, borrower is classified as Hispanic
- If NOT Hispanic, borrower categorized by race using all 5 race fields (applicant_race_1 through applicant_race_5)
- Race categories (non-Hispanic only):
  * Black: Race = 3
  * Asian: Race = 2 or 21-27 (Asian subgroups)
  * Native American: Race = 1
  * HoPI: Race = 4 or 41-44 (Pacific Islander subgroups)
  * White: Race = 5
  * Unknown: No race or ethnicity data provided

INCOME CALCULATIONS:
- LMIB (Low-to-Moderate Income Borrowers): Borrower income ≤80% of MSA median income
  * Calculated as: (income * 1000) / ffiec_msa_md_median_family_income * 100 ≤ 80%
  * Income field is in thousands, so multiplied by 1000 to get actual income
- LMICT (Low-to-Moderate Income Census Tracts): Tract median income ≤80% of MSA median income
  * Uses tract_to_msa_income_percentage field from HMDA data

PEER SELECTION METHODOLOGY:

HMDA and SMALL BUSINESS PEER MATCHING:
- **Peer Rule: 50% to 200% Volume Rule**
- Peers are identified within each:
  * CBSA (Core Based Statistical Area)
  * Year
  * Loan type/category (for HMDA)
- A lender qualifies as a peer if their origination/loan volume is between 50% and 200% 
  of the subject bank's volume in that CBSA-year combination
- Subject bank is EXCLUDED from peer group calculations
- Peer metrics are aggregated across all qualifying peer lenders
- Separate peer groups calculated for each CBSA-year combination

Example: If Huntington has 100 originations in Chicago in 2024:
- Included as peers: Lenders with 50-200 originations in Chicago 2024
- Excluded: Lenders with <50 or >200 originations, and Huntington itself

BRANCH PEER MATCHING:
- **NO peer rule applied**
- Compare subject bank to ALL other lenders in assessment areas
- Calculate market average across all lenders (excluding subject bank)
- Show gap: Subject % - Market Average %

CALCULATION METHODS:
- Subject Share: (Subject bank metric count / Subject bank total originations) × 100
- Peer Share: (Peer banks metric count / Peer banks total originations) × 100
- Gap: Subject Share - Peer Share (percentage points)
- Positive gap means subject bank outperforms peers
- Negative gap means subject bank underperforms peers

STATISTICAL SIGNIFICANCE:
- Two-proportion z-test performed for relevant metrics
- Tests whether gap is statistically different from zero (p < 0.05)
- Significant gaps indicate reliable patterns (not random variation)

SHEET STRUCTURE:

1. Mortgage Goals Sheet:
   - Combined goals for both banks post-merger
   - State-level aggregation
   - Separate columns for Home Purchase, Refinance, and Home Equity
   - Formulas preserved - DO NOT MODIFY

2. Assessment Areas Sheet:
   - One row per bank + CBSA + County combination
   - Lists all specific counties in each bank's assessment areas
   - Columns: Bank Name, CBSA Name, CBSA Code, State Name, County Name, County Code (GEOID)

3. [Bank]Mortgage Data Sheets (separate for each bank):
   - One row per State + CBSA + Metric combination
   - Shows subject bank values and peer values
   - Format: Counts (numbers), Percentages (##.##%), Dollars ($#,###)

4. [Bank]Small Business Data Sheets (separate for each bank):
   - Same structure as Mortgage Data sheets
   - SB-specific metrics

5. [Bank]Branch Data Sheets (separate for each bank):
   - 2025 data only
   - Comparison to market average (all other lenders)

INTERPRETATION:
- Focus on patterns across years, not single-year gaps
- Statistically significant gaps (marked with *) are more reliable
- Large, persistent negative gaps warrant closer review
- Peer comparisons help identify whether disparities are bank-specific or market-wide
- Goals should reflect peer performance benchmarks

LIMITATIONS:
- Does not account for differences in business models or market strategies
- Geographic selection based on assessment areas (may not cover all bank operations)
- Peer matching based solely on origination volume
- Does not include all factors affecting lending decisions
- Correlation does not imply causation
- Does not assess individual loan applications for discrimination

CONTACT:
NCRC Research Department
National Community Reinvestment Coalition
For questions about methodology or data, contact: research@ncrc.org

Report Generated: {date}
""".format(date=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

def update_ticket_sheets(ticket_file):
    """Update ticket with Methods and Assessment Areas sheets"""
    ticket_path = Path(ticket_file)
    
    if not ticket_path.exists():
        print(f"ERROR: Ticket file not found: {ticket_path}")
        return False
    
    print(f"Loading ticket: {ticket_path}")
    wb = load_workbook(ticket_path)
    
    # Define styles
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    title_font = Font(bold=True, size=14)
    data_font = Font(size=11)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Notes/Methods Sheet
    if "Notes" in wb.sheetnames:
        ws_notes = wb["Notes"]
    else:
        ws_notes = wb.create_sheet("Notes")
    
    # Clear existing content
    ws_notes.delete_rows(1, ws_notes.max_row)
    
    # Add title
    ws_notes.cell(row=1, column=1, value="HUNTINGTON NATIONAL BANK + CADENCE BANK MERGER ANALYSIS")
    ws_notes.cell(row=1, column=1).font = title_font
    
    ws_notes.cell(row=2, column=1, value="METHODOLOGY AND DATA SOURCES")
    ws_notes.cell(row=2, column=1).font = Font(bold=True, size=12)
    
    # Add methodology text
    methodology = create_methodology_text()
    lines = methodology.strip().split('\n')
    
    current_row = 4
    for line in lines:
        if line.strip():
            ws_notes.cell(row=current_row, column=1, value=line.strip())
            ws_notes.cell(row=current_row, column=1).font = data_font
            ws_notes.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
            current_row += 1
        else:
            current_row += 1  # Add blank line
    
    # Set column width
    ws_notes.column_dimensions['A'].width = 120
    
    # Sheet 2: Assessment Areas Sheet
    if "Assessment Areas" in wb.sheetnames:
        ws_aa = wb["Assessment Areas"]
        # Clear existing data (keep headers if they exist)
        if ws_aa.max_row > 1:
            ws_aa.delete_rows(2, ws_aa.max_row)
    else:
        ws_aa = wb.create_sheet("Assessment Areas")
    
    # Headers
    headers = ['Bank Name', 'CBSA Name', 'CBSA Code', 'State Name', 'County Name', 'County Code (GEOID)']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_aa.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    
    # Add example rows with instructions
    example_row = 2
    ws_aa.cell(row=example_row, column=1, value="The Huntington National Bank")
    ws_aa.cell(row=example_row, column=2, value="[CBSA Name - Example: Chicago-Naperville-Elgin]")
    ws_aa.cell(row=example_row, column=3, value="[CBSA Code - Example: 16980]")
    ws_aa.cell(row=example_row, column=4, value="[State Name - Example: Illinois]")
    ws_aa.cell(row=example_row, column=5, value="[County Name - Example: Cook]")
    ws_aa.cell(row=example_row, column=6, value="[GEOID5 - Example: 17031]")
    
    # Format example row (gray, italic)
    example_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    example_font = Font(size=10, italic=True)
    for col in range(1, 7):
        cell = ws_aa.cell(row=example_row, column=col)
        cell.fill = example_fill
        cell.font = example_font
        cell.border = border_thin
    
    # Add instruction note
    note_row = example_row + 2
    ws_aa.cell(row=note_row, column=1, value="INSTRUCTIONS:")
    ws_aa.cell(row=note_row, column=1).font = Font(bold=True)
    note_row += 1
    ws_aa.cell(row=note_row, column=1, value="1. Delete the example row above and replace with actual assessment area data")
    note_row += 1
    ws_aa.cell(row=note_row, column=1, value="2. Add one row per county in each bank's CBSA assessment areas")
    note_row += 1
    ws_aa.cell(row=note_row, column=1, value="3. Use exact bank names: 'The Huntington National Bank' or 'Cadence Bank'")
    note_row += 1
    ws_aa.cell(row=note_row, column=1, value="4. County Code (GEOID) is 5-digit: State Code (2) + County Code (3)")
    note_row += 1
    ws_aa.cell(row=note_row, column=1, value="5. Only include specific counties listed in each bank's assessment areas (not all counties in CBSA)")
    
    # Set column widths
    ws_aa.column_dimensions['A'].width = 35
    ws_aa.column_dimensions['B'].width = 40
    ws_aa.column_dimensions['C'].width = 15
    ws_aa.column_dimensions['D'].width = 20
    ws_aa.column_dimensions['E'].width = 30
    ws_aa.column_dimensions['F'].width = 25
    
    # Save
    wb.save(ticket_path)
    
    print("\n" + "="*80)
    print("TICKET UPDATED")
    print("="*80)
    print(f"\nFile: {ticket_path}")
    print("\nSheets updated/created:")
    print("  1. Notes Sheet - Complete methodology documentation")
    print("  2. Assessment Areas Sheet - Formatted with headers and example row")
    print("\nNext Steps:")
    print("  1. Review the Notes sheet for methodology")
    print("  2. Fill in the Assessment Areas sheet with county data")
    print("  3. Run: python scripts/goal_setting_analysis_main.py \"Huntington+Cadence merger research ticket.xlsx\"")
    print("="*80)
    
    return True

if __name__ == "__main__":
    ticket_file = Path(__file__).parent.parent / "Huntington+Cadence merger research ticket.xlsx"
    update_ticket_sheets(ticket_file)

