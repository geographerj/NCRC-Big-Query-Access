"""Debug Bank Details sheet - check all columns"""

import openpyxl

ticket_file = r"C:\Users\edite\OneDrive - Nat'l Community Reinvestment Coaltn\Desktop\DREAM Analysis\PNC+FirstBank Research Ticket_CBA.xlsx"

wb = openpyxl.load_workbook(ticket_file, data_only=True)
ws = wb["Bank Details"]

print("="*80)
print("FULL BANK DETAILS SHEET INSPECTION")
print("="*80)

print("\nRow 3 (headers) - all columns:")
row3 = []
for col_idx in range(1, 11):
    cell_value = ws.cell(3, col_idx).value
    if cell_value:
        row3.append(f"Col {col_idx}: {cell_value}")
print(f"  {' | '.join(row3)}")

print("\nRow 4 - all columns with data:")
row4 = []
for col_idx in range(1, 11):
    cell_value = ws.cell(4, col_idx).value
    if cell_value:
        row4.append(f"Col {col_idx}: {repr(cell_value)}")
print(f"  {' | '.join(row4)}")

print("\nRow 6 - all columns with data:")
row6 = []
for col_idx in range(1, 11):
    cell_value = ws.cell(6, col_idx).value
    if cell_value:
        row6.append(f"Col {col_idx}: {repr(cell_value)}")
print(f"  {' | '.join(row6)}")

print("\nRow 7 - all columns with data:")
row7 = []
for col_idx in range(1, 11):
    cell_value = ws.cell(7, col_idx).value
    if cell_value:
        row7.append(f"Col {col_idx}: {repr(cell_value)}")
print(f"  {' | '.join(row7)}")

print("\n" + "="*80)
print("Looking for Target bank columns...")
print("="*80)
print("Checking columns 5-10 for row 4-8:")

for row_idx in range(4, 9):
    target_data = []
    for col_idx in range(5, 11):
        cell_value = ws.cell(row_idx, col_idx).value
        if cell_value:
            target_data.append(f"Col {col_idx}: {repr(cell_value)}")
    if target_data:
        print(f"  Row {row_idx}: {' | '.join(target_data)}")

