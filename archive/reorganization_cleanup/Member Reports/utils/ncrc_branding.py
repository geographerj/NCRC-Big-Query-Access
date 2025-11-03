"""
NCRC Branding and Style Guidelines

This module contains NCRC branding standards for report generation including:
- Colors
- Fonts
- Logo usage
- Formatting standards

Based on:
- NCRC Brand Guidelines V19b
- NCRC Style Guide FP
- NCRC Letterhead standards
"""

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Optional, Tuple
import os


# =============================================================================
# NCRC BRAND COLORS
# =============================================================================

# Primary Colors (based on NCRC brand guidelines)
NCRC_BLUE = "003366"  # Primary NCRC blue
NCRC_BLUE_LIGHT = "4A90A4"  # Light blue accent
NCRC_ORANGE = "FF6B35"  # Primary accent color
NCRC_ORANGE_LIGHT = "FFA07A"  # Light orange

# Status Colors (for gap analysis)
COLOR_POSITIVE = "C6EFCE"  # Light green - positive gap (bank outperforms)
COLOR_WARNING_MINOR = "FFEB9C"  # Yellow - minor negative gap (0 to -2.0pp)
COLOR_WARNING_MODERATE = "FFC000"  # Orange - moderate negative gap (-2.0 to -5.0pp)
COLOR_WARNING_LARGE = "FFC7CE"  # Light red - large negative gap (-5.0 to -10.0pp)
COLOR_WARNING_SEVERE = "FF0000"  # Red - severe negative gap (<-10.0pp)

# Neutral Colors
COLOR_HEADER = "D3D3D3"  # Light gray for headers
COLOR_WHITE = "FFFFFF"
COLOR_BLACK = "000000"

# Alternative color schemes (used in some reports)
COLOR_GREEN_ALT = "90EE90"  # Alternative green
COLOR_YELLOW_ALT = "FFFFE0"  # Alternative yellow
COLOR_ORANGE_ALT = "FFA500"  # Alternative orange
COLOR_RED_ALT = "FF6B6B"  # Alternative red


# =============================================================================
# FONTS
# =============================================================================

# Primary Font Family: Helvetica Neue (or Arial/Calibri as fallback)
# Based on brand guidelines showing HelveticaNeueLTStd usage

FONT_FAMILY = "Calibri"  # Excel-compatible sans-serif (similar to Helvetica Neue)
FONT_FAMILY_HEADER = "Calibri"  # For headers
FONT_FAMILY_BODY = "Calibri"  # For body text

# Font Sizes
FONT_SIZE_TITLE = 16
FONT_SIZE_HEADER = 12
FONT_SIZE_SUBHEADER = 11
FONT_SIZE_BODY = 10
FONT_SIZE_FOOTNOTE = 9

# =============================================================================
# NUMBER FORMATTING (Report Writing Guidelines)
# =============================================================================

# Format specifications per report writing guidelines
PERCENTAGE_FORMAT = "0.0"  # One decimal place: 12.5%
INTEGER_FORMAT = "#,##0"   # Comma-separated: 1,234
GAP_FORMAT = "0.0"         # Percentage points, one decimal: -5.2pp

# Font Styles - Initialize as functions to avoid import-time errors
def get_font_bold():
    return Font(bold=True, name=FONT_FAMILY, size=FONT_SIZE_BODY)

def get_font_header():
    return Font(bold=True, name=FONT_FAMILY_HEADER, size=FONT_SIZE_HEADER)

def get_font_title():
    return Font(bold=True, name=FONT_FAMILY_HEADER, size=FONT_SIZE_TITLE)

def get_font_footnote():
    return Font(name=FONT_FAMILY, size=FONT_SIZE_FOOTNOTE)

def get_font_white_bold():
    return Font(bold=True, color="FFFFFF", name=FONT_FAMILY, size=FONT_SIZE_BODY)

# Create instances for backward compatibility
FONT_BOLD = get_font_bold()
FONT_HEADER = get_font_header()
FONT_TITLE = get_font_title()
FONT_FOOTNOTE = get_font_footnote()
FONT_WHITE_BOLD = get_font_white_bold()


# =============================================================================
# FILLS (Background Colors)
# =============================================================================

def get_positive_fill() -> PatternFill:
    """Green fill for positive gaps"""
    return PatternFill(start_color=COLOR_POSITIVE, end_color=COLOR_POSITIVE, fill_type='solid')


def get_warning_minor_fill() -> PatternFill:
    """Yellow fill for minor negative gaps"""
    return PatternFill(start_color=COLOR_WARNING_MINOR, end_color=COLOR_WARNING_MINOR, fill_type='solid')


def get_warning_moderate_fill() -> PatternFill:
    """Orange fill for moderate negative gaps"""
    return PatternFill(start_color=COLOR_WARNING_MODERATE, end_color=COLOR_WARNING_MODERATE, fill_type='solid')


def get_warning_large_fill() -> PatternFill:
    """Light red fill for large negative gaps"""
    return PatternFill(start_color=COLOR_WARNING_LARGE, end_color=COLOR_WARNING_LARGE, fill_type='solid')


def get_warning_severe_fill() -> PatternFill:
    """Red fill for severe negative gaps"""
    return PatternFill(start_color=COLOR_WARNING_SEVERE, end_color=COLOR_WARNING_SEVERE, fill_type='solid')


def get_header_fill() -> PatternFill:
    """Gray fill for header rows"""
    return PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')


def get_gap_fill(gap_value: float, is_significant: bool = False) -> Tuple[PatternFill, Optional[Font]]:
    """
    Get appropriate fill color for gap value
    
    Args:
        gap_value: Gap in percentage points (negative = underperformance)
        is_significant: Whether gap is statistically significant
    
    Returns:
        Tuple of (PatternFill, Font or None)
    """
    if gap_value >= 0:
        return get_positive_fill(), None
    elif gap_value >= -2.0:
        return get_warning_minor_fill(), None
    elif gap_value >= -5.0:
        return get_warning_moderate_fill(), None
    elif gap_value >= -10.0:
        if is_significant:
            return get_warning_severe_fill(), FONT_WHITE_BOLD
        return get_warning_large_fill(), None
    else:
        # Severe negative gap
        return get_warning_severe_fill(), FONT_WHITE_BOLD


# =============================================================================
# BORDERS
# =============================================================================

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

MEDIUM_BORDER = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

HEADER_BORDER = THIN_BORDER


# =============================================================================
# ALIGNMENTS
# =============================================================================

ALIGN_HEADER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')


# =============================================================================
# REPORT HEADER/FOOTER
# =============================================================================

def get_ncrc_header_info() -> Dict[str, str]:
    """
    Get NCRC header information for reports
    
    Returns:
        Dictionary with header fields
    """
    return {
        "organization": "National Community Reinvestment Coalition",
        "organization_short": "NCRC",
        "contact_email": "research@ncrc.org",
        "website": "www.ncrc.org"
    }


def create_report_header(worksheet, report_title: str, report_subtitle: Optional[str] = None, 
                        member_org: Optional[str] = None, report_date: Optional[str] = None):
    """
    Create standardized NCRC report header
    
    Args:
        worksheet: Excel worksheet
        report_title: Main report title
        report_subtitle: Optional subtitle
        member_org: Member organization name
        report_date: Report date (ISO format)
    """
    header_info = get_ncrc_header_info()
    
    # Row 1: NCRC header
    worksheet.merge_cells('A1:E1')
    cell = worksheet['A1']
    cell.value = header_info['organization']
    cell.font = FONT_HEADER
    cell.alignment = ALIGN_CENTER
    
    # Row 2: Report title
    if report_subtitle:
        worksheet.merge_cells('A2:E2')
        cell = worksheet['A2']
        cell.value = f"{report_title}: {report_subtitle}"
    else:
        worksheet.merge_cells('A2:E2')
        cell = worksheet['A2']
        cell.value = report_title
        cell.font = get_font_title()
    cell.alignment = ALIGN_CENTER
    
    # Row 3: Member organization (if provided)
    if member_org:
        worksheet.merge_cells('A3:E3')
        cell = worksheet['A3']
        cell.value = f"Prepared for: {member_org}"
        cell.font = get_font_bold()
        cell.alignment = ALIGN_CENTER
    
    # Row 4: Report date (if provided)
    if report_date:
        worksheet.merge_cells('A4:E4')
        cell = worksheet['A4']
        cell.value = f"Report Date: {report_date}"
        cell.font = get_font_footnote()
        cell.alignment = ALIGN_CENTER
    
    # Row 5: Spacer
    worksheet.append([])
    
    return worksheet.max_row


def create_report_footer(worksheet, row: int, include_contact: bool = True):
    """
    Create standardized NCRC report footer
    
    Args:
        worksheet: Excel worksheet
        row: Starting row for footer
        include_contact: Whether to include contact information
    """
    header_info = get_ncrc_header_info()
    
    # Add spacer
    worksheet.append([])
    footer_row = worksheet.max_row
    
    if include_contact:
        worksheet.merge_cells(f'A{footer_row}:E{footer_row}')
        cell = worksheet[f'A{footer_row}']
        cell.value = f"For questions about this report, contact: {header_info['contact_email']}"
        cell.font = get_font_footnote()
        cell.alignment = ALIGN_CENTER
        cell.font = Font(name=FONT_FAMILY, size=FONT_SIZE_FOOTNOTE, italic=True)
    
    return worksheet.max_row


# =============================================================================
# SHEET FORMATTING FUNCTIONS
# =============================================================================

def format_header_row(worksheet, row: int = 1):
    """
    Format a header row with NCRC standard styling
    
    Args:
        worksheet: Excel worksheet
        row: Row number (default 1)
    """
    header_fill = get_header_fill()
    
    for cell in worksheet[row]:
        if cell.value:  # Only format non-empty cells
            cell.font = get_font_header()
            cell.fill = header_fill
            cell.alignment = ALIGN_HEADER
            cell.border = HEADER_BORDER


def format_table_headers(worksheet, start_row: int, end_col: int):
    """
    Format table headers in a range
    
    Args:
        worksheet: Excel worksheet
        start_row: Starting row number
        end_col: Ending column number (as integer)
    """
    header_fill = get_header_fill()
    
    for col in range(1, end_col + 1):
        cell = worksheet.cell(row=start_row, column=col)
        if cell.value:  # Only format non-empty cells
            cell.font = get_font_header()
            cell.fill = header_fill
            cell.alignment = ALIGN_HEADER
            cell.border = HEADER_BORDER


def apply_gap_coloring(worksheet, gap_column: int, start_row: int, end_row: int,
                      dataframe, gap_column_name: str):
    """
    Apply color coding to gap column based on gap values
    
    Args:
        worksheet: Excel worksheet
        gap_column: Column number for gap values
        start_row: Starting data row (after header)
        end_row: Ending data row
        dataframe: DataFrame with gap values
        gap_column_name: Name of gap column in dataframe
    """
    col_letter = get_column_letter(gap_column)
    
    for row_idx in range(start_row, end_row + 1):
        df_row = row_idx - start_row
        if df_row < len(dataframe):
            row_data = dataframe.iloc[df_row]
            gap_value = row_data.get(gap_column_name)
            
            # Check for significance flag
            significant_key = gap_column_name.replace('_gap', '_significant')
            is_significant = row_data.get(significant_key, False) if significant_key in row_data else False
            
            if gap_value is not None and isinstance(gap_value, (int, float)):
                fill, font = get_gap_fill(gap_value, is_significant)
                cell = worksheet[f'{col_letter}{row_idx}']
                cell.fill = fill
                if font:
                    cell.font = font


# =============================================================================
# LOGO AND BRANDING ASSETS
# =============================================================================

def get_logo_path() -> Optional[str]:
    """
    Get path to NCRC logo file if available
    
    Returns:
        Path to logo file or None
    """
    # Check common locations for logo
    possible_paths = [
        "Member Reports/supporting_files/NCRC_Logo.png",
        "Member Reports/supporting_files/NCRC_Logo.jpg",
        "supporting_files/NCRC_Logo.png",
        "supporting_files/NCRC_Logo.jpg"
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            return path
    
    return None


# =============================================================================
# REPORT METADATA TEMPLATE
# =============================================================================

def get_methodology_template() -> str:
    """
    Get standard NCRC methodology template text
    
    Returns:
        Methodology text template
    """
    return """
NCRC MEMBER REPORT - METHODOLOGY

PURPOSE:
This report analyzes mortgage lending patterns to identify potential fair lending concerns
through borrower demographics analysis, income-based metrics, and geographic (redlining) analysis.

DATA SOURCES:
- HMDA Data: Federal Financial Institutions Examination Council (FFIEC)
- Geographic Data: Census tract demographics
- Years: As specified in report configuration

LOAN FILTERS:
Standard HMDA filters applied:
- Owner Occupied: Yes (occupancy_type = '1')
- Reverse Mortgage: Excluded (reverse_mortgage != '1')
- Construction Method: Site-built (construction_method = '1')
- Number of Units: 1-4 units (total_units IN '1','2','3','4')
- Action Taken: Originations only (action_taken = '1')
- Loan Purpose: As specified in report configuration

RACE/ETHNICITY CLASSIFICATION (NCRC Methodology):
- Hierarchical classification: Hispanic ethnicity checked FIRST (regardless of race)
- If any ethnicity field contains codes 1, 11, 12, 13, or 14, borrower is classified as Hispanic
- If NOT Hispanic, borrower categorized by race using all 5 race fields
- Race categories (non-Hispanic only): Black, Asian, Native American, HoPI, White

PEER COMPARISON:
- Peers defined as lenders with 50%-200% of subject lender's origination volume in each CBSA-year
- Analysis excludes subject lender from peer group
- Each CBSA-year has its own peer group

STATISTICAL SIGNIFICANCE:
- Two-proportion z-test performed for each metric
- Asterisk (*) indicates statistically significant negative gap (p < 0.05)
- Only negative gaps are flagged for significance

COLOR CODING:
- Green: Positive gap (bank outperforms peers)
- Yellow: Minor negative gap (0 to -2.0pp)
- Orange: Moderate negative gap (-2.0 to -5.0pp)
- Light Red: Large negative gap (-5.0 to -10.0pp)
- Red: Severe negative gap (<-10.0pp)
- Red with white bold text: Severe negative gap + statistically significant

REPORT PRODUCTION AND DISCLAIMER:
This report was produced using Cursor, an AI-powered code editor, and other AI language models 
to assist with data analysis, narrative generation, and report formatting. While NCRC staff have 
reviewed this report for accuracy and quality, there may be errors or omissions. If readers have 
questions about the data, methodology, or findings presented in this report, or if they identify 
any potential errors, please contact the NCRC Research Department for review and clarification.

CONTACT:
NCRC Research Department
National Community Reinvestment Coalition
research@ncrc.org
www.ncrc.org
"""

