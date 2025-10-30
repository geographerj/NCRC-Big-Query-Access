"""
NCRC CBA Banks Disparate Impact Analysis - Excel Generator v2.0

Processes BigQuery output from cba_banks_analysis_v2.sql to create Excel workbooks
analyzing lending patterns of Community Benefits Agreement signatory banks.

Analyzes TWO loan purpose categories:
  1. Home Purchase (loan_purpose = '1')
  2. All Purposes (loan_purpose IN '1','2','31','32')

Input: CSV from BigQuery (cba_banks_analysis_v2.sql output)
Output: Excel workbook with:
  - One sheet per bank showing ratios and patterns (years as columns)
  - One raw data sheet per bank showing counts and shares by year
  - Summary sheet with bank performance overview

Author: NCRC Research Department
Date: October 2025
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
import sys
import os
import glob
from datetime import datetime
import argparse
from scipy import stats

warnings.filterwarnings('ignore')

# =============================================================================
# CONFIGURATION
# =============================================================================

# Metrics to analyze - now includes detailed race/ethnicity categories
METRICS = {
    'lmict': 'LMICT%',
    'lmib': 'LMIB%', 
    'lmib_amount': 'LMIB$',
    'mmct': 'MMCT%',
    'hispanic': 'Hispanic%',
    'black': 'Black%',
    'asian': 'Asian%',
    'native_american': 'Native American%',
    'hopi': 'HoPI%',
    'minb': 'MINB%'
}

# Data years expected
EXPECTED_YEARS = ['2018', '2019', '2020', '2021', '2022', '2023', '2024']

# Loan purpose categories (from SQL)
LOAN_PURPOSE_CATEGORIES = ['Home Purchase', 'All Purposes']

# Statistical significance threshold
SIGNIFICANCE_LEVEL = 0.05

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def find_input_file(specified_file=None):
    """Find the most recent BigQuery export CSV file"""
    if specified_file and os.path.exists(specified_file):
        print(f"Using specified input file: {specified_file}")
        return specified_file
    
    # Look for CSV files matching pattern
    csv_files = glob.glob('cba_banks_*.csv')
    if not csv_files:
        csv_files = glob.glob('bquxjob_*.csv')
    
    if not csv_files:
        print("ERROR: No input CSV file found!")
        print("Expected: cba_banks_*.csv or bquxjob_*.csv")
        sys.exit(1)
    
    # Get most recent file
    csv_files.sort(key=os.path.getmtime, reverse=True)
    input_file = csv_files[0]
    print(f"Using most recent file: {input_file}")
    return input_file

def two_proportion_z_test(x1, n1, x2, n2):
    """
    Perform two-proportion z-test
    Returns: (z_statistic, p_value)
    x1, n1: successes and trials for group 1 (subject bank)
    x2, n2: successes and trials for group 2 (peer banks)
    """
    if n1 == 0 or n2 == 0:
        return None, None
    
    p1 = x1 / n1
    p2 = x2 / n2
    
    # Pooled proportion
    p_pool = (x1 + x2) / (n1 + n2)
    
    # Standard error
    se = np.sqrt(p_pool * (1 - p_pool) * (1/n1 + 1/n2))
    
    if se == 0:
        return None, None
    
    # Z-statistic
    z = (p1 - p2) / se
    
    # Two-tailed p-value
    p_value = 2 * (1 - stats.norm.cdf(abs(z)))
    
    return z, p_value

# =============================================================================
# STEP 1: LOAD AND PREPARE DATA
# =============================================================================

def load_data(input_file):
    """Load CSV data from BigQuery export"""
    print(f"\nLoading data from: {input_file}")
    
    df = pd.read_csv(input_file)
    print(f"  Loaded {len(df):,} rows")
    
    # Convert year to string for consistency
    df['activity_year'] = df['activity_year'].astype(str)
    
    # Verify loan_purpose_category exists and has expected values
    if 'loan_purpose_category' not in df.columns:
        print("ERROR: 'loan_purpose_category' column not found in CSV!")
        print("Expected columns from SQL: loan_purpose_category with values 'Home Purchase' or 'All Purposes'")
        sys.exit(1)
    
    # Clean CBSA names
    if 'cbsa_name' in df.columns:
        df['cbsa_name'] = df['cbsa_name'].fillna('Unknown')
    
    print(f"  Banks: {df['lender_name'].nunique()}")
    print(f"  CBSAs: {df['cbsa_code'].nunique()}")
    print(f"  Years: {sorted(df['activity_year'].unique())}")
    print(f"  Loan Purpose Categories: {sorted(df['loan_purpose_category'].unique())}")
    
    return df

# =============================================================================
# STEP 2: CALCULATE SHARES AND GAPS
# =============================================================================

def calculate_metrics(row, metric_base):
    """
    Calculate share and gap for a metric
    Returns: (subject_share, peer_share, gap, is_significant)
    
    For amount metrics (LMIB$), uses dollar amounts
    For percentage metrics, uses counts
    """
    
    # Handle LMIB amount metric differently
    if metric_base == 'lmib_amount':
        subject_total = row.get('subject_total_originations', 0)
        peer_total = row.get('peer_total_originations', 0)
        
        if subject_total == 0 or peer_total == 0:
            return None, None, None, False
        
        subject_amount = row.get('subject_lmib_amount_originations', 0)
        peer_amount = row.get('peer_lmib_amount_originations', 0)
        
        # Calculate as average loan amount to LMIB borrowers
        subject_share = (subject_amount / subject_total) if subject_total > 0 else 0
        peer_share = (peer_amount / peer_total) if peer_total > 0 else 0
        gap = subject_share - peer_share
        
        # For amount metrics, we don't do statistical testing
        return subject_share, peer_share, gap, False
    
    # For percentage metrics, use originations counts
    subject_total = row.get('subject_total_originations', 0)
    peer_total = row.get('peer_total_originations', 0)
    
    if subject_total == 0 or peer_total == 0:
        return None, None, None, False
    
    subject_count = row.get(f'subject_{metric_base}_originations', 0)
    peer_count = row.get(f'peer_{metric_base}_originations', 0)
    
    subject_share = (subject_count / subject_total * 100) if subject_total > 0 else 0
    peer_share = (peer_count / peer_total * 100) if peer_total > 0 else 0
    gap = subject_share - peer_share
    
    # Statistical significance test
    z_stat, p_value = two_proportion_z_test(
        subject_count, subject_total,
        peer_count, peer_total
    )
    
    is_significant = (p_value is not None and 
                     p_value < SIGNIFICANCE_LEVEL and 
                     gap < 0)  # Only flag negative gaps
    
    return subject_share, peer_share, gap, is_significant

def add_calculated_metrics(df):
    """Add calculated shares and gaps for all metrics"""
    print("\nCalculating shares and gaps...")
    
    for metric_base, metric_name in METRICS.items():
        print(f"  Processing {metric_name}...")
        
        # Calculate for each row
        results = df.apply(lambda row: calculate_metrics(row, metric_base), axis=1)
        
        # Unpack results into separate columns
        df[f'{metric_base}_subject_share'] = results.apply(lambda x: x[0])
        df[f'{metric_base}_peer_share'] = results.apply(lambda x: x[1])
        df[f'{metric_base}_gap'] = results.apply(lambda x: x[2])
        df[f'{metric_base}_significant'] = results.apply(lambda x: x[3])
    
    return df

# =============================================================================
# PATTERN CLASSIFICATION (NOT USED IN RATIO-BASED ANALYSIS)
# =============================================================================
# These functions are preserved for reference but not used in current analysis
# Ratio sheets don't classify patterns - they show ratios with color coding

# def classify_pattern(significant_years):
#     """
#     Classify the pattern based on number of statistically significant years
#     Returns pattern classification string
#     """
#     if significant_years >= 6:
#         return 'Severe'
#     elif significant_years == 5:
#         return 'Persistent'
#     elif significant_years >= 3:
#         return 'Frequent'
#     elif significant_years == 2:
#         return 'Moderate'
#     elif significant_years == 1:
#         return 'Minimal'
#     else:
#         return 'None'

# def add_pattern_classification(df):
#     """Add pattern classification for each CBSA/metric combination"""
#     print("\nClassifying patterns...")
#     
#     # Group by bank, CBSA, loan purpose category, and metric
#     pattern_data = []
#     
#     for (lender, cbsa_code, cbsa_name, loan_purpose_cat), group in df.groupby([
#         'lender_name', 'cbsa_code', 'cbsa_name', 'loan_purpose_category'
#     ]):
#         
#         for metric_base, metric_name in METRICS.items():
#             sig_col = f'{metric_base}_significant'
#             
#             if sig_col in group.columns:
#                 significant_years = group[sig_col].sum()
#                 pattern = classify_pattern(significant_years)
#                 
#                 pattern_data.append({
#                     'lender_name': lender,
#                     'cbsa_code': cbsa_code,
#                     'cbsa_name': cbsa_name,
#                     'loan_purpose_category': loan_purpose_cat,
#                     'metric': metric_name,
#                     'significant_years': significant_years,
#                     'pattern': pattern
#                 })
#     
#     pattern_df = pd.DataFrame(pattern_data)
#     return pattern_df

# =============================================================================
# STEP 4: PREPARE RATIO SHEET DATA (Years as Columns)
# =============================================================================

def prepare_ratio_sheet_data(df, bank_name):
    """
    Prepare data for ratio analysis sheet
    Format: Metrics as rows, years as columns, showing RATIOS (subject_share / peer_share)
    NO significance testing on ratio sheets
    Ratio >= 2.0 indicates underperformance (colored red/orange)
    """
    bank_data = df[df['lender_name'] == bank_name].copy()
    
    if len(bank_data) == 0:
        return pd.DataFrame()
    
    result_rows = []
    
    # Group by CBSA and loan purpose category
    for (cbsa_code, cbsa_name, loan_purpose_cat), group in bank_data.groupby([
        'cbsa_code', 'cbsa_name', 'loan_purpose_category'
    ]):
        
        # For each metric, create a row
        for metric_base, metric_name in METRICS.items():
            subject_share_col = f'{metric_base}_subject_share'
            peer_share_col = f'{metric_base}_peer_share'
            
            row = {
                'CBSA_Name': cbsa_name,
                'Loan_Purpose_Category': loan_purpose_cat,
                'Metric': metric_name
            }
            
            # Check if we have data for all years
            has_all_years = True
            year_values = {}
            
            for year in EXPECTED_YEARS:
                year_data = group[group['activity_year'] == year]
                
                if len(year_data) > 0:
                    subject_share = year_data[subject_share_col].iloc[0]
                    peer_share = year_data[peer_share_col].iloc[0]
                    
                    if pd.notna(subject_share) and pd.notna(peer_share):
                        # Calculate ratio: subject / peer
                        # Ratio > 1 means underperformance (subject has lower share)
                        if peer_share > 0:
                            ratio = peer_share / subject_share if subject_share > 0 else 999.0
                            year_values[year] = round(ratio, 2)
                        else:
                            has_all_years = False
                            break
                    else:
                        has_all_years = False
                        break
                else:
                    has_all_years = False
                    break
            
            # Only add row if it has data for all years
            if has_all_years:
                row.update(year_values)
                result_rows.append(row)
    
    result_df = pd.DataFrame(result_rows)
    
    # Sort by CBSA, Loan Purpose Category, Metric
    if len(result_df) > 0:
        result_df = result_df.sort_values([
            'CBSA_Name', 'Loan_Purpose_Category', 'Metric'
        ])
    
    return result_df

# =============================================================================
# STEP 5: PREPARE RAW DATA SHEET (Counts and Shares)
# =============================================================================

def prepare_raw_data_sheet(df, bank_name):
    """
    Prepare raw data sheet with counts and shares by year
    Wide format: Years as column groups with Subject Count, Subject Share, Peer Share, Gap*
    WITH significance testing - asterisk (*) marks significant negative gaps
    """
    bank_data = df[df['lender_name'] == bank_name].copy()
    
    if len(bank_data) == 0:
        return pd.DataFrame()
    
    result_rows = []
    
    # Group by CBSA, loan purpose category, and metric
    for (cbsa_code, cbsa_name, loan_purpose_cat), group in bank_data.groupby([
        'cbsa_code', 'cbsa_name', 'loan_purpose_category'
    ]):
        
        # For each metric, create ONE row with all years as columns
        for metric_base, metric_name in METRICS.items():
            
            row = {
                'CBSA_Name': cbsa_name,
                'Loan_Purpose_Category': loan_purpose_cat,
                'Metric': metric_name
            }
            
            # For each year, add four columns: Count, Subject Share, Peer Share, Gap
            has_all_years = True
            
            for year in EXPECTED_YEARS:
                year_data = group[group['activity_year'] == year]
                
                if len(year_data) > 0:
                    row_data = year_data.iloc[0]
                    
                    # Get base counts
                    if metric_base == 'lmib_amount':
                        # For LMIB$, use origination counts
                        subject_count = row_data.get('subject_total_originations', 0)
                        peer_count = row_data.get('peer_total_originations', 0)
                    else:
                        # For other metrics, use metric-specific origination counts
                        subject_count = row_data.get(f'subject_{metric_base}_originations', 0)
                        peer_count = row_data.get(f'peer_{metric_base}_originations', 0)
                    
                    # Get shares
                    subject_share = row_data.get(f'{metric_base}_subject_share', 0)
                    peer_share = row_data.get(f'{metric_base}_peer_share', 0)
                    gap = row_data.get(f'{metric_base}_gap', 0)
                    is_significant = row_data.get(f'{metric_base}_significant', False)
                    
                    # Format values
                    row[f'{year}_Subject_Count'] = int(subject_count) if pd.notna(subject_count) else 0
                    row[f'{year}_Subject_Share'] = f"{subject_share:.1f}%" if pd.notna(subject_share) else "0.0%"
                    row[f'{year}_Peer_Share'] = f"{peer_share:.1f}%" if pd.notna(peer_share) else "0.0%"
                    
                    # Format gap with asterisk if significant
                    if pd.notna(gap):
                        gap_str = f"{gap:.1f}pp"
                        if is_significant and gap < 0:
                            gap_str += "*"
                        row[f'{year}_Gap'] = gap_str
                    else:
                        row[f'{year}_Gap'] = "0.0pp"
                else:
                    has_all_years = False
                    break
            
            if has_all_years:
                result_rows.append(row)
    
    result_df = pd.DataFrame(result_rows)
    
    # Sort by CBSA, Loan Purpose Category, Metric
    if len(result_df) > 0:
        result_df = result_df.sort_values([
            'CBSA_Name', 'Loan_Purpose_Category', 'Metric'
        ])
    
    return result_df

# =============================================================================
# STEP 6: CREATE SUMMARY SHEET
# =============================================================================

def create_summary_sheet(df):
    """Create summary sheet showing bank performance overview"""
    summary_rows = []
    
    for bank in df['lender_name'].unique():
        bank_data = df[df['lender_name'] == bank]
        
        # Count total CBSAs analyzed
        total_cbsas = bank_data['cbsa_code'].nunique()
        
        # Count CBSAs with bad ratios (>= 2.0)
        # This requires checking all metrics across all years
        cbsas_with_problems = set()
        
        for (cbsa_code, loan_purpose_cat), group in bank_data.groupby(['cbsa_code', 'loan_purpose_category']):
            for metric_base in METRICS.keys():
                subject_share_col = f'{metric_base}_subject_share'
                peer_share_col = f'{metric_base}_peer_share'
                
                for _, row in group.iterrows():
                    subject_share = row.get(subject_share_col, 0)
                    peer_share = row.get(peer_share_col, 0)
                    
                    if pd.notna(subject_share) and pd.notna(peer_share) and subject_share > 0 and peer_share > 0:
                        ratio = peer_share / subject_share
                        if ratio >= 2.0:
                            cbsas_with_problems.add(cbsa_code)
                            break
        
        summary_rows.append({
            'Bank_Name': bank,
            'Total_CBSAs_Analyzed': total_cbsas,
            'CBSAs_With_Bad_Ratios': len(cbsas_with_problems)
        })
    
    summary_df = pd.DataFrame(summary_rows)
    summary_df = summary_df.sort_values('CBSAs_With_Bad_Ratios', ascending=False)
    
    return summary_df

# =============================================================================
# STEP 7: CREATE EXCEL WORKBOOK
# =============================================================================

def create_excel_workbook(df, output_file):
    """Create Excel workbook with multiple sheets per bank"""
    print(f"\nCreating Excel workbook: {output_file}")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # Sheet 1: Summary
        print("  Creating summary sheet...")
        summary_df = create_summary_sheet(df)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format summary sheet
        ws_summary = writer.sheets['Summary']
        format_header_row(ws_summary)
        
        # Sheets for each bank
        banks = sorted(df['lender_name'].unique())
        
        for i, bank in enumerate(banks, 1):
            # Ratio analysis sheet (with ratio coloring)
            ratio_sheet_name = f"{i}_{bank[:20]}_Ratios"
            print(f"  Creating sheet: {ratio_sheet_name}")
            
            ratio_data = prepare_ratio_sheet_data(df, bank)
            if len(ratio_data) > 0:
                ratio_data.to_excel(writer, sheet_name=ratio_sheet_name, index=False)
                ws_ratio = writer.sheets[ratio_sheet_name]
                format_header_row(ws_ratio)
                ws_ratio.auto_filter.ref = ws_ratio.dimensions
                ws_ratio.freeze_panes = 'A2'
                apply_ratio_coloring(ws_ratio, ratio_data)
            
            # Raw data sheet (with significance coloring)
            raw_sheet_name = f"{i}_{bank[:20]}_Data"
            print(f"  Creating sheet: {raw_sheet_name}")
            
            raw_data = prepare_raw_data_sheet(df, bank)
            if len(raw_data) > 0:
                raw_data.to_excel(writer, sheet_name=raw_sheet_name, index=False)
                ws_raw = writer.sheets[raw_sheet_name]
                format_header_row(ws_raw)
                ws_raw.auto_filter.ref = ws_raw.dimensions
                ws_raw.freeze_panes = 'A2'
                apply_data_sheet_coloring(ws_raw, raw_data)
        
        # Methodology sheet
        print("  Creating methodology sheet...")
        create_methodology_sheet(writer)
    
    print(f"\n✓ Excel workbook created: {output_file}")

def format_header_row(worksheet):
    """Apply formatting to header row"""
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def apply_ratio_coloring(worksheet, dataframe):
    """Apply color coding to year columns based on RATIO values"""
    year_cols = [col for col in dataframe.columns if col in EXPECTED_YEARS]
    
    col_indices = {}
    for col in year_cols:
        col_idx = dataframe.columns.get_loc(col) + 1
        col_indices[col] = col_idx
    
    # Define fills for ratios (matching worst lender analysis)
    fill_green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # < 1.0 (good)
    fill_yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # 1.0-1.5
    fill_orange = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')  # 1.5-2.0
    fill_light_red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # 2.0-3.0
    fill_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # >= 3.0
    
    # Apply colors based on ratio thresholds
    for row_idx in range(2, len(dataframe) + 2):
        for col_name, col_idx in col_indices.items():
            cell = worksheet.cell(row=row_idx, column=col_idx)
            
            try:
                value = float(cell.value)
                
                if pd.isna(value):
                    continue
                elif value < 1.0:
                    cell.fill = fill_green
                elif value < 1.5:
                    cell.fill = fill_yellow
                elif value < 2.0:
                    cell.fill = fill_orange
                elif value < 3.0:
                    cell.fill = fill_light_red
                else:
                    cell.fill = fill_red
                    cell.font = Font(color='FFFFFF', bold=True)
            except (ValueError, TypeError):
                continue

def apply_data_sheet_coloring(worksheet, dataframe):
    """Apply red fill to negative significant gaps (marked with *) in data sheets"""
    # Find gap columns (end with "_Gap")
    gap_cols = [col for col in dataframe.columns if col.endswith('_Gap')]
    
    col_indices = {}
    for col in gap_cols:
        col_idx = dataframe.columns.get_loc(col) + 1
        col_indices[col] = col_idx
    
    # Define fill for significant negative gaps
    fill_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    
    # Apply red fill to cells with asterisk (significant negative gaps)
    for row_idx in range(2, len(dataframe) + 2):
        for col_name, col_idx in col_indices.items():
            cell = worksheet.cell(row=row_idx, column=col_idx)
            
            try:
                value_str = str(cell.value)
                # Check if value contains asterisk (significant) and is negative
                if '*' in value_str and '-' in value_str.split('pp')[0]:
                    cell.fill = fill_red
                    cell.font = Font(color='FFFFFF', bold=True)
            except (ValueError, TypeError):
                continue

def create_methodology_sheet(writer):
    """Create methodology documentation sheet"""
    methodology_text = """
NCRC CBA BANKS DISPARATE IMPACT ANALYSIS - METHODOLOGY

DATA SOURCE:
- Home Mortgage Disclosure Act (HMDA) data, years 2018-2024
- Extracted via BigQuery using cba_banks_analysis_v2.sql

BANKS ANALYZED:
Community Benefits Agreement (CBA) signatory banks including:
- Wells Fargo Bank, US Bank, PNC Bank, Truist Bank, Morgan Stanley Private Bank
- Fifth Third Bank, KeyBank, Huntington National Bank, BMO Harris Bank
- Santander Bank, Flagstar Bank, Cadence Bank, Old National Bank
- And 13 additional CBA signatory institutions

GEOGRAPHIC SCOPE:
- Top 10 CBSAs per bank (by application volume in 2024)
- OR CBSAs representing ≥1% of bank's total applications
- Maximum 10 CBSAs per bank

LOAN PURPOSE CATEGORIES:
Two separate analyses conducted:
1. HOME PURCHASE - Home purchase loans only (loan_purpose = '1')
2. ALL PURPOSES - All residential mortgage purposes combined:
   - Home Purchase (loan_purpose = '1')
   - Home Improvement (loan_purpose = '2')
   - Cash-Out Refinancing (loan_purpose = '31')
   - No Cash-Out Refinancing (loan_purpose = '32')

METRICS ANALYZED:
1. LMICT% - Share of loans in Low-to-Moderate Income Census Tracts (tract income ≤80% of MSA median)
2. LMIB% - Share of loans to Low-to-Moderate Income Borrowers (borrower income ≤80% of MSA median)
3. LMIB$ - Dollar volume to LMIB borrowers as share of total originations
4. MMCT% - Share of loans in Majority-Minority Census Tracts (>50% minority population)
5. Hispanic% - Share of loans to Hispanic borrowers (any ethnicity field = 1, 11-14)
6. Black% - Share of loans to Black borrowers (non-Hispanic, race = 3)
7. Asian% - Share of loans to Asian borrowers (non-Hispanic, race = 2 or 21-27)
8. Native American% - Share of loans to Native American borrowers (non-Hispanic, race = 1)
9. HoPI% - Share of loans to Native Hawaiian/Pacific Islander borrowers (non-Hispanic, race = 4 or 41-44)
10. MINB% - Share of loans to Minority Borrowers (Hispanic OR any non-White race)

RACE/ETHNICITY LOGIC:
- Hierarchical classification: Hispanic checked FIRST (regardless of race)
- If NOT Hispanic, borrower categorized by race
- All five ethnicity fields checked for Hispanic codes (1, 11-14)
- All five race fields checked for racial category codes
- Records missing both race and ethnicity excluded from race/ethnicity metrics

PEER BANK DEFINITION:
- Other lenders in same CBSA, year, and loan purpose category
- With application volume between 50% and 200% of subject bank
- Aggregated across all qualifying peer banks
- Separate peer groups for Home Purchase vs. All Purposes

GAP CALCULATION:
- Gap = Subject Bank Share - Peer Banks Share (in percentage points)
- Negative gaps indicate underperformance relative to peers
- Positive gaps indicate overperformance relative to peers

STATISTICAL SIGNIFICANCE:
- Two-proportion z-test (p < 0.05)
- Tests whether gap is statistically different from zero
- Marked with asterisk (*) in ratio sheets
- Only negative gaps flagged as potentially problematic

PATTERN CLASSIFICATION:
Based on number of years with statistically significant negative gaps:
- Severe: 6-7 years of significant negative gaps
- Persistent: 5 years of significant negative gaps
- Frequent: 3-4 years of significant negative gaps
- Moderate: 2 years of significant negative gaps
- Minimal: 1 year of significant negative gaps
- None: 0 years of significant negative gaps

SHEET TYPES:

1. SUMMARY SHEET
   - Overview of all banks analyzed
   - Count of CBSAs with problems
   - Count of severe/persistent/frequent patterns

2. RATIO SHEETS (One per bank)
   - Years as columns (2018-2024)
   - Metrics as rows
   - Shows gaps (subject - peer) in percentage points
   - Asterisk (*) indicates statistical significance
   - Pattern classification in rightmost column
   - Separate rows for Home Purchase and All Purposes
   - Color coding: Green (positive gap), Yellow/Orange/Red (increasingly negative gaps)

3. RAW DATA SHEETS (One per bank)
   - Year-by-year counts and shares
   - Shows actual application/origination numbers
   - Subject bank share, peer bank share, and gap
   - Separate rows for Home Purchase and All Purposes
   - Enables detailed review of underlying data

COLOR CODING (Ratio Sheets):
- Green: Positive gap (bank outperforms peers)
- Yellow: Small negative gap (0 to -2.0pp)
- Orange: Moderate negative gap (-2.0 to -5.0pp)
- Light Red: Large negative gap (-5.0 to -10.0pp)
- Red: Severe negative gap (<-10.0pp)
- Red with white bold text: Severe negative gap + statistically significant

HOW TO READ THE DATA:
- Focus on patterns across years, not single-year gaps
- Statistically significant gaps (marked with *) are more reliable
- Severe/Persistent patterns warrant closer review
- Compare Home Purchase vs. All Purposes to identify product-specific issues
- Consider market context and business model differences
- Use raw data sheets to understand scale and volume

LIMITATIONS:
- Does not account for differences in business models or market strategies
- Geographic selection biased toward high-volume markets
- Peer matching based solely on application volume
- Does not include all factors affecting lending decisions
- Correlation does not imply causation
- Aggregating across loan purposes may mask product-specific patterns

CONTACT:
NCRC Research Department
National Community Reinvestment Coalition
For questions about methodology or data, contact: research@ncrc.org
"""
    
    # Create methodology sheet
    ws = writer.book.create_sheet('Methodology')
    
    # Write methodology text
    for i, line in enumerate(methodology_text.strip().split('\n'), 1):
        ws.cell(row=i, column=1, value=line)
    
    # Format
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    ws.column_dimensions['A'].width = 120

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='NCRC CBA Banks Disparate Impact Analysis v2.0'
    )
    parser.add_argument('--input', type=str, 
                       help='Input CSV file from BigQuery')
    parser.add_argument('--output', type=str, 
                       default='NCRC_CBA_Banks_Analysis.xlsx',
                       help='Output Excel file name')
    
    args = parser.parse_args()
    
    print("="*80)
    print("NCRC CBA BANKS DISPARATE IMPACT ANALYSIS v2.0")
    print("="*80)
    print(f"Working Directory: {os.getcwd()}")
    
    # Load and process data
    input_file = find_input_file(args.input)
    df = load_data(input_file)
    df = add_calculated_metrics(df)
    
    # Create Excel workbook
    create_excel_workbook(df, args.output)
    
    print("\n" + "="*80)
    print("ANALYSIS COMPLETE!")
    print("="*80)
    print(f"\nOutput: {args.output}")
    print(f"\nSheets created:")
    print(f"  - Summary: Overview of all CBA banks")
    print(f"  - Ratio sheets: One per bank showing RATIOS (years as columns)")
    print(f"    * Ratio >= 2.0 indicates underperformance (red/orange)")
    print(f"    * NO significance testing on ratio sheets")
    print(f"  - Data sheets: One per bank showing counts, shares, and gaps")
    print(f"    * WITH significance testing (asterisks on gaps)")
    print(f"    * Red cells for negative significant gaps")
    print(f"  - Methodology: Documentation of analysis approach")
    print(f"\nLoan Purpose Categories:")
    print(f"  - Home Purchase (loan_purpose = '1')")
    print(f"  - All Purposes (loan_purpose IN '1','2','31','32')")
    print(f"\nMetrics analyzed:")
    for metric_name in METRICS.values():
        print(f"  - {metric_name}")

if __name__ == "__main__":
    main()