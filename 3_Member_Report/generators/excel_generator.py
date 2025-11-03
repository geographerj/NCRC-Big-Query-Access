"""
Excel Report Generator for NCRC Member Reports

Generates Excel workbooks with:
- One sheet per table
- Methods sheet (similar to CBA reports)
- NCRC branding and formatting
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Any
import sys
import os

# Import branding utilities
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.ncrc_branding import (
    get_header_fill, format_header_row, get_gap_fill,
    FONT_HEADER, FONT_BOLD, COLOR_HEADER
)
from utils.report_formatting import (
    format_percentage, format_integer, format_gap
)


class ExcelReportGenerator:
    """Generate Excel reports with NCRC branding"""
    
    def __init__(self):
        """Initialize Excel generator"""
        self.header_fill = get_header_fill()
    
    def generate_report(self,
                       output_path: str,
                       report_title: str,
                       report_data: Dict[str, Any],
                       member_org: Optional[str] = None,
                       report_date: Optional[str] = None) -> str:
        """
        Generate complete Excel report
        
        Args:
            output_path: Path to save Excel file
            report_title: Report title
            report_data: Dictionary with tables and data
            member_org: Member organization name
            report_date: Report date
        
        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create sheet for each table
        if 'tables' in report_data:
            for table_name, table_data in report_data['tables'].items():
                self._create_table_sheet(wb, table_name, table_data)
        
        # Create methods sheet
        self._create_methods_sheet(wb, report_data.get('methods', {}), 
                                   report_title, member_org, report_date)
        
        # Save workbook
        wb.save(output_path)
        return output_path
    
    def _create_table_sheet(self, wb: Workbook, sheet_name: str, table_data: pd.DataFrame):
        """
        Create a sheet for a data table
        
        Args:
            wb: Workbook
            sheet_name: Name for the sheet
            table_data: DataFrame with table data
        """
        # Clean sheet name (Excel limit: 31 characters)
        clean_name = sheet_name[:31].replace('/', '_').replace('\\', '_')
        ws = wb.create_sheet(clean_name)
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(table_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Format header row
        format_header_row(ws, row=1)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Apply gap coloring if gap columns exist
        self._apply_gap_coloring(ws, table_data)
    
    def _apply_gap_coloring(self, ws, df: pd.DataFrame):
        """Apply color coding to gap columns"""
        if df is None or len(df) == 0:
            return
        
        # Find gap columns
        gap_columns = [col for col in df.columns if 'gap' in col.lower() or '_Gap' in col]
        
        for col_name in gap_columns:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_idx)
                
                # Apply coloring to each gap value
                for row_idx in range(2, len(df) + 2):
                    cell = ws[f'{col_letter}{row_idx}']
                    gap_value = cell.value
                    
                    if gap_value is not None:
                        try:
                            # Extract numeric value
                            if isinstance(gap_value, str):
                                gap_num = float(gap_value.replace('pp', '').replace('*', '').replace('%', ''))
                            else:
                                gap_num = float(gap_value)
                            
                            # Get fill and font
                            fill, font = get_gap_fill(gap_num, False)  # TODO: Check significance
                            
                            cell.fill = fill
                            if font:
                                cell.font = font
                        except (ValueError, TypeError):
                            continue
    
    def _create_methods_sheet(self, wb: Workbook, methods_data: Dict,
                             report_title: str, member_org: Optional[str],
                             report_date: Optional[str]):
        """
        Create methods sheet similar to CBA reports
        
        Args:
            wb: Workbook
            methods_data: Methods data dictionary
            report_title: Report title
            member_org: Member organization
            report_date: Report date
        """
        ws = wb.create_sheet('Methods')
        
        # Header information
        header_info = {
            "organization": "National Community Reinvestment Coalition",
            "contact_email": "research@ncrc.org",
            "website": "www.ncrc.org"
        }
        
        # Title
        ws.append([f"{report_title.upper()} - METHODOLOGY"])
        ws.append([])
        
        # Report metadata
        if member_org:
            ws.append([f"Prepared for: {member_org}"])
        if report_date:
            ws.append([f"Report Date: {report_date}"])
        ws.append([])
        
        # Methods text
        methods_text = methods_data.get('text', self._get_default_methodology())
        
        # Split into lines and write
        for line in methods_text.split('\n'):
            if line.strip():
                ws.append([line.strip()])
        
        ws.append([])
        ws.append(["CONTACT:"])
        ws.append([header_info['organization']])
        ws.append([header_info['contact_email']])
        ws.append([header_info['website']])
        
        # Format
        ws.column_dimensions['A'].width = 120
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Format title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = self.header_fill
    
    def _get_default_methodology(self) -> str:
        """Get default methodology text"""
        return """PURPOSE:
This report analyzes mortgage lending patterns to identify potential fair lending concerns
through borrower demographics analysis, income-based metrics, and geographic (redlining) analysis.

DATA SOURCES:
- HMDA Data: Federal Financial Institutions Examination Council (FFIEC)
- Geographic Data: Census tract demographics
- Years: As specified in report configuration

LOAN FILTERS (STANDARD HMDA FILTERS):
- Owner Occupied: Yes (occupancy_type = '1')
- Reverse Mortgage: Excluded (reverse_mortgage != '1')
- Construction Method: Site-built (construction_method = '1')
- Number of Units: 1-4 units (total_units IN '1','2','3','4')
- Action Taken: Originations only (action_taken = '1')
- Loan Purpose: As specified in report configuration

RACE/ETHNICITY CLASSIFICATION (NCRC Methodology):
- Hierarchical classification: Hispanic ethnicity checked FIRST (regardless of race)
- If any ethnicity field contains codes 1, 11, 12, 13, or 14, borrower is classified as Hispanic
- If NOT Hispanic, borrower categorized by race using all 5 race fields
- Race categories (non-Hispanic only): Black, Asian, Native American, HoPI, White

PEER COMPARISON:
- Peers defined as lenders with 50%-200% of subject lender's origination volume in each CBSA-year
- Analysis excludes subject lender from peer group
- Each CBSA-year has its own peer group

STATISTICAL SIGNIFICANCE:
- Two-proportion z-test performed for each metric
- Asterisk (*) indicates statistically significant negative gap (p < 0.05)
- Only negative gaps are flagged for significance

COLOR CODING:
- Green: Positive gap (bank outperforms peers)
- Yellow: Minor negative gap (0 to -2.0pp)
- Orange: Moderate negative gap (-2.0 to -5.0pp)
- Light Red: Large negative gap (-5.0 to -10.0pp)
- Red: Severe negative gap (<-10.0pp)
- Red with white bold text: Severe negative gap + statistically significant

INTERPRETATION:
- Focus on patterns across years, not single-year gaps
- Statistically significant gaps (marked with *) are more reliable
- Large, persistent negative gaps warrant closer review

LIMITATIONS:
- Does not account for differences in business models or market strategies
- Geographic selection may be biased toward high-volume markets
- Peer matching based solely on application volume
- Does not include all factors affecting lending decisions
- Correlation does not imply causation"""

